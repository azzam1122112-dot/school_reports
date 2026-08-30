# reports/services_export.py
# -*- coding: utf-8 -*-
"""تصدير بيانات المدرسة الكاملة كملف Excel احترافي (.xlsx).

يشمل: ملخص عام + التقارير + ملفات الإنجاز + الاجتماعات + المبادرات + الوثائق
والمعلمون + الأقسام. يصدّر البيانات الوصفية في Excel، وتضيف حزمة ZIP
الملفات الأصلية وملفات PDF الرسمية حيث توجد.
"""
from __future__ import annotations

import json
from io import BytesIO

from django.utils import timezone

from .gender_labels import school_gender_labels
from .model_parts.approvals import ApprovalState
from .models import (
    Assignment,
    AchievementEvidenceImage,
    AchievementEvidenceReport,
    Department,
    DepartmentMembership,
    Document,
    LeadershipEvidenceImage,
    LabAsset,
    LabAssetHandover,
    LabExperiment,
    Initiative,
    Meeting,
    Notification,
    Plan,
    Report,
    SchoolLeadershipPortfolio,
    SchoolMembership,
    TeacherAchievementFile,
    Ticket,
    TicketImage,
)

# ألوان الهوية
_BRAND_GREEN = "006C35"
_BRAND_BLUE = "0072BC"
_HEADER_FILL = "0F8F6B"
_ZEBRA = "F1F8F4"


def _membership_role_label(role_type: str, labels: dict) -> str:
    """اسم الدور بصيغته المؤنثة/المذكرة حسب نوع المدرسة.

    نقطة واحدة تخدم ورقة المنسوبين في التصدير وفي أرشيف السنة معاً، فلا يظهر
    الدور الواحد باسمين في ملفين يصدران من المدرسة نفسها. والقيمة غير المعروفة
    تعود إلى تسمية Django بدل أن تظهر فارغة.
    """
    mapping = {
        SchoolMembership.RoleType.MANAGER: labels.get("manager"),
        SchoolMembership.RoleType.DEPUTY: labels.get("deputy"),
        SchoolMembership.RoleType.ADMIN_STAFF: labels.get("admin_staff"),
        SchoolMembership.RoleType.TEACHER: labels.get("teacher_indefinite"),
    }
    resolved = mapping.get(role_type)
    if resolved:
        return str(resolved)
    return dict(SchoolMembership.RoleType.choices).get(role_type, role_type or "—")


def _counts(school) -> dict:
    return {
        "reports": Report.objects.filter(school=school).count(),
        "achievements": TeacherAchievementFile.objects.filter(school=school).count(),
        "leadership": SchoolLeadershipPortfolio.objects.filter(school=school).count(),
        "tickets": Ticket.objects.filter(school=school).count(),
        "circulars": Notification.objects.filter(
            school=school, requires_signature=True
        ).count(),
        "notifications": Notification.objects.filter(
            school=school, requires_signature=False
        ).count(),
        "documents": Document.objects.filter(school=school).count(),
        "meetings": Meeting.objects.filter(school=school).count(),
        "plans": Plan.objects.filter(school=school).count(),
        "initiatives": Initiative.objects.filter(school=school).count(),
        "teachers": SchoolMembership.seats_used(school),
        "departments": Department.objects.filter(school=school).count(),
    }


def export_filename(school) -> str:
    code = (getattr(school, "code", "") or "school").strip() or "school"
    stamp = timezone.localtime().strftime("%Y%m%d-%H%M")
    return f"school-data-{code}-{stamp}.xlsx"


def build_school_export_workbook(school):
    """يبني ويُعيد Workbook لبيانات المدرسة الكاملة."""
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Font, PatternFill, Border, Side
    from openpyxl.utils import get_column_letter

    wb = Workbook()
    labels = school_gender_labels(school)
    wb.properties.creator = "منصة توثيق"

    header_font = Font(name="Calibri", bold=True, color="FFFFFF", size=11)
    header_fill = PatternFill("solid", fgColor=_HEADER_FILL)
    title_font = Font(name="Calibri", bold=True, color=_BRAND_GREEN, size=16)
    label_font = Font(name="Calibri", bold=True, color=_BRAND_BLUE, size=11)
    right = Alignment(horizontal="right", vertical="center", wrap_text=True, readingOrder=2)
    center = Alignment(horizontal="center", vertical="center")
    thin = Side(style="thin", color="D6E4DC")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    def _style_sheet_rtl(ws):
        ws.sheet_view.rightToLeft = True

    def _write_table(ws, headers, rows, *, start_row=1):
        # رؤوس الأعمدة
        for col, head in enumerate(headers, start=1):
            cell = ws.cell(row=start_row, column=col, value=head)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = center
            cell.border = border
        # الصفوف
        for r, row in enumerate(rows, start=start_row + 1):
            for col, value in enumerate(row, start=1):
                cell = ws.cell(row=r, column=col, value=value)
                cell.alignment = right
                cell.border = border
                if (r - start_row) % 2 == 0:
                    cell.fill = PatternFill("solid", fgColor=_ZEBRA)
        # عرض الأعمدة (تقدير تلقائي)
        for col in range(1, len(headers) + 1):
            letter = get_column_letter(col)
            max_len = len(str(headers[col - 1]))
            for row in rows:
                try:
                    max_len = max(max_len, len(str(row[col - 1])))
                except Exception:
                    pass
            ws.column_dimensions[letter].width = min(60, max(12, max_len + 4))
        ws.freeze_panes = ws.cell(row=start_row + 1, column=1)
        ws.row_dimensions[start_row].height = 24

    # ---------------- ورقة الملخص ----------------
    ws = wb.active
    ws.title = "ملخص"
    _style_sheet_rtl(ws)
    counts = _counts(school)
    ws.cell(row=1, column=1, value="ملف بيانات المدرسة").font = title_font
    ws.cell(row=2, column=1, value="منصة توثيق").font = label_font

    summary_rows = [
        ("اسم المدرسة", getattr(school, "name", "") or ""),
        ("المعرّف (code)", getattr(school, "code", "") or ""),
        ("المدينة", getattr(school, "city", "") or "—"),
        ("المرحلة", getattr(school, "get_stage_display", lambda: "")() or "—"),
        ("النوع", getattr(school, "get_gender_display", lambda: "")() or "—"),
        ("السنة الدراسية الحالية", getattr(school, "current_academic_year", "") or "—"),
        ("عدد التقارير", counts["reports"]),
        ("عدد ملفات الإنجاز", counts["achievements"]),
        ("عدد ملفات الأداء القيادي", counts["leadership"]),
        ("عدد التذاكر والطلبات", counts["tickets"]),
        ("عدد التعاميم", counts["circulars"]),
        ("عدد الإشعارات", counts["notifications"]),
        (f"عدد {labels['teachers_object']}", counts["teachers"]),
        ("عدد الأقسام", counts["departments"]),
        ("تاريخ التصدير", timezone.localtime().strftime("%Y-%m-%d %H:%M")),
    ]
    for i, (label, value) in enumerate(summary_rows, start=4):
        lc = ws.cell(row=i, column=1, value=label)
        lc.font = label_font
        lc.alignment = right
        vc = ws.cell(row=i, column=2, value=value)
        vc.alignment = right
    ws.column_dimensions["A"].width = 28
    ws.column_dimensions["B"].width = 44

    # ---------------- ورقة التقارير ----------------
    ws_reports = wb.create_sheet("التقارير")
    _style_sheet_rtl(ws_reports)
    report_rows = []
    reports_qs = (
        Report.objects.filter(school=school)
        .select_related("teacher", "category")
        .prefetch_related("evidences")
        .order_by("-report_date", "-id")
    )
    for rep in reports_qs.iterator(chunk_size=200):
        evidence_count = rep.evidences.count()
        images = evidence_count or sum(1 for f in (rep.image1, rep.image2, rep.image3, rep.image4) if getattr(f, "name", ""))
        report_rows.append(
            [
                rep.id,
                rep.title or "",
                getattr(rep.category, "name", "") or "—",
                rep.teacher_display_name,
                rep.report_date.strftime("%Y-%m-%d") if rep.report_date else "",
                rep.academic_year or "—",
                rep.beneficiaries_count if rep.beneficiaries_count is not None else "—",
                images,
                (rep.idea or "").strip(),
            ]
        )
    _write_table(
        ws_reports,
        ["#", "العنوان", "النوع", labels["executor"], "التاريخ", "السنة", labels["beneficiaries"], "عدد الصور", "التفاصيل"],
        report_rows,
    )

    # ---------------- ورقة ملفات الإنجاز ----------------
    ws_ach = wb.create_sheet("ملفات الإنجاز")
    _style_sheet_rtl(ws_ach)
    ach_rows = []
    ach_qs = (
        TeacherAchievementFile.objects.filter(school=school)
        .select_related("teacher")
        .order_by("-academic_year", "teacher__name")
    )
    for ach in ach_qs.iterator(chunk_size=100):
        ach_rows.append(
            [
                ach.id,
                getattr(ach.teacher, "name", "") or "—",
                ach.academic_year or "—",
                ach.get_status_display() if hasattr(ach, "get_status_display") else "",
                ach.created_at.strftime("%Y-%m-%d") if getattr(ach, "created_at", None) else "",
            ]
        )
    _write_table(
        ws_ach,
        ["#", labels["teacher"], "السنة الدراسية", "الحالة", "تاريخ الإنشاء"],
        ach_rows,
    )

    # ---------------- ورقتا القيادة المدرسية ----------------
    ws_leadership = wb.create_sheet("الأداء القيادي")
    _style_sheet_rtl(ws_leadership)
    leadership_rows = []
    leadership_qs = _leadership_portfolios_qs(school)
    for portfolio in leadership_qs:
        sections = list(portfolio.sections.all())
        leadership_rows.append(
            [
                portfolio.id,
                portfolio.manager_name or getattr(portfolio.manager, "name", "") or "—",
                portfolio.academic_year,
                portfolio.get_status_display(),
                sum(1 for section in sections if section.is_completed),
                sum(len(section.evidence_images.all()) for section in sections),
                (portfolio.leadership_vision or "").strip(),
                (portfolio.executive_summary or "").strip(),
                (portfolio.notable_achievements or "").strip(),
                (portfolio.improvement_priorities or "").strip(),
            ]
        )
    _write_table(
        ws_leadership,
        [
            "#",
            labels["leader"],
            "السنة الدراسية",
            "الحالة",
            "المحاور المكتملة",
            "عدد الشواهد",
            "الرؤية القيادية",
            "الملخص التنفيذي",
            "أبرز المنجزات",
            "أولويات التحسين",
        ],
        leadership_rows,
    )

    ws_leadership_evidence = wb.create_sheet("شواهد الأداء القيادي")
    _style_sheet_rtl(ws_leadership_evidence)
    leadership_evidence_rows = []
    for evidence in (
        LeadershipEvidenceImage.objects.filter(section__portfolio__school=school)
        .select_related("section", "section__portfolio")
        .order_by("section__portfolio__academic_year", "section__code", "id")
    ):
        leadership_evidence_rows.append(
            [
                evidence.section.portfolio_id,
                evidence.section.portfolio.academic_year,
                evidence.section.get_code_display(),
                evidence.caption or "",
                getattr(evidence.image, "name", "") or "",
            ]
        )
    _write_table(
        ws_leadership_evidence,
        ["معرف ملف الأداء", "السنة الدراسية", "المحور", "وصف الشاهد", "اسم الملف"],
        leadership_evidence_rows,
    )

    # ---------------- ورقة الطلبات والتذاكر ----------------
    ws_tickets = wb.create_sheet("الطلبات والتذاكر")
    _style_sheet_rtl(ws_tickets)
    ticket_rows = []
    for ticket in _tickets_qs(school).iterator(chunk_size=100):
        ticket_rows.append(
            [
                ticket.id,
                ticket.title or "",
                getattr(ticket.creator, "name", "") or "—",
                getattr(ticket.department, "name", "") or "—",
                ticket.get_status_display(),
                ticket.created_at.strftime("%Y-%m-%d %H:%M"),
                "نعم" if getattr(ticket.attachment, "name", "") else "لا",
                (ticket.body or "").strip(),
            ]
        )
    _write_table(
        ws_tickets,
        ["#", "العنوان", "المرسل", "القسم", "الحالة", "التاريخ", "مرفق", "التفاصيل"],
        ticket_rows,
    )

    # ---------------- ورقة التعاميم والإشعارات ----------------
    ws_notifications = wb.create_sheet("التعاميم والإشعارات")
    _style_sheet_rtl(ws_notifications)
    notification_rows = []
    for notification in _notifications_qs(school).iterator(chunk_size=100):
        notification_rows.append(
            [
                notification.id,
                "تعميم" if notification.requires_signature else "إشعار",
                notification.title or "",
                getattr(notification.created_by, "name", "") or "—",
                notification.created_at.strftime("%Y-%m-%d %H:%M"),
                notification.recipients.count(),
                "نعم" if getattr(notification.attachment, "name", "") else "لا",
                (notification.message or "").strip(),
            ]
        )
    _write_table(
        ws_notifications,
        ["#", "النوع", "العنوان", "المرسل", "التاريخ", "المستلمون", "مرفق", "النص"],
        notification_rows,
    )

    # ---------------- ورقة الوثائق ----------------
    ws_documents = wb.create_sheet("الوثائق")
    _style_sheet_rtl(ws_documents)
    document_rows = []
    for document in _documents_qs(school).iterator(chunk_size=100):
        document_rows.append(
            [
                document.id,
                document.title or "",
                document.get_kind_display(),
                document.academic_year or "—",
                getattr(document.department, "name", "") or "—",
                document.owner_name or getattr(document.owner, "name", "") or "—",
                document.get_approval_state_display(),
                getattr(document.file, "name", "") or "",
                document.created_at.strftime("%Y-%m-%d") if document.created_at else "",
            ]
        )
    _write_table(
        ws_documents,
        ["#", "العنوان", "النوع", "السنة", "القسم", "المالك", "الحالة", "اسم الملف", "تاريخ الرفع"],
        document_rows,
    )

    # ---------------- ورقة الاجتماعات ومحاضرها ----------------
    ws_meetings = wb.create_sheet("الاجتماعات")
    _style_sheet_rtl(ws_meetings)
    meeting_rows = []
    for meeting in _meetings_qs(school).iterator(chunk_size=100):
        minutes = getattr(meeting, "minutes", None)
        attendance = meeting.attendance_summary
        meeting_rows.append(
            [
                meeting.id,
                meeting.title or "",
                getattr(meeting.department, "name", "") or "—",
                meeting.organizer_name or getattr(meeting.organizer, "name", "") or "—",
                meeting.scheduled_at.strftime("%Y-%m-%d %H:%M") if meeting.scheduled_at else "",
                meeting.get_status_display(),
                attendance["total"],
                attendance["present"],
                meeting.agenda_items.count(),
                meeting.decisions.count(),
                minutes.get_approval_state_display() if minutes else "لا يوجد محضر",
            ]
        )
    _write_table(
        ws_meetings,
        [
            "#",
            "العنوان",
            "اللجنة / القسم",
            "المنظم",
            "الموعد",
            "حالة الاجتماع",
            "المدعوون",
            "الحضور",
            "بنود جدول الأعمال",
            "القرارات",
            "حالة المحضر",
        ],
        meeting_rows,
    )

    # ---------------- ورقتا الخطط والمبادرات ----------------
    ws_plans = wb.create_sheet("الخطط")
    _style_sheet_rtl(ws_plans)
    plan_rows = []
    for plan in _plans_qs(school).iterator(chunk_size=100):
        summary = plan.task_summary
        plan_rows.append(
            [
                plan.id,
                plan.title or "",
                plan.academic_year or "—",
                plan.get_stage_display(),
                plan.owner_name or getattr(plan.owner, "name", "") or "—",
                plan.get_approval_state_display(),
                summary["total"],
                summary["done"],
                plan.progress_percent,
                plan.description or "",
            ]
        )
    _write_table(
        ws_plans,
        ["#", "العنوان", "السنة", "مرحلة التنفيذ", "المعد", "حالة الاعتماد", "المهام", "المنجز", "نسبة الإنجاز", "الوصف"],
        plan_rows,
    )

    ws_initiatives = wb.create_sheet("المبادرات")
    _style_sheet_rtl(ws_initiatives)
    initiative_rows = []
    for initiative in _initiatives_qs(school).iterator(chunk_size=100):
        initiative_rows.append(
            [
                initiative.id,
                initiative.title or "",
                initiative.teacher_name or getattr(initiative.teacher, "name", "") or "—",
                getattr(initiative.plan, "title", "") or "—",
                initiative.get_approval_state_display(),
                "نعم" if initiative.is_best_practice else "لا",
                "نعم" if initiative.is_shared else "لا",
                initiative.created_at.strftime("%Y-%m-%d") if initiative.created_at else "",
                initiative.summary or "",
            ]
        )
    _write_table(
        ws_initiatives,
        ["#", "العنوان", "المقدم", "الخطة المرتبطة", "حالة الاعتماد", "ممارسة ناجحة", "مشاركة", "تاريخ الإنشاء", "الفكرة والأثر"],
        initiative_rows,
    )

    # ---------------- ورقة المعلمين ----------------
    ws_teachers = wb.create_sheet(str(labels["teachers"]))
    _style_sheet_rtl(ws_teachers)
    teacher_rows = []
    memberships = (
        SchoolMembership.objects.filter(school=school)
        .select_related("teacher")
        .order_by("role_type", "teacher__name")
    )
    for m in memberships.iterator(chunk_size=100):
        role_label = _membership_role_label(m.role_type, labels)
        job_label = m.get_job_title_display()
        if m.job_title == SchoolMembership.JobTitle.TEACHER:
            job_label = labels["teacher_indefinite"]
        elif m.job_title == SchoolMembership.JobTitle.ADMIN_STAFF:
            job_label = labels["admin_staff"]
        elif m.job_title == SchoolMembership.JobTitle.LAB_TECH:
            job_label = labels["lab_tech"]
        teacher_rows.append(
            [
                getattr(m.teacher, "name", "") or "—",
                getattr(m.teacher, "phone", "") or "—",
                role_label,
                job_label,
                "نشط" if m.is_active else "موقوف",
            ]
        )
    _write_table(
        ws_teachers,
        ["الاسم", "الجوال", "الدور", "المسمى الوظيفي", "الحالة"],
        teacher_rows,
    )

    # ---------------- ورقة الأقسام ----------------
    ws_depts = wb.create_sheet("الأقسام")
    _style_sheet_rtl(ws_depts)
    dept_rows = []
    for dep in Department.objects.filter(school=school).order_by("id").iterator(chunk_size=100):
        members = DepartmentMembership.objects.filter(department=dep).count()
        dept_rows.append(
            [
                dep.name or "—",
                dep.role_label or "—",
                "نشط" if dep.is_active else "موقوف",
                members,
            ]
        )
    _write_table(
        ws_depts,
        ["اسم القسم", "الدور الظاهر", "الحالة", "عدد الأعضاء"],
        dept_rows,
    )

    _append_extended_school_work_sheets(wb, school, academic_year=None)

    return wb


def build_school_export_bytes(school) -> bytes:
    wb = build_school_export_workbook(school)
    buffer = BytesIO()
    wb.save(buffer)
    return buffer.getvalue()


def build_year_archive_index_bytes(
    school,
    academic_year,
    *,
    teacher=None,
    school_wide=True,
) -> bytes:
    """Build the Excel index promised inside every yearly archive package."""
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Font, PatternFill
    from .services_archive import UNCLASSIFIED_YEAR, archive_year_label

    labels = school_gender_labels(school)

    reports = _reports_qs(
        school,
        academic_year=academic_year,
        teacher=teacher,
        school_wide=school_wide,
    )
    achievements = (
        TeacherAchievementFile.objects.none()
        if academic_year == UNCLASSIFIED_YEAR
        else _achievement_files_qs(
            school,
            academic_year=academic_year,
            teacher=teacher,
            school_wide=school_wide,
        )
    )
    leadership_portfolios = (
        _leadership_portfolios_qs(school, academic_year=academic_year)
        if school_wide and academic_year != UNCLASSIFIED_YEAR
        else SchoolLeadershipPortfolio.objects.none()
    )

    workbook = Workbook()
    summary_sheet = workbook.active
    summary_sheet.title = "ملخص النسخة"
    report_sheet = workbook.create_sheet("التقارير")
    achievement_sheet = workbook.create_sheet("ملفات الإنجاز")
    evidence_sheet = workbook.create_sheet("شواهد الإنجاز")
    sheets = [summary_sheet, report_sheet, achievement_sheet, evidence_sheet]
    ticket_sheet = None
    notification_sheet = None
    team_sheet = None
    department_sheet = None
    leadership_sheet = None
    leadership_evidence_sheet = None
    document_sheet = None
    meeting_sheet = None
    plan_sheet = None
    initiative_sheet = None
    if school_wide:
        leadership_sheet = workbook.create_sheet("الأداء القيادي")
        leadership_evidence_sheet = workbook.create_sheet("شواهد الأداء القيادي")
        ticket_sheet = workbook.create_sheet("الطلبات والتذاكر")
        notification_sheet = workbook.create_sheet("التعاميم والإشعارات")
        document_sheet = workbook.create_sheet("الوثائق")
        meeting_sheet = workbook.create_sheet("الاجتماعات والمحاضر")
        plan_sheet = workbook.create_sheet("الخطط")
        initiative_sheet = workbook.create_sheet("المبادرات")
        team_sheet = workbook.create_sheet("فريق المدرسة")
        department_sheet = workbook.create_sheet("الأقسام")
        sheets.extend(
            [
                leadership_sheet,
                leadership_evidence_sheet,
                ticket_sheet,
                notification_sheet,
                document_sheet,
                meeting_sheet,
                plan_sheet,
                initiative_sheet,
                team_sheet,
                department_sheet,
            ]
        )
    for sheet in sheets:
        sheet.sheet_view.rightToLeft = True

    summary_sheet.append(["البيان", "القيمة"])
    summary_rows = [
        ("اسم المدرسة", getattr(school, "name", "") or ""),
        ("رمز المدرسة", getattr(school, "code", "") or ""),
        ("السنة المحددة", archive_year_label(academic_year)),
        ("نطاق النسخة", "أرشيف مستندات وسجلات المدرسة"),
        (
            "ملاحظة السجلات الإدارية",
            "الطلبات والتعاميم والإشعارات تحفظ بحالتها الكاملة لحظة إنشاء النسخة؛ لأنها غير مرتبطة بسنة دراسية.",
        ),
        ("تاريخ إنشاء الفهرس", timezone.localtime().isoformat()),
        ("عدد التقارير", reports.count()),
        ("عدد ملفات الإنجاز", achievements.count()),
        ("عدد ملفات الأداء القيادي", leadership_portfolios.count()),
    ]
    if school_wide:
        summary_rows.extend(
            [
                ("عدد الطلبات والتذاكر", _tickets_qs(school).count()),
                (
                    "عدد التعاميم",
                    Notification.objects.filter(
                        school=school,
                        requires_signature=True,
                    ).count(),
                ),
                (
                    "عدد الإشعارات",
                    Notification.objects.filter(
                        school=school,
                        requires_signature=False,
                    ).count(),
                ),
                (
                    "منها إشعارات آلية",
                    Notification.objects.filter(
                        school=school,
                        requires_signature=False,
                        created_by__isnull=True,
                    ).count(),
                ),
                ("عدد الوثائق", _documents_qs(school, academic_year=academic_year).count()),
                ("عدد الاجتماعات والمحاضر", _meetings_qs(school).count()),
                ("عدد الخطط", _plans_qs(school, academic_year=academic_year).count()),
                ("عدد المبادرات", _initiatives_qs(school, academic_year=academic_year).count()),
            ]
        )
    for label, value in summary_rows:
        summary_sheet.append([label, value])

    report_sheet.append(["المعرف", "العنوان", labels["teacher"], "التاريخ", "التصنيف", "السنة الدراسية"])
    for report in reports.iterator(chunk_size=200):
        report_sheet.append(
            [
                report.pk,
                report.title,
                getattr(report, "teacher_display_name", "")
                or getattr(getattr(report, "teacher", None), "name", ""),
                report.report_date.isoformat() if report.report_date else "",
                getattr(getattr(report, "category", None), "name", ""),
                report.academic_year or "غير مصنف",
            ]
        )

    achievement_sheet.append(["المعرف", labels["teacher"], "السنة الدراسية", "الحالة"])
    for achievement in achievements.iterator(chunk_size=100):
        achievement_sheet.append(
            [
                achievement.pk,
                getattr(achievement, "teacher_name", "")
                or getattr(getattr(achievement, "teacher", None), "name", ""),
                achievement.academic_year,
                achievement.get_status_display(),
            ]
        )

    evidence_sheet.append(
        [
            "معرف ملف الإنجاز",
            labels["teacher"],
            "المحور",
            "نوع الشاهد",
            "معرف السجل",
            "اسم الملف",
            "تاريخ التجميد",
            "بيانات التقرير المجمدة",
        ]
    )
    evidence_images = (
        AchievementEvidenceImage.objects.filter(
            section__file__school=school,
            section__file__academic_year=academic_year,
        )
        .select_related("section", "section__file", "section__file__teacher")
        .order_by("section__file_id", "section__code", "id")
    )
    evidence_reports = (
        AchievementEvidenceReport.objects.filter(
            section__file__school=school,
            section__file__academic_year=academic_year,
        )
        .select_related("section", "section__file", "section__file__teacher", "report")
        .order_by("section__file_id", "section__code", "id")
    )
    if not school_wide and teacher is not None:
        evidence_images = evidence_images.filter(section__file__teacher=teacher)
        evidence_reports = evidence_reports.filter(section__file__teacher=teacher)
    if academic_year == UNCLASSIFIED_YEAR:
        evidence_images = evidence_images.none()
        evidence_reports = evidence_reports.none()
    for evidence in evidence_images.iterator(chunk_size=100):
        achievement_file = evidence.section.file
        evidence_sheet.append(
            [
                achievement_file.id,
                achievement_file.teacher_name or achievement_file.teacher.name,
                evidence.section.title or str(evidence.section.code),
                "صورة شاهد",
                evidence.id,
                getattr(evidence.image, "name", "") or "",
                "",
                "",
            ]
        )
    for evidence in evidence_reports.iterator(chunk_size=100):
        achievement_file = evidence.section.file
        archived_names = [
            getattr(field, "name", "") or ""
            for field in (
                evidence.archived_image1,
                evidence.archived_image2,
                evidence.archived_image3,
                evidence.archived_image4,
            )
            if getattr(field, "name", "")
        ]
        evidence_sheet.append(
            [
                achievement_file.id,
                achievement_file.teacher_name or achievement_file.teacher.name,
                evidence.section.title or str(evidence.section.code),
                "تقرير شاهد مجمد",
                evidence.report_id or evidence.id,
                "، ".join(archived_names),
                evidence.frozen_at.isoformat() if evidence.frozen_at else "",
                json.dumps(evidence.frozen_data or {}, ensure_ascii=False, default=str),
            ]
        )

    if leadership_sheet is not None and leadership_evidence_sheet is not None:
        leadership_sheet.append(
            [
                "المعرف",
                labels["leader"],
                "السنة الدراسية",
                "الحالة",
                "المحاور المكتملة",
                "عدد الشواهد",
                "الرؤية القيادية",
                "الملخص التنفيذي",
                "أبرز المنجزات",
                "أولويات التحسين",
            ]
        )
        leadership_evidence_sheet.append(
            ["معرف ملف الأداء", "السنة الدراسية", "المحور", "وصف الشاهد", "اسم الملف"]
        )
        for portfolio in leadership_portfolios:
            sections = list(portfolio.sections.all())
            leadership_sheet.append(
                [
                    portfolio.id,
                    portfolio.manager_name or getattr(portfolio.manager, "name", "") or "",
                    portfolio.academic_year,
                    portfolio.get_status_display(),
                    sum(1 for section in sections if section.is_completed),
                    sum(len(section.evidence_images.all()) for section in sections),
                    portfolio.leadership_vision or "",
                    portfolio.executive_summary or "",
                    portfolio.notable_achievements or "",
                    portfolio.improvement_priorities or "",
                ]
            )
            for section in sections:
                for evidence in section.evidence_images.all():
                    leadership_evidence_sheet.append(
                        [
                            portfolio.id,
                            portfolio.academic_year,
                            section.get_code_display(),
                            evidence.caption or "",
                            getattr(evidence.image, "name", "") or "",
                        ]
                    )

    if ticket_sheet is not None:
        ticket_sheet.append(
            [
                "المعرف",
                "العنوان",
                "التفاصيل",
                "المرسل",
                "المسؤول الرئيسي",
                "المستلمون",
                "القسم",
                "الحالة",
                "تاريخ الإنشاء",
                "آخر تحديث",
                "نوع الطلب",
                "مرفق",
                "عدد الصور",
                "عدد الملاحظات",
            ]
        )
        for ticket in _tickets_qs(school).iterator(chunk_size=100):
            ticket_sheet.append(
                [
                    ticket.pk,
                    ticket.title,
                    ticket.body or "",
                    getattr(ticket.creator, "name", ""),
                    getattr(ticket.assignee, "name", ""),
                    "، ".join(recipient.name for recipient in ticket.recipients.all()),
                    getattr(ticket.department, "name", ""),
                    ticket.get_status_display(),
                    ticket.created_at.isoformat() if ticket.created_at else "",
                    ticket.updated_at.isoformat() if ticket.updated_at else "",
                    "دعم فني للمنصة" if ticket.is_platform else "طلب مدرسي",
                    "نعم" if getattr(ticket.attachment, "name", "") else "لا",
                    ticket.images.count(),
                    ticket.notes.count(),
                ]
            )

    if notification_sheet is not None:
        notification_sheet.append(
            [
                "المعرف",
                "النوع",
                "العنوان",
                "النص",
                "المرسل",
                "مصدر السجل",
                "تاريخ الإنشاء",
                "مهم",
                "انتهاء العرض",
                "بث عام",
                "آخر موعد للتوقيع",
                "المستلمون",
                "تمت القراءة",
                "تم التوقيع",
                "مرفق",
            ]
        )
        for notification in _notifications_qs(school).iterator(chunk_size=100):
            recipient_rows = list(notification.recipients.all())
            notification_sheet.append(
                [
                    notification.pk,
                    "تعميم" if notification.requires_signature else "إشعار",
                    notification.title,
                    notification.message or "",
                    getattr(notification.created_by, "name", "") or "النظام",
                    "آلي" if notification.created_by_id is None else "مستخدم",
                    notification.created_at.isoformat() if notification.created_at else "",
                    "نعم" if notification.is_important else "لا",
                    notification.expires_at.isoformat() if notification.expires_at else "",
                    "نعم" if notification.is_broadcast else "لا",
                    (
                        notification.signature_deadline_at.isoformat()
                        if notification.signature_deadline_at
                        else ""
                    ),
                    len(recipient_rows),
                    sum(1 for row in recipient_rows if row.is_read),
                    sum(1 for row in recipient_rows if row.is_signed),
                    "نعم" if getattr(notification.attachment, "name", "") else "لا",
                ]
            )

    if document_sheet is not None:
        document_sheet.append(
            ["المعرف", "العنوان", "النوع", "السنة الدراسية", "القسم", "المالك", "الحالة", "اسم الملف"]
        )
        document_qs = _documents_qs(
            school,
            academic_year=None if academic_year == UNCLASSIFIED_YEAR else academic_year,
        )
        if academic_year == UNCLASSIFIED_YEAR:
            from django.db.models import Q

            document_qs = document_qs.filter(Q(academic_year="") | Q(academic_year__isnull=True))
        for document in document_qs.iterator(chunk_size=100):
            document_sheet.append(
                [
                    document.pk,
                    document.title or "",
                    document.get_kind_display(),
                    document.academic_year or "غير مصنف",
                    getattr(document.department, "name", "") or "",
                    document.owner_name or getattr(document.owner, "name", "") or "",
                    document.get_approval_state_display(),
                    getattr(document.file, "name", "") or "",
                ]
            )

    if meeting_sheet is not None:
        meeting_sheet.append(
            [
                "المعرف",
                "العنوان",
                "اللجنة / القسم",
                "المنظم",
                "الموعد",
                "حالة الاجتماع",
                "حالة المحضر",
                "المدعوون",
                "الحضور",
                "بنود جدول الأعمال",
                "القرارات",
            ]
        )
        for meeting in _meetings_qs(school).iterator(chunk_size=100):
            minutes = getattr(meeting, "minutes", None)
            attendance = meeting.attendance_summary
            meeting_sheet.append(
                [
                    meeting.pk,
                    meeting.title or "",
                    getattr(meeting.department, "name", "") or "",
                    meeting.organizer_name or getattr(meeting.organizer, "name", "") or "",
                    meeting.scheduled_at.isoformat() if meeting.scheduled_at else "",
                    meeting.get_status_display(),
                    minutes.get_approval_state_display() if minutes else "لا يوجد محضر",
                    attendance["total"],
                    attendance["present"],
                    meeting.agenda_items.count(),
                    meeting.decisions.count(),
                ]
            )

    if plan_sheet is not None:
        plan_sheet.append(
            ["المعرف", "العنوان", "السنة الدراسية", "مرحلة التنفيذ", "المعد", "حالة الاعتماد", "المهام", "المنجز", "نسبة الإنجاز", "الوصف"]
        )
        for plan in _plans_qs(school, academic_year=academic_year).iterator(chunk_size=100):
            summary = plan.task_summary
            plan_sheet.append(
                [
                    plan.pk,
                    plan.title or "",
                    plan.academic_year or "",
                    plan.get_stage_display(),
                    plan.owner_name or getattr(plan.owner, "name", "") or "",
                    plan.get_approval_state_display(),
                    summary["total"],
                    summary["done"],
                    plan.progress_percent,
                    plan.description or "",
                ]
            )

    if initiative_sheet is not None:
        initiative_sheet.append(
            ["المعرف", "العنوان", "المقدم", "الخطة المرتبطة", "حالة الاعتماد", "ممارسة ناجحة", "مشاركة", "الفكرة والأثر"]
        )
        for initiative in _initiatives_qs(school, academic_year=academic_year).iterator(chunk_size=100):
            initiative_sheet.append(
                [
                    initiative.pk,
                    initiative.title or "",
                    initiative.teacher_name or getattr(initiative.teacher, "name", "") or "",
                    getattr(initiative.plan, "title", "") or "",
                    initiative.get_approval_state_display(),
                    "نعم" if initiative.is_best_practice else "لا",
                    "نعم" if initiative.is_shared else "لا",
                    initiative.summary or "",
                ]
            )

    if team_sheet is not None:
        team_sheet.append(["الاسم", "الجوال", "الدور", "المسمى الوظيفي", "الحالة"])
        memberships = (
            SchoolMembership.objects.filter(school=school)
            .select_related("teacher")
            .order_by("role_type", "teacher__name", "id")
        )
        for membership in memberships.iterator(chunk_size=100):
            role_label = _membership_role_label(membership.role_type, labels)
            job_label = membership.get_job_title_display()
            if membership.job_title == SchoolMembership.JobTitle.TEACHER:
                job_label = labels["teacher_indefinite"]
            elif membership.job_title == SchoolMembership.JobTitle.ADMIN_STAFF:
                job_label = labels["admin_staff"]
            elif membership.job_title == SchoolMembership.JobTitle.LAB_TECH:
                job_label = labels["lab_tech"]
            team_sheet.append(
                [
                    getattr(membership.teacher, "name", "") or "",
                    getattr(membership.teacher, "phone", "") or "",
                    role_label,
                    job_label,
                    "نشط" if membership.is_active else "موقوف",
                ]
            )

    if department_sheet is not None:
        department_sheet.append(["اسم القسم", "الدور الظاهر", "الحالة", "عدد الأعضاء"])
        departments = Department.objects.filter(school=school).order_by("name", "id")
        for department in departments.iterator(chunk_size=100):
            department_sheet.append(
                [
                    department.name or "",
                    department.role_label or "",
                    "نشط" if department.is_active else "موقوف",
                    DepartmentMembership.objects.filter(department=department).count(),
                ]
            )

    for sheet in sheets:
        for cell in sheet[1]:
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill("solid", fgColor=_HEADER_FILL)
            cell.alignment = Alignment(horizontal="center")
        sheet.freeze_panes = "A2"
        for column, width in {
            "A": 16,
            "B": 34,
            "C": 34,
            "D": 28,
            "E": 24,
            "F": 24,
            "G": 20,
            "H": 18,
            "I": 22,
            "J": 22,
            "K": 20,
            "L": 18,
            "M": 18,
            "N": 18,
            "O": 18,
        }.items():
            sheet.column_dimensions[column].width = width

    if school_wide:
        _append_extended_school_work_sheets(
            workbook,
            school,
            academic_year=academic_year,
        )

    buffer = BytesIO()
    workbook.save(buffer)
    return buffer.getvalue()


def export_summary_counts(school) -> dict:
    """تُستخدم في صفحة المعاينة قبل التنزيل."""
    counts = _counts(school)
    counts["files"] = count_export_files(school)
    return counts


# =====================================================================
# تصدير الملفات الفعلية داخل أرشيف ZIP
# =====================================================================
import os
import re
import zipfile
import tempfile


def export_zip_filename(school) -> str:
    code = (getattr(school, "code", "") or "school").strip() or "school"
    stamp = timezone.localtime().strftime("%Y%m%d-%H%M")
    return f"school-files-{code}-{stamp}.zip"


def _safe_segment(text, *, fallback: str = "ملف", maxlen: int = 60) -> str:
    """يُنظّف نصًّا ليصلح كاسم مجلد/ملف داخل الأرشيف (يدعم العربية)."""
    text = (str(text) if text is not None else "").strip()
    text = re.sub(r'[\\/:*?"<>|\r\n\t]+', " ", text)
    text = re.sub(r"\s+", " ", text).strip(" .")
    if not text:
        text = fallback
    return text[:maxlen].strip() or fallback


def _reports_qs(school, *, academic_year=None, teacher=None, school_wide=True):
    """قائمة تقارير المدرسة ضمن النطاق المطلوب (مدرسة/سنة/معلّم)."""
    from .services_archive import UNCLASSIFIED_YEAR

    qs = (
        Report.objects.filter(school=school)
        .select_related("teacher", "category")
        .prefetch_related("evidences")
        .order_by("-report_date", "-id")
    )
    if not school_wide and teacher is not None:
        qs = qs.filter(teacher=teacher)
    if academic_year == UNCLASSIFIED_YEAR:
        from django.db.models import Q

        qs = qs.filter(Q(academic_year="") | Q(academic_year__isnull=True))
    elif academic_year:
        qs = qs.filter(academic_year=academic_year)
    return qs


def _report_folder(rep) -> str:
    """مجلد ثابت داخل الأرشيف لتقرير (تُوضَع فيه صوره وملف PDF الخاص به)."""
    day = rep.report_date.strftime("%Y-%m-%d") if rep.report_date else "بدون-تاريخ"
    return "التقارير/" + _safe_segment(f"{day} - {rep.title} - {rep.id}", fallback=f"تقرير-{rep.id}")


def _achievement_files_qs(school, *, academic_year=None, teacher=None, school_wide=True):
    """قائمة ملفات الإنجاز ضمن النطاق المطلوب (مدرسة/سنة/معلّم)."""
    qs = (
        TeacherAchievementFile.objects.filter(school=school)
        .select_related("teacher")
        .order_by("-academic_year", "teacher__name")
    )
    if not school_wide and teacher is not None:
        qs = qs.filter(teacher=teacher)
    if academic_year:
        qs = qs.filter(academic_year=academic_year)
    return qs


def _achievement_arc_name(ach) -> str:
    """مسار ثابت داخل الأرشيف لملف إنجاز (سواء كان مخزّنًا أو مُولّدًا)."""
    tname = _safe_segment(getattr(ach.teacher, "name", "") or "معلم")
    year = _safe_segment(ach.academic_year or "بدون-سنة")
    return f"ملفات-الإنجاز/{tname} - {year}.pdf"


_LEADERSHIP_ARCHIVE_ROOT = "منصة توثيق · القيادة المدرسية"


def _leadership_portfolios_qs(school, *, academic_year=None):
    qs = (
        SchoolLeadershipPortfolio.objects.filter(school=school)
        .select_related("manager", "school")
        .prefetch_related("sections__evidence_images")
        .order_by("-academic_year", "-id")
    )
    if academic_year:
        qs = qs.filter(academic_year=academic_year)
    return qs


def _leadership_portfolio_folder(portfolio) -> str:
    year = _safe_segment(portfolio.academic_year or "بدون-سنة")
    return f"{_LEADERSHIP_ARCHIVE_ROOT}/{year}"


def _leadership_pdf_arc_name(portfolio) -> str:
    return f"{_leadership_portfolio_folder(portfolio)}/ملف الأداء القيادي.pdf"


def _tickets_qs(school):
    return (
        Ticket.objects.filter(school=school)
        .select_related("creator", "assignee", "department", "school")
        .prefetch_related("recipients", "notes__author", "images")
        .order_by("-created_at", "-id")
    )


def _notifications_qs(school):
    return (
        Notification.objects.filter(school=school)
        .select_related("created_by", "school")
        .prefetch_related("recipients__teacher")
        .order_by("-created_at", "-id")
    )


def _documents_qs(school, *, academic_year=None):
    qs = (
        Document.objects.filter(school=school)
        .select_related("owner", "department")
        .order_by("-created_at", "-id")
    )
    if academic_year:
        qs = qs.filter(academic_year=academic_year)
    return qs


def _meetings_qs(school):
    return (
        Meeting.objects.filter(school=school)
        .select_related("organizer", "department")
        .prefetch_related("agenda_items", "attendees__person", "decisions", "minutes")
        .order_by("-scheduled_at", "-id")
    )


def _plans_qs(school, *, academic_year=None):
    qs = (
        Plan.objects.filter(school=school)
        .select_related("owner")
        .prefetch_related("tasks__assignment__targets")
        .order_by("-created_at", "-id")
    )
    if academic_year:
        qs = qs.filter(academic_year=academic_year)
    return qs


def _initiatives_qs(school, *, academic_year=None):
    """All initiatives, including drafts, so the historical archive loses nothing."""
    qs = (
        Initiative.objects.filter(school=school)
        .select_related("teacher", "plan")
        .order_by("-created_at", "-id")
    )
    if academic_year:
        from django.db.models import Q

        qs = qs.filter(
            Q(plan__academic_year=academic_year) | Q(plan__isnull=True),
        )
    return qs


def _assignments_qs(school):
    """All assignments that belong to, or were issued to, this school.

    Group assignments have no ``school`` on the parent, so the target school is
    part of the archive boundary.  ``distinct`` prevents one parent appearing
    once per target.
    """
    from django.db.models import Q

    return (
        Assignment.objects.filter(Q(school=school) | Q(targets__school=school))
        .select_related("issuer", "department", "school", "group", "cancelled_by")
        .prefetch_related(
            "targets__assignee",
            "targets__school",
            "targets__evidence__uploaded_by",
        )
        .distinct()
        .order_by("-created_at", "-id")
    )


def _lab_assets_qs(school):
    return (
        LabAsset.objects.filter(school=school)
        .select_related("department", "custodian", "recorded_by")
        .prefetch_related("handovers")
        .order_by("lab_kind", "name", "id")
    )


def _lab_handovers_qs(school):
    return (
        LabAssetHandover.objects.filter(school=school)
        .select_related("asset", "person", "recorded_by")
        .order_by("-happened_at", "-id")
    )


def _lab_experiments_qs(school):
    return (
        LabExperiment.objects.filter(school=school)
        .select_related("department", "recorder", "requested_by", "report")
        .prefetch_related("assets")
        .order_by("-experiment_date", "-id")
    )


def school_archive_source_counts(school, *, academic_year) -> dict[str, int]:
    """Counts every live domain that makes a yearly snapshot non-empty."""

    return {
        "reports": _reports_qs(school, academic_year=academic_year).count(),
        "achievements": _achievement_files_qs(
            school, academic_year=academic_year
        ).count(),
        "leadership": _leadership_portfolios_qs(
            school, academic_year=academic_year
        ).count(),
        "tickets": _tickets_qs(school).count(),
        "notifications": _notifications_qs(school).count(),
        "documents": _documents_qs(school, academic_year=academic_year).count(),
        "meetings": _meetings_qs(school).count(),
        "plans": _plans_qs(school, academic_year=academic_year).count(),
        "initiatives": _initiatives_qs(
            school, academic_year=academic_year
        ).count(),
        "assignments": _assignments_qs(school).count(),
        "lab_assets": _lab_assets_qs(school).count(),
        "lab_handovers": _lab_handovers_qs(school).count(),
        "lab_experiments": _lab_experiments_qs(school).count(),
    }


def _append_extended_school_work_sheets(workbook, school, *, academic_year=None) -> None:
    """Add detailed, searchable sheets for execution and laboratory history."""
    from openpyxl.styles import Alignment, Font, PatternFill

    def add_sheet(name: str, headers: list[str], rows) -> None:
        sheet = workbook.create_sheet(name)
        sheet.sheet_view.rightToLeft = True
        sheet.append(headers)
        for row in rows:
            sheet.append(list(row))
        for cell in sheet[1]:
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill("solid", fgColor=_HEADER_FILL)
            cell.alignment = Alignment(horizontal="center", vertical="center")
        sheet.freeze_panes = "A2"
        sheet.auto_filter.ref = sheet.dimensions
        for column in sheet.columns:
            letter = column[0].column_letter
            values = [str(cell.value or "") for cell in column[:200]]
            sheet.column_dimensions[letter].width = min(
                44,
                max(14, max((len(value) for value in values), default=12) + 2),
            )

    assignments = list(_assignments_qs(school))
    school_targets = []
    evidence_rows = []
    target_counts: dict[int, int] = {}
    completed_target_counts: dict[int, int] = {}
    for assignment in assignments:
        targets = [
            target
            for target in assignment.targets.all()
            if target.school_id == school.pk
            or (target.school_id is None and assignment.school_id == school.pk)
        ]
        target_counts[assignment.id] = len(targets)
        completed_target_counts[assignment.id] = sum(
            1 for target in targets if target.approval_state == ApprovalState.APPROVED
        )
        for target in targets:
            school_targets.append((assignment, target))
            for evidence in target.evidence.all():
                evidence_rows.append((assignment, target, evidence))

    add_sheet(
        "التكليفات",
        [
            "المعرف", "العنوان", "المكلِّف", "المصدر", "الأولوية",
            "القسم", "موعد التسليم", "يتطلب شواهد", "الحد الأدنى",
            "المكلَّفون في المدرسة", "نسبة الإنجاز", "ملغى", "سبب الإلغاء",
            "المطلوب", "تاريخ الإصدار",
        ],
        (
            (
                assignment.id,
                assignment.title,
                assignment.issuer_name or getattr(assignment.issuer, "name", ""),
                assignment.get_source_display(),
                assignment.get_priority_display(),
                getattr(assignment.department, "name", "") or "—",
                assignment.due_at.isoformat() if assignment.due_at else "",
                "نعم" if assignment.requires_evidence else "لا",
                assignment.min_evidence_count,
                target_counts[assignment.id],
                (
                    round(completed_target_counts[assignment.id] * 100 / target_counts[assignment.id])
                    if target_counts[assignment.id]
                    else 0
                ),
                "نعم" if assignment.is_cancelled else "لا",
                assignment.cancel_reason or "",
                assignment.description or "",
                assignment.created_at.isoformat() if assignment.created_at else "",
            )
            for assignment in assignments
        ),
    )
    add_sheet(
        "تنفيذ التكليفات",
        [
            "معرف التكليف", "التكليف", "المكلَّف", "المدرسة", "الحالة",
            "نسبة الإنجاز", "متأخر", "قُبل في", "طلب التوضيح",
            "ملاحظات التنفيذ", "عدد الشواهد", "تاريخ الإسناد",
        ],
        (
            (
                assignment.id,
                assignment.title,
                getattr(target.assignee, "name", "") or "—",
                getattr(target.school, "name", "") or getattr(school, "name", ""),
                target.get_approval_state_display(),
                target.progress_percent,
                "نعم" if target.is_overdue else "لا",
                target.accepted_at.isoformat() if target.accepted_at else "",
                target.clarification_note or "",
                target.progress_note or "",
                target.evidence_count,
                target.created_at.isoformat() if target.created_at else "",
            )
            for assignment, target in school_targets
        ),
    )
    add_sheet(
        "شواهد التكليفات",
        [
            "معرف التكليف", "التكليف", "المكلَّف", "الوصف", "اسم الملف",
            "الحجم بالبايت", "رفعه", "تاريخ الرفع",
        ],
        (
            (
                assignment.id,
                assignment.title,
                getattr(target.assignee, "name", "") or "—",
                evidence.note or "",
                getattr(evidence.file, "name", "") or "",
                evidence.storage_bytes,
                getattr(evidence.uploaded_by, "name", "") or "—",
                evidence.created_at.isoformat() if evidence.created_at else "",
            )
            for assignment, target, evidence in evidence_rows
        ),
    )

    plans = list(_plans_qs(school, academic_year=academic_year))
    add_sheet(
        "أهداف الخطط",
        ["معرف الخطة", "الخطة", "الترتيب", "الهدف", "مؤشر القياس", "المستهدف", "نسبة الإنجاز"],
        (
            (
                plan.id, plan.title, goal.order, goal.title, goal.indicator,
                goal.target, goal.progress_percent,
            )
            for plan in plans
            for goal in plan.goals.all()
        ),
    )
    add_sheet(
        "مهام الخطط",
        [
            "معرف الخطة", "الخطة", "الترتيب", "المهمة", "التفصيل", "الهدف",
            "المسؤول", "القسم", "الموعد", "الحالة", "معرف التكليف",
        ],
        (
            (
                plan.id,
                plan.title,
                task.order,
                task.title,
                task.description or "",
                getattr(task.goal, "title", "") or "—",
                getattr(task.responsible, "name", "") or "—",
                getattr(task.department, "name", "") or "—",
                task.due_at.isoformat() if task.due_at else "",
                task.state,
                task.assignment_id or "",
            )
            for plan in plans
            for task in plan.tasks.all()
        ),
    )

    assets = list(_lab_assets_qs(school))
    add_sheet(
        "أصول المختبر",
        [
            "المعرف", "الصنف", "رقم العهدة", "المختبر", "القسم", "النوع",
            "الكمية", "الوحدة", "الخارج", "المتاح", "الحالة", "الموقع",
            "مسؤول العهدة", "نشط", "الملاحظات", "تاريخ الإضافة", "آخر تحديث",
        ],
        (
            (
                asset.id, asset.name, asset.code, asset.get_lab_kind_display() if asset.lab_kind else "—",
                getattr(asset.department, "name", "") or "—", asset.get_category_display(),
                asset.quantity, asset.unit, asset.out_quantity, asset.available_quantity,
                asset.get_condition_display(), asset.location,
                getattr(asset.custodian, "name", "") or "—",
                "نعم" if asset.is_active else "لا", asset.notes,
                asset.created_at.isoformat() if asset.created_at else "",
                asset.updated_at.isoformat() if asset.updated_at else "",
            )
            for asset in assets
        ),
    )
    handovers = list(_lab_handovers_qs(school))
    add_sheet(
        "حركات العهدة",
        ["المعرف", "الصنف", "الحركة", "المستلم", "الكمية", "تاريخ الحركة", "سجلها", "الملاحظة"],
        (
            (
                row.id, getattr(row.asset, "name", "") or "—", row.get_direction_display(),
                row.person_name or getattr(row.person, "name", "") or "—", row.quantity,
                row.happened_at.isoformat() if row.happened_at else "",
                getattr(row.recorded_by, "name", "") or "—", row.note,
            )
            for row in handovers
        ),
    )
    experiments = list(_lab_experiments_qs(school))
    add_sheet(
        "تجارب المختبر",
        [
            "المعرف", "العنوان", "المختبر", "القسم", "المحضر", "المعلم الطالب",
            "تاريخ التنفيذ", "المادة", "الصف", "عدد الطلاب", "الأهداف", "الخطوات",
            "المواد والأدوات", "إجراءات السلامة", "العهد المستخدمة", "التقرير المرتبط",
            "حالة الاعتماد", "تاريخ الإنشاء", "آخر تحديث",
        ],
        (
            (
                experiment.id, experiment.title, experiment.get_lab_kind_display() if experiment.lab_kind else "—",
                getattr(experiment.department, "name", "") or "—",
                getattr(experiment.recorder, "name", "") or "—",
                getattr(experiment.requested_by, "name", "") or "—",
                experiment.experiment_date.isoformat() if experiment.experiment_date else "",
                experiment.subject, experiment.class_name, experiment.students_count,
                experiment.objectives, experiment.procedure, experiment.materials_note,
                experiment.safety_notes, "، ".join(asset.name for asset in experiment.assets.all()),
                experiment.report_id or "", experiment.get_approval_state_display(),
                experiment.created_at.isoformat() if experiment.created_at else "",
                experiment.updated_at.isoformat() if experiment.updated_at else "",
            )
            for experiment in experiments
        ),
    )


def _ticket_folder(ticket) -> str:
    day = ticket.created_at.strftime("%Y-%m-%d") if ticket.created_at else "بدون-تاريخ"
    name = _safe_segment(
        f"{day} - {ticket.title} - {ticket.id}",
        fallback=f"طلب-{ticket.id}",
    )
    return f"الطلبات-والتذاكر/{name}"


def _notification_folder(notification) -> str:
    root = "التعاميم" if notification.requires_signature else "الإشعارات"
    day = (
        notification.created_at.strftime("%Y-%m-%d")
        if notification.created_at
        else "بدون-تاريخ"
    )
    title = notification.title or ((notification.message or "")[:40])
    name = _safe_segment(
        f"{day} - {title} - {notification.id}",
        fallback=f"سجل-{notification.id}",
    )
    return f"{root}/{name}"


def _administrative_file_fields(school):
    """Original attachments/images for school-wide administrative records."""
    for ticket in _tickets_qs(school).iterator(chunk_size=100):
        folder = _ticket_folder(ticket)
        field = getattr(ticket, "attachment", None)
        name = getattr(field, "name", "") or ""
        if name:
            ext = os.path.splitext(name)[1].lower() or ".bin"
            yield f"{folder}/المرفق-الرئيسي{ext}", field
        for index, ticket_image in enumerate(ticket.images.all(), start=1):
            image = getattr(ticket_image, "image", None)
            image_name = getattr(image, "name", "") or ""
            if image_name:
                ext = os.path.splitext(image_name)[1].lower() or ".jpg"
                yield f"{folder}/صورة-{index}{ext}", image

    for notification in _notifications_qs(school).iterator(chunk_size=100):
        field = getattr(notification, "attachment", None)
        name = getattr(field, "name", "") or ""
        if not name:
            continue
        ext = os.path.splitext(name)[1].lower() or ".bin"
        yield f"{_notification_folder(notification)}/المرفق{ext}", field


def _file_fields_for_school(school, *, academic_year=None, teacher=None, school_wide=True):
    """مولّد ينتج (مسار_داخل_الأرشيف، حقل_الملف) لكل ملف فعلي للمدرسة.

    - ``academic_year``: قصر النطاق على سنة محددة (أو السنتينل UNCLASSIFIED لغير المصنّفة).
    - ``teacher`` + ``school_wide=False``: قصر النطاق على ملفات معلّم واحد.
    """
    from .models import AchievementEvidenceImage, AchievementEvidenceReport
    from .services_archive import UNCLASSIFIED_YEAR

    scope_teacher = teacher if not school_wide else None

    # 1) صور التقارير (ملف PDF لكل تقرير يُولَّد لاحقًا في باني الـ ZIP)
    for rep in _reports_qs(
        school, academic_year=academic_year, teacher=scope_teacher, school_wide=school_wide
    ).iterator(chunk_size=200):
        folder = _report_folder(rep)
        evidences = list(rep.evidences.all())
        if evidences:
            for idx, evidence in enumerate(evidences, start=1):
                field = evidence.image
                name = getattr(field, "name", "") or ""
                if not name:
                    continue
                ext = os.path.splitext(name)[1].lower() or ".webp"
                yield f"{folder}/شاهد-{idx}{ext}", field
            continue
        for idx, field in enumerate((rep.image1, rep.image2, rep.image3, rep.image4), start=1):
            name = getattr(field, "name", "") or ""
            if not name:
                continue
            ext = os.path.splitext(name)[1].lower() or ".jpg"
            yield f"{folder}/صورة-{idx}{ext}", field

    # ملفات الإنجاز والشواهد لا تنطبق على "غير المصنّفة بسنة"
    if academic_year == UNCLASSIFIED_YEAR:
        return

    # 2) ملفات الإنجاز (PDF المخزّن فقط؛ غير المخزّن يُولَّد لاحقًا في باني الـ ZIP)
    achievements = _achievement_files_qs(
        school, academic_year=academic_year, teacher=scope_teacher, school_wide=school_wide
    )
    for ach in achievements.iterator(chunk_size=100):
        field = getattr(ach, "pdf_file", None)
        if not getattr(field, "name", ""):
            continue
        yield _achievement_arc_name(ach), field

    # 3) شواهد ملفات الإنجاز (صور)
    evidence = (
        AchievementEvidenceImage.objects
        .filter(section__file__school=school)
        .select_related("section", "section__file", "section__file__teacher")
    )
    if scope_teacher is not None:
        evidence = evidence.filter(section__file__teacher=scope_teacher)
    if academic_year:
        evidence = evidence.filter(section__file__academic_year=academic_year)
    for ev in evidence.iterator(chunk_size=100):
        field = getattr(ev, "image", None)
        name = getattr(field, "name", "") or ""
        if not name:
            continue
        try:
            sec = ev.section
            ach_file = sec.file
            tname = _safe_segment(getattr(ach_file.teacher, "name", "") or "معلم")
            year = _safe_segment(getattr(ach_file, "academic_year", "") or "بدون-سنة")
            section = _safe_segment(getattr(sec, "title", "") or getattr(sec, "code", "") or "قسم")
        except Exception:
            tname, year, section = "معلم", "بدون-سنة", "قسم"
        base = os.path.basename(name) or f"شاهد-{ev.id}.jpg"
        yield f"ملفات-الإنجاز/شواهد/{tname} - {year}/{section}/{base}", field

    # 4) الصور المجمدة لتقارير الشواهد. هذه نسخ مستقلة عن صور التقرير
    # الأصلي ويجب حفظها حتى يبقى ملف الإنجاز صالحًا بعد حذف التقرير الحي.
    evidence_reports = (
        AchievementEvidenceReport.objects.filter(section__file__school=school)
        .select_related("section", "section__file", "section__file__teacher", "report")
    )
    if scope_teacher is not None:
        evidence_reports = evidence_reports.filter(section__file__teacher=scope_teacher)
    if academic_year:
        evidence_reports = evidence_reports.filter(section__file__academic_year=academic_year)
    for evidence_report in evidence_reports.iterator(chunk_size=100):
        section_obj = evidence_report.section
        achievement_file = section_obj.file
        teacher_name = _safe_segment(
            getattr(achievement_file.teacher, "name", "") or "معلم"
        )
        year = _safe_segment(achievement_file.academic_year or "بدون-سنة")
        section_name = _safe_segment(
            section_obj.title or str(section_obj.code) or "قسم"
        )
        report_id = evidence_report.report_id or evidence_report.id
        report_folder = f"تقرير-شاهد-{report_id}"
        for index, field in enumerate(
            (
                evidence_report.archived_image1,
                evidence_report.archived_image2,
                evidence_report.archived_image3,
                evidence_report.archived_image4,
            ),
            start=1,
        ):
            name = getattr(field, "name", "") or ""
            if not name:
                continue
            extension = os.path.splitext(name)[1].lower() or ".jpg"
            yield (
                f"ملفات-الإنجاز/شواهد/{teacher_name} - {year}/"
                f"{section_name}/{report_folder}/صورة-مؤرشفة-{index}{extension}",
                field,
            )

    # 5) شواهد ملف الأداء القيادي. يضاف الملف القيادي للنسخة المدرسية فقط،
    # ويرشح بحسب السنة في الأرشيف السنوي، بينما يشمل كل السنوات في التنزيل الكامل.
    if school_wide:
        leadership_evidence = (
            LeadershipEvidenceImage.objects.filter(section__portfolio__school=school)
            .select_related("section", "section__portfolio")
            .order_by("section__portfolio__academic_year", "section__code", "id")
        )
        if academic_year:
            leadership_evidence = leadership_evidence.filter(
                section__portfolio__academic_year=academic_year
            )
        for evidence in leadership_evidence.iterator(chunk_size=100):
            field = getattr(evidence, "image", None)
            name = getattr(field, "name", "") or ""
            if not name:
                continue
            portfolio = evidence.section.portfolio
            section_name = _safe_segment(evidence.section.get_code_display(), fallback="محور")
            caption = _safe_segment(evidence.caption, fallback=f"شاهد-{evidence.id}")
            extension = os.path.splitext(name)[1].lower() or ".jpg"
            yield (
                f"{_leadership_portfolio_folder(portfolio)}/شواهد/"
                f"{section_name}/{caption}-{evidence.id}{extension}",
                field,
            )

    # 6) الوثائق المدرسية تحفظ بصيغتها الأصلية داخل الحزمة.
    if school_wide:
        documents = _documents_qs(
            school,
            academic_year=None if academic_year == UNCLASSIFIED_YEAR else academic_year,
        )
        if academic_year == UNCLASSIFIED_YEAR:
            from django.db.models import Q

            documents = documents.filter(Q(academic_year="") | Q(academic_year__isnull=True))
        for document in documents.iterator(chunk_size=100):
            field = getattr(document, "file", None)
            name = getattr(field, "name", "") or ""
            if not name:
                continue
            year = _safe_segment(document.academic_year or "بدون-سنة")
            kind = _safe_segment(document.get_kind_display(), fallback="وثيقة")
            title = _safe_segment(document.title, fallback=f"وثيقة-{document.id}")
            extension = os.path.splitext(name)[1].lower() or ".bin"
            yield f"الوثائق/{year}/{kind}/{title}-{document.id}{extension}", field

    # 7) شواهد تنفيذ التكليفات. التكليفات الإدارية لا تحمل سنة دراسية، لذلك
    # تحفظ النسخة السنوية حالتها الكاملة وقت الإنشاء، مع قصر تكليفات المجموعة
    # على المكلَّفين داخل المدرسة الحالية منعاً لتسريب مدارس شقيقة.
    if school_wide:
        from django.db.models import Q

        for assignment in _assignments_qs(school):
            assignment_folder = _safe_segment(
                f"{assignment.title} - {assignment.id}",
                fallback=f"تكليف-{assignment.id}",
            )
            targets = assignment.targets.filter(
                Q(school=school)
                | Q(school__isnull=True, assignment__school=school)
            ).select_related("assignee")
            for target in targets:
                assignee = _safe_segment(
                    getattr(target.assignee, "name", "") or f"مكلف-{target.id}"
                )
                for index, evidence in enumerate(target.evidence.all(), start=1):
                    field = getattr(evidence, "file", None)
                    name = getattr(field, "name", "") or ""
                    if not name:
                        continue
                    extension = os.path.splitext(name)[1].lower() or ".bin"
                    note = _safe_segment(
                        evidence.note,
                        fallback=f"شاهد-{index}",
                    )
                    yield (
                        f"التكليفات/{assignment_folder}/المكلفون/{assignee}/"
                        f"شواهد/{note}-{evidence.id}{extension}",
                        field,
                    )

    # 8) صور التقارير المرتبطة بالتجارب تُنسخ أيضاً تحت سجل التجربة. وجودها
    # في مجلد التقرير وحده لا يكفي عند قراءة أرشيف المختبر مستقلاً.
    if school_wide:
        for experiment in _lab_experiments_qs(school):
            report = getattr(experiment, "report", None)
            if report is None:
                continue
            experiment_folder = _safe_segment(
                f"{experiment.experiment_date or 'بدون-تاريخ'} - "
                f"{experiment.title or 'تجربة'} - {experiment.id}",
                fallback=f"تجربة-{experiment.id}",
            )
            evidence_fields = [item.image for item in report.evidences.all()]
            if not evidence_fields:
                evidence_fields = [
                    report.image1,
                    report.image2,
                    report.image3,
                    report.image4,
                ]
            for index, field in enumerate(evidence_fields, start=1):
                name = getattr(field, "name", "") or ""
                if not name:
                    continue
                extension = os.path.splitext(name)[1].lower() or ".jpg"
                yield (
                    f"المختبرات/التجارب/{experiment_folder}/"
                    f"شاهد-{index}{extension}",
                    field,
                )

    # السجلات الإدارية لا تحمل سنة دراسية؛ يضيفها أرشيف السنة صراحةً
    # بوصفها "الحالة حتى لحظة إنشاء النسخة". وهنا نضيفها للتصدير الكامل.
    if academic_year or not school_wide:
        return

    yield from _administrative_file_fields(school)


def count_export_files(school, *, academic_year=None, teacher=None, school_wide=True) -> int:
    """عدد الملفات الفعلية التي ستُصدَّر (للعرض في صفحة المعاينة)."""
    try:
        count = sum(
            1 for _ in _file_fields_for_school(
                school, academic_year=academic_year, teacher=teacher, school_wide=school_wide
            )
        )
        if academic_year and school_wide:
            count += sum(1 for _ in _administrative_file_fields(school))
        return count
    except Exception:
        return 0


def archive_zip_filename(school, academic_year=None) -> str:
    code = (getattr(school, "code", "") or "school").strip() or "school"
    stamp = timezone.localtime().strftime("%Y%m%d-%H%M")
    year_part = ""
    if academic_year:
        year_part = "-" + re.sub(r"[^0-9A-Za-z\-]+", "", str(academic_year)) or ""
    return f"archive-{code}{year_part}-{stamp}.zip"


def build_school_export_zip_file(
    school,
    *,
    academic_year=None,
    teacher=None,
    school_wide=True,
    request=None,
    return_metadata=False,
):
    """يبني أرشيف ZIP يحوي ملفات المدرسة الفعلية + فهرس + ملف تحقّق سلامة (SHA-256).

    - عند تمرير ``academic_year`` يصبح أرشيف سنة محددة (مع manifest سلامة للسنة).
    - عند عدم تمريرها يصبح تصدير المدرسة الكامل (مع فهرس Excel).
    - ``request`` (اختياري): يُمرَّر لتوليد PDF ملفات الإنجاز غير المخزّنة لحظيًا.
    - عند ``return_metadata=True`` يعيد ``(file, metadata)`` لإظهار تقرير اكتمال صريح.
    """
    import hashlib
    from .services_archive import UNCLASSIFIED_YEAR, archive_year_label

    is_year_archive = bool(academic_year)
    is_unclassified = academic_year == UNCLASSIFIED_YEAR
    generated_at = timezone.localtime().strftime("%Y-%m-%d %H:%M:%S %Z")
    manifest_lines = []  # (sha256, size, path)
    missing_file_count = 0
    report_count = _reports_qs(
        school,
        academic_year=academic_year,
        teacher=teacher,
        school_wide=school_wide,
    ).count()
    achievement_count = (
        0
        if is_unclassified
        else _achievement_files_qs(
            school,
            academic_year=academic_year,
            teacher=teacher,
            school_wide=school_wide,
        ).count()
    )
    leadership_count = (
        _leadership_portfolios_qs(school, academic_year=academic_year).count()
        if school_wide and not is_unclassified
        else 0
    )
    ticket_count = _tickets_qs(school).count() if school_wide else 0
    circular_count = (
        Notification.objects.filter(school=school, requires_signature=True).count()
        if school_wide
        else 0
    )
    notification_count = (
        Notification.objects.filter(school=school, requires_signature=False).count()
        if school_wide
        else 0
    )
    document_count = (
        _documents_qs(school, academic_year=None if is_unclassified else academic_year).count()
        if school_wide
        else 0
    )
    meeting_count = _meetings_qs(school).count() if school_wide else 0
    plan_count = _plans_qs(school, academic_year=academic_year).count() if school_wide else 0
    initiative_count = (
        _initiatives_qs(school, academic_year=academic_year).count()
        if school_wide
        else 0
    )
    assignment_count = _assignments_qs(school).count() if school_wide else 0
    lab_asset_count = _lab_assets_qs(school).count() if school_wide else 0
    lab_handover_count = _lab_handovers_qs(school).count() if school_wide else 0
    lab_experiment_count = _lab_experiments_qs(school).count() if school_wide else 0

    tmp = tempfile.SpooledTemporaryFile(max_size=24 * 1024 * 1024)
    with zipfile.ZipFile(tmp, "w", zipfile.ZIP_DEFLATED, compresslevel=1) as zf:
        added = 0

        def write_bytes(archive_path: str, data: bytes) -> None:
            nonlocal added
            zf.writestr(archive_path, data)
            manifest_lines.append((hashlib.sha256(data).hexdigest(), len(data), archive_path))
            added += 1

        try:
            if is_year_archive:
                index_bytes = build_year_archive_index_bytes(
                    school,
                    academic_year,
                    teacher=teacher,
                    school_wide=school_wide,
                )
                write_bytes("فهرس-السنة.xlsx", index_bytes)
            else:
                write_bytes("ملخص-البيانات.xlsx", build_school_export_bytes(school))
        except Exception:
            missing_file_count += 1

        for arc_path, field in _file_fields_for_school(
            school, academic_year=academic_year, teacher=teacher, school_wide=school_wide
        ):
            try:
                field.open("rb")
                try:
                    data = field.read()
                finally:
                    try:
                        field.close()
                    except Exception:
                        pass
                write_bytes(arc_path, data)
            except Exception:
                missing_file_count += 1

        # السجلات الإدارية غير مرتبطة بسنة دراسية في قاعدة البيانات. في النسخة
        # السنوية للمدرسة نحفظ حالتها كاملة حتى لحظة إنشاء النسخة.
        if is_year_archive and school_wide:
            for arc_path, field in _administrative_file_fields(school):
                try:
                    field.open("rb")
                    try:
                        data = field.read()
                    finally:
                        try:
                            field.close()
                        except Exception:
                            pass
                    write_bytes(arc_path, data)
                except Exception:
                    missing_file_count += 1

        weasy_unavailable = False
        gen_ok = 0
        gen_failed = 0
        rep_ok = 0
        rep_failed = 0
        ticket_pdf_ok = 0
        ticket_pdf_failed = 0
        notification_pdf_ok = 0
        notification_pdf_failed = 0
        leadership_pdf_ok = 0
        leadership_pdf_failed = 0
        meeting_pdf_ok = 0
        meeting_pdf_failed = 0
        assignment_pdf_ok = 0
        assignment_pdf_failed = 0
        plan_pdf_ok = 0
        plan_pdf_failed = 0
        initiative_pdf_ok = 0
        initiative_pdf_failed = 0
        lab_inventory_pdf_ok = 0
        lab_inventory_pdf_failed = 0
        lab_experiment_pdf_ok = 0
        lab_experiment_pdf_failed = 0

        # (أ) ملف PDF لكل تقرير (لكل النطاقات بما فيها غير المصنّفة بسنة)
        for rep in _reports_qs(
            school, academic_year=academic_year, teacher=teacher, school_wide=school_wide
        ).iterator(chunk_size=100):
            if weasy_unavailable:
                rep_failed += 1
                continue
            try:
                from .pdf_report import generate_report_pdf

                pdf_bytes, _fn = generate_report_pdf(request=request, report=rep)
            except OSError:
                weasy_unavailable = True
                rep_failed += 1
                continue
            except Exception:
                rep_failed += 1
                continue
            arc = f"{_report_folder(rep)}/التقرير.pdf"
            write_bytes(arc, pdf_bytes)
            rep_ok += 1

        # (ب) ملفات الإنجاز غير المخزّنة (لا تنطبق على غير المصنّفة بسنة)
        if not is_unclassified:
            ach_qs = _achievement_files_qs(
                school, academic_year=academic_year, teacher=teacher, school_wide=school_wide
            )
            for ach in ach_qs.iterator(chunk_size=100):
                if getattr(getattr(ach, "pdf_file", None), "name", ""):
                    continue  # مخزّن مسبقًا وأُدرِج في المرور الأساسي
                if weasy_unavailable:
                    gen_failed += 1
                    continue
                try:
                    from .pdf_achievement import generate_achievement_pdf

                    pdf_bytes, _fn = generate_achievement_pdf(request=request, ach_file=ach)
                except OSError:
                    weasy_unavailable = True
                    gen_failed += 1
                    continue
                except Exception:
                    gen_failed += 1
                    continue
                arc = _achievement_arc_name(ach)
                write_bytes(arc, pdf_bytes)
                gen_ok += 1

        # (ج) PDF موثق لكل طلب/تذكرة ولكل تعميم/إشعار، مع سجل الحالات
        # والمستلمين والتوقيعات. تُحفظ المرفقات الأصلية بجانبه.
        if school_wide:
            from .pdf_archive_records import (
                generate_notification_archive_pdf,
                generate_ticket_archive_pdf,
            )

            for ticket in _tickets_qs(school).iterator(chunk_size=100):
                if weasy_unavailable:
                    ticket_pdf_failed += 1
                    continue
                try:
                    pdf_bytes = generate_ticket_archive_pdf(ticket, request=request)
                except OSError:
                    weasy_unavailable = True
                    ticket_pdf_failed += 1
                    continue
                except Exception:
                    ticket_pdf_failed += 1
                    continue
                write_bytes(f"{_ticket_folder(ticket)}/سجل-الطلب.pdf", pdf_bytes)
                ticket_pdf_ok += 1

            for notification in _notifications_qs(school).iterator(chunk_size=100):
                if weasy_unavailable:
                    notification_pdf_failed += 1
                    continue
                try:
                    pdf_bytes = generate_notification_archive_pdf(
                        notification,
                        request=request,
                    )
                except OSError:
                    weasy_unavailable = True
                    notification_pdf_failed += 1
                    continue
                except Exception:
                    notification_pdf_failed += 1
                    continue
                write_bytes(
                    f"{_notification_folder(notification)}/السجل.pdf",
                    pdf_bytes,
                )
                notification_pdf_ok += 1

            for meeting in _meetings_qs(school).iterator(chunk_size=100):
                try:
                    from .pdf_meeting import generate_meeting_pdf

                    pdf_bytes, _filename = generate_meeting_pdf(
                        request=request,
                        meeting=meeting,
                    )
                except Exception:
                    meeting_pdf_failed += 1
                    continue
                if meeting.scheduled_at:
                    folder_name = f"{meeting.scheduled_at:%Y-%m-%d} - {meeting.title} - {meeting.id}"
                else:
                    folder_name = f"{meeting.title} - {meeting.id}"
                folder = _safe_segment(folder_name, fallback=f"اجتماع-{meeting.id}")
                write_bytes(
                    f"الاجتماعات-والمحاضر/{folder}/محضر-الاجتماع.pdf",
                    pdf_bytes,
                )
                meeting_pdf_ok += 1

            # (هـ) ملفات التنفيذ المدرسي التي كانت سابقاً مجرد صفوف Excel.
            # تستخدم مولداً حتمياً لا يعتمد على مكتبات HTML الأصلية، فلا يحول
            # تعطل Pango نسخة المدرسة إلى أرشيف ناقص.
            from .pdf_school_archive import (
                generate_assignment_archive_pdf,
                generate_initiative_archive_pdf,
                generate_lab_experiment_archive_pdf,
                generate_lab_inventory_archive_pdf,
                generate_plan_archive_pdf,
            )

            for assignment in _assignments_qs(school):
                try:
                    pdf_bytes = generate_assignment_archive_pdf(
                        assignment,
                        school=school,
                    )
                except Exception:
                    assignment_pdf_failed += 1
                    continue
                folder = _safe_segment(
                    f"{assignment.title} - {assignment.id}",
                    fallback=f"تكليف-{assignment.id}",
                )
                write_bytes(f"التكليفات/{folder}/سجل-التكليف.pdf", pdf_bytes)
                assignment_pdf_ok += 1

            for plan in _plans_qs(school, academic_year=academic_year):
                try:
                    pdf_bytes = generate_plan_archive_pdf(plan)
                except Exception:
                    plan_pdf_failed += 1
                    continue
                folder = _safe_segment(
                    f"{plan.title} - {plan.id}",
                    fallback=f"خطة-{plan.id}",
                )
                write_bytes(f"الخطط/{folder}/الخطة.pdf", pdf_bytes)
                plan_pdf_ok += 1

            for initiative in _initiatives_qs(
                school,
                academic_year=academic_year,
            ):
                try:
                    pdf_bytes = generate_initiative_archive_pdf(initiative)
                except Exception:
                    initiative_pdf_failed += 1
                    continue
                folder = _safe_segment(
                    f"{initiative.title} - {initiative.id}",
                    fallback=f"مبادرة-{initiative.id}",
                )
                write_bytes(f"المبادرات/{folder}/المبادرة.pdf", pdf_bytes)
                initiative_pdf_ok += 1

            lab_assets = list(_lab_assets_qs(school))
            lab_handovers = list(_lab_handovers_qs(school))
            if lab_assets or lab_handovers:
                try:
                    pdf_bytes = generate_lab_inventory_archive_pdf(
                        school=school,
                        assets=lab_assets,
                        handovers=lab_handovers,
                    )
                except Exception:
                    lab_inventory_pdf_failed += 1
                else:
                    write_bytes(
                        "المختبرات/كشف-العهدة-وحركاتها.pdf",
                        pdf_bytes,
                    )
                    lab_inventory_pdf_ok += 1

            for experiment in _lab_experiments_qs(school):
                try:
                    pdf_bytes = generate_lab_experiment_archive_pdf(experiment)
                except Exception:
                    lab_experiment_pdf_failed += 1
                    continue
                folder = _safe_segment(
                    f"{experiment.experiment_date or 'بدون-تاريخ'} - "
                    f"{experiment.title or 'تجربة'} - {experiment.id}",
                    fallback=f"تجربة-{experiment.id}",
                )
                write_bytes(
                    f"المختبرات/التجارب/{folder}/سجل-التجربة.pdf",
                    pdf_bytes,
                )
                lab_experiment_pdf_ok += 1

        # (د) ملف الأداء القيادي الرسمي للمدرسة، مع شواهده الأصلية التي أضيفت
        # في المرور الأساسي تحت «منصة توثيق · القيادة المدرسية».
        if school_wide and not is_unclassified:
            leadership_qs = _leadership_portfolios_qs(
                school,
                academic_year=academic_year,
            )
            for portfolio in leadership_qs:
                if weasy_unavailable:
                    leadership_pdf_failed += 1
                    continue
                try:
                    from .pdf_leadership import generate_leadership_portfolio_pdf

                    pdf_bytes = generate_leadership_portfolio_pdf(
                        portfolio,
                        request=request,
                    )
                except OSError:
                    weasy_unavailable = True
                    leadership_pdf_failed += 1
                    continue
                except Exception:
                    leadership_pdf_failed += 1
                    continue
                write_bytes(_leadership_pdf_arc_name(portfolio), pdf_bytes)
                leadership_pdf_ok += 1

        content_notes = [
            "تقرير اكتمال الحزمة:",
            f"  • التقارير في النطاق: {report_count}",
            f"  • ملفات الإنجاز في النطاق: {achievement_count}",
            f"  • ملفات الأداء القيادي في النطاق: {leadership_count}",
            f"  • الطلبات والتذاكر حتى لحظة إنشاء النسخة: {ticket_count}",
            f"  • التعاميم حتى لحظة إنشاء النسخة: {circular_count}",
            f"  • الإشعارات حتى لحظة إنشاء النسخة: {notification_count}",
            f"  • الوثائق في النطاق: {document_count}",
            f"  • الاجتماعات والمحاضر حتى لحظة إنشاء النسخة: {meeting_count}",
            f"  • الخطط في النطاق: {plan_count}",
            f"  • المبادرات في النطاق: {initiative_count}",
            f"  • التكليفات حتى لحظة إنشاء النسخة: {assignment_count}",
            f"  • أصول المختبر حتى لحظة إنشاء النسخة: {lab_asset_count}",
            f"  • حركات العهدة حتى لحظة إنشاء النسخة: {lab_handover_count}",
            f"  • تجارب المختبر حتى لحظة إنشاء النسخة: {lab_experiment_count}",
            f"  • ملفات PDF للتقارير التي تم توليدها: {rep_ok}",
            f"  • ملفات PDF لملفات الإنجاز التي تم توليدها: {gen_ok}",
            f"  • ملفات PDF للأداء القيادي التي تم توليدها: {leadership_pdf_ok}",
            f"  • ملفات PDF للطلبات والتذاكر التي تم توليدها: {ticket_pdf_ok}",
            f"  • ملفات PDF للتعاميم والإشعارات التي تم توليدها: {notification_pdf_ok}",
            f"  • ملفات PDF لمحاضر الاجتماعات التي تم توليدها: {meeting_pdf_ok}",
            f"  • ملفات PDF للتكليفات التي تم توليدها: {assignment_pdf_ok}",
            f"  • ملفات PDF للخطط التي تم توليدها: {plan_pdf_ok}",
            f"  • ملفات PDF للمبادرات التي تم توليدها: {initiative_pdf_ok}",
            f"  • ملفات PDF لكشف عهدة المختبر التي تم توليدها: {lab_inventory_pdf_ok}",
            f"  • ملفات PDF لتجارب المختبر التي تم توليدها: {lab_experiment_pdf_ok}",
        ]
        if missing_file_count:
            content_notes.append(f"  ⚠ ملفات أصلية تعذرت قراءتها: {missing_file_count}")
        if rep_failed:
            content_notes.append(f"  ⚠ ملفات PDF للتقارير تعذر توليدها: {rep_failed}")
        if gen_failed:
            content_notes.append(f"  ⚠ ملفات PDF لملفات الإنجاز تعذر توليدها: {gen_failed}")
        if ticket_pdf_failed:
            content_notes.append(
                f"  ⚠ ملفات PDF للطلبات والتذاكر تعذر توليدها: {ticket_pdf_failed}"
            )
        if notification_pdf_failed:
            content_notes.append(
                "  ⚠ ملفات PDF للتعاميم والإشعارات تعذر توليدها: "
                f"{notification_pdf_failed}"
            )
        if leadership_pdf_failed:
            content_notes.append(
                f"  ⚠ ملفات PDF للأداء القيادي تعذر توليدها: {leadership_pdf_failed}"
            )
        if meeting_pdf_failed:
            content_notes.append(
                f"  ⚠ ملفات PDF لمحاضر الاجتماعات تعذر توليدها: {meeting_pdf_failed}"
            )
        for label, failed in (
            ("التكليفات", assignment_pdf_failed),
            ("الخطط", plan_pdf_failed),
            ("المبادرات", initiative_pdf_failed),
            ("كشف عهدة المختبر", lab_inventory_pdf_failed),
            ("تجارب المختبر", lab_experiment_pdf_failed),
        ):
            if failed:
                content_notes.append(
                    f"  ⚠ ملفات PDF لـ{label} تعذر توليدها: {failed}"
                )
        if not (
            missing_file_count
            or rep_failed
            or gen_failed
            or ticket_pdf_failed
            or notification_pdf_failed
            or leadership_pdf_failed
            or meeting_pdf_failed
            or assignment_pdf_failed
            or plan_pdf_failed
            or initiative_pdf_failed
            or lab_inventory_pdf_failed
            or lab_experiment_pdf_failed
        ):
            content_notes.append("  ✓ اكتمل إنشاء الحزمة دون ملفات مفقودة أو أخطاء توليد.")
        if school_wide:
            content_notes.append(
                "  • ملاحظة النطاق: السجلات الإدارية لا تحمل سنة دراسية؛ "
                "لذلك تم حفظ حالتها كاملة حتى لحظة إنشاء النسخة."
            )
        content_notes.append("=" * 60)

        scope = archive_year_label(academic_year) if is_year_archive else "كل السنوات"
        header = [
            "ملف تحقّق وسلامة الأرشيف — منصة توثيق",
            "=" * 60,
            f"المدرسة      : {getattr(school, 'name', '') or ''} ({getattr(school, 'code', '') or ''})",
            f"النطاق       : {scope}",
            f"تاريخ الإنشاء: {generated_at}",
            f"عدد الملفات  : {added}",
            "خوارزمية البصمة: SHA-256",
            "ملاحظة: هذا الأرشيف نسخة للقراءة فقط؛ يمكن التحقق من سلامة كل ملف عبر بصمته أدناه.",
            "=" * 60,
            *content_notes,
            "",
            "البصمة (SHA-256)".ljust(66) + "الحجم(بايت)".ljust(14) + "المسار",
        ]
        body = [f"{h}  {str(size).ljust(12)}  {path}" for (h, size, path) in manifest_lines]
        zf.writestr("الفهرس-والتحقق.txt", "\n".join(header + body) + "\n")

        if (
            report_count == 0
            and achievement_count == 0
            and leadership_count == 0
            and ticket_count == 0
            and circular_count == 0
            and notification_count == 0
            and document_count == 0
            and meeting_count == 0
            and plan_count == 0
            and initiative_count == 0
            and assignment_count == 0
            and lab_asset_count == 0
            and lab_handover_count == 0
            and lab_experiment_count == 0
        ):
            zf.writestr(
                "اقرأني.txt",
                "لا توجد سجلات مدرسية ضمن هذه السنة وقت إنشاء النسخة.\n",
            )

    tmp.seek(0)
    digest = hashlib.sha256()
    archive_size = 0
    while True:
        chunk = tmp.read(1024 * 1024)
        if not chunk:
            break
        digest.update(chunk)
        archive_size += len(chunk)
    tmp.seek(0)
    metadata = {
        "generated_at": generated_at,
        "file_count": added,
        "missing_file_count": missing_file_count,
        "failed_pdf_count": (
            rep_failed
            + gen_failed
            + ticket_pdf_failed
            + notification_pdf_failed
            + leadership_pdf_failed
            + meeting_pdf_failed
            + assignment_pdf_failed
            + plan_pdf_failed
            + initiative_pdf_failed
            + lab_inventory_pdf_failed
            + lab_experiment_pdf_failed
        ),
        "generated_report_pdf_count": rep_ok,
        "generated_achievement_pdf_count": gen_ok,
        "generated_leadership_pdf_count": leadership_pdf_ok,
        "generated_meeting_pdf_count": meeting_pdf_ok,
        "generated_assignment_pdf_count": assignment_pdf_ok,
        "generated_plan_pdf_count": plan_pdf_ok,
        "generated_initiative_pdf_count": initiative_pdf_ok,
        "generated_lab_inventory_pdf_count": lab_inventory_pdf_ok,
        "generated_lab_experiment_pdf_count": lab_experiment_pdf_ok,
        "report_count": report_count,
        "achievement_count": achievement_count,
        "leadership_count": leadership_count,
        "ticket_count": ticket_count,
        "circular_count": circular_count,
        "notification_count": notification_count,
        "document_count": document_count,
        "meeting_count": meeting_count,
        "plan_count": plan_count,
        "initiative_count": initiative_count,
        "assignment_count": assignment_count,
        "lab_asset_count": lab_asset_count,
        "lab_handover_count": lab_handover_count,
        "lab_experiment_count": lab_experiment_count,
        "archive_size_bytes": archive_size,
        "archive_sha256": digest.hexdigest(),
        "is_partial": bool(
            missing_file_count
            or rep_failed
            or gen_failed
            or ticket_pdf_failed
            or notification_pdf_failed
            or leadership_pdf_failed
            or meeting_pdf_failed
            or assignment_pdf_failed
            or plan_pdf_failed
            or initiative_pdf_failed
            or lab_inventory_pdf_failed
            or lab_experiment_pdf_failed
        ),
        "notes": "\n".join(content_notes),
    }
    return (tmp, metadata) if return_metadata else tmp
