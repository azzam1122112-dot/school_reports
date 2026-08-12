# -*- coding: utf-8 -*-
"""تقرير المجموعة المجمَّع — Excel و PDF.

يُجيب عن البند الذي كان بلا مقابل في توصيف المدير التنفيذي: «استخراج تقارير
المجموعة بصيغتي PDF وExcel».

**المصدر واحد للصيغتين.** :func:`build_group_snapshot` تبني الأرقام مرة واحدة،
ثم تُصاغ في جدول أو في صفحة. ولو بُنيت في كل مُصدِّر على حدة لاختلف رقمٌ بين
الملفين المستخرجين في الدقيقة نفسها — وهو أسوأ عيب يمكن أن يحمله تقرير.

**الأرقام محسوبة لا مخزَّنة**، وبعدد ثابت من الاستعلامات لا استعلام لكل مدرسة:
مجموعةٌ من عشرين مدرسة لا يجوز أن تعني ثمانين استعلاماً عند كل تصدير.
"""
from __future__ import annotations

from datetime import timedelta
from io import BytesIO

from django.db.models import Count, Q
from django.utils import timezone

from .model_parts.approvals import ApprovalState, PENDING_REVIEW_STATES
from .models import (
    Assignment,
    AssignmentTarget,
    Meeting,
    Report,
    SchoolMembership,
    TeacherAchievementFile,
)

__all__ = [
    "build_group_snapshot",
    "build_group_workbook_bytes",
    "group_export_filename",
]

# نافذة النشاط — نفس ثابت لوحة المجموعة، فلا يفترق رقم الشاشة عن رقم الملف.
ACTIVITY_WINDOW_DAYS = 30

_BRAND_GREEN = "006C35"
_HEADER_FILL = "0F8F6B"
_ZEBRA = "F1F8F4"


def build_group_snapshot(group) -> dict:
    """لقطة رقمية للمجموعة — مصدر الحقيقة الوحيد لكل صيغة تصدير."""
    now = timezone.now()
    since = now - timedelta(days=ACTIVITY_WINDOW_DAYS)

    schools = list(group.schools.filter(is_active=True).select_related("subscription").order_by("name"))
    school_ids = [school.pk for school in schools]

    reports = {
        row["school_id"]: row
        for row in Report.objects.filter(school_id__in=school_ids)
        .values("school_id")
        .annotate(
            total=Count("id"),
            recent=Count("id", filter=Q(created_at__gte=since)),
            approved=Count("id", filter=Q(approval_state=ApprovalState.APPROVED)),
            pending=Count("id", filter=Q(approval_state__in=PENDING_REVIEW_STATES)),
        )
    }
    achievements = {
        row["school_id"]: row["total"]
        for row in TeacherAchievementFile.objects.filter(school_id__in=school_ids)
        .values("school_id")
        .annotate(total=Count("id"))
    }
    seats = SchoolMembership.seats_used_by_school(school_ids)

    # التكليفات تُقرأ من حصص المكلَّفين لا من التكليف نفسه: مدرسةٌ شُملت في
    # تكليف مجموعة قد تكون أنجزته وغيرها لم يُنجز.
    targets = {
        row["school_id"]: row
        for row in AssignmentTarget.objects.filter(school_id__in=school_ids)
        .values("school_id")
        .annotate(
            total=Count("id"),
            done=Count("id", filter=Q(approval_state=ApprovalState.APPROVED)),
        )
    }
    overdue = {
        row["school_id"]: row["late"]
        for row in AssignmentTarget.objects.filter(
            school_id__in=school_ids,
            assignment__due_at__lt=now,
            assignment__cancelled_at__isnull=True,
        )
        .exclude(approval_state=ApprovalState.APPROVED)
        .values("school_id")
        .annotate(late=Count("id"))
    }
    meetings = {
        row["school_id"]: row["total"]
        for row in Meeting.objects.filter(
            school_id__in=school_ids, status=Meeting.Status.HELD
        )
        .values("school_id")
        .annotate(total=Count("id"))
    }

    rows = []
    for school in schools:
        report_row = reports.get(school.pk, {}) or {}
        target_row = targets.get(school.pk, {}) or {}
        assigned = int(target_row.get("total") or 0)
        done = int(target_row.get("done") or 0)
        subscription = getattr(school, "subscription", None)
        try:
            days_left = None if subscription is None else int(subscription.days_remaining or 0)
        except Exception:
            days_left = None

        row = {
                "school": school,
                "name": school.name,
                "seats": int(seats.get(school.pk) or 0),
                "reports_total": int(report_row.get("total") or 0),
                "reports_recent": int(report_row.get("recent") or 0),
                "reports_pending": int(report_row.get("pending") or 0),
                "achievements": int(achievements.get(school.pk) or 0),
                "assignments": assigned,
                "assignments_done": done,
                "assignments_overdue": int(overdue.get(school.pk) or 0),
                "completion": round(done * 100 / assigned) if assigned else 0,
                "meetings": int(meetings.get(school.pk) or 0),
                "subscription_days": days_left,
            }

        risk_score = 0
        risk_reasons: list[str] = []
        recommendations: list[str] = []
        if days_left is None or days_left <= 0:
            risk_score += 50
            risk_reasons.append("لا يوجد اشتراك سارٍ")
            recommendations.append("التنسيق مع مدير المدرسة لتفعيل الاشتراك")
        elif days_left <= 7:
            risk_score += 35
            risk_reasons.append(f"الاشتراك ينتهي خلال {days_left} أيام")
            recommendations.append("بدء التجديد قبل توقف وصول الفريق")
        elif days_left <= 30:
            risk_score += 20
            risk_reasons.append(f"الاشتراك يقترب من الانتهاء ({days_left} يومًا)")
            recommendations.append("إدراج التجديد ضمن خطة الشهر")
        if row["assignments_overdue"]:
            risk_score += 25
            risk_reasons.append(f"{row['assignments_overdue']} تكليف متأخر")
            recommendations.append("مراجعة التكليفات المتأخرة مع مدير المدرسة")
        if row["reports_pending"]:
            risk_score += 10
            risk_reasons.append(f"{row['reports_pending']} تقرير ينتظر الاعتماد")
            recommendations.append("إغلاق دورة اعتماد التقارير المعلقة")
        if not row["reports_recent"]:
            risk_score += 10
            risk_reasons.append(f"لا تقارير خلال {ACTIVITY_WINDOW_DAYS} يومًا")
            recommendations.append("التحقق من تفعيل أنواع التقارير واستخدام الفريق")
        if not row["seats"]:
            risk_score += 15
            risk_reasons.append("لا يوجد فريق نشط")
            recommendations.append("إضافة منسوبي المدرسة وتحديد أدوارهم")

        row["risk_score"] = min(risk_score, 100)
        row["risk_reasons"] = risk_reasons
        row["recommendations"] = recommendations
        row["risk_reasons_text"] = "، ".join(risk_reasons) or "لا توجد مؤشرات خطر"
        row["recommendation_text"] = recommendations[0] if recommendations else "الاستمرار في المتابعة الدورية"
        if row["risk_score"] >= 50:
            row["risk_tone"], row["risk_label"] = "danger", "تحتاج تدخلاً"
        elif row["risk_score"] >= 20:
            row["risk_tone"], row["risk_label"] = "warning", "تحتاج متابعة"
        else:
            row["risk_tone"], row["risk_label"] = "stable", "مستقرة"
        rows.append(row)

    group_assignments = Assignment.objects.filter(group=group).count()
    group_councils = Meeting.objects.filter(
        group=group, scope=Meeting.Scope.GROUP, status=Meeting.Status.HELD
    ).count()

    totals = {
        "schools": len(rows),
        "seats": sum(row["seats"] for row in rows),
        "reports_total": sum(row["reports_total"] for row in rows),
        "reports_recent": sum(row["reports_recent"] for row in rows),
        "reports_pending": sum(row["reports_pending"] for row in rows),
        "achievements": sum(row["achievements"] for row in rows),
        "assignments": sum(row["assignments"] for row in rows),
        "assignments_done": sum(row["assignments_done"] for row in rows),
        "assignments_overdue": sum(row["assignments_overdue"] for row in rows),
        "meetings": sum(row["meetings"] for row in rows),
        "group_assignments": group_assignments,
        "group_councils": group_councils,
        "schools_at_risk": sum(1 for row in rows if row["risk_score"] >= 50),
        "schools_to_watch": sum(1 for row in rows if 20 <= row["risk_score"] < 50),
    }
    totals["completion"] = (
        round(totals["assignments_done"] * 100 / totals["assignments"])
        if totals["assignments"]
        else 0
    )

    # الترتيب بالإنجاز ثم بالتأخر: المقارنة هي غرض التقرير، وجدولٌ مرتّب
    # أبجدياً يخفيها.
    ranked = sorted(
        rows, key=lambda row: (-row["completion"], row["assignments_overdue"], row["name"])
    )
    attention = sorted(rows, key=lambda row: (-row["risk_score"], row["name"]))

    return {
        "group": group,
        "rows": rows,
        "ranked": ranked,
        "attention": attention,
        "totals": totals,
        "window_days": ACTIVITY_WINDOW_DAYS,
        "generated_at": timezone.localtime(),
    }


_COLUMNS = (
    ("المدرسة", "name"),
    ("حالة الصحة", "risk_label"),
    ("درجة الخطر / 100", "risk_score"),
    ("أسباب الخطر", "risk_reasons_text"),
    ("التوصية الأولى", "recommendation_text"),
    ("المنسوبون", "seats"),
    ("التقارير", "reports_total"),
    (f"تقارير آخر {ACTIVITY_WINDOW_DAYS} يوماً", "reports_recent"),
    ("تقارير تنتظر الاعتماد", "reports_pending"),
    ("ملفات الإنجاز", "achievements"),
    ("التكليفات", "assignments"),
    ("المنجز منها", "assignments_done"),
    ("المتأخر", "assignments_overdue"),
    ("نسبة الإنجاز %", "completion"),
    ("الاجتماعات المنعقدة", "meetings"),
    ("أيام الاشتراك المتبقية", "subscription_days"),
)


def build_group_workbook_bytes(snapshot: dict) -> bytes:
    """ملف Excel بورقتين: ملخص المجموعة، ومقارنة مدارسها."""
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
    from openpyxl.utils import get_column_letter

    wb = Workbook()
    wb.properties.creator = "منصة توثيق"

    header_font = Font(name="Calibri", bold=True, color="FFFFFF", size=11)
    header_fill = PatternFill("solid", fgColor=_HEADER_FILL)
    title_font = Font(name="Calibri", bold=True, color=_BRAND_GREEN, size=16)
    label_font = Font(name="Calibri", bold=True, size=11)
    right = Alignment(horizontal="right", vertical="center", wrap_text=True, readingOrder=2)
    center = Alignment(horizontal="center", vertical="center")
    thin = Side(style="thin", color="D6E4DC")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    totals = snapshot["totals"]

    # ── ورقة الملخص ────────────────────────────────────────────────
    ws = wb.active
    ws.title = "ملخص المجموعة"
    ws.sheet_view.rightToLeft = True
    ws.cell(row=1, column=1, value="التقرير التنفيذي المجمَّع").font = title_font
    ws.cell(row=2, column=1, value=snapshot["group"].name).font = label_font

    summary = [
        ("عدد المدارس", totals["schools"]),
        ("مدارس تحتاج تدخلاً", totals["schools_at_risk"]),
        ("مدارس تحتاج متابعة", totals["schools_to_watch"]),
        ("إجمالي المنسوبين", totals["seats"]),
        ("إجمالي التقارير", totals["reports_total"]),
        (f"تقارير آخر {snapshot['window_days']} يوماً", totals["reports_recent"]),
        ("تقارير تنتظر الاعتماد", totals["reports_pending"]),
        ("ملفات الإنجاز", totals["achievements"]),
        ("حصص التكليفات", totals["assignments"]),
        ("المنجز منها", totals["assignments_done"]),
        ("المتأخر", totals["assignments_overdue"]),
        ("نسبة الإنجاز العامة %", totals["completion"]),
        ("اجتماعات المدارس المنعقدة", totals["meetings"]),
        ("تكليفات المجموعة", totals["group_assignments"]),
        ("جلسات المجلس المنعقدة", totals["group_councils"]),
        ("تاريخ الاستخراج", snapshot["generated_at"].strftime("%Y-%m-%d %H:%M")),
    ]
    for index, (label, value) in enumerate(summary, start=4):
        cell = ws.cell(row=index, column=1, value=label)
        cell.font = label_font
        cell.alignment = right
        ws.cell(row=index, column=2, value=value).alignment = right
    ws.column_dimensions["A"].width = 32
    ws.column_dimensions["B"].width = 22

    # ── ورقة المقارنة ──────────────────────────────────────────────
    ws2 = wb.create_sheet("مقارنة المدارس")
    ws2.sheet_view.rightToLeft = True

    headers = [label for label, _key in _COLUMNS]
    for column, label in enumerate(headers, start=1):
        cell = ws2.cell(row=1, column=column, value=label)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center
        cell.border = border

    for row_index, row in enumerate(snapshot["ranked"], start=2):
        for column, (_label, key) in enumerate(_COLUMNS, start=1):
            value = row.get(key)
            if value is None:
                value = "—"
            cell = ws2.cell(row=row_index, column=column, value=value)
            cell.alignment = right
            cell.border = border
            if row_index % 2 == 0:
                cell.fill = PatternFill("solid", fgColor=_ZEBRA)

    for column in range(1, len(headers) + 1):
        longest = len(str(headers[column - 1]))
        for row in snapshot["ranked"]:
            longest = max(longest, len(str(row.get(_COLUMNS[column - 1][1], ""))))
        ws2.column_dimensions[get_column_letter(column)].width = min(40, max(12, longest + 4))
    ws2.freeze_panes = ws2.cell(row=2, column=1)
    ws2.row_dimensions[1].height = 30

    buffer = BytesIO()
    wb.save(buffer)
    return buffer.getvalue()


def group_export_filename(group, *, extension: str) -> str:
    code = (getattr(group, "code", "") or "group").strip() or "group"
    stamp = timezone.localtime().strftime("%Y%m%d-%H%M")
    return f"group-report-{code}-{stamp}.{extension}"
