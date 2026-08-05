import re
from io import BytesIO
from unittest.mock import patch

import openpyxl
from django.contrib import messages
from django.contrib.messages import get_messages
from django.core.files.uploadedfile import SimpleUploadedFile
from django.test import TestCase, override_settings
from django.urls import reverse

from reports import teacher_onboarding
from reports.forms import TeacherCreateForm
from reports.models import (
    Department,
    DepartmentMembership,
    School,
    SchoolMembership,
    SchoolSubscription,
    SubscriptionPlan,
    Teacher,
)
from reports.teacher_onboarding import PREVIEW_SESSION_KEY


@override_settings(ALLOWED_HOSTS=["testserver"])
class TeacherOnboardingTests(TestCase):
    # The preview carries a 30-minute wall-clock TTL. None of the tests below
    # are about that expiry, yet every one of them would inherit it: the two
    # requests they make are milliseconds apart normally, but on a loaded
    # machine the gap can cross the limit and the confirmation is then rejected
    # as expired. That failed as a bare "False is not true" on whichever
    # side effect the test happened to assert. Expiry gets its own test below;
    # here it is held far enough away to stop timing from deciding the result.
    def setUp(self):
        ttl_patch = patch.object(teacher_onboarding, "PREVIEW_MAX_AGE_SECONDS", 24 * 60 * 60)
        ttl_patch.start()
        self.addCleanup(ttl_patch.stop)
        self._setup_school()

    def _setup_school(self):
        self.school = School.objects.create(
            name="مدرسة الإضافة السريعة",
            code="teacher-onboarding",
        )
        self.plan = SubscriptionPlan.objects.create(
            name="خطة الإضافة السريعة",
            price=0,
            days_duration=30,
            max_teachers=20,
        )
        SchoolSubscription.objects.create(school=self.school, plan=self.plan)
        self.manager = Teacher.objects.create_user(
            phone="0500000001",
            name="مدير المدرسة",
            password="manager-safe-password",
            is_staff=True,
        )
        SchoolMembership.objects.create(
            school=self.school,
            teacher=self.manager,
            role_type=SchoolMembership.RoleType.MANAGER,
        )
        self.department = Department.objects.create(
            school=self.school,
            name="النشاط الطلابي",
            slug="activities",
        )
        self.client.force_login(self.manager)
        session = self.client.session
        session["active_school_id"] = self.school.pk
        session.save()

    def _quick_preview(self, *, phone="0551234567", name="محمد المعلم"):
        return self.client.post(
            reverse("reports:bulk_import_teachers"),
            {
                "action": "quick_preview",
                "name": [name],
                "phone": [phone],
                "national_id": ["1012345678"],
                "job_title": [SchoolMembership.JobTitle.TEACHER],
                "department": [str(self.department.pk)],
            },
        )

    def _confirm_current_preview(self, *, expect_success=True):
        preview = self.client.session[PREVIEW_SESSION_KEY]
        response = self.client.post(
            reverse("reports:bulk_import_teachers"),
            {
                "action": "confirm",
                "preview_token": preview["token"],
            },
        )
        if expect_success:
            # A rejected confirmation writes nothing, so without this the caller
            # only sees its own side-effect assertion fail with no clue why.
            # Surface the reason the server gave instead.
            errors = [
                str(message)
                for message in get_messages(response.wsgi_request)
                if message.level >= messages.WARNING
            ]
            self.assertEqual(errors, [], f"لم تُقبل المعاينة: {errors}")
        return response

    def test_unified_onboarding_page_renders_with_one_clear_heading(self):
        response = self.client.get(reverse("reports:bulk_import_teachers"))
        html = response.content.decode("utf-8")

        self.assertEqual(response.status_code, 200)
        self.assertContains(response, "إضافة منسوبي المدرسة بسهولة")
        self.assertContains(response, "إضافة سريعة")
        self.assertContains(response, "رفع ملف Excel")
        self.assertEqual(len(re.findall(r"<h1\b", html, re.IGNORECASE)), 1)
        self.assertContains(response, "static/js/teacher-onboarding.js")
        self.assertNotIn(
            "document.getElementById('addQuickRow')",
            html,
        )

    def test_add_row_has_a_server_fallback_that_preserves_entered_values(self):
        response = self.client.post(
            reverse("reports:bulk_import_teachers"),
            {
                "action": "add_quick_row",
                "name": ["معلم محفوظ", "", ""],
                "phone": ["0553332211", "", ""],
                "national_id": ["", "", ""],
                "job_title": ["teacher", "teacher", "teacher"],
                "department": [str(self.department.pk), "", ""],
            },
        )

        self.assertEqual(response.status_code, 200)
        self.assertEqual(len(response.context["quick_rows"]), 4)
        self.assertEqual(response.context["quick_rows"][0]["name"], "معلم محفوظ")
        self.assertEqual(response.context["quick_rows"][0]["phone"], "0553332211")
        self.assertContains(response, 'value="معلم محفوظ"')
        self.assertContains(response, 'value="0553332211"')
        self.assertNotIn(PREVIEW_SESSION_KEY, self.client.session)

    def test_remove_row_has_a_server_fallback_that_preserves_other_rows(self):
        response = self.client.post(
            reverse("reports:bulk_import_teachers"),
            {
                "action": "quick_preview",
                "remove_row": "1",
                "name": ["الأول", "الثاني", "الثالث"],
                "phone": ["0551111111", "0552222222", "0553333333"],
                "national_id": ["", "", ""],
                "job_title": ["teacher", "teacher", "teacher"],
                "department": ["", "", ""],
            },
        )

        self.assertEqual(response.status_code, 200)
        self.assertEqual(
            [row["name"] for row in response.context["quick_rows"]],
            ["الأول", "الثالث"],
        )
        self.assertNotContains(response, 'value="الثاني"')
        self.assertNotIn(PREVIEW_SESSION_KEY, self.client.session)

    def test_individual_form_uses_phone_as_temporary_password(self):
        form = TeacherCreateForm(
            data={
                "name": "معلم فردي",
                "phone": "0552223344",
                "national_id": "",
                "job_title": SchoolMembership.JobTitle.TEACHER,
                "is_active": "on",
            },
            active_school=self.school,
        )

        self.assertNotIn("password", form.fields)
        self.assertTrue(form.is_valid(), form.errors)
        teacher = form.save()
        self.assertTrue(teacher.check_password("0552223344"))

    def test_individual_add_view_creates_phone_password_and_school_membership(self):
        response = self.client.post(
            reverse("reports:add_teacher"),
            {
                "name": "معلم من الإضافة الفردية",
                "phone": "0552223355",
                "national_id": "",
                "job_title": SchoolMembership.JobTitle.ADMIN_STAFF,
                "is_active": "on",
            },
        )

        self.assertRedirects(
            response,
            reverse("reports:manage_teachers"),
            fetch_redirect_response=False,
        )
        teacher = Teacher.objects.get(phone="0552223355")
        self.assertTrue(teacher.check_password(teacher.phone))
        membership = SchoolMembership.objects.get(
            school=self.school,
            teacher=teacher,
            role_type=SchoolMembership.RoleType.TEACHER,
        )
        self.assertEqual(membership.job_title, SchoolMembership.JobTitle.ADMIN_STAFF)

    def test_blank_ready_rows_in_quick_table_are_not_treated_as_users(self):
        self.client.post(
            reverse("reports:bulk_import_teachers"),
            {
                "action": "quick_preview",
                "name": ["مستخدم واحد", "", ""],
                "phone": ["0551010101", "", ""],
                "national_id": ["", "", ""],
                "job_title": ["teacher", "teacher", "teacher"],
                "department": ["", "", ""],
            },
        )

        preview = self.client.session[PREVIEW_SESSION_KEY]
        self.assertEqual(preview["summary"]["total"], 1)
        self.assertTrue(preview["can_confirm"])

    def test_quick_preview_writes_nothing_then_confirmation_creates_everything(self):
        response = self._quick_preview()

        self.assertRedirects(
            response,
            f"{reverse('reports:bulk_import_teachers')}?step=preview",
            fetch_redirect_response=False,
        )
        self.assertFalse(Teacher.objects.filter(phone="0551234567").exists())
        preview = self.client.session[PREVIEW_SESSION_KEY]
        self.assertTrue(preview["can_confirm"])
        self.assertEqual(preview["summary"]["new"], 1)

        response = self._confirm_current_preview()

        self.assertRedirects(
            response,
            f"{reverse('reports:bulk_import_teachers')}?completed=1",
            fetch_redirect_response=False,
        )
        teacher = Teacher.objects.get(phone="0551234567")
        self.assertTrue(teacher.check_password(teacher.phone))
        membership = SchoolMembership.objects.get(
            school=self.school,
            teacher=teacher,
            role_type=SchoolMembership.RoleType.TEACHER,
        )
        self.assertEqual(membership.job_title, SchoolMembership.JobTitle.TEACHER)
        self.assertTrue(
            DepartmentMembership.objects.filter(
                department=self.department,
                teacher=teacher,
            ).exists()
        )

    def test_stale_preview_is_refused_and_says_so(self):
        """The TTL the other tests hold away must still be enforced for real."""
        self._quick_preview(phone="0559998877", name="معلم معاينة قديمة")

        # Age the stored preview past the limit rather than waiting it out.
        # The bound is inclusive, so zero would still count as fresh within the
        # same second — it has to be negative to land on the far side of it.
        with patch.object(teacher_onboarding, "PREVIEW_MAX_AGE_SECONDS", -1):
            response = self._confirm_current_preview(expect_success=False)

        self.assertFalse(
            SchoolMembership.objects.filter(
                school=self.school,
                teacher__phone="0559998877",
            ).exists()
        )
        shown = [str(message) for message in get_messages(response.wsgi_request)]
        self.assertTrue(
            any("انتهت صلاحية المعاينة" in message for message in shown),
            f"لم تظهر رسالة انتهاء الصلاحية: {shown}",
        )

    def test_created_teacher_must_change_phone_password_on_first_login(self):
        self._quick_preview(phone="0557654321", name="معلم دخول أول")
        self._confirm_current_preview()
        self.client.logout()

        response = self.client.post(
            reverse("reports:login"),
            {"phone": "0557654321", "password": "0557654321"},
        )

        self.assertRedirects(
            response,
            reverse("reports:my_profile"),
            fetch_redirect_response=False,
        )

    def test_existing_account_is_linked_without_overwriting_identity_or_password(self):
        existing = Teacher.objects.create_user(
            phone="0553334455",
            name="الاسم الأصلي",
            national_id="1023456789",
            password="existing-safe-password",
        )

        self._quick_preview(phone=existing.phone, name="اسم مختلف في الملف")
        preview = self.client.session[PREVIEW_SESSION_KEY]
        self.assertEqual(preview["summary"]["link"], 1)
        self.assertTrue(preview["rows"][0]["warnings"])
        self._confirm_current_preview()

        existing.refresh_from_db()
        self.assertEqual(existing.name, "الاسم الأصلي")
        self.assertEqual(existing.national_id, "1023456789")
        self.assertTrue(existing.check_password("existing-safe-password"))
        self.assertTrue(
            SchoolMembership.objects.filter(
                school=self.school,
                teacher=existing,
                role_type=SchoolMembership.RoleType.TEACHER,
            ).exists()
        )

    def test_duplicate_or_over_capacity_rows_cannot_be_confirmed(self):
        response = self.client.post(
            reverse("reports:bulk_import_teachers"),
            {
                "action": "quick_preview",
                "name": ["الأول", "الثاني"],
                "phone": ["0554445566", "0554445566"],
                "national_id": ["", ""],
                "job_title": ["teacher", "teacher"],
                "department": ["", ""],
            },
        )
        self.assertEqual(response.status_code, 302)
        preview = self.client.session[PREVIEW_SESSION_KEY]
        self.assertFalse(preview["can_confirm"])
        self.assertEqual(preview["summary"]["invalid"], 1)

        response = self.client.post(
            reverse("reports:bulk_import_teachers"),
            {
                "action": "quick_preview",
                "name": ["الأول", "الثاني"],
                "phone": ["0554445566", "0554445577"],
                "national_id": ["1010101010", "1010101010"],
                "job_title": ["teacher", "teacher"],
                "department": ["", ""],
            },
        )
        self.assertEqual(response.status_code, 302)
        preview = self.client.session[PREVIEW_SESSION_KEY]
        self.assertFalse(preview["can_confirm"])
        self.assertEqual(preview["summary"]["invalid"], 1)

        self.plan.max_teachers = 1
        self.plan.save(update_fields=["max_teachers"])
        response = self.client.post(
            reverse("reports:bulk_import_teachers"),
            {
                "action": "quick_preview",
                "name": ["الأول", "الثاني"],
                "phone": ["0554445566", "0557778899"],
                "national_id": ["", ""],
                "job_title": ["teacher", "teacher"],
                "department": ["", ""],
            },
        )
        self.assertEqual(response.status_code, 302)
        preview = self.client.session[PREVIEW_SESSION_KEY]
        self.assertFalse(preview["can_confirm"])
        self.assertIn("المتبقي في الباقة", preview["capacity_error"])

    def test_excel_template_is_empty_and_upload_is_preview_first(self):
        template_response = self.client.get(reverse("reports:bulk_import_teachers_template"))
        self.assertEqual(template_response.status_code, 200)
        template_book = openpyxl.load_workbook(BytesIO(template_response.content))
        self.assertEqual(template_book["المعلمون"].max_row, 1)
        self.assertIn("التعليمات والأمثلة", template_book.sheetnames)

        book = openpyxl.Workbook()
        sheet = book.active
        sheet.append(["الاسم الكامل", "رقم الجوال", "رقم الهوية", "المسمى الوظيفي", "القسم"])
        sheet.append(["معلم من Excel", "0558889900", "", "معلم", "النشاط الطلابي"])
        payload = BytesIO()
        book.save(payload)
        upload = SimpleUploadedFile(
            "teachers.xlsx",
            payload.getvalue(),
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )

        response = self.client.post(
            reverse("reports:bulk_import_teachers"),
            {"action": "file_preview", "excel_file": upload},
        )

        self.assertEqual(response.status_code, 302)
        self.assertFalse(Teacher.objects.filter(phone="0558889900").exists())
        self.assertEqual(self.client.session[PREVIEW_SESSION_KEY]["source"], "file")
        self._confirm_current_preview()
        self.assertTrue(Teacher.objects.filter(phone="0558889900").exists())

    def test_issues_and_new_login_exports_are_available_to_same_school_manager(self):
        self.client.post(
            reverse("reports:bulk_import_teachers"),
            {
                "action": "quick_preview",
                "name": ["صف غير صالح"],
                "phone": ["123"],
                "national_id": [""],
                "job_title": ["teacher"],
                "department": [""],
            },
        )
        issues_response = self.client.get(reverse("reports:bulk_import_teachers_issues"))
        self.assertEqual(issues_response.status_code, 200)
        self.assertEqual(
            issues_response["Content-Type"],
            "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )

        self._quick_preview(phone="0559990011", name="معلم جديد")
        self._confirm_current_preview()
        result_response = self.client.get(reverse("reports:bulk_import_teachers_result"))
        self.assertEqual(result_response.status_code, 200)
        result_book = openpyxl.load_workbook(BytesIO(result_response.content))
        self.assertEqual(result_book["بيانات الدخول"]["B2"].value, "0559990011")
        self.assertEqual(result_book["بيانات الدخول"]["C2"].value, "0559990011")

    def test_manager_page_makes_unified_onboarding_primary_and_supports_filters(self):
        self._quick_preview(phone="0551112233", name="معلم للبحث")
        self._confirm_current_preview()

        response = self.client.get(
            reverse("reports:manage_teachers"),
            {
                "q": "معلم للبحث",
                "status": "active",
                "job_title": "teacher",
                "department": str(self.department.pk),
            },
        )

        self.assertEqual(response.status_code, 200)
        self.assertContains(response, "إضافة منسوبي المدرسة")
        self.assertContains(response, "معلم للبحث")
        self.assertContains(response, "كلمة مرور مؤقتة هي رقم الجوال")
