from io import BytesIO

import openpyxl
from django.test import TestCase, override_settings
from django.urls import reverse

from reports.gender_labels import school_gender_labels, school_gender_template_context
from reports.models import (
    School,
    SchoolMembership,
    SchoolSubscription,
    SubscriptionPlan,
    Teacher,
)
from reports.permissions import effective_user_role_label, _school_role_labels
from reports.forms import TeacherCreateForm, _school_job_title_choices
from reports.services_export import build_school_export_workbook
from reports.staff_assignments import assignment_choices


class GenderedRoleLabelTests(TestCase):
    def setUp(self):
        self.boys = School.objects.create(
            name="مدرسة البنين", code="boys-school", gender=School.Gender.BOYS
        )
        self.girls = School.objects.create(
            name="مدرسة البنات", code="girls-school", gender=School.Gender.GIRLS
        )
        plan = SubscriptionPlan.objects.create(
            name="Plan", price=0, days_duration=30, max_teachers=0
        )
        SchoolSubscription.objects.create(school=self.boys, plan=plan)
        SchoolSubscription.objects.create(school=self.girls, plan=plan)

    def _make_member(self, school, role_type, job_title=None):
        t = Teacher.objects.create_user(
            phone=f"5{school.id}{role_type}{job_title or ''}"[:15],
            name="عضو",
            password="pass",
        )
        kwargs = dict(school=school, teacher=t, role_type=role_type)
        if job_title is not None:
            kwargs["job_title"] = job_title
        SchoolMembership.objects.create(**kwargs)
        return t

    def test_role_labels_map_boys(self):
        labels = _school_role_labels(self.boys)
        self.assertEqual(labels["manager"], "مدير المدرسة")
        self.assertEqual(labels["teacher"], "المعلم")
        self.assertEqual(labels["admin_staff"], "موظف إداري")
        self.assertEqual(labels["lab_tech"], "محضر مختبر")

    def test_role_labels_map_girls(self):
        labels = _school_role_labels(self.girls)
        self.assertEqual(labels["manager"], "مديرة المدرسة")
        self.assertEqual(labels["teacher"], "المعلمة")
        self.assertEqual(labels["admin_staff"], "موظفة إدارية")
        self.assertEqual(labels["lab_tech"], "محضرة مختبر")

    def test_complete_girls_terminology_covers_grammar_and_report_labels(self):
        labels = school_gender_labels(self.girls)
        self.assertEqual(labels["manager"], "مديرة المدرسة")
        self.assertEqual(labels["manager_short"], "المديرة")
        self.assertEqual(labels["teacher"], "المعلمة")
        self.assertEqual(labels["teachers"], "المعلمات")
        self.assertEqual(labels["executor"], "المنفّذة")
        self.assertEqual(labels["head_of_department"], "رئيسة القسم")
        self.assertEqual(labels["students"], "الطالبات")
        self.assertEqual(labels["beneficiaries"], "المستفيدات")
        self.assertEqual(labels["beneficiaries_object"], "المستفيدات")
        self.assertEqual(labels["teacher_dative"], "للمعلمة")
        self.assertEqual(labels["manager_dative"], "لمديرة المدرسة")

        context = school_gender_template_context(self.girls)
        self.assertTrue(context["IS_GIRLS_SCHOOL"])
        self.assertEqual(context["SCHOOL_EXECUTOR_LABEL"], "المنفّذة")
        self.assertEqual(context["SCHOOL_BENEFICIARIES_OBJ_LABEL"], "المستفيدات")

    def test_effective_label_manager(self):
        boss_b = self._make_member(self.boys, SchoolMembership.RoleType.MANAGER)
        boss_g = self._make_member(self.girls, SchoolMembership.RoleType.MANAGER)
        self.assertEqual(
            effective_user_role_label(boss_b, active_school=self.boys), "مدير المدرسة"
        )
        self.assertEqual(
            effective_user_role_label(boss_g, active_school=self.girls), "مديرة المدرسة"
        )

    def test_effective_label_job_titles(self):
        for jt, boys_label, girls_label in [
            (SchoolMembership.JobTitle.TEACHER, "المعلم", "المعلمة"),
            (SchoolMembership.JobTitle.ADMIN_STAFF, "موظف إداري", "موظفة إدارية"),
            (SchoolMembership.JobTitle.LAB_TECH, "محضر مختبر", "محضرة مختبر"),
        ]:
            tb = self._make_member(self.boys, SchoolMembership.RoleType.TEACHER, jt)
            tg = self._make_member(self.girls, SchoolMembership.RoleType.TEACHER, jt)
            self.assertEqual(
                effective_user_role_label(tb, active_school=self.boys), boys_label
            )
            self.assertEqual(
                effective_user_role_label(tg, active_school=self.girls), girls_label
            )

    def test_job_title_choice_labels(self):
        boys_choices = dict(_school_job_title_choices(self.boys))
        girls_choices = dict(_school_job_title_choices(self.girls))
        self.assertEqual(boys_choices[SchoolMembership.JobTitle.TEACHER], "معلم")
        self.assertEqual(girls_choices[SchoolMembership.JobTitle.TEACHER], "معلمة")
        self.assertEqual(
            boys_choices[SchoolMembership.JobTitle.ADMIN_STAFF], "موظف إداري"
        )
        self.assertEqual(
            girls_choices[SchoolMembership.JobTitle.ADMIN_STAFF], "موظفة إدارية"
        )
        self.assertEqual(boys_choices[SchoolMembership.JobTitle.LAB_TECH], "محضر مختبر")
        self.assertEqual(
            girls_choices[SchoolMembership.JobTitle.LAB_TECH], "محضرة مختبر"
        )

    def test_staff_assignment_choices_switch_to_feminine_labels(self):
        boys_choices = dict(assignment_choices(self.boys))
        girls_choices = dict(assignment_choices(self.girls))

        self.assertEqual(boys_choices[SchoolMembership.RoleType.TEACHER], "معلم")
        self.assertEqual(girls_choices[SchoolMembership.RoleType.TEACHER], "معلمة")
        self.assertEqual(boys_choices[SchoolMembership.RoleType.DEPUTY], "وكيل المدرسة")
        self.assertEqual(girls_choices[SchoolMembership.RoleType.DEPUTY], "وكيلة المدرسة")
        self.assertEqual(boys_choices[SchoolMembership.RoleType.ADMIN_STAFF], "موظف إداري")
        self.assertEqual(girls_choices[SchoolMembership.RoleType.ADMIN_STAFF], "موظفة إدارية")
        self.assertEqual(boys_choices[SchoolMembership.JobTitle.LAB_TECH], "محضر مختبر")
        self.assertEqual(girls_choices[SchoolMembership.JobTitle.LAB_TECH], "محضرة مختبر")

        form = TeacherCreateForm(active_school=self.girls)
        self.assertEqual(
            dict(form.fields["job_title"].choices)[SchoolMembership.RoleType.DEPUTY],
            "وكيلة المدرسة",
        )

    @override_settings(ALLOWED_HOSTS=["testserver"])
    def test_girls_school_pages_render_feminine_report_and_workspace_labels(self):
        teacher = self._make_member(self.girls, SchoolMembership.RoleType.TEACHER)
        self.client.force_login(teacher)
        session = self.client.session
        session["active_school_id"] = self.girls.pk
        session.save()

        report_response = self.client.get(reverse("reports:add_report"))
        home_response = self.client.get(reverse("reports:home"))

        self.assertEqual(report_response.status_code, 200)
        self.assertContains(report_response, "المنفّذة")
        self.assertContains(report_response, "عدد المستفيدات")
        self.assertNotContains(report_response, "> المنفذ</label>", html=False)
        self.assertEqual(home_response.status_code, 200)
        self.assertContains(home_response, "مساحة عمل المعلمة")

    @override_settings(ALLOWED_HOSTS=["testserver"])
    def test_girls_school_excel_outputs_use_feminine_labels(self):
        manager = self._make_member(self.girls, SchoolMembership.RoleType.MANAGER)
        self.client.force_login(manager)
        session = self.client.session
        session["active_school_id"] = self.girls.pk
        session.save()

        response = self.client.get(reverse("reports:bulk_import_teachers_template"))
        self.assertEqual(response.status_code, 200)
        template_book = openpyxl.load_workbook(BytesIO(response.content))
        self.assertIn("المعلمات", template_book.sheetnames)
        instructions = template_book["التعليمات والأمثلة"]
        self.assertEqual(instructions["C5"].value, "معلمة")
        self.assertIn("وكيلة المدرسة", instructions["D5"].value)
        self.assertIn("محضرة مختبر", instructions["D5"].value)

        export_book = build_school_export_workbook(self.girls)
        self.assertIn("المعلمات", export_book.sheetnames)
        summary = export_book["ملخص"]
        summary_labels = [summary.cell(row=row, column=1).value for row in range(1, summary.max_row + 1)]
        self.assertIn("عدد المعلمات", summary_labels)

    @override_settings(ALLOWED_HOSTS=["testserver"])
    def test_girls_school_import_page_renders_feminine_assignment_choices(self):
        manager = self._make_member(self.girls, SchoolMembership.RoleType.MANAGER)
        self.client.force_login(manager)
        session = self.client.session
        session["active_school_id"] = self.girls.pk
        session.save()

        response = self.client.get(reverse("reports:bulk_import_teachers"))

        self.assertEqual(response.status_code, 200)
        self.assertContains(response, "معلمة")
        self.assertContains(response, "وكيلة المدرسة")
        self.assertContains(response, "موظفة إدارية")
        self.assertContains(response, "محضرة مختبر")
        self.assertNotContains(response, ">وكيل المدرسة</option>", html=False)
