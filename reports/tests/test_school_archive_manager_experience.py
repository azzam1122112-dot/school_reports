import re
import tempfile
import zipfile
from datetime import timedelta
from io import BytesIO
from unittest.mock import patch

from django.core.exceptions import ValidationError
from django.core.files.uploadedfile import SimpleUploadedFile
from django.template.loader import render_to_string
from django.test import TestCase, override_settings
from django.urls import reverse
from django.utils import timezone
from openpyxl import load_workbook

from maintenance.services import collect_reset_summary
from reports.models import (
    Assignment,
    AssignmentEvidence,
    AssignmentTarget,
    AchievementEvidenceReport,
    AchievementSection,
    Department,
    DepartmentMembership,
    Document,
    GeneratedExportJob,
    Initiative,
    LabAsset,
    LabAssetHandover,
    LabExperiment,
    Payment,
    Meeting,
    MeetingMinutes,
    Notification,
    NotificationRecipient,
    Plan,
    PlanGoal,
    PlanTask,
    Report,
    LeadershipEvidenceImage,
    LeadershipPortfolioSection,
    School,
    SchoolArchiveAddon,
    SchoolMembership,
    SchoolLeadershipPortfolio,
    SchoolSubscription,
    SchoolYearArchive,
    SchoolYearArchiveDownload,
    SubscriptionPlan,
    Teacher,
    TeacherAchievementFile,
    Ticket,
    TicketImage,
    TicketNote,
)
from reports.pdf_report import build_report_print_context
from reports.model_parts.approvals import ApprovalState
from reports.services_archive import (
    archive_available_years,
    calculate_school_archive_storage_bytes,
    school_storage_breakdown,
)
from reports.services_export import (
    build_school_export_bytes,
    build_school_export_zip_file,
    build_year_archive_index_bytes,
)


@override_settings(ALLOWED_HOSTS=["testserver"])
class SchoolArchiveManagerExperienceTests(TestCase):
    def setUp(self):
        self.media_dir = tempfile.TemporaryDirectory()
        self.addCleanup(self.media_dir.cleanup)
        self.settings_override = override_settings(MEDIA_ROOT=self.media_dir.name)
        self.settings_override.enable()
        self.addCleanup(self.settings_override.disable)

        self.school = School.objects.create(
            name="مدرسة الأرشيف الاحترافي",
            code="archive-manager-ux",
            current_academic_year="1447-1448",
            allowed_academic_years=["1446-1447", "1447-1448"],
        )
        plan = SubscriptionPlan.objects.create(
            name="خطة الأرشيف",
            price=500,
            days_duration=365,
            max_teachers=0,
        )
        SchoolSubscription.objects.create(school=self.school, plan=plan)
        self.addon = SchoolArchiveAddon.objects.create(
            school=self.school,
            is_enabled=True,
            start_date=timezone.localdate(),
            end_date=timezone.localdate() + timedelta(days=20),
            storage_limit_gb=2,
        )
        self.manager = Teacher.objects.create_user(
            phone="0500700001",
            name="مدير الأرشيف",
            password="safe-manager-password",
        )
        SchoolMembership.objects.create(
            school=self.school,
            teacher=self.manager,
            role_type=SchoolMembership.RoleType.MANAGER,
        )
        self.teacher = Teacher.objects.create_user(
            phone="0500700002",
            name="معلم الأرشيف",
            password="safe-teacher-password",
        )
        SchoolMembership.objects.create(
            school=self.school,
            teacher=self.teacher,
            role_type=SchoolMembership.RoleType.TEACHER,
        )
        self.report = Report.objects.create(
            school=self.school,
            teacher=self.teacher,
            title="تقرير ثابت للأرشيف",
            report_date=timezone.localdate(),
            academic_year="1447-1448",
        )
        self.client.force_login(self.manager)
        session = self.client.session
        session["active_school_id"] = self.school.pk
        session.save()

    def _create_snapshot(self):
        with (
            patch(
                "reports.pdf_report.generate_report_pdf",
                return_value=(b"%PDF-archive-report", "report.pdf"),
            ),
            patch(
                "reports.pdf_leadership.generate_leadership_portfolio_pdf",
                return_value=b"%PDF-leadership",
            ),
        ):
            return self.client.post(
                reverse("reports:school_archive_create"),
                {"year": "1447-1448"},
            )

    def test_archive_uses_complete_fallback_pdf_when_weasyprint_is_unavailable(self):
        self.report.idea = "وصف التقرير الذي يجب أن يبقى داخل ملف PDF المؤرشف."
        self.report.image1 = SimpleUploadedFile(
            "report.png",
            bytes.fromhex(
                "89504e470d0a1a0a0000000d49484452000000010000000108060000001f15c4"
                "890000000a49444154789c6360000002000154a24f6f0000000049454e44ae426082"
            ),
            content_type="image/png",
        )
        self.report.save(update_fields=["idea", "image1"])

        context = build_report_print_context(self.report)
        self.assertTrue(context["PDF_IMAGE1_URL"].startswith("data:image/png;base64,"))

        with patch(
            "reports.pdf_report._generate_report_pdf_weasy",
            side_effect=OSError("native PDF libraries unavailable"),
        ):
            package, metadata = build_school_export_zip_file(
                self.school,
                academic_year="1447-1448",
                teacher=self.manager,
                school_wide=True,
                request=self.client.request().wsgi_request,
                return_metadata=True,
            )
        try:
            self.assertEqual(metadata["generated_report_pdf_count"], 1)
            self.assertEqual(metadata["failed_pdf_count"], 0)
            self.assertFalse(metadata["is_partial"])
            with zipfile.ZipFile(package) as zipped:
                pdf_name = next(
                    name for name in zipped.namelist() if name.endswith("/التقرير.pdf")
                )
                pdf_bytes = zipped.read(pdf_name)
                self.assertTrue(pdf_bytes.startswith(b"%PDF-"))
                self.assertGreater(len(pdf_bytes), 1000)
                self.assertTrue(
                    any(name.endswith("/صورة-1.png") for name in zipped.namelist())
                )
        finally:
            package.close()

    def test_manager_archive_page_is_clear_searchable_and_warns_before_expiry(self):
        response = self.client.get(reverse("reports:school_archive"))
        html = response.content.decode("utf-8")

        self.assertEqual(response.status_code, 200)
        self.assertEqual(len(re.findall(r"<h1\b", html, re.IGNORECASE)), 1)
        self.assertContains(response, "نسخة ZIP ثابتة")
        self.assertContains(response, "تبقى 20 يومًا")
        self.assertContains(response, "اسم التقرير أو المعلم أو الجوال")
        self.assertContains(response, "إنشاء نسخة ثابتة الآن")
        self.assertNotContains(response, "1446-1447 هـ")

    def test_creation_persists_versioned_zip_with_excel_manifest_and_audit_download(self):
        response = self._create_snapshot()

        archive = SchoolYearArchive.objects.get()
        self.assertRedirects(
            response,
            f"{reverse('reports:school_archive')}?year=1447-1448&snapshot={archive.pk}",
            fetch_redirect_response=False,
        )
        self.assertEqual(archive.status, SchoolYearArchive.Status.READY)
        self.assertEqual(archive.version, 1)
        self.assertTrue(archive.archive_sha256)
        self.assertGreater(archive.storage_bytes, 0)
        self.school.refresh_from_db()
        self.assertEqual(self.school.storage_used_bytes, archive.storage_bytes)

        archive.archive_file.open("rb")
        package = archive.archive_file.read()
        archive.archive_file.close()
        with zipfile.ZipFile(BytesIO(package)) as zipped:
            names = set(zipped.namelist())
            self.assertIn("فهرس-السنة.xlsx", names)
            self.assertIn("الفهرس-والتحقق.txt", names)
            self.assertTrue(any(name.endswith("/التقرير.pdf") for name in names))
            manifest = zipped.read("الفهرس-والتحقق.txt").decode("utf-8")
            self.assertIn("اكتمل إنشاء الحزمة دون ملفات مفقودة", manifest)

        download = self.client.get(
            reverse("reports:school_archive_download", args=[archive.pk])
        )
        self.assertEqual(download.status_code, 200)
        self.assertEqual(download["Content-Type"], "application/zip")
        download.close()
        self.assertEqual(
            SchoolYearArchiveDownload.objects.filter(
                archive=archive,
                downloaded_by=self.manager,
            ).count(),
            1,
        )

    def test_archive_and_full_download_include_leadership_portfolio_and_evidence(self):
        portfolio = SchoolLeadershipPortfolio.objects.create(
            school=self.school,
            manager=self.manager,
            academic_year="1447-1448",
            leadership_vision="رؤية قيادة مدرسية موثقة",
            executive_summary="ملخص أثر القيادة",
        )
        section = LeadershipPortfolioSection.objects.create(
            portfolio=portfolio,
            code=LeadershipPortfolioSection.Code.PLANNING,
            notes="ممارسات التخطيط والتشغيل",
            is_completed=True,
        )
        evidence = LeadershipEvidenceImage.objects.create(
            section=section,
            caption="شاهد الخطة التشغيلية",
            image=SimpleUploadedFile(
                "leadership.png",
                bytes.fromhex(
                    "89504e470d0a1a0a0000000d49484452000000010000000108060000001f15c4"
                    "890000000a49444154789c6360000002000154a24f6f0000000049454e44ae426082"
                ),
                content_type="image/png",
            ),
        )
        self.assertGreater(calculate_school_archive_storage_bytes(self.school), 0)
        self.assertGreater(school_storage_breakdown(self.school)["leadership"], 0)

        index_bytes = build_year_archive_index_bytes(
            self.school,
            "1447-1448",
            teacher=self.manager,
            school_wide=True,
        )
        workbook = load_workbook(BytesIO(index_bytes), read_only=True, data_only=True)
        try:
            self.assertIn("الأداء القيادي", workbook.sheetnames)
            self.assertIn("شواهد الأداء القيادي", workbook.sheetnames)
            leadership_rows = list(workbook["الأداء القيادي"].iter_rows(values_only=True))
            self.assertTrue(
                any(
                    row[0] == portfolio.id
                    and row[2] == "1447-1448"
                    and row[4] == 1
                    and row[5] == 1
                    for row in leadership_rows[1:]
                )
            )
        finally:
            workbook.close()

        response = self._create_snapshot()
        self.assertEqual(response.status_code, 302)
        archive = SchoolYearArchive.objects.get()
        self.assertEqual(archive.leadership_count, 1)
        archive.archive_file.open("rb")
        try:
            with zipfile.ZipFile(BytesIO(archive.archive_file.read())) as zipped:
                names = set(zipped.namelist())
                self.assertIn(
                    "منصة توثيق · القيادة المدرسية/1447-1448/ملف الأداء القيادي.pdf",
                    names,
                )
                self.assertTrue(
                    any(
                        name.startswith(
                            "منصة توثيق · القيادة المدرسية/1447-1448/شواهد/"
                        )
                        and name.endswith(f"-{evidence.id}.png")
                        for name in names
                    )
                )
                manifest = zipped.read("الفهرس-والتحقق.txt").decode("utf-8")
                self.assertIn("ملفات الأداء القيادي في النطاق: 1", manifest)
        finally:
            archive.archive_file.close()

        with (
            patch(
                "reports.pdf_report.generate_report_pdf",
                return_value=(b"%PDF-report", "report.pdf"),
            ),
            patch(
                "reports.pdf_leadership.generate_leadership_portfolio_pdf",
                return_value=b"%PDF-leadership-full",
            ),
        ):
            full_package, metadata = build_school_export_zip_file(
                self.school,
                school_wide=True,
                return_metadata=True,
            )
        try:
            self.assertEqual(metadata["leadership_count"], 1)
            self.assertEqual(metadata["generated_leadership_pdf_count"], 1)
            with zipfile.ZipFile(full_package) as zipped:
                self.assertIn(
                    "منصة توثيق · القيادة المدرسية/1447-1448/ملف الأداء القيادي.pdf",
                    zipped.namelist(),
                )
        finally:
            full_package.close()

        full_workbook = load_workbook(
            BytesIO(build_school_export_bytes(self.school)),
            read_only=True,
            data_only=True,
        )
        try:
            self.assertIn("الأداء القيادي", full_workbook.sheetnames)
            self.assertIn("شواهد الأداء القيادي", full_workbook.sheetnames)
        finally:
            full_workbook.close()

    def test_each_snapshot_is_a_new_immutable_version(self):
        self._create_snapshot()
        self.report.title = "عنوان حي جديد"
        self.report.save(update_fields=["title"])
        self._create_snapshot()

        versions = list(
            SchoolYearArchive.objects.order_by("version").values_list("version", flat=True)
        )
        self.assertEqual(versions, [1, 2])
        first = SchoolYearArchive.objects.get(version=1)
        first.archive_sha256 = "0" * 64
        with self.assertRaises(ValidationError):
            first.save()

    def test_partial_snapshot_is_saved_with_visible_completion_counts(self):
        with patch(
            "reports.pdf_report.generate_report_pdf",
            side_effect=RuntimeError("PDF renderer unavailable"),
        ):
            response = self.client.post(
                reverse("reports:school_archive_create"),
                {"year": "1447-1448"},
            )

        self.assertEqual(response.status_code, 302)
        archive = SchoolYearArchive.objects.get()
        self.assertEqual(archive.status, SchoolYearArchive.Status.PARTIAL)
        self.assertEqual(archive.failed_pdf_count, 1)

        page = self.client.get(
            reverse("reports:school_archive"),
            {"year": "1447-1448"},
        )
        self.assertContains(page, "مكتمل مع ملاحظات")
        self.assertContains(page, "1 PDF متعذر")

    def test_saved_snapshot_remains_visible_downloadable_and_reset_protected_after_expiry(self):
        self._create_snapshot()
        archive = SchoolYearArchive.objects.get()
        self.addon.end_date = timezone.localdate() - timedelta(days=1)
        self.addon.save(update_fields=["end_date"])

        page = self.client.get(reverse("reports:school_archive"))
        self.assertEqual(page.status_code, 200)
        self.assertContains(page, "النسخ محفوظة")
        self.assertContains(page, "طلب التجديد")
        self.assertContains(page, "تنزيل هذه النسخة")

        download = self.client.get(
            reverse("reports:school_archive_download", args=[archive.pk])
        )
        self.assertEqual(download.status_code, 200)
        download.close()
        summary = collect_reset_summary(
            [self.school],
            {
                "reports": True,
                "achievements": True,
                "tickets": False,
                "notifications": False,
                "share_links": True,
            },
        )
        self.assertEqual(summary["archive_protected_schools_count"], 1)
        self.assertEqual(summary["reports_count"], 0)

    def test_configured_but_empty_years_are_not_presented_as_archived(self):
        self.report.delete()

        years = archive_available_years(
            school=self.school,
            teacher=self.manager,
            school_wide=True,
        )

        self.assertEqual(years, [])

    def test_school_with_only_administrative_records_can_create_current_year_snapshot(self):
        self.report.delete()
        Ticket.objects.create(
            school=self.school,
            creator=self.teacher,
            title="طلب إداري وحيد",
        )

        years = archive_available_years(
            school=self.school,
            teacher=self.manager,
            school_wide=True,
        )
        self.assertIn("1447-1448", years)

        with (
            patch(
                "reports.pdf_archive_records.generate_ticket_archive_pdf",
                return_value=b"%PDF-ticket-record",
            ),
            patch(
                "reports.pdf_archive_records.generate_notification_archive_pdf",
                return_value=b"%PDF-notification-record",
            ),
        ):
            response = self.client.post(
                reverse("reports:school_archive_create"),
                {"year": "1447-1448"},
            )

        self.assertEqual(response.status_code, 302)
        archive = SchoolYearArchive.objects.get()
        self.assertEqual(archive.report_count, 0)
        self.assertEqual(archive.ticket_count, 1)

    def test_school_with_only_plans_and_laboratory_work_can_create_snapshot(self):
        self.report.delete()
        Plan.objects.create(
            school=self.school,
            owner=self.manager,
            title="خطة هي المحتوى الوحيد",
            academic_year="1447-1448",
        )
        LabAsset.objects.create(
            school=self.school,
            name="أصل مختبر وحيد",
            quantity=1,
            recorded_by=self.manager,
        )

        years = archive_available_years(
            school=self.school,
            teacher=self.manager,
            school_wide=True,
        )
        self.assertIn("1447-1448", years)

        response = self.client.post(
            reverse("reports:school_archive_create"),
            {"year": "1447-1448"},
        )

        self.assertEqual(response.status_code, 302)
        archive = SchoolYearArchive.objects.get()
        self.assertEqual(archive.report_count, 0)
        self.assertEqual(archive.plan_count, 1)
        self.assertEqual(archive.lab_asset_count, 1)
        self.assertEqual(archive.status, SchoolYearArchive.Status.READY)

    def test_one_time_year_export_also_contains_excel_index(self):
        with patch(
            "reports.pdf_report.generate_report_pdf",
            return_value=(b"%PDF-archive-report", "report.pdf"),
        ):
            package, metadata = build_school_export_zip_file(
                self.school,
                academic_year="1447-1448",
                teacher=self.manager,
                school_wide=True,
                return_metadata=True,
            )
        try:
            with zipfile.ZipFile(package) as zipped:
                self.assertIn("فهرس-السنة.xlsx", zipped.namelist())
            self.assertFalse(metadata["is_partial"])
            self.assertEqual(metadata["report_count"], 1)
        finally:
            package.close()

    def test_year_archive_includes_documents_meeting_minutes_plans_and_initiatives(self):
        document = Document.objects.create(
            school=self.school,
            uploaded_by=self.manager,
            owner=self.manager,
            title="محضر لجنة الجودة المرفوع",
            academic_year="1447-1448",
            kind=Document.Kind.MINUTES,
            file=SimpleUploadedFile(
                "quality-minutes.pdf",
                b"%PDF-quality-minutes",
                content_type="application/pdf",
            ),
        )
        meeting = Meeting.objects.create(
            school=self.school,
            organizer=self.manager,
            title="اجتماع متابعة المبادرات",
            purpose="متابعة المبادرات النوعية.",
            scheduled_at=timezone.now(),
            status=Meeting.Status.HELD,
            held_at=timezone.now(),
        )
        MeetingMinutes.objects.create(
            meeting=meeting,
            recorder=self.manager,
            body="تمت مناقشة المبادرات واعتماد إجراءات المتابعة.",
        )
        plan = Plan.objects.create(
            school=self.school,
            owner=self.manager,
            title="خطة تحسين المبادرات",
            academic_year="1447-1448",
            description="خطة تشغيلية مرتبطة بالمبادرات.",
        )
        initiative = Initiative.objects.create(
            school=self.school,
            teacher=self.teacher,
            plan=plan,
            title="مبادرة القراءة اليومية",
            summary="رفع أثر القراءة اليومية داخل المدرسة.",
            is_best_practice=True,
            approval_state=ApprovalState.APPROVED,
        )
        draft_initiative = Initiative.objects.create(
            school=self.school,
            teacher=self.teacher,
            plan=plan,
            title="مسودة مبادرة محفوظة تاريخياً",
            summary="مقترح لم يعتمد بعد.",
        )

        with (
            patch(
                "reports.pdf_report.generate_report_pdf",
                return_value=(b"%PDF-report", "report.pdf"),
            ),
            patch(
                "reports.pdf_meeting.generate_meeting_pdf",
                return_value=(b"%PDF-meeting-minutes", "meeting.pdf"),
            ),
        ):
            package, metadata = build_school_export_zip_file(
                self.school,
                academic_year="1447-1448",
                teacher=self.manager,
                school_wide=True,
                return_metadata=True,
            )
        try:
            self.assertEqual(metadata["document_count"], 1)
            self.assertEqual(metadata["meeting_count"], 1)
            self.assertEqual(metadata["plan_count"], 1)
            self.assertEqual(metadata["initiative_count"], 2)
            self.assertEqual(metadata["generated_meeting_pdf_count"], 1)
            self.assertEqual(metadata["generated_initiative_pdf_count"], 2)
            with zipfile.ZipFile(package) as zipped:
                names = zipped.namelist()
                self.assertTrue(
                    any(
                        name.startswith("الوثائق/1447-1448/محضر/")
                        and name.endswith(f"-{document.id}.pdf")
                        for name in names
                    )
                )
                self.assertTrue(
                    any(
                        name.startswith("الاجتماعات-والمحاضر/")
                        and name.endswith("/محضر-الاجتماع.pdf")
                        for name in names
                    )
                )

                workbook = load_workbook(
                    BytesIO(zipped.read("فهرس-السنة.xlsx")),
                    read_only=True,
                    data_only=True,
                )
                self.addCleanup(workbook.close)
                self.assertTrue(
                    {
                        "الوثائق",
                        "الاجتماعات والمحاضر",
                        "الخطط",
                        "المبادرات",
                    }.issubset(set(workbook.sheetnames))
                )
                document_rows = list(workbook["الوثائق"].iter_rows(values_only=True))
                self.assertTrue(any(row[1] == document.title for row in document_rows[1:]))
                meeting_rows = list(
                    workbook["الاجتماعات والمحاضر"].iter_rows(values_only=True)
                )
                self.assertTrue(any(row[1] == meeting.title for row in meeting_rows[1:]))
                plan_rows = list(workbook["الخطط"].iter_rows(values_only=True))
                self.assertTrue(any(row[1] == plan.title for row in plan_rows[1:]))
                initiative_rows = list(workbook["المبادرات"].iter_rows(values_only=True))
                self.assertTrue(any(row[1] == initiative.title for row in initiative_rows[1:]))
                self.assertTrue(
                    any(row[1] == draft_initiative.title for row in initiative_rows[1:])
                )
        finally:
            package.close()

        full_workbook = load_workbook(
            BytesIO(build_school_export_bytes(self.school)),
            read_only=True,
            data_only=True,
        )
        self.addCleanup(full_workbook.close)
        self.assertIn("الاجتماعات", full_workbook.sheetnames)
        self.assertIn("المبادرات", full_workbook.sheetnames)

    def test_snapshot_fully_archives_assignments_plans_initiatives_and_labs(self):
        assignment = Assignment.objects.create(
            school=self.school,
            issuer=self.manager,
            issuer_name=self.manager.name,
            title="تكليف تنفيذ الخطة",
            description="تنفيذ المهمة ورفع شاهد الإنجاز.",
            due_at=timezone.now() + timedelta(days=5),
            requires_evidence=True,
        )
        target = AssignmentTarget.objects.create(
            assignment=assignment,
            assignee=self.teacher,
            school=self.school,
            progress_percent=80,
            progress_note="اكتملت الخطوات الأساسية.",
            approval_state=ApprovalState.APPROVED,
        )
        evidence = AssignmentEvidence.objects.create(
            target=target,
            file=SimpleUploadedFile(
                "assignment-proof.pdf",
                b"%PDF-assignment-proof",
                content_type="application/pdf",
            ),
            note="شاهد تنفيذ التكليف",
            uploaded_by=self.teacher,
        )
        other_school = School.objects.create(
            name="مدرسة أخرى لا تظهر بياناتها",
            code="archive-isolation-other-school",
        )
        other_teacher = Teacher.objects.create_user(
            phone="0500700099",
            name="معلم مدرسة أخرى",
            password="safe-other-school-password",
        )
        other_target = AssignmentTarget.objects.create(
            assignment=assignment,
            assignee=other_teacher,
            school=other_school,
            approval_state=ApprovalState.DRAFT,
        )
        other_evidence = AssignmentEvidence.objects.create(
            target=other_target,
            file=SimpleUploadedFile(
                "other-school-secret.pdf",
                b"%PDF-other-school-secret",
                content_type="application/pdf",
            ),
            note="شاهد خاص بمدرسة أخرى",
            uploaded_by=other_teacher,
        )
        plan = Plan.objects.create(
            school=self.school,
            owner=self.manager,
            title="الخطة التشغيلية المتكاملة",
            academic_year="1447-1448",
            description="خطة محفوظة بكل أهدافها ومهامها.",
        )
        goal = PlanGoal.objects.create(
            plan=plan,
            title="رفع جودة التنفيذ",
            indicator="نسبة المهام المعتمدة",
            target="100%",
        )
        PlanTask.objects.create(
            plan=plan,
            goal=goal,
            title="تنفيذ التكليف الموثق",
            description="تفصيل المهمة التي يجب ألا تضيع من الأرشيف.",
            responsible=self.teacher,
            due_at=assignment.due_at,
            assignment=assignment,
        )
        initiative = Initiative.objects.create(
            school=self.school,
            teacher=self.teacher,
            plan=plan,
            title="مبادرة الأثر المستدام",
            summary="وصف الفكرة والأثر السابق للرجوع إليه.",
            approval_state=ApprovalState.DRAFT,
        )
        asset = LabAsset.objects.create(
            school=self.school,
            name="مجهر رقمي",
            code="LAB-001",
            category=LabAsset.Category.DEVICE,
            quantity=3,
            custodian=self.teacher,
            recorded_by=self.manager,
        )
        handover = LabAssetHandover.objects.create(
            school=self.school,
            asset=asset,
            direction=LabAssetHandover.Direction.OUT,
            person=self.teacher,
            person_name=self.teacher.name,
            quantity=1,
            recorded_by=self.manager,
            note="تسليم لتنفيذ التجربة",
        )
        experiment = LabExperiment.objects.create(
            school=self.school,
            recorder=self.teacher,
            requested_by=self.manager,
            title="تجربة المجهر الرقمي",
            experiment_date=timezone.localdate(),
            subject="العلوم",
            class_name="الثالث المتوسط",
            students_count=20,
            objectives="تمييز مكونات الخلية.",
            procedure="تجهيز الشريحة ثم فحصها وتسجيل النتائج.",
            materials_note="شرائح وعينات محفوظة.",
            safety_notes="تعقيم الأدوات قبل الاستخدام وبعده.",
            report=self.report,
            approval_state=ApprovalState.APPROVED,
        )
        experiment.assets.add(asset)

        with patch(
            "reports.pdf_report.generate_report_pdf",
            return_value=(b"%PDF-report", "report.pdf"),
        ):
            package, metadata = build_school_export_zip_file(
                self.school,
                academic_year="1447-1448",
                teacher=self.manager,
                school_wide=True,
                return_metadata=True,
            )
        try:
            self.assertEqual(metadata["assignment_count"], 1)
            self.assertEqual(metadata["plan_count"], 1)
            self.assertEqual(metadata["initiative_count"], 1)
            self.assertEqual(metadata["lab_asset_count"], 1)
            self.assertEqual(metadata["lab_handover_count"], 1)
            self.assertEqual(metadata["lab_experiment_count"], 1)
            self.assertEqual(metadata["generated_assignment_pdf_count"], 1)
            self.assertEqual(metadata["generated_plan_pdf_count"], 1)
            self.assertEqual(metadata["generated_initiative_pdf_count"], 1)
            self.assertEqual(metadata["generated_lab_inventory_pdf_count"], 1)
            self.assertEqual(metadata["generated_lab_experiment_pdf_count"], 1)
            self.assertFalse(metadata["is_partial"])

            with zipfile.ZipFile(package) as zipped:
                names = zipped.namelist()
                expected_pdf_suffixes = (
                    "/سجل-التكليف.pdf",
                    "/الخطة.pdf",
                    "/المبادرة.pdf",
                    "المختبرات/كشف-العهدة-وحركاتها.pdf",
                    "/سجل-التجربة.pdf",
                )
                for suffix in expected_pdf_suffixes:
                    name = next(item for item in names if item.endswith(suffix))
                    self.assertTrue(zipped.read(name).startswith(b"%PDF-"))
                self.assertTrue(
                    any(
                        name.startswith("التكليفات/")
                        and "/شواهد/" in name
                        and name.endswith(f"-{evidence.id}.pdf")
                        for name in names
                    )
                )
                self.assertFalse(
                    any(name.endswith(f"-{other_evidence.id}.pdf") for name in names)
                )

                workbook = load_workbook(
                    BytesIO(zipped.read("فهرس-السنة.xlsx")),
                    read_only=True,
                    data_only=True,
                )
                self.addCleanup(workbook.close)
                self.assertTrue(
                    {
                        "التكليفات",
                        "تنفيذ التكليفات",
                        "شواهد التكليفات",
                        "أهداف الخطط",
                        "مهام الخطط",
                        "أصول المختبر",
                        "حركات العهدة",
                        "تجارب المختبر",
                    }.issubset(set(workbook.sheetnames))
                )
                self.assertTrue(
                    any(
                        row[1] == assignment.title
                        for row in list(workbook["التكليفات"].iter_rows(values_only=True))[1:]
                    )
                )
                assignment_row = next(
                    row
                    for row in list(workbook["التكليفات"].iter_rows(values_only=True))[1:]
                    if row[1] == assignment.title
                )
                self.assertEqual(assignment_row[9], 1)
                self.assertEqual(assignment_row[10], 100)
                self.assertNotIn(
                    other_teacher.name,
                    {
                        row[2]
                        for row in list(
                            workbook["تنفيذ التكليفات"].iter_rows(values_only=True)
                        )[1:]
                    },
                )
                self.assertTrue(
                    any(
                        row[1] == asset.name
                        for row in list(workbook["أصول المختبر"].iter_rows(values_only=True))[1:]
                    )
                )
                manifest = zipped.read("الفهرس-والتحقق.txt").decode("utf-8")
                self.assertIn("التكليفات حتى لحظة إنشاء النسخة: 1", manifest)
                self.assertIn("تجارب المختبر حتى لحظة إنشاء النسخة: 1", manifest)
        finally:
            package.close()

        with patch(
            "reports.pdf_report.generate_report_pdf",
            return_value=(b"%PDF-report", "report.pdf"),
        ):
            response = self.client.post(
                reverse("reports:school_archive_create"),
                {"year": "1447-1448"},
            )
        self.assertEqual(response.status_code, 302)
        snapshot = SchoolYearArchive.objects.get()
        self.assertEqual(snapshot.assignment_count, 1)
        self.assertEqual(snapshot.plan_count, 1)
        self.assertEqual(snapshot.initiative_count, 1)
        self.assertEqual(snapshot.lab_asset_count, 1)
        self.assertEqual(snapshot.lab_handover_count, 1)
        self.assertEqual(snapshot.lab_experiment_count, 1)

    def test_background_snapshot_persists_complete_school_work_counts(self):
        from reports.tasks import build_generated_export_task

        payload = tempfile.SpooledTemporaryFile(max_size=1024)
        payload.write(b"PK\x03\x04complete-archive-background")
        payload.seek(0)
        metadata = {
            "is_partial": False,
            "archive_sha256": "a" * 64,
            "file_count": 18,
            "missing_file_count": 0,
            "failed_pdf_count": 0,
            "report_count": 1,
            "achievement_count": 2,
            "leadership_count": 3,
            "ticket_count": 4,
            "circular_count": 5,
            "notification_count": 6,
            "assignment_count": 7,
            "plan_count": 8,
            "initiative_count": 9,
            "lab_asset_count": 10,
            "lab_handover_count": 11,
            "lab_experiment_count": 12,
            "notes": "نسخة مكتملة من العامل الخلفي.",
        }
        job = GeneratedExportJob.objects.create(
            school=self.school,
            requested_by=self.manager,
            kind=GeneratedExportJob.Kind.ARCHIVE_SNAPSHOT,
            parameters={"academic_year": "1447-1448", "school_wide": True},
        )

        with patch(
            "reports.services_export.build_school_export_zip_file",
            return_value=(payload, metadata),
        ):
            result = build_generated_export_task.apply(args=[job.pk]).get()

        self.assertTrue(result)
        job.refresh_from_db()
        self.assertEqual(job.status, GeneratedExportJob.Status.READY)
        snapshot = job.archive
        self.assertEqual(snapshot.assignment_count, 7)
        self.assertEqual(snapshot.plan_count, 8)
        self.assertEqual(snapshot.initiative_count, 9)
        self.assertEqual(snapshot.lab_asset_count, 10)
        self.assertEqual(snapshot.lab_handover_count, 11)
        self.assertEqual(snapshot.lab_experiment_count, 12)

    def test_snapshot_contains_pdf_and_original_files_for_tickets_and_circulars(self):
        ticket = Ticket.objects.create(
            school=self.school,
            creator=self.teacher,
            title="طلب تجهيز مختبر",
            body="تجهيز المختبر قبل الزيارة.",
            attachment=SimpleUploadedFile(
                "ticket.pdf",
                b"%PDF-ticket-attachment",
                content_type="application/pdf",
            ),
        )
        TicketImage.objects.create(
            ticket=ticket,
            image=SimpleUploadedFile(
                "ticket.png",
                bytes.fromhex(
                    "89504e470d0a1a0a0000000d49484452000000010000000108060000001f15c4"
                    "890000000a49444154789c6360000002000154a24f6f0000000049454e44ae426082"
                ),
                content_type="image/png",
            ),
        )
        circular = Notification.objects.create(
            school=self.school,
            created_by=self.manager,
            title="تعميم السلامة",
            message="الالتزام بإجراءات السلامة.",
            requires_signature=True,
            attachment=SimpleUploadedFile(
                "circular.pdf",
                b"%PDF-circular-attachment",
                content_type="application/pdf",
            ),
        )
        NotificationRecipient.objects.create(
            notification=circular,
            teacher=self.teacher,
        )
        notification = Notification.objects.create(
            school=self.school,
            created_by=self.manager,
            title="إشعار اجتماع",
            message="اجتماع قصير بعد الدوام.",
        )
        NotificationRecipient.objects.create(
            notification=notification,
            teacher=self.teacher,
        )

        with (
            patch(
                "reports.pdf_report.generate_report_pdf",
                return_value=(b"%PDF-report", "report.pdf"),
            ),
            patch(
                "reports.pdf_archive_records.generate_ticket_archive_pdf",
                return_value=b"%PDF-ticket-record",
            ),
            patch(
                "reports.pdf_archive_records.generate_notification_archive_pdf",
                return_value=b"%PDF-notification-record",
            ),
        ):
            response = self.client.post(
                reverse("reports:school_archive_create"),
                {"year": "1447-1448"},
            )

        self.assertEqual(response.status_code, 302)
        archive = SchoolYearArchive.objects.get()
        self.assertEqual(archive.ticket_count, 1)
        self.assertEqual(archive.circular_count, 1)
        self.assertEqual(
            archive.notification_count,
            Notification.objects.filter(
                school=self.school,
                requires_signature=False,
            ).count(),
        )

        archive.archive_file.open("rb")
        package = archive.archive_file.read()
        archive.archive_file.close()
        with zipfile.ZipFile(BytesIO(package)) as zipped:
            names = zipped.namelist()
            self.assertTrue(
                any(
                    name.startswith("الطلبات-والتذاكر/")
                    and name.endswith("/سجل-الطلب.pdf")
                    for name in names
                )
            )
            self.assertTrue(
                any(
                    name.startswith("الطلبات-والتذاكر/")
                    and name.endswith("/المرفق-الرئيسي.pdf")
                    for name in names
                )
            )
            self.assertTrue(
                any(
                    name.startswith("الطلبات-والتذاكر/")
                    and "/صورة-1." in name
                    for name in names
                )
            )
            self.assertTrue(
                any(
                    name.startswith("التعاميم/")
                    and name.endswith("/السجل.pdf")
                    for name in names
                )
            )
            self.assertTrue(
                any(
                    name.startswith("التعاميم/")
                    and name.endswith("/المرفق.pdf")
                    for name in names
                )
            )
            self.assertTrue(
                any(
                    name.startswith("الإشعارات/")
                    and name.endswith("/السجل.pdf")
                    for name in names
                )
            )
            manifest = zipped.read("الفهرس-والتحقق.txt").decode("utf-8")
            self.assertIn("الطلبات والتذاكر حتى لحظة إنشاء النسخة: 1", manifest)
            self.assertIn("التعاميم حتى لحظة إنشاء النسخة: 1", manifest)

        page = self.client.get(
            reverse("reports:school_archive"),
            {"year": "1447-1448"},
        )
        self.assertContains(page, "التذاكر حتى الآن")
        self.assertContains(page, "تشمل الحزمة الحالات والمرفقات والشواهد الأصلية")

    def test_manager_archive_previews_and_searches_all_administrative_records(self):
        Ticket.objects.create(
            school=self.school,
            creator=self.teacher,
            title="طلب صيانة جهاز العرض",
            body="فحص الجهاز قبل بداية الحصة.",
        )
        Notification.objects.create(
            school=self.school,
            created_by=None,
            title="تنبيه آلي للأرشيف",
            message="تنبيه أنشأه النظام تلقائيًا.",
        )
        Notification.objects.create(
            school=self.school,
            created_by=self.manager,
            title="إشعار جدول الاختبارات",
            message="اعتماد الجدول النهائي.",
        )
        expected_notifications = Notification.objects.filter(
            school=self.school,
            requires_signature=False,
        ).count()
        expected_system_notifications = Notification.objects.filter(
            school=self.school,
            requires_signature=False,
            created_by__isnull=True,
        ).count()
        expected_snapshot_total = (
            Report.objects.filter(school=self.school, academic_year="1447-1448").count()
            + Ticket.objects.filter(school=self.school).count()
            + expected_notifications
        )

        response = self.client.get(
            reverse("reports:school_archive"),
            {"year": "1447-1448", "q": "آلي"},
        )

        self.assertEqual(response.status_code, 200)
        self.assertContains(response, "بحث في جميع سجلات المدرسة")
        self.assertContains(response, "السجلات المدرسية حتى لحظة إنشاء النسخة")
        self.assertContains(response, "تنبيه آلي للأرشيف")
        self.assertContains(response, "النظام — آلي")
        self.assertContains(
            response,
            f"{expected_notifications} سجل · {expected_system_notifications} آلي",
        )
        self.assertContains(
            response,
            f"إنشاء نسخة ثابتة الآن ({expected_snapshot_total} سجل)",
        )
        self.assertEqual(
            response.context["administrative_stats"]["notifications"],
            expected_notifications,
        )
        self.assertEqual(
            response.context["administrative_stats"]["system_notifications"],
            expected_system_notifications,
        )
        self.assertEqual(response.context["administrative_matches"]["notifications"], 1)
        self.assertEqual(response.context["snapshot_total_records"], expected_snapshot_total)

    def test_year_index_contains_school_directory_and_complete_record_metadata(self):
        department = Department.objects.create(
            school=self.school,
            name="قسم الجودة",
            slug="quality",
            role_label="مسؤول الجودة",
        )
        DepartmentMembership.objects.create(
            department=department,
            teacher=self.teacher,
        )
        achievement_file = TeacherAchievementFile.objects.create(
            school=self.school,
            teacher=self.teacher,
            academic_year="1447-1448",
        )
        achievement_section = AchievementSection.objects.create(
            file=achievement_file,
            code=AchievementSection.Code.SECTION_1,
        )
        frozen_evidence = AchievementEvidenceReport.objects.create(
            section=achievement_section,
            report=self.report,
            frozen_at=timezone.now(),
            frozen_data={"title": "تقرير شاهد مجمد"},
            archived_image1=SimpleUploadedFile(
                "frozen-report.png",
                bytes.fromhex(
                    "89504e470d0a1a0a0000000d49484452000000010000000108060000001f15c4"
                    "890000000a49444154789c6360000002000154a24f6f0000000049454e44ae426082"
                ),
                content_type="image/png",
            ),
        )
        ticket = Ticket.objects.create(
            school=self.school,
            creator=self.teacher,
            assignee=self.manager,
            department=department,
            title="طلب توثيق زيارة",
            body="توثيق كامل لزيارة فريق الجودة.",
        )
        TicketNote.objects.create(
            ticket=ticket,
            author=self.manager,
            body="ملاحظة داخلية للتدقيق.",
            is_public=False,
        )
        circular = Notification.objects.create(
            school=self.school,
            created_by=self.manager,
            title="تعميم الجودة",
            message="الالتزام بخطة الجودة.",
            requires_signature=True,
            is_important=True,
            is_broadcast=True,
            signature_deadline_at=timezone.now() + timedelta(days=3),
        )
        NotificationRecipient.objects.create(
            notification=circular,
            teacher=self.teacher,
            is_read=True,
            read_at=timezone.now(),
            is_signed=True,
            signed_at=timezone.now(),
        )

        index_bytes = build_year_archive_index_bytes(
            self.school,
            "1447-1448",
            teacher=self.manager,
            school_wide=True,
        )
        workbook = load_workbook(BytesIO(index_bytes), read_only=True, data_only=True)
        self.addCleanup(workbook.close)

        self.assertTrue(
            {
                "ملخص النسخة",
                "التقارير",
                "ملفات الإنجاز",
                "شواهد الإنجاز",
                "الطلبات والتذاكر",
                "التعاميم والإشعارات",
                "فريق المدرسة",
                "الأقسام",
            }.issubset(set(workbook.sheetnames))
        )
        team_rows = list(workbook["فريق المدرسة"].iter_rows(values_only=True))
        self.assertTrue(any(row[0] == self.manager.name and row[1] == self.manager.phone for row in team_rows[1:]))
        department_rows = list(workbook["الأقسام"].iter_rows(values_only=True))
        self.assertTrue(any(row[0] == "قسم الجودة" and row[3] == 1 for row in department_rows[1:]))
        evidence_rows = list(workbook["شواهد الإنجاز"].iter_rows(values_only=True))
        self.assertTrue(
            any(
                row[3] == "تقرير شاهد مجمد"
                and row[4] == frozen_evidence.report_id
                and "تقرير شاهد مجمد" in (row[7] or "")
                for row in evidence_rows[1:]
            )
        )

        ticket_rows = list(workbook["الطلبات والتذاكر"].iter_rows(values_only=True))
        ticket_headers = list(ticket_rows[0])
        ticket_record = dict(zip(ticket_headers, next(row for row in ticket_rows[1:] if row[0] == ticket.id), strict=True))
        self.assertEqual(ticket_record["التفاصيل"], ticket.body)
        self.assertEqual(ticket_record["المسؤول الرئيسي"], self.manager.name)
        self.assertEqual(ticket_record["عدد الملاحظات"], 1)

        notification_rows = list(workbook["التعاميم والإشعارات"].iter_rows(values_only=True))
        notification_headers = list(notification_rows[0])
        notification_record = dict(
            zip(
                notification_headers,
                next(row for row in notification_rows[1:] if row[0] == circular.id),
                strict=True,
            )
        )
        self.assertEqual(notification_record["النص"], circular.message)
        self.assertEqual(notification_record["مهم"], "نعم")
        self.assertEqual(notification_record["تمت القراءة"], 1)
        self.assertEqual(notification_record["تم التوقيع"], 1)

        with (
            patch(
                "reports.pdf_report.generate_report_pdf",
                return_value=(b"%PDF-report", "report.pdf"),
            ),
            patch(
                "reports.pdf_achievement.generate_achievement_pdf",
                return_value=(b"%PDF-achievement", "achievement.pdf"),
            ),
            patch(
                "reports.pdf_archive_records.generate_ticket_archive_pdf",
                return_value=b"%PDF-ticket",
            ),
            patch(
                "reports.pdf_archive_records.generate_notification_archive_pdf",
                return_value=b"%PDF-notification",
            ),
        ):
            package, _metadata = build_school_export_zip_file(
                self.school,
                academic_year="1447-1448",
                teacher=self.manager,
                school_wide=True,
                return_metadata=True,
            )
        try:
            with zipfile.ZipFile(package) as zipped:
                self.assertTrue(
                    any(
                        name.endswith("/صورة-مؤرشفة-1.png")
                        for name in zipped.namelist()
                    )
                )
        finally:
            package.close()

    def test_archive_record_pdf_template_shows_audit_metadata(self):
        ticket = Ticket.objects.create(
            school=self.school,
            creator=self.teacher,
            title="طلب تدقيق",
            body="تفاصيل طلب التدقيق.",
        )
        note = TicketNote.objects.create(
            ticket=ticket,
            author=self.manager,
            body="ملاحظة داخلية.",
            is_public=False,
        )
        ticket_html = render_to_string(
            "reports/archive_record_pdf.html",
            {
                "record_kind": "ticket",
                "record": ticket,
                "school": self.school,
                "recipients": [],
                "notes": [note],
                "generated_at": timezone.now(),
            },
        )
        self.assertIn("آخر تحديث", ticket_html)
        self.assertIn("داخلية", ticket_html)
        self.assertIn("عدد الصور", ticket_html)

        circular = Notification.objects.create(
            school=self.school,
            created_by=None,
            title="تعميم آلي",
            message="نص التعميم الآلي.",
            requires_signature=True,
            is_important=True,
            signature_deadline_at=timezone.now() + timedelta(days=2),
        )
        recipient = NotificationRecipient.objects.create(
            notification=circular,
            teacher=self.teacher,
            is_read=True,
            read_at=timezone.now(),
            signature_attempt_count=2,
            signature_last_attempt_at=timezone.now(),
        )
        circular_html = render_to_string(
            "reports/archive_record_pdf.html",
            {
                "record_kind": "circular",
                "record": circular,
                "school": self.school,
                "recipient_rows": [recipient],
                "signed_count": 0,
                "read_count": 1,
                "generated_at": timezone.now(),
            },
        )
        self.assertIn("إشعار آلي من النظام", circular_html)
        self.assertIn("وقت القراءة", circular_html)
        self.assertIn("محاولات التوقيع", circular_html)
        self.assertIn(">2<", circular_html)

    def test_inactive_archive_page_has_direct_activation_cta_and_price(self):
        self.addon.is_enabled = False
        self.addon.save(update_fields=["is_enabled"])

        response = self.client.get(reverse("reports:school_archive"))

        self.assertEqual(response.status_code, 200)
        self.assertContains(response, "طلب التفعيل الآن")
        self.assertContains(response, "سنويًا")
        self.assertContains(response, reverse("reports:my_subscription") + "#archiveOrder")

    def test_non_manager_cannot_create_school_snapshot(self):
        self.client.force_login(self.teacher)
        session = self.client.session
        session["active_school_id"] = self.school.pk
        session.save()

        response = self.client.post(
            reverse("reports:school_archive_create"),
            {"year": "1447-1448"},
        )

        self.assertEqual(response.status_code, 302)
        self.assertFalse(SchoolYearArchive.objects.exists())


@override_settings(ALLOWED_HOSTS=["testserver"])
class ArchivePaymentIdempotencyTests(TestCase):
    def test_reapproving_same_archive_payment_does_not_extend_twice(self):
        school = School.objects.create(name="مدرسة دفعة الأرشيف", code="archive-payment-once")
        manager = Teacher.objects.create_user(
            phone="0500700010",
            name="مدير الدفعة",
            password="manager-password",
        )
        SchoolMembership.objects.create(
            school=school,
            teacher=manager,
            role_type=SchoolMembership.RoleType.MANAGER,
        )
        payment = Payment.objects.create(
            school=school,
            purpose=Payment.Purpose.ARCHIVE_ADDON,
            amount=399,
            created_by=manager,
        )
        admin = Teacher.objects.create_superuser(
            phone="0500700011",
            name="مدير المنصة",
            password="admin-password",
        )
        self.client.force_login(admin)

        self.client.post(
            reverse("reports:platform_payment_detail", args=[payment.pk]),
            {"status": Payment.Status.APPROVED, "notes": ""},
        )
        addon = SchoolArchiveAddon.objects.get(school=school)
        first_end_date = addon.end_date
        payment.refresh_from_db()
        self.assertIsNotNone(payment.effects_applied_at)

        self.client.post(
            reverse("reports:platform_payment_detail", args=[payment.pk]),
            {"status": Payment.Status.REJECTED, "notes": ""},
        )
        self.client.post(
            reverse("reports:platform_payment_detail", args=[payment.pk]),
            {"status": Payment.Status.APPROVED, "notes": ""},
        )

        addon.refresh_from_db()
        self.assertEqual(addon.end_date, first_end_date)
