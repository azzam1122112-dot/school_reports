# reports/views/teachers.py
# -*- coding: utf-8 -*-
from __future__ import annotations

from ._helpers import *
from ._helpers import (
    _safe_next_url, _model_has_field,
    _get_active_school, _user_manager_schools,
)
from ..permissions import effective_user_role_label, is_school_manager
from ..gender_labels import school_gender_labels


def _decorate_manage_teacher_rows(teachers, *, active_school: Optional[School]) -> None:
    """Attach display-only labels that come from memberships, not legacy Role."""
    for teacher in teachers:
        role_label = effective_user_role_label(teacher, active_school=active_school)
        role_kind = ""

        try:
            is_manager_here = bool(getattr(teacher, "is_school_manager_in_active_school", False))
            if active_school is None:
                is_manager_here = is_school_manager(teacher)
        except Exception:
            is_manager_here = False

        job_title = (getattr(teacher, "school_job_title", "") or "").strip()
        if is_manager_here:
            role_kind = "manager"
        elif bool(getattr(teacher, "is_report_viewer", False)):
            role_kind = "report_viewer"
            role_label = "مشرف تقارير"
        elif job_title == SchoolMembership.JobTitle.ADMIN_STAFF:
            role_kind = "admin_staff"
        elif job_title == SchoolMembership.JobTitle.LAB_TECH:
            role_kind = "lab_tech"
        elif bool(getattr(teacher, "has_teacher_membership", False)):
            role_kind = "teacher"
        elif role_label and role_label != "مستخدم":
            role_kind = "other"

        teacher.manage_role_kind = role_kind
        teacher.manage_role_label = "" if role_label == "مستخدم" else role_label

# =========================
# إدارة المعلّمين (مدير فقط)
# =========================
@login_required(login_url="reports:login")
@role_required({"manager"})  # إن كنت تبغى السماح للسوبر دائمًا، خلي role_required يتجاوز للسوبر أو أضف دور admin
@require_http_methods(["GET"])
def manage_teachers(request: HttpRequest) -> HttpResponse:
    active_school = _get_active_school(request)

    # ✅ اجبار اختيار مدرسة لغير السوبر (أوضح وأأمن)
    if not request.user.is_superuser:
        if active_school is None:
            messages.error(request, "فضلاً اختر مدرسة أولاً.")
            return redirect("reports:select_school")

        if active_school not in _user_manager_schools(request.user):
            messages.error(request, "ليست لديك صلاحية على هذه المدرسة.")
            return redirect("reports:select_school")

    term = (request.GET.get("q") or "").strip()
    status_filter = (request.GET.get("status") or "").strip()
    job_title_filter = (request.GET.get("job_title") or "").strip()
    department_filter = (request.GET.get("department") or "").strip()

    qs = Teacher.objects.order_by("-id")

    # ✅ عزل حسب المدرسة (نُظهر المعلمين + مشرفي التقارير المرتبطين بالمدرسة)
    if active_school is not None:
        qs = qs.filter(
            school_memberships__school=active_school,
            school_memberships__role_type__in=[
                SchoolMembership.RoleType.TEACHER,
                SchoolMembership.RoleType.REPORT_VIEWER,
            ],
        ).distinct()

    # ✅ بحث
    if term:
        qs = qs.filter(
            Q(name__icontains=term) |
            Q(phone__icontains=term) |
            Q(national_id__icontains=term)
        )
    if status_filter == "active":
        qs = qs.filter(is_active=True)
    elif status_filter == "inactive":
        qs = qs.filter(is_active=False)
    if active_school is not None and job_title_filter in SchoolMembership.JobTitle.values:
        qs = qs.filter(
            school_memberships__school=active_school,
            school_memberships__role_type=SchoolMembership.RoleType.TEACHER,
            school_memberships__job_title=job_title_filter,
        )
    if active_school is not None and department_filter.isdigit():
        qs = qs.filter(
            dept_memberships__department_id=int(department_filter),
            dept_memberships__department__school=active_school,
        )
    qs = qs.distinct()

    # ✅ تمييز العضوية الحالية داخل المدرسة النشطة
    if active_school is not None:
        try:
            teacher_m = SchoolMembership.objects.filter(
                school=active_school,
                teacher=OuterRef("pk"),
                role_type=SchoolMembership.RoleType.TEACHER,
                is_active=True,
            )
            title_sq = (
                teacher_m
                .values("job_title")[:1]
            )
            viewer_m = SchoolMembership.objects.filter(
                school=active_school,
                teacher=OuterRef("pk"),
                role_type=SchoolMembership.RoleType.REPORT_VIEWER,
                is_active=True,
            )
            manager_m = SchoolMembership.objects.filter(
                school=active_school,
                teacher=OuterRef("pk"),
                role_type=SchoolMembership.RoleType.MANAGER,
                is_active=True,
            )
            qs = qs.annotate(
                has_teacher_membership=Exists(teacher_m),
                is_report_viewer=Exists(viewer_m),
                is_school_manager_in_active_school=Exists(manager_m),
                school_job_title=Subquery(title_sq),
            )
        except Exception:
            pass

    # ✅ منع N+1: Prefetch عضويات الأقسام مرة واحدة وبحقول أقل
    if DepartmentMembership is not None:
        dm_qs = (
            DepartmentMembership.objects
            .select_related("department")
            .only("id", "teacher_id", "role_type", "department__id", "department__name", "department__slug")
            .order_by("department__name")
        )
        if active_school is not None and _model_has_field(Department, "school"):
            dm_qs = dm_qs.filter(Q(department__school=active_school) | Q(department__school__isnull=True))

        qs = qs.prefetch_related(Prefetch("dept_memberships", queryset=dm_qs))

    page = Paginator(qs, 20).get_page(request.GET.get("page"))
    _decorate_manage_teacher_rows(page.object_list, active_school=active_school)
    departments = (
        Department.objects.filter(school=active_school, is_active=True).order_by("name", "id")
        if active_school is not None
        else Department.objects.none()
    )
    current_count = (
        SchoolMembership.objects.filter(
            school=active_school,
            role_type=SchoolMembership.RoleType.TEACHER,
        ).count()
        if active_school is not None
        else 0
    )
    plan = getattr(getattr(active_school, "subscription", None), "plan", None)
    maximum = int(getattr(plan, "max_teachers", 0) or 0)
    filters_query = request.GET.copy()
    filters_query.pop("page", None)
    return render(
        request,
        "reports/manage_teachers.html",
        {
            "teachers_page": page,
            "term": term,
            "status_filter": status_filter,
            "job_title_filter": job_title_filter,
            "department_filter": department_filter,
            "departments": departments,
            "job_title_choices": SchoolMembership.JobTitle.choices,
            "capacity": {
                "current": current_count,
                "maximum": maximum,
                "remaining": max(maximum - current_count, 0) if maximum else None,
            },
            "filters_query": filters_query.urlencode(),
        },
    )

@login_required(login_url="reports:login")
@role_required({"manager"})
@ratelimit(key="user", rate="5/h", method="POST", block=True)
@require_http_methods(["GET", "POST"])
def _legacy_bulk_import_teachers(request: HttpRequest) -> HttpResponse:
    active_school = _get_active_school(request)
    labels = school_gender_labels(active_school)
    if active_school is None:
        messages.error(request, "فضلاً اختر مدرسة أولاً.")
        return redirect("reports:select_school")

    # Defense-in-depth: تأكد أن المدير يملك صلاحية على المدرسة النشطة
    try:
        if (not request.user.is_superuser) and active_school not in _user_manager_schools(request.user):
            messages.error(request, "ليست لديك صلاحية على هذه المدرسة.")
            return redirect("reports:select_school")
    except Exception:
        pass

    # الاستيراد يُنشئ عضويات TEACHER؛ نتحقق من وجود اشتراك فعّال لتجنب ValidationError العام
    sub = getattr(active_school, "subscription", None)
    try:
        if sub is None or bool(getattr(sub, "is_expired", True)):
            messages.error(request, "لا يوجد اشتراك فعّال لهذه المدرسة.")
            return redirect("reports:my_subscription")
    except Exception:
        messages.error(request, "لا يوجد اشتراك فعّال لهذه المدرسة.")
        return redirect("reports:my_subscription")

    if request.method == "POST":
        excel_file = request.FILES.get("excel_file")
        if not excel_file:
            messages.error(request, "الرجاء اختيار ملف للاستيراد.")
            return render(request, "reports/bulk_import_teachers.html")

        # دعم صيغتين: Excel (.xlsx) و CSV (.csv)
        fname = (getattr(excel_file, "name", "") or "").lower()
        is_csv = fname.endswith(".csv")
        is_xlsx = fname.endswith(".xlsx")
        if not (is_csv or is_xlsx):
            messages.error(request, "صيغة غير مدعومة. الرجاء اختيار ملف بصيغة .xlsx أو .csv")
            return render(request, "reports/bulk_import_teachers.html")

        try:
            import re
            from django.core.exceptions import ValidationError

            def _norm_str(v) -> str:
                return (str(v).strip() if v is not None else "").strip()

            def _normalize_phone(v) -> str:
                if v is None:
                    return ""
                # openpyxl يعيد int/float للأرقام
                try:
                    if isinstance(v, bool):
                        return ""
                    if isinstance(v, int):
                        s = str(v)
                    elif isinstance(v, float):
                        s = str(int(v)) if float(v).is_integer() else str(v)
                    else:
                        s = str(v)
                except Exception:
                    s = str(v)
                s = s.strip()
                # إزالة المسافات والرموز الشائعة (نحتفظ بالأرقام فقط)
                digits = re.sub(r"\D+", "", s)
                if not digits:
                    return s

                # تطبيع أرقام الجوال الشائعة (السعودية)
                # - 9665XXXXXXXX  -> 05XXXXXXXX
                # - 5XXXXXXXX     -> 05XXXXXXXX
                try:
                    if digits.startswith("966"):
                        digits = digits[3:]
                    if digits.startswith("5") and len(digits) == 9:
                        digits = "0" + digits
                    if digits.startswith("5") and len(digits) == 10:
                        digits = "0" + digits[-9:]
                except Exception:
                    pass
                return digits

            def _normalize_national_id(v) -> str:
                s = _norm_str(v)
                if not s:
                    return ""
                digits = re.sub(r"\D+", "", s)
                return digits or s

            # قراءة موحّدة للصفوف من xlsx أو csv
            if is_xlsx:
                wb = openpyxl.load_workbook(excel_file, read_only=True, data_only=True)
                sheet = wb.active
                all_rows = list(sheet.iter_rows(values_only=True))
            else:
                import csv as _csv
                import io as _io

                raw = excel_file.read()
                if isinstance(raw, bytes):
                    # utf-8-sig يتعامل مع BOM الذي يضيفه Excel عند حفظ CSV
                    text = raw.decode("utf-8-sig", errors="replace")
                else:
                    text = str(raw)
                all_rows = [tuple(r) for r in _csv.reader(_io.StringIO(text))]

            def _norm_header(v) -> str:
                s = _norm_str(v).lower()
                s = re.sub(r"\s+", "", s)
                s = re.sub(r"[\-_/\\]+", "", s)
                return s

            header_row = (all_rows[0] if all_rows else ()) or ()
            data_rows = all_rows[1:] if len(all_rows) > 1 else []
            header_norm = [_norm_header(h) for h in (header_row or ())]

            def _find_col_idx(candidates: tuple[str, ...]) -> int | None:
                for i, h in enumerate(header_norm):
                    if not h:
                        continue
                    for c in candidates:
                        if c and c in h:
                            return i
                return None

            # نحدد الأعمدة حسب العناوين لتفادي ملفات فيها أعمدة فارغة/غير متجاورة
            name_idx = _find_col_idx(("الاسمالكامل", "اسم", "الاسم"))
            phone_idx = _find_col_idx(("رقمالجوال", "الجوال", "رقمالهاتف", "الهاتف"))
            nat_idx = _find_col_idx(("رقمالهوية", "الهوية", "السجلالمدني", "رقمالسجل"))

            # توقع الأعمدة: الاسم، رقم الجوال، رقم الهوية (اختياري)
            # الصف الأول عناوين
            parsed_rows: list[tuple[int, str, str, str | None]] = []
            phones_in_file: set[str] = set()
            nat_ids_in_file: set[str] = set()

            max_rows_guard = 2000
            for row_idx, row in enumerate(data_rows, start=2):
                if len(parsed_rows) >= max_rows_guard:
                    messages.error(request, f"الملف يحتوي على عدد كبير من الصفوف (>{max_rows_guard}). الرجاء تقسيم الملف.")
                    return render(request, "reports/bulk_import_teachers.html")

                row = row or ()
                # إن استطعنا تحديد الأعمدة من العناوين: نقرأ حسب الفهارس
                if name_idx is not None or phone_idx is not None or nat_idx is not None:
                    name = row[name_idx] if name_idx is not None and name_idx < len(row) else None
                    phone = row[phone_idx] if phone_idx is not None and phone_idx < len(row) else None
                    national_id = row[nat_idx] if nat_idx is not None and nat_idx < len(row) else None
                else:
                    # fallback للملفات التي بلا عناوين واضحة
                    name, phone, national_id = (row + (None, None, None))[:3]

                name_s = _norm_str(name)
                phone_s = _normalize_phone(phone)
                nat_s = _normalize_national_id(national_id) or None

                if nat_s:
                    nat_ids_in_file.add(nat_s)

                if not name_s or not phone_s:
                    # نؤجل الأخطاء إلى مرحلة الرسائل حتى لا نقطع المعالجة مبكرًا
                    parsed_rows.append((row_idx, name_s, phone_s, nat_s))
                    continue

                if phone_s in phones_in_file:
                    parsed_rows.append((row_idx, name_s, phone_s, nat_s))
                    continue
                phones_in_file.add(phone_s)
                parsed_rows.append((row_idx, name_s, phone_s, nat_s))

            if not parsed_rows:
                messages.error(request, "الملف فارغ أو لا يحتوي على بيانات.")
                return render(request, "reports/bulk_import_teachers.html")

            # التحقق من حد الباقة (نحسب فقط العضويات الجديدة الفعلية)
            max_teachers = int(getattr(getattr(sub, "plan", None), "max_teachers", 0) or 0)
            current_count = SchoolMembership.objects.filter(
                school=active_school,
                role_type=SchoolMembership.RoleType.TEACHER,
            ).count()

            phones_unique = {p for p in phones_in_file if p}
            existing_phones_in_school: set[str] = set()
            if phones_unique:
                existing_phones_in_school = set(
                    SchoolMembership.objects.filter(
                        school=active_school,
                        role_type=SchoolMembership.RoleType.TEACHER,
                        teacher__phone__in=phones_unique,
                    ).values_list("teacher__phone", flat=True)
                )

            expected_new = len([p for p in phones_unique if p not in existing_phones_in_school])
            if max_teachers > 0 and (current_count + expected_new) > max_teachers:
                remaining = max_teachers - current_count
                messages.error(request, f"لا يمكن استيراد {expected_new} حسابات جديدة. الحد المتبقي في باقتك هو {remaining}.")
                return render(request, "reports/bulk_import_teachers.html")

            created_count = 0
            updated_count = 0
            reactivated_count = 0
            errors: list[str] = []
            seen_phone_rows: set[str] = set()

            # ملاحظة مهمة: أي IntegrityError داخل atomic قد يكسر المعاملة في Postgres.
            # لذلك نستخدم savepoint لكل صف حتى نستمر في الصفوف التالية.
            for row_idx, name_s, phone_s, nat_s in parsed_rows:
                with transaction.atomic():
                    if not name_s or not phone_s:
                        errors.append(f"الصف {row_idx}: الاسم ورقم الجوال مطلوبان.")
                        continue

                    if phone_s in seen_phone_rows:
                        # نعتبره تكرار داخل الملف ونتجاهله بدون تحذير (سلوك متوقع حسب الطلب)
                        continue
                    seen_phone_rows.add(phone_s)

                    # Upsert: تحديد المعلم الموجود ثم تحديث بياناته عند اللزوم
                    teacher = Teacher.objects.filter(phone=phone_s).first()
                    if teacher is None and nat_s:
                        teacher = Teacher.objects.filter(national_id=nat_s).first()

                    if teacher is None:
                        try:
                            teacher = Teacher.objects.create(
                                name=name_s,
                                phone=phone_s,
                                national_id=nat_s,
                                password=make_password(phone_s),  # كلمة المرور الافتراضية هي رقم الجوال
                            )
                            created_count += 1
                        except (IntegrityError, ValidationError):
                            errors.append(f"الصف {row_idx}: تعذّر إنشاء المستخدم بسبب تعارض في رقم الجوال/الهوية.")
                            continue
                    else:
                        changed_fields: list[str] = []

                        # ✅ تحديث الاسم إذا اختلف
                        try:
                            if name_s and (getattr(teacher, "name", "") or "").strip() != name_s:
                                teacher.name = name_s
                                changed_fields.append("name")
                        except Exception:
                            pass

                        # ✅ تحديث رقم الهوية (إن وُجد)
                        if nat_s:
                            try:
                                current_nat = (getattr(teacher, "national_id", None) or "").strip() or None
                                if current_nat != nat_s:
                                    # تأكد أن الهوية ليست مرتبطة بمستخدم آخر
                                    nat_owner = Teacher.objects.filter(national_id=nat_s).exclude(pk=teacher.pk).first()
                                    if nat_owner is None:
                                        teacher.national_id = nat_s
                                        changed_fields.append("national_id")
                                    else:
                                        errors.append(
                                            f"الصف {row_idx}: رقم الهوية مرتبط بمستخدم آخر، لا يمكن تحديثه تلقائياً."
                                        )
                                        continue
                            except Exception:
                                pass

                        # ✅ تحديث رقم الجوال إذا تم العثور على المعلم عبر الهوية (أو اختلاف الجوال)
                        try:
                            current_phone = (getattr(teacher, "phone", "") or "").strip()
                            if phone_s and current_phone != phone_s:
                                phone_owner = Teacher.objects.filter(phone=phone_s).exclude(pk=teacher.pk).first()
                                if phone_owner is None:
                                    teacher.phone = phone_s
                                    changed_fields.append("phone")
                                    # لو تغير الجوال نحدّث كلمة المرور الافتراضية لتبقى متوافقة (اختياري)
                                    try:
                                        teacher.password = make_password(phone_s)
                                        changed_fields.append("password")
                                    except Exception:
                                        pass
                                else:
                                    errors.append(
                                        f"الصف {row_idx}: رقم الجوال مرتبط بمستخدم آخر، لا يمكن تحديثه تلقائياً."
                                    )
                                    continue
                        except Exception:
                            pass

                        if changed_fields:
                            try:
                                # حفظ الحقول التي تغيرت فقط
                                teacher.save(update_fields=list(dict.fromkeys(changed_fields)))
                                updated_count += 1
                            except Exception:
                                # إن فشل التحديث لسبب غير متوقع، نعتبره خطأ صف ونكمل
                                errors.append(f"الصف {row_idx}: تعذّر تحديث بيانات المستخدم.")
                                continue

                    # ربط المعلم بالمدرسة
                    try:
                        membership, created = SchoolMembership.objects.get_or_create(
                            school=active_school,
                            teacher=teacher,
                            role_type=SchoolMembership.RoleType.TEACHER,
                            defaults={
                                "is_active": True,
                            },
                        )
                    except ValidationError as ve:
                        msg = " ".join(getattr(ve, "messages", []) or []) or str(ve)
                        errors.append(f"الصف {row_idx}: {msg}")
                        continue

                    if not created:
                        # إن كانت العضوية موجودة لكنها غير نشطة، فعّلها
                        try:
                            if hasattr(membership, "is_active") and not bool(getattr(membership, "is_active", True)):
                                membership.is_active = True
                                membership.save(update_fields=["is_active"])
                                reactivated_count += 1
                        except Exception:
                            pass

            if created_count > 0:
                messages.success(request, f"✅ تم إنشاء {created_count} من حسابات {labels['teachers_object']}.")
            if updated_count > 0:
                messages.info(request, f"تم تحديث بيانات {updated_count} من حسابات {labels['teachers_object']}.")
            if reactivated_count > 0:
                messages.info(request, f"تم تفعيل {reactivated_count} عضوية موجودة سابقاً.")
            if errors:
                for err in errors[:10]:
                    messages.warning(request, err)
                if len(errors) > 10:
                    messages.warning(request, f"... وهناك {len(errors)-10} أخطاء أخرى.")

            return redirect("reports:manage_teachers")

        except Exception:
            logger.exception("Bulk import failed")
            messages.error(request, "تعذّر معالجة الملف. تأكد أنه ملف .xlsx صحيح ومطابق للتعليمات.")

    return render(request, "reports/bulk_import_teachers.html")


@login_required(login_url="reports:login")
@role_required({"manager"})
@ratelimit(key="user", rate="30/h", method="POST", block=True)
@require_http_methods(["GET", "POST"])
def teacher_onboarding(request: HttpRequest) -> HttpResponse:
    """Unified, preview-first onboarding for one or many school users."""
    from ..teacher_onboarding import (
        RESULT_SESSION_KEY,
        build_preview,
        clear_preview,
        confirm_preview,
        load_preview,
        load_result,
        rows_from_quick_post,
        rows_from_uploaded_file,
        save_preview,
    )

    active_school = _get_active_school(request)
    if active_school is None:
        messages.error(request, "فضلاً اختر مدرسة أولاً.")
        return redirect("reports:select_school")
    if (not request.user.is_superuser) and active_school not in _user_manager_schools(request.user):
        messages.error(request, "ليست لديك صلاحية على هذه المدرسة.")
        return redirect("reports:select_school")

    subscription = getattr(active_school, "subscription", None)
    if subscription is None or bool(getattr(subscription, "is_expired", True)):
        messages.error(request, "لا يوجد اشتراك فعّال لهذه المدرسة.")
        return redirect("reports:my_subscription")

    def posted_quick_rows() -> list[dict[str, str]]:
        fields = {
            "name": request.POST.getlist("name"),
            "phone": request.POST.getlist("phone"),
            "national_id": request.POST.getlist("national_id"),
            "job_title": request.POST.getlist("job_title"),
            "department_id": request.POST.getlist("department"),
        }
        count = max((len(values) for values in fields.values()), default=0)
        return [
            {
                field: values[index] if index < len(values) else ""
                for field, values in fields.items()
            }
            for index in range(count)
        ]

    quick_rows: list[dict[str, Any]] | None = None
    if request.method == "POST":
        action = (request.POST.get("action") or "").strip()
        remove_row = (request.POST.get("remove_row") or "").strip()

        if action == "add_quick_row":
            quick_rows = posted_quick_rows()
            if len(quick_rows) >= 2000:
                messages.error(request, "وصلت إلى الحد الأقصى وهو 2000 صف.")
            else:
                quick_rows.append(
                    {
                        "name": "",
                        "phone": "",
                        "national_id": "",
                        "job_title": SchoolMembership.JobTitle.TEACHER,
                        "department_id": "",
                    }
                )

        elif remove_row:
            quick_rows = posted_quick_rows()
            try:
                remove_index = int(remove_row)
            except (TypeError, ValueError):
                remove_index = -1
            if len(quick_rows) > 1 and 0 <= remove_index < len(quick_rows):
                quick_rows.pop(remove_index)
            elif quick_rows:
                quick_rows[0] = {
                    "name": "",
                    "phone": "",
                    "national_id": "",
                    "job_title": SchoolMembership.JobTitle.TEACHER,
                    "department_id": "",
                }

        elif action == "cancel":
            clear_preview(request)
            request.session.pop(RESULT_SESSION_KEY, None)
            messages.info(request, "تم إلغاء العملية دون حفظ أي بيانات.")
            return redirect("reports:bulk_import_teachers")

        elif action == "quick_preview":
            try:
                rows = rows_from_quick_post(request.POST)
                if not rows:
                    messages.error(request, "أدخل مستخدمًا واحدًا على الأقل قبل المتابعة.")
                else:
                    preview = build_preview(rows, active_school)
                    save_preview(request, preview, active_school, source="quick")
                    return redirect(f"{reverse('reports:bulk_import_teachers')}?step=preview")
            except ValueError as exc:
                if str(exc) == "too_many_rows":
                    messages.error(request, "القائمة تتجاوز الحد الأقصى وهو 2000 مستخدم.")
                else:
                    messages.error(request, "تعذر قراءة البيانات المدخلة. راجعها ثم حاول مجددًا.")

        elif action == "file_preview":
            uploaded_file = request.FILES.get("excel_file")
            if uploaded_file is None:
                messages.error(request, "اختر ملف Excel أو CSV أولاً.")
            else:
                try:
                    rows = rows_from_uploaded_file(uploaded_file)
                    if not rows:
                        messages.error(request, "الملف لا يحتوي على صفوف بيانات.")
                    else:
                        preview = build_preview(rows, active_school)
                        save_preview(request, preview, active_school, source="file")
                        return redirect(f"{reverse('reports:bulk_import_teachers')}?step=preview")
                except ValueError as exc:
                    errors = {
                        "unsupported_file": "صيغة الملف غير مدعومة. استخدم Excel أو CSV.",
                        "required_headers_missing": "يجب أن يحتوي الملف على عمودي الاسم الكامل ورقم الجوال.",
                        "too_many_rows": "الملف يتجاوز الحد الأقصى وهو 2000 صف.",
                        "file_too_large": "حجم الملف يتجاوز 10 ميجابايت. قسّمه إلى ملفات أصغر ثم حاول مجددًا.",
                    }
                    messages.error(request, errors.get(str(exc), "تعذر قراءة الملف. تأكد من سلامته ثم حاول مجددًا."))
                except Exception:
                    logger.exception("Teacher onboarding preview failed")
                    messages.error(request, "تعذر قراءة الملف. تأكد من سلامته ثم حاول مجددًا.")

        elif action == "confirm":
            try:
                result = confirm_preview(
                    request,
                    active_school,
                    request.POST.get("preview_token") or "",
                )
                messages.success(
                    request,
                    f"اكتملت العملية بنجاح لـ {result['total']} مستخدم.",
                )
                return redirect(f"{reverse('reports:bulk_import_teachers')}?completed=1")
            except ValueError as exc:
                if str(exc) == "preview_changed":
                    messages.warning(request, "تغيرت بعض البيانات منذ المعاينة. راجع النتائج المحدثة ثم أكد مرة أخرى.")
                    return redirect(f"{reverse('reports:bulk_import_teachers')}?step=preview")
                messages.error(request, "انتهت صلاحية المعاينة. أعد إدخال البيانات أو رفع الملف.")
            except Exception:
                logger.exception("Teacher onboarding confirmation failed")
                messages.error(request, "تعذر إتمام العملية ولم تُحفظ بيانات جزئية. حاول مرة أخرى.")

    preview = load_preview(request, active_school)
    result = load_result(request, active_school)
    plan = getattr(subscription, "plan", None)
    current_count = SchoolMembership.objects.filter(
        school=active_school,
        role_type=SchoolMembership.RoleType.TEACHER,
    ).count()
    maximum = int(getattr(plan, "max_teachers", 0) or 0)
    capacity = {
        "current": current_count,
        "maximum": maximum,
        "remaining": max(maximum - current_count, 0) if maximum else None,
    }
    departments = Department.objects.filter(
        school=active_school,
        is_active=True,
    ).order_by("name", "id")
    if quick_rows is None:
        if preview and preview.get("source") == "quick":
            quick_rows = list(preview.get("rows") or [])
        else:
            quick_rows = [
                {
                    "name": "",
                    "phone": "",
                    "national_id": "",
                    "job_title": SchoolMembership.JobTitle.TEACHER,
                    "department_id": "",
                }
                for _ in range(3)
            ]
    return render(
        request,
        "reports/bulk_import_teachers.html",
        {
            "preview": preview,
            "result": result,
            "capacity": capacity,
            "departments": departments,
            "job_title_choices": SchoolMembership.JobTitle.choices,
            "active_mode": (request.GET.get("mode") or "quick").strip(),
            "quick_rows": quick_rows,
        },
    )


@login_required(login_url="reports:login")
@role_required({"manager"})
@require_http_methods(["GET"])
def bulk_import_teachers_issues(request: HttpRequest) -> HttpResponse:
    from io import BytesIO

    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Font, PatternFill

    from ..teacher_onboarding import load_preview

    active_school = _get_active_school(request)
    if active_school is None:
        return redirect("reports:select_school")
    preview = load_preview(request, active_school)
    if preview is None:
        messages.error(request, "لا توجد معاينة متاحة لتنزيل ملاحظاتها.")
        return redirect("reports:bulk_import_teachers")

    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "ملاحظات الاستيراد"
    sheet.sheet_view.rightToLeft = True
    headers = ["الصف", "الاسم", "رقم الجوال", "الحالة", "الأخطاء", "التنبيهات"]
    sheet.append(headers)
    for cell in sheet[1]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor="006C35")
        cell.alignment = Alignment(horizontal="center")
    for row in preview.get("rows") or []:
        if row.get("errors") or row.get("warnings"):
            sheet.append(
                [
                    row.get("row_number"),
                    row.get("name"),
                    row.get("phone"),
                    row.get("state_label"),
                    " | ".join(row.get("errors") or []),
                    " | ".join(row.get("warnings") or []),
                ]
            )
    for column, width in {"A": 10, "B": 28, "C": 18, "D": 22, "E": 55, "F": 55}.items():
        sheet.column_dimensions[column].width = width
    buffer = BytesIO()
    workbook.save(buffer)
    response = HttpResponse(
        buffer.getvalue(),
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )
    response["Content-Disposition"] = 'attachment; filename="teacher-import-issues.xlsx"'
    return response


@login_required(login_url="reports:login")
@role_required({"manager"})
@require_http_methods(["GET"])
def bulk_import_teachers_result(request: HttpRequest) -> HttpResponse:
    from io import BytesIO

    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Font, PatternFill

    from ..teacher_onboarding import load_result

    active_school = _get_active_school(request)
    if active_school is None:
        return redirect("reports:select_school")
    result = load_result(request, active_school)
    if result is None or not result.get("created_rows"):
        messages.error(request, "لا توجد حسابات جديدة لتنزيل بيانات دخولها.")
        return redirect("reports:bulk_import_teachers")

    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "بيانات الدخول"
    sheet.sheet_view.rightToLeft = True
    sheet.append(["الاسم", "رقم الجوال", "كلمة المرور المؤقتة", "تعليمات"])
    for cell in sheet[1]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor="006C35")
        cell.alignment = Alignment(horizontal="center")
    for row in result["created_rows"]:
        sheet.append(
            [
                row["name"],
                row["phone"],
                row["temporary_password"],
                "يجب تغيير كلمة المرور عند تسجيل الدخول لأول مرة.",
            ]
        )
    for column, width in {"A": 30, "B": 18, "C": 22, "D": 52}.items():
        sheet.column_dimensions[column].width = width
    buffer = BytesIO()
    workbook.save(buffer)
    response = HttpResponse(
        buffer.getvalue(),
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )
    response["Content-Disposition"] = 'attachment; filename="new-teachers-login.xlsx"'
    return response


@login_required(login_url="reports:login")
@role_required({"manager"})
@require_http_methods(["GET"])
def bulk_import_teachers_template(request: HttpRequest) -> HttpResponse:
    """Download a clean template; examples live on a separate instructions sheet."""
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Font, PatternFill, Border, Side
    from openpyxl.worksheet.datavalidation import DataValidation

    active_school = _get_active_school(request)
    labels = school_gender_labels(active_school)
    sample_name = "نورة أحمد الغامدي" if labels["is_girls"] else "محمد أحمد الغامدي"
    wb = Workbook()
    ws = wb.active
    ws.title = str(labels["teachers"])
    ws.sheet_view.rightToLeft = True

    headers = ["الاسم الكامل", "رقم الجوال", "رقم الهوية", "المسمى الوظيفي", "القسم"]

    header_font = Font(name="Calibri", bold=True, color="FFFFFF", size=12)
    header_fill = PatternFill("solid", fgColor="006C35")
    center = Alignment(horizontal="center", vertical="center")
    right = Alignment(horizontal="right", vertical="center", readingOrder=2)
    thin = Side(style="thin", color="D6E4DC")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    for col, head in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col, value=head)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center
        cell.border = border

    ws.column_dimensions["A"].width = 30
    ws.column_dimensions["B"].width = 18
    ws.column_dimensions["C"].width = 18
    ws.column_dimensions["D"].width = 22
    ws.column_dimensions["E"].width = 24
    ws.freeze_panes = "A2"

    job_validation = DataValidation(
        type="list",
        formula1=f'"{labels["teacher_indefinite"]},{labels["admin_staff"]},{labels["lab_tech"]}"',
        allow_blank=True,
    )
    job_validation.error = "اختر مسمى وظيفيًا من القائمة."
    job_validation.errorTitle = "مسمى غير صحيح"
    ws.add_data_validation(job_validation)
    job_validation.add("D2:D2001")

    instructions = wb.create_sheet("التعليمات والأمثلة")
    instructions.sheet_view.rightToLeft = True
    instruction_rows = [
        ["الحقل", "هل هو مطلوب؟", "مثال", "ملاحظة"],
        ["الاسم الكامل", "مطلوب", sample_name, "يظهر بهذا الشكل داخل النظام."],
        ["رقم الجوال", "مطلوب", "0551234567", "اسم الدخول وكلمة المرور المؤقتة."],
        ["رقم الهوية", "اختياري", "1012345678", "10 أرقام عند إدخاله."],
        ["المسمى الوظيفي", "اختياري", labels["teacher_indefinite"], f"{labels['teacher_indefinite']} أو {labels['admin_staff']} أو {labels['lab_tech']}."],
        ["القسم", "اختياري", "النشاط الطلابي", "اكتب اسم القسم كما يظهر في النظام."],
        ["تنبيه", "", "", f"لا تنسخ صفوف الأمثلة إلى ورقة {labels['teachers']}."],
    ]
    for row in instruction_rows:
        instructions.append(row)
    for cell in instructions[1]:
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center
    for row in instructions.iter_rows(min_row=2):
        for cell in row:
            cell.alignment = right
    for column, width in {"A": 24, "B": 17, "C": 28, "D": 55}.items():
        instructions.column_dimensions[column].width = width

    if active_school is not None:
        department_names = list(
            Department.objects.filter(
                school=active_school,
                is_active=True,
            ).order_by("name").values_list("name", flat=True)
        )
        if department_names:
            instructions.append([])
            instructions.append(["الأقسام المتاحة في المدرسة"])
            instructions.cell(row=instructions.max_row, column=1).font = Font(bold=True, color="006C35")
            for name in department_names:
                instructions.append([name])

    from io import BytesIO
    buf = BytesIO()
    wb.save(buf)
    response = HttpResponse(
        buf.getvalue(),
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )
    response["Content-Disposition"] = 'attachment; filename="teachers-import-template.xlsx"'
    return response


@login_required(login_url="reports:login")
@role_required({"manager"})
@require_http_methods(["GET", "POST"])
def add_teacher(request: HttpRequest) -> HttpResponse:
    # كل معلم جديد يُربط تلقائياً بالمدرسة النشطة لهذا المدير
    active_school = _get_active_school(request)
    labels = school_gender_labels(active_school)
    if School.objects.filter(is_active=True).exists():
        if active_school is None:
            messages.error(request, "فضلاً اختر مدرسة أولاً.")
            return redirect("reports:select_school")
        if (not request.user.is_superuser) and active_school not in _user_manager_schools(request.user):
            messages.error(request, "ليست لديك صلاحية على هذه المدرسة.")
            return redirect("reports:select_school")

    if request.method == "POST":
        # إنشاء معلّم فقط: بدون قسم/بدون دور داخل قسم. التكاليف تتم من صفحة أعضاء القسم.
        form = TeacherCreateForm(request.POST, active_school=active_school)
        continue_adding = request.POST.get("save_and_add_another") == "1"
        job_title = None
        try:
            # يحدد المسمى الوظيفي داخل المدرسة (بنفس الصلاحيات)
            job_title = (request.POST.get("job_title") or "").strip() or None
        except Exception:
            job_title = None
        if job_title not in SchoolMembership.JobTitle.values:
            messages.error(request, "اختر مسمى وظيفيًا صحيحًا.")
            return render(request, "reports/add_teacher.html", {"form": form, "title": "إضافة مستخدم"})

        # ✅ إذا كان رقم الجوال موجودًا مسبقًا: لا ننشئ مستخدمًا جديدًا، بل نربطه بهذه المدرسة
        try:
            phone_raw = (request.POST.get("phone") or "").strip()
            existing_teacher = None
            if phone_raw:
                existing_teacher = Teacher.objects.filter(phone=phone_raw).first()
            if existing_teacher is not None and active_school is not None:
                if not existing_teacher.is_active:
                    messages.error(
                        request,
                        "الحساب الموجود موقوف على مستوى المنصة، ولا يمكن ربطه قبل إعادة تفعيله.",
                    )
                    return render(request, "reports/add_teacher.html", {"form": form, "title": "إضافة مستخدم"})
                # هل هو مرتبط فعلاً بهذه المدرسة كـ TEACHER؟
                already = SchoolMembership.objects.filter(
                    school=active_school,
                    teacher=existing_teacher,
                    role_type=SchoolMembership.RoleType.TEACHER,
                    is_active=True,
                ).exists()
                if already:
                    messages.info(request, "المستخدم مرتبط بالفعل بهذه المدرسة.")
                    next_url = _safe_next_url(request.POST.get("next") or request.GET.get("next"))
                    return redirect("reports:add_teacher" if continue_adding else (next_url or "reports:manage_teachers"))

                # نفس منطق حد الباقة الحالي (مع ترك الضمان النهائي للموديل)
                try:
                    sub = getattr(active_school, "subscription", None)
                    if sub is None or bool(getattr(sub, "is_expired", True)):
                        messages.error(request, "لا يوجد اشتراك فعّال لهذه المدرسة.")
                        return render(request, "reports/add_teacher.html", {"form": form, "title": "إضافة مستخدم"})

                    max_teachers = int(getattr(getattr(sub, "plan", None), "max_teachers", 0) or 0)
                    if max_teachers > 0:
                        current_count = SchoolMembership.objects.filter(
                            school=active_school,
                            role_type=SchoolMembership.RoleType.TEACHER,
                        ).count()
                        if current_count >= max_teachers:
                            messages.error(request, f"لا يمكن إضافة أكثر من {max_teachers} من حسابات {labels['teachers_object']} لهذه المدرسة حسب الباقة.")
                            return render(request, "reports/add_teacher.html", {"form": form, "title": "إضافة مستخدم"})
                except Exception:
                    pass

                try:
                    with transaction.atomic():
                        SchoolMembership.objects.update_or_create(
                            school=active_school,
                            teacher=existing_teacher,
                            role_type=SchoolMembership.RoleType.TEACHER,
                            defaults={
                                "is_active": True,
                                **({"job_title": job_title} if job_title else {}),
                            },
                        )
                    messages.success(request, "✅ تم ربط المستخدم الموجود بهذه المدرسة بنجاح (بدون إنشاء حساب جديد).")
                    next_url = _safe_next_url(request.POST.get("next") or request.GET.get("next"))
                    return redirect("reports:add_teacher" if continue_adding else (next_url or "reports:manage_teachers"))
                except ValidationError as e:
                    messages.error(request, " ".join(getattr(e, "messages", []) or [str(e)]))
                except Exception:
                    logger.exception("add_teacher link existing failed")
                    messages.error(request, "حدث خطأ غير متوقع أثناء الربط. جرّب لاحقًا.")
        except Exception:
            # لو فشل هذا المسار لأي سبب نكمل التدفق الطبيعي (وقد يظهر خطأ unique من الفورم)
            pass

        # ✅ منع إضافة معلّم إذا تجاوزت المدرسة حد الباقة (يشمل غير النشط)
        try:
            if active_school is not None:
                sub = getattr(active_school, "subscription", None)
                if sub is None or bool(getattr(sub, "is_expired", True)):
                    messages.error(request, "لا يوجد اشتراك فعّال لهذه المدرسة.")
                    return render(request, "reports/add_teacher.html", {"form": form, "title": "إضافة مستخدم"})

                max_teachers = int(getattr(getattr(sub, "plan", None), "max_teachers", 0) or 0)
                if max_teachers > 0:
                    current_count = SchoolMembership.objects.filter(
                        school=active_school,
                        role_type=SchoolMembership.RoleType.TEACHER,
                    ).count()
                    if current_count >= max_teachers:
                        messages.error(request, f"لا يمكن إضافة أكثر من {max_teachers} من حسابات {labels['teachers_object']} لهذه المدرسة حسب الباقة.")
                        return render(request, "reports/add_teacher.html", {"form": form, "title": "إضافة مستخدم"})
        except Exception:
            # في حال خطأ غير متوقع، نكمل المسار الطبيعي (وسيمنعنا model validation عند الحفظ)
            pass

        if form.is_valid():
            try:
                with transaction.atomic():
                    teacher = form.save(commit=True)
                    # ربط المعلّم بالمدرسة الحالية كـ TEACHER
                    if active_school is not None:
                        SchoolMembership.objects.update_or_create(
                            school=active_school,
                            teacher=teacher,
                            role_type=SchoolMembership.RoleType.TEACHER,
                            defaults={
                                "is_active": True,
                                **({"job_title": job_title} if job_title else {}),
                            },
                        )
                messages.success(
                    request,
                    "تمت إضافة المستخدم. كلمة المرور المؤقتة هي رقم الجوال، وسيُطلب تغييرها عند أول دخول.",
                )
                next_url = _safe_next_url(request.POST.get("next") or request.GET.get("next"))
                return redirect("reports:add_teacher" if continue_adding else (next_url or "reports:manage_teachers"))
            except IntegrityError:
                messages.error(request, "تعذّر الحفظ: قد يكون رقم الجوال أو الهوية مستخدمًا مسبقًا.")
            except ValidationError as e:
                # مثال: تجاوز حد المعلمين حسب الباقة أو عدم وجود اشتراك فعّال
                messages.error(request, " ".join(getattr(e, "messages", []) or [str(e)]))
            except Exception:
                logger.exception("add_teacher failed")
                messages.error(request, "حدث خطأ غير متوقع أثناء الحفظ. جرّب لاحقًا.")
        else:
            messages.error(request, "الرجاء تصحيح الأخطاء الظاهرة.")
    else:
        form = TeacherCreateForm(active_school=active_school)
    return render(request, "reports/add_teacher.html", {"form": form, "title": "إضافة مستخدم"})

@login_required(login_url="reports:login")
@role_required({"manager"})
@require_http_methods(["GET", "POST"])
def edit_teacher(request: HttpRequest, pk: int) -> HttpResponse:
    active_school = _get_active_school(request)
    labels = school_gender_labels(active_school)
    teacher = get_object_or_404(Teacher, pk=pk)

    # لا يُسمح للمدير بتعديل معلّم غير مرتبط بمدرسته
    if not getattr(request.user, "is_superuser", False) and active_school is not None:
        has_membership = SchoolMembership.objects.filter(
            school=active_school,
            teacher=teacher,
            role_type=SchoolMembership.RoleType.TEACHER,
            is_active=True,
        ).exists()
        if not has_membership:
            messages.error(request, f"لا يمكنك تعديل بيانات {labels['teacher']} لأنها غير مرتبطة بمدرستك.")
            return redirect("reports:manage_teachers")
    if request.method == "POST":
        # تعديل بيانات المعلّم فقط — التكاليف تتم من صفحة أعضاء القسم
        form = TeacherEditForm(request.POST, instance=teacher, active_school=active_school)
        if form.is_valid():
            try:
                with transaction.atomic():
                    form.save(commit=True)
                messages.success(request, "✏️ تم تحديث بيانات المستخدم بنجاح.")
                next_url = _safe_next_url(request.POST.get("next") or request.GET.get("next"))
                return redirect(next_url or "reports:manage_teachers")
            except Exception:
                logger.exception("edit_teacher failed")
                messages.error(request, "حدث خطأ غير متوقع أثناء التحديث.")
        else:
            messages.error(request, "الرجاء تصحيح الأخطاء الظاهرة.")
    else:
        form = TeacherEditForm(instance=teacher, active_school=active_school)

    return render(request, "reports/edit_teacher.html", {"form": form, "teacher": teacher, "title": "تعديل مستخدم"})

@login_required(login_url="reports:login")
@role_required({"manager"})
@require_http_methods(["POST"])
def delete_teacher(request: HttpRequest, pk: int) -> HttpResponse:
    active_school = _get_active_school(request)
    labels = school_gender_labels(active_school)
    teacher = get_object_or_404(Teacher, pk=pk)

    # لا يُسمح للمدير بحذف معلّم غير مرتبط بمدرسته
    if not getattr(request.user, "is_superuser", False) and active_school is not None:
        has_membership = SchoolMembership.objects.filter(
            school=active_school,
            teacher=teacher,
            role_type=SchoolMembership.RoleType.TEACHER,
            is_active=True,
        ).exists()
        if not has_membership:
            messages.error(request, f"لا يمكنك حذف {labels['teacher']} لأنها غير مرتبطة بمدرستك.")
            return redirect("reports:manage_teachers")
    try:
        with transaction.atomic():
            if active_school is not None and not getattr(request.user, "is_superuser", False):
                # ✅ في وضع تعدد المدارس: لا نحذف الحساب عالميًا، بل نفصل عضويته عن هذه المدرسة فقط
                SchoolMembership.objects.filter(
                    school=active_school,
                    teacher=teacher,
                    role_type__in=[
                        SchoolMembership.RoleType.TEACHER,
                        SchoolMembership.RoleType.REPORT_VIEWER,
                    ],
                ).delete()
                messages.success(request, "🗑️ تم إزالة المستخدم من المدرسة الحالية.")
            else:
                teacher.delete()
                messages.success(request, "🗑️ تم حذف المستخدم.")
    except Exception:
        logger.exception("delete_teacher failed")
        messages.error(request, "تعذّر حذف المستخدم. حاول لاحقًا.")
    next_url = _safe_next_url(request.POST.get("next") or request.GET.get("next"))
    return redirect(next_url or "reports:manage_teachers")
