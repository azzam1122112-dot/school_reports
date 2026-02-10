# reports/views.py
# -*- coding: utf-8 -*-
from __future__ import annotations

from datetime import date, timedelta
import logging
import os
import traceback
from typing import Optional, Tuple
from urllib.parse import quote, urlencode, urlparse

import openpyxl

from django import forms
from django.conf import settings
from django.contrib import messages
from django.contrib.auth import authenticate, login, logout, update_session_auth_hash
from django.contrib.auth.decorators import login_required, user_passes_test
from django.contrib.auth.hashers import make_password
from django.core.cache import cache
from django.core.paginator import Paginator
from django.db import IntegrityError, transaction
from django.db.models import (
    Count,
    Exists,
    F,
    Prefetch,
    Q,
    ManyToManyField,
    ForeignKey,
    OuterRef,
    Subquery,
    Sum,
)
from django.db.models.functions import TruncWeek, TruncMonth
from django.core.exceptions import ValidationError
from django.http import FileResponse, Http404, HttpRequest, HttpResponse, JsonResponse
from django.shortcuts import get_object_or_404, redirect, render
from django.urls import reverse
from django.utils import timezone
from django.utils.dateparse import parse_date
from django.utils.http import url_has_allowed_host_and_scheme
from django.utils.safestring import mark_safe
from django.views.decorators.http import require_http_methods
from django.db.models.deletion import ProtectedError

from django.templatetags.static import static

from django_ratelimit.decorators import ratelimit


def _user_guide_md_path() -> str:
    return os.path.join(settings.BASE_DIR, "docs", "user_guide_complete_ar.md")


@require_http_methods(["GET"])
def user_guide(request: HttpRequest) -> HttpResponse:
    """Public HTML page rendering the Arabic user guide Markdown."""

    md_path = _user_guide_md_path()
    if not os.path.exists(md_path):
        raise Http404("User guide not found")

    with open(md_path, "r", encoding="utf-8") as fp:
        md_text = fp.read()

    # Remove the first top-level title to avoid duplication with the page header.
    try:
        lines = md_text.splitlines()
        while lines and not lines[0].strip():
            lines.pop(0)
        if lines and lines[0].startswith("# "):
            lines.pop(0)
            while lines and not lines[0].strip():
                lines.pop(0)
        md_text = "\n".join(lines)
    except Exception:
        pass

    try:
        import markdown as md
    except Exception:
        return HttpResponse("Markdown renderer is not installed.", status=500, content_type="text/plain")

    guide_html = md.markdown(
        md_text,
        extensions=["extra", "fenced_code", "tables"],
        output_format="html5",
    )

    ctx = {
        "guide_html": mark_safe(guide_html),
        "download_url": reverse("reports:user_guide_download"),
    }
    return render(request, "reports/user_guide.html", ctx)


@require_http_methods(["GET"])
def user_guide_download(request: HttpRequest) -> HttpResponse:
    """Download the raw Markdown file for the user guide."""

    md_path = _user_guide_md_path()
    if not os.path.exists(md_path):
        raise Http404("User guide not found")

    return FileResponse(
        open(md_path, "rb"),
        as_attachment=True,
        filename="user_guide_complete_ar.md",
        content_type="text/markdown; charset=utf-8",
    )

# ===== فورمات =====
from .forms import (
    ReportForm,
    TeacherCreateForm,
    TeacherEditForm,
    MyProfilePhoneForm,
    MyPasswordChangeForm,
    TicketActionForm,
    TicketCreateForm,
    DepartmentForm,  # إن لم تكن موجودة في مشروعك سيتم استخدام بديل داخلي
    ManagerCreateForm,
    SubscriptionPlanForm,
    SchoolSubscriptionForm,
    AchievementCreateYearForm,
    TeacherAchievementFileForm,
    AchievementSectionNotesForm,
    AchievementEvidenceUploadForm,
    AchievementManagerNotesForm,
    PlatformAdminCreateForm,
    PlatformSchoolNotificationForm,
    PrivateCommentForm,
    TicketNoteEditForm,
)

# إشعارات (اختياري)
try:
    from .forms import NotificationCreateForm  # type: ignore
except Exception:
    NotificationCreateForm = None  # type: ignore

# ===== موديلات =====
from .models import (
    Report,
    PlatformSettings,
    ShareLink,
    get_share_link_default_days,
    Teacher,
    Ticket,
    TicketNote,
    TicketImage,
    Role,
    School,
    SchoolMembership,
    MANAGER_SLUG,
    SubscriptionPlan,
    SchoolSubscription,
    Payment,
    AuditLog,
    TeacherAchievementFile,
    AchievementSection,
    AchievementEvidenceImage,
    AchievementEvidenceReport,
    TeacherPrivateComment,
)

from .services_achievement import (
    achievement_picker_reports_qs,
    add_report_evidence,
    freeze_achievement_report_evidences,
    remove_report_evidence,
)

# موديلات الإشعارات (اختياري)
try:
    from .models import Notification, NotificationRecipient  # type: ignore
except Exception:
    Notification = None  # type: ignore
    NotificationRecipient = None  # type: ignore

# موديلات مرجعية اختيارية
try:
    from .models import ReportType  # type: ignore
except Exception:  # pragma: no cover
    ReportType = None  # type: ignore

try:
    from .models import Department  # type: ignore
except Exception:  # pragma: no cover
    Department = None  # type: ignore

try:
    from .models import DepartmentMembership  # type: ignore
except Exception:  # pragma: no cover
    DepartmentMembership = None  # type: ignore

# ===== صلاحيات =====
from .permissions import (
    allowed_categories_for,
    role_required,
    restrict_queryset_for_user,
    is_platform_admin,
    platform_allowed_schools_qs,
    platform_can_access_school,
)
try:
    from .permissions import is_officer  # type: ignore
except Exception:
    # بديل مرن إن لم تتوفر الدالة في permissions
    def is_officer(user) -> bool:
        try:
            if not getattr(user, "is_authenticated", False):
                return False
            from .models import DepartmentMembership  # import محلي
            role_type = getattr(DepartmentMembership, "OFFICER", "officer")
            return DepartmentMembership.objects.filter(
                teacher=user, role_type=role_type, department__is_active=True
            ).exists()
        except Exception:
            return False

# ===== خدمات التقارير (تنظيم منطق العرض/التصفية) =====
from .services_reports import (
    apply_admin_report_filters,
    apply_teacher_report_filters,
    get_admin_reports_queryset,
    get_report_for_user_or_404 as svc_get_report_for_user_or_404,
    get_reporttype_choices,
    get_teacher_reports_queryset,
    paginate as svc_paginate,
    teacher_report_stats,
)

from .permissions import (
    can_delete_report,
    can_edit_report,
    can_share_report,
)

# ===== إعدادات محلية =====
HAS_RTYPE: bool = ReportType is not None
DM_TEACHER = getattr(DepartmentMembership, "TEACHER", "teacher") if DepartmentMembership else "teacher"
DM_OFFICER = getattr(DepartmentMembership, "OFFICER", "officer") if DepartmentMembership else "officer"

# إيقاف/تشغيل الرجوع التلقائي للحالة عند ملاحظة المرسل (افتراضي معطّل)
AUTO_REOPEN_ON_SENDER_NOTE: bool = getattr(settings, "TICKETS_AUTO_REOPEN_ON_SENDER_NOTE", False)

logger = logging.getLogger(__name__)

# =========================
# أدوات مساعدة عامة
# =========================
def _is_staff(user) -> bool:
    # ✅ دعم مدير المدرسة (School Manager)
    if not getattr(user, "is_authenticated", False):
        return False
    if getattr(user, "is_staff", False):
        return True
    try:
        return SchoolMembership.objects.filter(
            teacher=user, 
            role_type=SchoolMembership.RoleType.MANAGER,
            is_active=True
        ).exists()
    except Exception:
        return False


def _is_staff_or_officer(user) -> bool:
    """يسمح للموظّفين (is_staff) أو لمسؤولي الأقسام (Officer)."""
    return bool(
        getattr(user, "is_authenticated", False)
        and (_is_staff(user) or is_officer(user) or is_platform_admin(user))
    )


def _safe_next_url(next_url: str | None) -> str | None:
    if not next_url:
        return None
    next_url = (next_url or "").strip()
    if not next_url:
        return None
    # حماية من قيم template الشائعة عند وجود None
    if next_url.lower() in {"none", "null", "undefined"}:
        return None

    # نسمح فقط بمسارات داخلية تبدأ بـ / (ونمنع //)
    if not next_url.startswith("/") or next_url.startswith("//"):
        return None

    parsed = urlparse(next_url)
    if parsed.scheme == "" and parsed.netloc == "":
        return next_url
    return None


def _role_display_map(active_school: Optional[School] = None) -> dict:
    """خريطة عرض عربية للأدوار/الأقسام.

    ملاحظة مهمة للتوسع (Multi-tenant): قد تتكرر slugs للأقسام بين المدارس،
    لذا عندما تتوفر مدرسة نشطة نُقيّد القراءة عليها (مع السماح بالأقسام العامة school=NULL).
    """
    base = {"teacher": "المعلم", "manager": "المدير", "officer": "مسؤول قسم"}
    if Department is not None:
        try:
            qs = Department.objects.filter(is_active=True).only("slug", "role_label", "name")
            if active_school is not None and _model_has_field(Department, "school"):
                qs = qs.filter(Q(school=active_school) | Q(school__isnull=True))
            for d in qs:
                base[d.slug] = d.role_label or d.name or d.slug
        except Exception:
            pass
    return base


def _is_manager_in_school(user, active_school: Optional[School]) -> bool:
    """هل المستخدم مدير داخل المدرسة النشطة؟

    - السوبر: نعم
    - role.slug == manager: نعم (توافق خلفي)
    - أو SchoolMembership(RoleType.MANAGER) داخل active_school
    """
    if getattr(user, "is_superuser", False):
        return True
    try:
        if getattr(getattr(user, "role", None), "slug", None) == "manager":
            return True
    except Exception:
        pass

    if active_school is None:
        return False
    try:
        return SchoolMembership.objects.filter(
            teacher=user,
            school=active_school,
            role_type=SchoolMembership.RoleType.MANAGER,
            is_active=True,
        ).exists()
    except Exception:
        return False


def _safe_redirect(request: HttpRequest, fallback_name: str) -> HttpResponse:
    nxt = request.POST.get("next") or request.GET.get("next")
    if nxt and url_has_allowed_host_and_scheme(nxt, {request.get_host()}):
        return redirect(nxt)
    return redirect(fallback_name)

def _parse_date_safe(value: str | None) -> date | None:
    if not value:
        return None
    return parse_date(value)


def _filter_by_school(qs, school: Optional[School]):
    """تطبيق فلتر المدرسة إذا كان للموديل حقل school وكان هناك مدرسة نشطة."""
    if not school:
        return qs
    try:
        if "school" in [f.name for f in qs.model._meta.get_fields()]:
            return qs.filter(school=school)
    except Exception:
        return qs
    return qs


def _private_comment_role_label(author, school: Optional[School]) -> str:
    return _canonical_role_label(author, school)


def _canonical_role_label(user, school: Optional[School]) -> str:
    """إرجاع تسمية دور موحّدة (بدون تداخلات).

    الأدوار المعتمدة فقط:
    - معلم
    - مدير المدرسة
    - المشرف العام
    - مدير النظام
    """
    if user is None:
        return ""

    # مدير النظام (الأعلى أولوية): سوبر يوزر دائمًا
    try:
        if getattr(user, "is_superuser", False):
            return "مدير النظام"
    except Exception:
        pass

    # المشرف العام
    try:
        if is_platform_admin(user) or getattr(user, "is_platform_admin", False):
            scope = getattr(user, "platform_scope", None)
            role_obj = getattr(scope, "role", None) if scope is not None else None
            role_name = (getattr(role_obj, "name", "") or "").strip()
            return role_name or "المشرف العام"
    except Exception:
        pass

    # مدير المدرسة
    try:
        if school is not None and _is_manager_in_school(user, school):
            return "مدير المدرسة"
    except Exception:
        pass
    try:
        role = getattr(user, "role", None)
        if role is not None and (getattr(role, "slug", "") or "").strip().lower() == "manager":
            return "مدير المدرسة"
    except Exception:
        pass

    # مدير النظام (is_staff فقط) — لا نستخدم _is_staff هنا لأنه يُعيد True لمدير المدرسة
    try:
        if getattr(user, "is_staff", False):
            return "مدير النظام"
    except Exception:
        pass

    return "معلم"


def _canonical_sender_name(user) -> str:
    if user is None:
        return "الإدارة"
    return (
        (getattr(user, "name", None) or "").strip()
        or (getattr(user, "phone", None) or "").strip()
        or (getattr(user, "username", None) or "").strip()
        or "الإدارة"
    )


def _model_has_field(model, field_name: str) -> bool:
    """تحقق آمن: هل الموديل يحتوي على حقل باسم معين؟"""
    if model is None:
        return False
    try:
        return field_name in {f.name for f in model._meta.get_fields()}
    except Exception:
        return False


def _get_active_school(request: HttpRequest) -> Optional[School]:
    """إرجاع المدرسة المختارة حالياً من الجلسة (إن وُجدت).

    تحسين احترافي:
    - إذا لم تُحدَّد مدرسة في الجلسة، وكان للمستخدم مدرسة واحدة فقط → نعتبرها المدرسة النشطة تلقائياً.
    - للمشرف العام: إن لم يكن لديه عضويات ومدرسة واحدة فقط مفعّلة في النظام → نختارها تلقائياً.
    """
    sid = request.session.get("active_school_id")
    try:
        if sid:
            return School.objects.filter(pk=sid, is_active=True).first()

        user = getattr(request, "user", None)
        # مستخدم عادي: مدرسة واحدة فقط ضمن عضوياته
        if user is not None and getattr(user, "is_authenticated", False):
            schools = _user_schools(user)
            if len(schools) == 1:
                school = schools[0]
                _set_active_school(request, school)
                return school

            # مشرف عام مع مدرسة واحدة فقط في النظام
            if getattr(user, "is_superuser", False):
                qs = School.objects.filter(is_active=True)
                if qs.count() == 1:
                    school = qs.first()
                    if school is not None:
                        _set_active_school(request, school)
                        return school
    except Exception:
        return None
    return None


def _set_active_school(request: HttpRequest, school: Optional[School]) -> None:
    """تحديث المدرسة المختارة في الجلسة للمستخدم الحالي."""
    if school is None:
        request.session.pop("active_school_id", None)
    else:
        request.session["active_school_id"] = school.pk


def _user_schools(user) -> list[School]:
    """إرجاع المدارس المرتبطة بالمستخدم عبر عضويات SchoolMembership."""
    if not getattr(user, "is_authenticated", False):
        return []
    try:
        qs = (
            School.objects.filter(memberships__teacher=user, memberships__is_active=True)
            .distinct()
            .order_by("name")
        )
        return list(qs)
    except Exception:
        return []


def _user_manager_schools(user) -> list[School]:
    """المدارس التي يكون فيها المستخدم مدير مدرسة."""
    if not getattr(user, "is_authenticated", False):
        return []

    # توافق خلفي: بعض الحسابات القديمة تعتمد على Role(slug='manager')
    role_slug = None
    try:
        role_slug = getattr(getattr(user, "role", None), "slug", None)
    except Exception:
        role_slug = None

    try:
        if role_slug == "manager":
            qs = (
                School.objects.filter(
                    memberships__teacher=user,
                    memberships__is_active=True,
                    is_active=True,
                )
                .distinct()
                .order_by("name")
            )
        else:
            qs = (
                School.objects.filter(
                    memberships__teacher=user,
                    memberships__role_type=SchoolMembership.RoleType.MANAGER,
                    memberships__is_active=True,
                    is_active=True,
                )
                .distinct()
                .order_by("name")
            )
        return list(qs)
    except Exception:
        return []


def _is_report_viewer(user, active_school: Optional[School] = None) -> bool:
    """(تم إلغاء دور مشرف التقارير)"""
    return False

# =========================
# الدخول / الخروج
# =========================
@ratelimit(key="ip", rate="10/m", method="POST", block=True)
@require_http_methods(["GET", "POST"])
def login_view(request: HttpRequest) -> HttpResponse:
    if request.user.is_authenticated:
        # إن كان المستخدم موظّف لوحة (مدير/سوبر أدمن) نوجّهه للوحة المناسبة
        if getattr(request.user, "is_superuser", False):
            return redirect("reports:platform_admin_dashboard")
        if is_platform_admin(request.user):
            return redirect("reports:platform_schools_directory")
        if _is_staff(request.user):
            return redirect("reports:admin_dashboard")
        return redirect("reports:home")

    if request.method == "POST":
        identifier = (
            request.POST.get("phone")
            or request.POST.get("username")
            or request.POST.get("identifier")
            or ""
        ).strip()
        password = request.POST.get("password") or ""

        # يدعم تسجيل الدخول عبر:
        # - رقم الجوال (المعرف الافتراضي USERNAME_FIELD)
        # - رقم الهوية (نبحث عنه ثم نستخدم phone)
        # مع بعض التطبيع الخفيف لأشكال رقم الجوال الشائعة.
        attempts: list[str] = []
        if identifier:
            attempts.append(identifier)
            ident_no_plus = identifier.lstrip("+")
            if ident_no_plus != identifier:
                attempts.append(ident_no_plus)
            if identifier.isdigit() and len(identifier) == 9:
                attempts.append("0" + identifier)
            if ident_no_plus.isdigit() and ident_no_plus.startswith("966") and len(ident_no_plus) >= 12:
                # +9665XXXXXXXX -> 05XXXXXXXX
                attempts.append("0" + ident_no_plus[-9:])

        # إزالة التكرارات مع الحفاظ على الترتيب
        seen: set[str] = set()
        attempts = [a for a in attempts if a and not (a in seen or seen.add(a))]

        user = None
        for phone_candidate in attempts:
            user = authenticate(request, username=phone_candidate, password=password)
            if user is not None:
                break

        if user is None and identifier:
            try:
                potential_by_national = Teacher.objects.filter(national_id=identifier).only("phone").first()
                if potential_by_national is not None and getattr(potential_by_national, "phone", None):
                    user = authenticate(request, username=potential_by_national.phone, password=password)
            except Exception:
                user = None
        if user is not None:
            # ✅ قواعد الاشتراك عند تسجيل الدخول:
            # - السوبر: يتجاوز دائمًا.
            # - مدير المدرسة: يُسمح له بالدخول حتى لو انتهى الاشتراك، لكن يُوجّه لصفحة (انتهاء الاشتراك)
            #   ولا يستطيع استخدام المنصة إلا لصفحات التجديد (يُفرض ذلك عبر SubscriptionMiddleware).
            # - بقية المستخدمين: إن لم توجد أي مدرسة باشتراك ساري → نمنع تسجيل الدخول.

            if not getattr(user, "is_superuser", False):
                try:
                    memberships = (
                        SchoolMembership.objects.filter(teacher=user, is_active=True)
                        .select_related("school")
                        .order_by("id")
                    )

                    # إن لم تكن هناك أي عضوية مدرسة، لا نمنع تسجيل الدخول برسالة اشتراك (لأننا لا نستطيع ربطه بمدرسة).
                    # هذا يحدث أحياناً لحسابات قديمة أو حسابات لم تُربط بعد.
                    if not memberships.exists():
                        login(request, user)
                        messages.warning(request, "تنبيه: حسابك غير مرتبط بمدرسة فعّالة. تواصل مع إدارة النظام لربط الحساب بالمدرسة.")
                        next_url = _safe_next_url(request.POST.get("next") or request.GET.get("next"))
                        if getattr(user, "is_superuser", False):
                            default_name = "reports:platform_admin_dashboard"
                        elif is_platform_admin(user):
                            default_name = "reports:platform_schools_directory"
                        elif _is_staff(user):
                            default_name = "reports:admin_dashboard"
                        else:
                            default_name = "reports:home"
                        return redirect(next_url or default_name)

                    active_school = None
                    any_active_subscription = False
                    is_any_manager = False
                    manager_school = None
                    first_school_name = None

                    role_slug = getattr(getattr(user, "role", None), "slug", None)

                    for m in memberships:
                        if first_school_name is None:
                            first_school_name = getattr(getattr(m, "school", None), "name", None)
                        if m.role_type == SchoolMembership.RoleType.MANAGER:
                            is_any_manager = True
                            if manager_school is None:
                                manager_school = m.school

                        # دعم حسابات مدير قديمة تعتمد على Role(slug='manager') حتى لو role_type مختلف.
                        if not is_any_manager and role_slug == "manager":
                            is_any_manager = True
                            if manager_school is None:
                                manager_school = m.school

                        sub = None
                        try:
                            sub = getattr(m.school, 'subscription', None)
                        except Exception:
                            sub = None

                        # عدم وجود اشتراك = منتهي
                        if sub is not None and not bool(sub.is_expired) and bool(getattr(m.school, "is_active", True)):
                            any_active_subscription = True
                            if active_school is None:
                                active_school = m.school

                    if not any_active_subscription:
                        if is_any_manager and manager_school is not None:
                            # المدير يُسمح له بالدخول للتجديد فقط
                            login(request, user)
                            _set_active_school(request, manager_school)
                            return redirect("reports:subscription_expired")

                        school_label = f" ({first_school_name})" if first_school_name else ""
                        messages.error(request, f"عذرًا، اشتراك المدرسة{school_label} منتهي. لا يمكن الدخول حتى يتم تجديد الاشتراك.")
                        return redirect("reports:login")

                    # هناك اشتراك ساري واحد على الأقل → نكمل تسجيل الدخول ونثبت مدرسة نشطة مناسبة
                    login(request, user)
                    if active_school is not None:
                        _set_active_school(request, active_school)
                except Exception:
                    # في حال أي مشكلة في تحقق الاشتراك، لا نكسر تسجيل الدخول (سيتولى Middleware المنع لاحقاً)
                    login(request, user)
            else:
                login(request, user)

            # بعد تسجيل الدخول مباشرةً: اختيار مدرسة افتراضية عند توفر مدرسة واحدة فقط
            try:
                # إن كان للمستخدم مدرسة واحدة فقط ضمن عضوياته نعتبرها المدرسة النشطة
                schools = _user_schools(user)
                if len(schools) == 1:
                    _set_active_school(request, schools[0])
                # أو إن كان مشرفاً عاماً وهناك مدرسة واحدة فقط مفعّلة في النظام
                elif user.is_superuser:
                    qs = School.objects.filter(is_active=True)
                    if qs.count() == 1:
                        s = qs.first()
                        if s is not None:
                            _set_active_school(request, s)
            except Exception:
                pass

            next_url = _safe_next_url(request.POST.get("next") or request.GET.get("next"))
            # الوجهة الافتراضية حسب الدور
            if getattr(user, "is_superuser", False):
                default_name = "reports:platform_admin_dashboard"
            elif is_platform_admin(user):
                default_name = "reports:platform_schools_directory"
            elif _is_staff(user):
                default_name = "reports:admin_dashboard"
            else:
                default_name = "reports:home"
            return redirect(next_url or default_name)

        # فشل المصادقة: نتحقق هل السبب هو أن الحساب موقوف (is_active=False)
        try:
            from django.db.models import Q

            q = Q()
            if attempts:
                q |= Q(phone__in=attempts)
            if identifier:
                q |= Q(national_id=identifier)

            potential_user = Teacher.objects.filter(q).first() if q else None
            if potential_user is not None and (not potential_user.is_active) and potential_user.check_password(password):
                messages.error(request, "عذرًا، حسابك موقوف. يرجى التواصل مع الإدارة.")
            else:
                messages.error(request, "رقم الجوال/الهوية أو كلمة المرور غير صحيحة")
        except Exception:
            messages.error(request, "رقم الجوال/الهوية أو كلمة المرور غير صحيحة")

    context = {"next": _safe_next_url(request.GET.get("next"))}
    return render(request, "reports/login.html", context)


@login_required(login_url="reports:login")
@require_http_methods(["GET"])
def logout_view(request: HttpRequest) -> HttpResponse:
    _set_active_school(request, None)
    logout(request)
    return redirect("reports:login")


@login_required(login_url="reports:login")
@require_http_methods(["GET", "POST"])
def my_profile(request: HttpRequest) -> HttpResponse:
    """بروفايل المستخدم الحالي.

    - متاح لكل المستخدمين ما عدا (مشرف تقارير - عرض فقط).
    - يعرض الاسم + المدارس المسندة.
    - يسمح بتغيير رقم الجوال + تغيير كلمة المرور.
    """

    active_school = _get_active_school(request)
    if _is_report_viewer(request.user, active_school) or _is_report_viewer(request.user):
        messages.error(request, "هذا الحساب للعرض فقط ولا يملك صفحة بروفايل.")
        return redirect("reports:school_reports_readonly")

    memberships = (
        SchoolMembership.objects.filter(teacher=request.user, is_active=True)
        .select_related("school")
        .order_by("school__name", "id")
    )

    phone_form = MyProfilePhoneForm(instance=request.user, prefix="phone")
    pwd_form = MyPasswordChangeForm(request.user, prefix="pwd")

    if request.method == "POST":
        if "update_phone" in request.POST:
            phone_form = MyProfilePhoneForm(request.POST, instance=request.user, prefix="phone")
            if phone_form.is_valid():
                try:
                    phone_form.save()
                    messages.success(request, "تم تحديث رقم الجوال بنجاح.")
                    return redirect("reports:my_profile")
                except IntegrityError:
                    messages.error(request, "تعذر تحديث رقم الجوال (قد يكون مستخدمًا بالفعل).")
        elif "update_password" in request.POST:
            pwd_form = MyPasswordChangeForm(request.user, request.POST, prefix="pwd")
            if pwd_form.is_valid():
                user = pwd_form.save()
                update_session_auth_hash(request, user)
                messages.success(request, "تم تحديث كلمة المرور بنجاح.")
                return redirect("reports:my_profile")

    ctx = {
        "active_school": active_school,
        "memberships": memberships,
        "phone_form": phone_form,
        "pwd_form": pwd_form,
    }
    return render(request, "reports/my_profile.html", ctx)


@require_http_methods(["GET"])
def platform_landing(request: HttpRequest) -> HttpResponse:
    """الصفحة الرئيسية العامة للمنصة (تعريف + مميزات + زر دخول).

    - المستخدِم المسجّل بالفعل يُعاد توجيهه مباشرةً للواجهة المناسبة.
    - الزر الأساسي يقود إلى شاشة تسجيل الدخول العادية.
    """

    if getattr(request.user, "is_authenticated", False):
        if getattr(request.user, "is_superuser", False):
            return redirect("reports:platform_admin_dashboard")
        if is_platform_admin(request.user):
            return redirect("reports:platform_schools_directory")
        if _is_staff(request.user):
            return redirect("reports:admin_dashboard")
        return redirect("reports:home")

    return render(request, "reports/landing.html")


# =========================
# المشرف العام (عرض + تواصل فقط)
# =========================


def _require_platform_admin_or_superuser(request: HttpRequest) -> bool:
    return bool(getattr(request.user, "is_superuser", False) or is_platform_admin(request.user))


def _require_platform_school_access(request: HttpRequest, school: Optional[School]) -> bool:
    if getattr(request.user, "is_superuser", False):
        return True
    return bool(is_platform_admin(request.user) and platform_can_access_school(request.user, school))


@login_required(login_url="reports:login")
@require_http_methods(["GET"])
def platform_schools_directory(request: HttpRequest) -> HttpResponse:
    user = request.user
    if not _require_platform_admin_or_superuser(request):
        messages.error(request, "لا تملك صلاحية الوصول إلى شاشة المدارس.")
        return redirect("reports:home")

    # السوبر يوزر يرى كل المدارس، المشرف العام يرى المدارس ضمن نطاقه.
    base_qs = School.objects.all().order_by("name") if getattr(user, "is_superuser", False) else platform_allowed_schools_qs(user)

    q = (request.GET.get("q") or "").strip()
    gender = (request.GET.get("gender") or "").strip().lower()
    city = (request.GET.get("city") or "").strip()

    # قائمة المدن من كامل النطاق (قبل فلترة city) حتى تبقى القائمة مفيدة.
    try:
        cities = (
            base_qs.exclude(city__isnull=True)
            .exclude(city__exact="")
            .values_list("city", flat=True)
            .distinct()
            .order_by("city")
        )
        cities = list(cities)
    except Exception:
        cities = []

    qs = base_qs
    if q:
        qs = qs.filter(Q(name__icontains=q) | Q(code__icontains=q) | Q(city__icontains=q))
    if gender in {"boys", "girls"}:
        qs = qs.filter(gender=gender)
    if city:
        qs = qs.filter(city=city)

    ctx = {
        "schools": list(qs.order_by("name")),
        "cities": cities,
        "q": q,
        "gender": gender,
        "city": city,
    }
    return render(request, "reports/platform_schools_directory.html", ctx)


@login_required(login_url="reports:login")
@require_http_methods(["GET"])
def platform_enter_school(request: HttpRequest, pk: int) -> HttpResponse:
    user = request.user
    if not _require_platform_admin_or_superuser(request):
        raise Http404("ليس لديك صلاحية")

    if getattr(user, "is_superuser", False):
        school = get_object_or_404(School, pk=pk)
    else:
        school = get_object_or_404(platform_allowed_schools_qs(user), pk=pk)

    _set_active_school(request, school)
    return redirect("reports:platform_school_dashboard")


@login_required(login_url="reports:login")
@require_http_methods(["GET"])
def platform_school_dashboard(request: HttpRequest) -> HttpResponse:
    if not _require_platform_admin_or_superuser(request):
        messages.error(request, "لا تملك صلاحية الوصول إلى لوحة المدرسة.")
        return redirect("reports:home")

    active_school = _get_active_school(request)
    if active_school is None:
        return redirect("reports:platform_schools_directory")

    if not _require_platform_school_access(request, active_school):
        try:
            request.session.pop("active_school_id", None)
        except Exception:
            pass
        messages.error(request, "هذه المدرسة خارج نطاق صلاحياتك.")
        return redirect("reports:platform_schools_directory")

    subscription = (
        SchoolSubscription.objects.filter(school=active_school)
        .select_related("plan")
        .first()
    )

    return render(
        request,
        "reports/platform_school_dashboard.html",
        {"school": active_school, "subscription": subscription},
    )


@login_required(login_url="reports:login")
@require_http_methods(["GET"])
def platform_school_reports(request: HttpRequest) -> HttpResponse:
    if not _require_platform_admin_or_superuser(request):
        messages.error(request, "لا تملك صلاحية الوصول.")
        return redirect("reports:home")

    active_school = _get_active_school(request)
    if active_school is None:
        messages.error(request, "فضلاً اختر مدرسة أولاً.")
        return redirect("reports:platform_schools_directory")

    if not _require_platform_school_access(request, active_school):
        messages.error(request, "هذه المدرسة خارج نطاق صلاحياتك.")
        return redirect("reports:platform_schools_directory")

    cats = allowed_categories_for(request.user, active_school)
    qs = get_admin_reports_queryset(user=request.user, active_school=active_school)

    start_date = _parse_date_safe(request.GET.get("start_date"))
    end_date = _parse_date_safe(request.GET.get("end_date"))
    teacher_name = (request.GET.get("teacher_name") or "").strip()
    category = (request.GET.get("category") or "").strip().lower()

    qs = apply_admin_report_filters(
        qs,
        start_date=start_date,
        end_date=end_date,
        teacher_name=teacher_name,
        category=category,
        cats=cats,
    )

    allowed_choices = get_reporttype_choices(active_school=active_school) if (HAS_RTYPE and ReportType is not None) else []
    reports_page = svc_paginate(qs, per_page=20, page=request.GET.get("page", 1))

    context = {
        "reports": reports_page,
        "start_date": request.GET.get("start_date", ""),
        "end_date": request.GET.get("end_date", ""),
        "teacher_name": teacher_name,
        "category": category if (not cats or "all" in cats or category in cats) else "",
        "categories": allowed_choices,
        "can_delete": False,
    }
    return render(request, "reports/admin_reports.html", context)


@login_required(login_url="reports:login")
@require_http_methods(["GET"])
def platform_school_tickets(request: HttpRequest) -> HttpResponse:
    if not _require_platform_admin_or_superuser(request):
        messages.error(request, "لا تملك صلاحية الوصول.")
        return redirect("reports:home")

    active_school = _get_active_school(request)
    if active_school is None:
        messages.error(request, "فضلاً اختر مدرسة أولاً.")
        return redirect("reports:platform_schools_directory")

    if not _require_platform_school_access(request, active_school):
        messages.error(request, "هذه المدرسة خارج نطاق صلاحياتك.")
        return redirect("reports:platform_schools_directory")

    qs = (
        Ticket.objects.select_related("creator", "assignee", "department")
        .prefetch_related("recipients")
        .filter(school=active_school, is_platform=False)
        .order_by("-created_at")
    )

    status = (request.GET.get("status") or "").strip()
    q = (request.GET.get("q") or "").strip()
    mine = request.GET.get("mine") == "1"

    if status:
        qs = qs.filter(status=status)
    if mine:
        qs = qs.filter(Q(assignee=request.user) | Q(recipients=request.user)).distinct()
    if q:
        for kw in q.split():
            qs = qs.filter(Q(title__icontains=kw) | Q(body__icontains=kw))

    ctx = {
        "tickets": list(qs[:200]),
        "status": status,
        "q": q,
        "mine": mine,
        "status_choices": Ticket.Status.choices,
    }
    return render(request, "reports/tickets_inbox.html", ctx)


@login_required(login_url="reports:login")
@user_passes_test(_is_staff, login_url="reports:login")
@role_required({"manager"})
@require_http_methods(["GET"])
def manager_school_tickets(request: HttpRequest) -> HttpResponse:
    """قائمة جميع طلبات المدرسة للمدير (مع فلترة وبحث)."""
    active_school = _get_active_school(request)

    if School.objects.filter(is_active=True).exists():
        if active_school is None:
            messages.error(request, "فضلاً اختر مدرسة أولاً.")
            return redirect("reports:select_school")
        if (not request.user.is_superuser) and active_school not in _user_manager_schools(request.user):
            messages.error(request, "ليست لديك صلاحية كمدير على هذه المدرسة.")
            return redirect("reports:select_school")

    qs = (
        Ticket.objects.select_related("creator", "assignee", "department")
        .prefetch_related("recipients")
        .filter(school=active_school, is_platform=False)
        .order_by("-created_at")
    )

    status = (request.GET.get("status") or "").strip()
    q = (request.GET.get("q") or "").strip()
    mine = request.GET.get("mine") == "1"

    if status:
        qs = qs.filter(status=status)
    if mine:
        qs = qs.filter(Q(assignee=request.user) | Q(recipients=request.user)).distinct()
    if q:
        for kw in q.split():
            qs = qs.filter(Q(title__icontains=kw) | Q(body__icontains=kw))

    ctx = {
        "tickets": list(qs[:200]),
        "status": status,
        "q": q,
        "mine": mine,
        "status_choices": Ticket.Status.choices,
        "page_title": "طلبات المدرسة",
        "page_heading": "📌 طلبات المدرسة",
        "page_subtitle": "استعرض جميع الطلبات التابعة للمدرسة، ويمكنك إضافة ملاحظات وتغيير الحالة من داخل الطلب.",
    }
    return render(request, "reports/tickets_inbox.html", ctx)


@login_required(login_url="reports:login")
@require_http_methods(["GET", "POST"])
def platform_school_notify(request: HttpRequest) -> HttpResponse:
    if not _require_platform_admin_or_superuser(request):
        messages.error(request, "لا تملك صلاحية الوصول.")
        return redirect("reports:home")

    active_school = _get_active_school(request)
    if active_school is None:
        messages.error(request, "فضلاً اختر مدرسة أولاً.")
        return redirect("reports:platform_schools_directory")

    if not _require_platform_school_access(request, active_school):
        messages.error(request, "هذه المدرسة خارج نطاق صلاحياتك.")
        return redirect("reports:platform_schools_directory")

    form = PlatformSchoolNotificationForm(request.POST or None)
    if request.method == "POST" and form.is_valid():
        title = (form.cleaned_data.get("title") or "").strip()
        message_text = form.cleaned_data["message"]
        is_important = bool(form.cleaned_data.get("is_important"))

        try:
            with transaction.atomic():
                n = Notification.objects.create(
                    title=title,
                    message=message_text,
                    is_important=is_important,
                    school=active_school,
                    created_by=request.user,
                )
                teacher_ids = list(
                    SchoolMembership.objects.filter(
                        school=active_school,
                        is_active=True,
                        teacher__is_active=True,
                    )
                    .values_list("teacher_id", flat=True)
                    .distinct()
                )
                recipients = [NotificationRecipient(notification=n, teacher_id=tid) for tid in teacher_ids]
                NotificationRecipient.objects.bulk_create(recipients, ignore_conflicts=True)

                # Push WS delta (bulk_create doesn't trigger signals)
                try:
                    from .realtime_notifications import push_new_notification_to_teachers

                    push_new_notification_to_teachers(notification=n, teacher_ids=teacher_ids)
                except Exception:
                    pass
            messages.success(request, "تم إرسال الإشعار إلى جميع مستخدمي المدرسة.")
            return redirect("reports:platform_school_dashboard")
        except Exception:
            logger.exception("Failed to send school notification")
            messages.error(request, "تعذّر إرسال الإشعار. حاول مرة أخرى.")

    return render(request, "reports/platform_school_notify.html", {"form": form, "school": active_school})


@login_required(login_url="reports:login")
@user_passes_test(lambda u: getattr(u, "is_superuser", False), login_url="reports:login")
@require_http_methods(["GET", "POST"])
def platform_admin_create(request: HttpRequest) -> HttpResponse:
    from .models import PlatformAdminScope

    form = PlatformAdminCreateForm(request.POST or None)
    if request.method == "POST" and form.is_valid():
        try:
            with transaction.atomic():
                admin_user = form.save(commit=True)

                role_obj = form.cleaned_data.get("role")
                gender_scope = (form.cleaned_data.get("gender_scope") or "all").strip().lower()
                cities_raw = (form.cleaned_data.get("cities") or "").strip()
                allowed_schools = form.cleaned_data.get("allowed_schools")

                cities_list = []
                if cities_raw:
                    for part in cities_raw.replace("؛", ",").split(","):
                        c = (part or "").strip()
                        if c and c not in cities_list:
                            cities_list.append(c)

                scope, _created = PlatformAdminScope.objects.get_or_create(admin=admin_user)
                scope.role = role_obj
                scope.gender_scope = gender_scope if gender_scope in {"all", "boys", "girls"} else "all"
                scope.allowed_cities = cities_list
                scope.save()
                if allowed_schools is not None:
                    scope.allowed_schools.set(list(allowed_schools))

            messages.success(request, "تم إنشاء مشرف المنصة بنجاح.")
            return redirect("reports:platform_admin_dashboard")
        except Exception:
            logger.exception("Failed to create platform admin")
            messages.error(request, "تعذّر إنشاء مشرف المنصة. تحقق من البيانات.")

    return render(request, "reports/platform_admin_create.html", {"form": form})


@login_required(login_url="reports:login")
@user_passes_test(lambda u: getattr(u, "is_superuser", False), login_url="reports:login")
@require_http_methods(["GET"])
def platform_admins_list(request: HttpRequest) -> HttpResponse:
    from .models import PlatformAdminScope

    q = (request.GET.get("q") or "").strip()
    qs = (
        Teacher.objects.filter(is_platform_admin=True)
        .select_related("platform_scope")
        .prefetch_related("platform_scope__allowed_schools")
        .order_by("name", "id")
    )
    if q:
        qs = qs.filter(Q(name__icontains=q) | Q(phone__icontains=q))

    # تأكيد وجود scope لكل مشرف (اختياري/مساعد)
    try:
        missing_ids = list(qs.filter(platform_scope__isnull=True).values_list("id", flat=True))
        if missing_ids:
            for tid in missing_ids:
                try:
                    PlatformAdminScope.objects.get_or_create(admin_id=tid)
                except Exception:
                    pass
            qs = (
                Teacher.objects.filter(is_platform_admin=True)
                .select_related("platform_scope")
                .prefetch_related("platform_scope__allowed_schools")
                .order_by("name", "id")
            )
            if q:
                qs = qs.filter(Q(name__icontains=q) | Q(phone__icontains=q))
    except Exception:
        pass

    return render(request, "reports/platform_admins_list.html", {"admins": list(qs), "q": q})


@login_required(login_url="reports:login")
@user_passes_test(lambda u: getattr(u, "is_superuser", False), login_url="reports:login")
@require_http_methods(["GET", "POST"])
def platform_admin_update(request: HttpRequest, pk: int) -> HttpResponse:
    from .models import PlatformAdminScope

    admin_user = get_object_or_404(Teacher, pk=pk, is_platform_admin=True)
    scope, _created = PlatformAdminScope.objects.get_or_create(admin=admin_user)

    form = PlatformAdminCreateForm(request.POST or None, instance=admin_user)

    if request.method == "POST" and form.is_valid():
        try:
            with transaction.atomic():
                updated_user = form.save(commit=True)

                role_obj = form.cleaned_data.get("role")
                gender_scope = (form.cleaned_data.get("gender_scope") or "all").strip().lower()
                cities_raw = (form.cleaned_data.get("cities") or "").strip()
                allowed_schools = form.cleaned_data.get("allowed_schools")

                cities_list = []
                if cities_raw:
                    for part in cities_raw.replace("؛", ",").split(","):
                        c = (part or "").strip()
                        if c and c not in cities_list:
                            cities_list.append(c)

                scope.admin = updated_user
                scope.role = role_obj
                scope.gender_scope = gender_scope if gender_scope in {"all", "boys", "girls"} else "all"
                scope.allowed_cities = cities_list
                scope.save()
                if allowed_schools is not None:
                    scope.allowed_schools.set(list(allowed_schools))

            messages.success(request, "تم تحديث بيانات مشرف المنصة.")
            return redirect("reports:platform_admins_list")
        except Exception:
            logger.exception("Failed to update platform admin")
            messages.error(request, "تعذّر حفظ التعديلات. حاول مرة أخرى.")

    return render(
        request,
        "reports/platform_admin_edit.html",
        {
            "form": form,
            "admin_user": admin_user,
        },
    )


@login_required(login_url="reports:login")
@user_passes_test(lambda u: getattr(u, "is_superuser", False), login_url="reports:login")
@require_http_methods(["GET", "POST"])
def platform_admin_delete(request: HttpRequest, pk: int) -> HttpResponse:
    admin_user = get_object_or_404(Teacher, pk=pk, is_platform_admin=True)

    # حماية: لا نحذف السوبر يوزر عبر هذه الشاشة
    if getattr(admin_user, "is_superuser", False):
        messages.error(request, "لا يمكن حذف مستخدم سوبر يوزر من هنا.")
        return redirect("reports:platform_admins_list")

    if request.method == "POST":
        try:
            admin_user.delete()
            messages.success(request, "تم حذف المشرف العام.")
        except Exception:
            logger.exception("Failed to delete platform admin")
            messages.error(request, "تعذّر حذف المشرف العام.")
        return redirect("reports:platform_admins_list")

    return render(request, "reports/platform_admin_delete.html", {"admin_user": admin_user})


@login_required(login_url="reports:login")
@user_passes_test(_is_staff, login_url="reports:login")
@require_http_methods(["GET", "POST"])
def select_school(request: HttpRequest) -> HttpResponse:
    """شاشة اختيار المدرسة للآدمن ومديري المدارس.

    - المستخدم السوبر يوزر يشاهد جميع المدارس.
    - مدير المدرسة يشاهد فقط المدارس التي هو مدير لها.
    """

    if request.user.is_superuser:
        schools_qs = School.objects.filter(is_active=True).order_by("name")
    else:
        manager_schools = _user_manager_schools(request.user)
        schools_qs = School.objects.filter(id__in=[s.id for s in manager_schools], is_active=True).order_by("name")

    # إن لم يكن للمستخدم أي مدارس مرتبطة به نسمح له برؤية لا شيء

    if request.method == "POST":
        sid = request.POST.get("school_id")
        try:
            school = schools_qs.get(pk=sid)
            _set_active_school(request, school)
            messages.success(request, f"تم اختيار المدرسة: {school.name}")
            return redirect("reports:admin_dashboard")
        except (School.DoesNotExist, ValueError, TypeError):
            messages.error(request, "تعذّر اختيار المدرسة. فضلاً اختر مدرسة صحيحة.")

    context = {
        "schools": list(schools_qs),
        "current_school": _get_active_school(request),
    }
    return render(request, "reports/select_school.html", context)


@login_required(login_url="reports:login")
@require_http_methods(["GET", "POST"])
def switch_school(request: HttpRequest) -> HttpResponse:
    """تبديل المدرسة النشطة بسرعة من الهيدر/القائمة."""
    if request.method == "POST":
        sid = request.POST.get("school_id")
        next_raw = request.POST.get("next")
    else:
        sid = request.GET.get("school_id")
        next_raw = request.GET.get("next")

    default_next = "reports:admin_dashboard" if _is_staff(request.user) or getattr(request.user, "is_superuser", False) else "reports:home"
    next_url = _safe_next_url(next_raw) or default_next

    if not sid:
        return redirect(next_url)

    if request.user.is_superuser:
        schools_qs = School.objects.filter(is_active=True)
    else:
        # ✅ أي مستخدم يملك عضوية نشطة في المدرسة يمكنه التبديل إليها
        schools_qs = (
            School.objects.filter(
                is_active=True,
                memberships__teacher=request.user,
                memberships__is_active=True,
            )
            .distinct()
        )

    try:
        school = schools_qs.get(pk=sid)
        _set_active_school(request, school)
        messages.success(request, f"تم اختيار المدرسة: {school.name}")
    except (School.DoesNotExist, ValueError, TypeError):
        messages.error(request, "تعذّر تبديل المدرسة. فضلاً اختر مدرسة صحيحة.")

    return redirect(next_url)

# =========================
# الرئيسية (لوحة المعلم)
# =========================
@login_required(login_url="reports:login")
@require_http_methods(["GET"])
def home(request: HttpRequest) -> HttpResponse:
    # --- Platform Admin Redirect (Not Superuser) ---
    # إذا كان مشرف منصة (وليس سوبر يوزر)، وجهه لصفحة مدارسه أو المدرسة النشطة
    if is_platform_admin(request.user) and not getattr(request.user, "is_superuser", False):
        active_school = _get_active_school(request)
        if active_school:
            # توجيه للوحة المدرسة الخاصة بالمشرف
            return redirect("reports:platform_school_dashboard")
        # توجيه لدليل المدارس لاختيار مدرسة
        return redirect("reports:platform_schools_directory")
    # -----------------------------------------------

    active_school = _get_active_school(request)
    stats = {"today_count": 0, "total_count": 0, "last_title": "—"}
    req_stats = {"open": 0, "in_progress": 0, "done": 0, "rejected": 0, "total": 0}

    # إشعار التحفيز: اعرض أحدث إشعار غير مقروء فقط.
    # (يُعلّم كمقروء فقط بعد إغلاق المستخدم للرسالة من الواجهة.)
    home_notification = None
    home_notification_recipient_id: int | None = None
    try:
        if NotificationRecipient is not None and Notification is not None:
            now = timezone.now()
            nqs = (
                NotificationRecipient.objects.select_related("notification", "notification__created_by")
                .filter(teacher=request.user)
            )

            # غير مقروء فقط
            try:
                if hasattr(NotificationRecipient, "is_read"):
                    nqs = nqs.filter(is_read=False)
                elif hasattr(NotificationRecipient, "read_at"):
                    nqs = nqs.filter(read_at__isnull=True)
            except Exception:
                pass

            # عزل حسب المدرسة النشطة (مع السماح بإشعارات عامة school=NULL)
            try:
                if active_school is not None and hasattr(Notification, "school"):
                    nqs = nqs.filter(Q(notification__school=active_school) | Q(notification__school__isnull=True))
            except Exception:
                pass

            # استبعاد المنتهي
            try:
                if hasattr(Notification, "expires_at"):
                    nqs = nqs.filter(Q(notification__expires_at__gt=now) | Q(notification__expires_at__isnull=True))
            except Exception:
                pass

            rec = nqs.order_by("-created_at", "-id").first()
            if rec is not None:
                home_notification = getattr(rec, "notification", None)
                try:
                    home_notification_recipient_id = int(getattr(rec, "pk"))
                except Exception:
                    home_notification_recipient_id = None
    except Exception:
        home_notification = None
        home_notification_recipient_id = None

    try:
        my_qs = _filter_by_school(
            Report.objects.filter(teacher=request.user).only(
                "id", "title", "report_date", "day_name", "beneficiaries_count"
            ),
            active_school,
        )
        today = timezone.localdate()
        stats["total_count"] = my_qs.count()
        stats["today_count"] = my_qs.filter(report_date=today).count()
        last_report = my_qs.order_by("-report_date", "-id").first()
        stats["last_title"] = (last_report.title if last_report else "—")
        recent_reports = list(my_qs.order_by("-report_date", "-id")[:5])

        my_tickets_qs = _filter_by_school(
            Ticket.objects.filter(creator=request.user)
            .select_related("assignee", "department")
            .only("id", "title", "status", "department", "created_at", "assignee__name")
            .order_by("-created_at", "-id"),
            active_school,
        )
        agg = my_tickets_qs.aggregate(
            open=Count("id", filter=Q(status="open")),
            in_progress=Count("id", filter=Q(status="in_progress")),
            done=Count("id", filter=Q(status="done")),
            rejected=Count("id", filter=Q(status="rejected")),
            total=Count("id"),
        )
        for k in req_stats.keys():
            req_stats[k] = int(agg.get(k) or 0)
        recent_tickets = list(my_tickets_qs[:5])

        return render(
            request,
            "reports/home.html",
            {
                "stats": stats,
                "recent_reports": recent_reports[:2],
                "req_stats": req_stats,
                "recent_tickets": recent_tickets[:2],
                "home_notification": home_notification,
                "home_notification_recipient_id": home_notification_recipient_id,
            },
        )
    except Exception:
        logger.exception("Home view failed")
        if settings.DEBUG or os.getenv("SHOW_ERRORS") == "1":
            html = "<h2>Home exception</h2><pre>{}</pre>".format(traceback.format_exc())
            return HttpResponse(html, status=500)
    return redirect("reports:home")

# =========================
# التقارير: إضافة/عرض/إدارة
# =========================
@login_required(login_url="reports:login")
@require_http_methods(["GET", "POST"])
def add_report(request: HttpRequest) -> HttpResponse:
    active_school = _get_active_school(request)
    if request.method == "POST":
        form = ReportForm(request.POST, request.FILES, active_school=active_school)
        if form.is_valid():
            report = form.save(commit=False)
            report.teacher = request.user
            if hasattr(report, "school") and active_school is not None:
                report.school = active_school

            # حماية حقل "المنفذ": يُحفظ دائمًا باسم المستخدم الحالي ولا نقبل أي قيمة مرسلة من الفورم.
            teacher_name_final = (getattr(request.user, "name", "") or "").strip()
            if not teacher_name_final:
                teacher_name_final = (getattr(request.user, "username", "") or str(request.user) or "").strip()
            teacher_name_final = teacher_name_final[:120]
            if hasattr(report, "teacher_name"):
                report.teacher_name = teacher_name_final

            report.save()
            messages.success(request, "تم إضافة التقرير بنجاح ✅")
            return redirect("reports:my_reports")
        messages.error(request, "فضلاً تحقق من الحقول وأعد المحاولة.")
    else:
        form = ReportForm(active_school=active_school)

    return render(request, "reports/add_report.html", {"form": form})

@login_required(login_url="reports:login")
@require_http_methods(["GET"])
def my_reports(request: HttpRequest) -> HttpResponse:
    active_school = _get_active_school(request)
    qs = get_teacher_reports_queryset(user=request.user, active_school=active_school)
    start_date = _parse_date_safe(request.GET.get("start_date"))
    end_date = _parse_date_safe(request.GET.get("end_date"))
    q = request.GET.get("q", "").strip()

    qs = apply_teacher_report_filters(qs, start_date=start_date, end_date=end_date, q=q)
    stats = teacher_report_stats(qs)
    reports_page = svc_paginate(qs, per_page=10, page=request.GET.get("page", 1))

    params = request.GET.copy()
    if "page" in params:
        params.pop("page")
    qs_params = params.urlencode()

    return render(
        request,
        "reports/my_reports.html",
        {
            "reports": reports_page,
            "qs": qs_params,
            "start_date": request.GET.get("start_date", ""),
            "end_date": request.GET.get("end_date", ""),
            "q": q,
            "stats": stats,
        },
    )

@user_passes_test(_is_staff, login_url="reports:login")
@role_required({"manager"})
@require_http_methods(["GET"])
def admin_reports(request: HttpRequest) -> HttpResponse:
    active_school = _get_active_school(request)
    cats = allowed_categories_for(request.user, active_school)
    qs = get_admin_reports_queryset(user=request.user, active_school=active_school)

    start_date = _parse_date_safe(request.GET.get("start_date"))
    end_date = _parse_date_safe(request.GET.get("end_date"))
    teacher_name = (request.GET.get("teacher_name") or "").strip()
    category = (request.GET.get("category") or "").strip().lower()

    qs = apply_admin_report_filters(
        qs,
        start_date=start_date,
        end_date=end_date,
        teacher_name=teacher_name,
        category=category,
        cats=cats,
    )

    allowed_choices = get_reporttype_choices(active_school=active_school) if (HAS_RTYPE and ReportType is not None) else []
    reports_page = svc_paginate(qs, per_page=20, page=request.GET.get("page", 1))
    
    # ✅ إضافة صلاحيات الحذف والتعديل والمشاركة لكل تقرير باستخدام الدوال الصحيحة
    # التي تراعي رؤساء الأقسام (OFFICER) وأعضاء الأقسام (TEACHER)
    user = request.user
    for report in reports_page:
        report.user_can_delete = can_delete_report(user, report, active_school=active_school)
        report.user_can_edit = can_edit_report(user, report, active_school=active_school)
        report.user_can_share = can_share_report(user, report, active_school=active_school)

    context = {
        "reports": reports_page,
        "start_date": request.GET.get("start_date", ""),
        "end_date": request.GET.get("end_date", ""),
        "teacher_name": teacher_name,
        "category": category if (not cats or "all" in cats or category in cats) else "",
        "categories": allowed_choices,
        "can_delete": True,  # للتوافق الخلفي
    }
    return render(request, "reports/admin_reports.html", context)


@login_required(login_url="reports:login")
@require_http_methods(["GET"])
def school_reports_readonly(request: HttpRequest) -> HttpResponse:
    """عرض تقارير المدرسة (عرض فقط) لمشرف التقارير المرتبط بالمدرسة."""
    active_school = _get_active_school(request)
    if active_school is None:
        messages.error(request, "فضلاً اختر/حدّد مدرسة أولاً.")
        return redirect("reports:home")

    # لا نسمح بالسوبر أو الموظف هنا (لمنع خلط الصلاحيات/الحسابات)
    if getattr(request.user, "is_superuser", False) or _is_staff(request.user):
        return redirect("reports:admin_reports")

    if not _is_report_viewer(request.user, active_school):
        messages.error(request, "لا تملك صلاحية الاطلاع على تقارير هذه المدرسة.")
        return redirect("reports:home")

    cats = allowed_categories_for(request.user, active_school)
    qs = get_admin_reports_queryset(user=request.user, active_school=active_school)

    start_date = _parse_date_safe(request.GET.get("start_date"))
    end_date = _parse_date_safe(request.GET.get("end_date"))
    teacher_name = (request.GET.get("teacher_name") or "").strip()
    category = (request.GET.get("category") or "").strip().lower()

    qs = apply_admin_report_filters(
        qs,
        start_date=start_date,
        end_date=end_date,
        teacher_name=teacher_name,
        category=category,
        cats=cats,
    )

    allowed_choices = get_reporttype_choices(active_school=active_school) if (HAS_RTYPE and ReportType is not None) else []
    reports_page = svc_paginate(qs, per_page=20, page=request.GET.get("page", 1))

    context = {
        "reports": reports_page,
        "start_date": request.GET.get("start_date", ""),
        "end_date": request.GET.get("end_date", ""),
        "teacher_name": teacher_name,
        "category": category,
        "categories": allowed_choices,
        "can_delete": False,
    }
    return render(request, "reports/admin_reports.html", context)


# =========================
# ملف إنجاز المعلّم
# =========================
def _ensure_achievement_sections(ach_file: TeacherAchievementFile) -> None:
    """يضمن وجود 11 محورًا ثابتًا داخل الملف."""
    existing = set(
        AchievementSection.objects.filter(file=ach_file).values_list("code", flat=True)
    )
    to_create = []
    for code, title in AchievementSection.Code.choices:
        if int(code) in existing:
            continue
        to_create.append(
            AchievementSection(file=ach_file, code=int(code), title=str(title))
        )
    if to_create:
        AchievementSection.objects.bulk_create(to_create)


def _can_manage_achievement(user, active_school: Optional[School]) -> bool:
    if getattr(user, "is_superuser", False):
        return True
    if active_school is None:
        return False
    try:
        return SchoolMembership.objects.filter(
            teacher=user,
            school=active_school,
            role_type=SchoolMembership.RoleType.MANAGER,
            is_active=True,
        ).exists()
    except Exception:
        return False


def _can_view_achievement(user, active_school: Optional[School]) -> bool:
    if getattr(user, "is_superuser", False):
        return True
    if is_platform_admin(user) and platform_can_access_school(user, active_school):
        return True
    if _can_manage_achievement(user, active_school):
        return True
    return False


@login_required(login_url="reports:login")
@require_http_methods(["GET", "POST"])
def achievement_my_files(request: HttpRequest) -> HttpResponse:
    """قائمة ملفات الإنجاز الخاصة بالمعلّم + إنشاء سنة جديدة."""
    active_school = _get_active_school(request)
    if active_school is None:
        messages.error(request, "فضلاً اختر/حدّد مدرسة أولاً.")
        return redirect("reports:home")

    existing_years = list(
        TeacherAchievementFile.objects.filter(school=active_school)
        .values_list("academic_year", flat=True)
        .distinct()
    )
    
    # استخراج السنوات المسموحة من إعدادات المدرسة
    allowed = active_school.allowed_academic_years if active_school else []
    
    create_form = AchievementCreateYearForm(
        request.POST or None, 
        year_choices=existing_years,
        allowed_years=allowed
    )
    if request.method == "POST" and (request.POST.get("action") == "create"):
        if create_form.is_valid():
            year = create_form.cleaned_data["academic_year"]
            ach_file, created = TeacherAchievementFile.objects.get_or_create(
                teacher=request.user,
                school=active_school,
                academic_year=year,
                defaults={},
            )
            _ensure_achievement_sections(ach_file)
            if created:
                messages.success(request, "تم إنشاء ملف الإنجاز للسنة بنجاح ✅")
            return redirect("reports:achievement_file_detail", pk=ach_file.pk)
        messages.error(request, "تحقق من السنة الدراسية وأعد المحاولة.")

    files = (
        TeacherAchievementFile.objects.filter(teacher=request.user, school=active_school)
        .order_by("-academic_year", "-id")
    )
    return render(
        request,
        "reports/achievement_my_files.html",
        {"files": files, "create_form": create_form, "current_school": active_school},
    )


@login_required(login_url="reports:login")
@require_http_methods(["POST"])
def achievement_file_delete(request: HttpRequest, pk: int) -> HttpResponse:
    """حذف ملف إنجاز (للمالك فقط)."""
    file = get_object_or_404(TeacherAchievementFile, pk=pk, teacher=request.user)
    
    # يمكن إضافة شرط الحالة لو أردنا منع حذف المعتمد، لكن السؤال يوحي بالحرية للتصحيح
    file.delete()
    messages.success(request, "تم حذف ملف الإنجاز بنجاح ✅")
    return redirect("reports:achievement_my_files")


@login_required(login_url="reports:login")
@require_http_methods(["POST"])
def achievement_file_update_year(request: HttpRequest, pk: int) -> HttpResponse:
    """تصحيح السنة الدراسية لملف الإنجاز (للمالك فقط)."""
    file = get_object_or_404(TeacherAchievementFile, pk=pk, teacher=request.user)
    active_school = _get_active_school(request)

    # نموذج بسيط للتحقق من السنة (نستخدم نفس فورم الإنشاء للتحقق مع تمرير القيمة المرسلة كخيار مقبول)
    # هذا يسمح بقبول أي سنة صحيحة (هيئة + تتابع) حتى لو لم تكن في القائمة الافتراضية
    submitted_year = request.POST.get("academic_year", "")
    form = AchievementCreateYearForm(request.POST, year_choices=[submitted_year]) 
    
    if form.is_valid():
        new_year = form.cleaned_data["academic_year"]
        
        # 1. التحقق من عدم التكرار
        duplicate = TeacherAchievementFile.objects.filter(
            teacher=request.user, 
            school=file.school, 
            academic_year=new_year
        ).exclude(pk=file.pk).exists()

        if duplicate:
            messages.error(request, f" لا يمكن التعديل: لديك ملف آخر بالفعل للسنة {new_year}")
        else:
            file.academic_year = new_year
            file.save(update_fields=["academic_year", "updated_at"])
            messages.success(request, f"تم تعديل السنة الدراسية إلى {new_year} ✅")

    else:
        # استخراج أول خطأ
        err = next(iter(form.errors.values()))[0] if form.errors else "بيانات غير صالحة"
        messages.error(request, f"تعذر تحديث السنة: {err}")

    return redirect("reports:achievement_my_files")


@login_required(login_url="reports:login")
@require_http_methods(["GET", "POST"])
def achievement_school_files(request: HttpRequest) -> HttpResponse:
    """قائمة ملفات الإنجاز للمدرسة (مدير/مشرف عرض فقط).

    - تعرض جميع المعلمين في المدرسة النشطة.
    - بجانب كل معلم: فتح الملف + طباعة/حفظ PDF.
    - الاعتماد/الرفض يكون داخل صفحة الملف نفسها.
    """
    active_school = _get_active_school(request)
    if active_school is None:
        messages.error(request, "فضلاً اختر/حدّد مدرسة أولاً.")
        return redirect("reports:home")

    if not _can_view_achievement(request.user, active_school):
        messages.error(request, "لا تملك صلاحية الاطلاع على ملفات الإنجاز.")
        return redirect("reports:home")

    # اختيار سنة (اختياري): إن لم تُحدد، نأخذ آخر سنة موجودة في المدرسة
    year = (request.GET.get("year") or request.POST.get("year") or "").strip()
    try:
        year = year.replace("–", "-").replace("—", "-")
    except Exception:
        pass

    existing_years = list(
        TeacherAchievementFile.objects.filter(school=active_school)
        .values_list("academic_year", flat=True)
        .distinct()
    )
    # نفس منطق الاختيارات في نموذج الإنشاء (بدون إدخال يدوي)
    tmp_form = AchievementCreateYearForm(year_choices=existing_years)
    year_choices = [c[0] for c in tmp_form.fields["academic_year"].choices]

    if not year and year_choices:
        year = year_choices[0]
    if year and year_choices and year not in year_choices:
        year = year_choices[0]

    base_url = reverse("reports:achievement_school_files")

    def _redirect_with_year(year_value: str) -> HttpResponse:
        year_value = (year_value or "").strip()
        if not year_value:
            return redirect(base_url)
        return redirect(f"{base_url}?{urlencode({'year': year_value})}")

    # إنشاء ملف إنجاز من صفحة المدرسة غير مسموح: المعلّم هو من ينشئ ملفه من (ملف الإنجاز)
    if request.method == "POST" and (request.POST.get("action") == "create"):
        messages.error(request, "إنشاء ملف الإنجاز متاح للمعلّم فقط.")
        return _redirect_with_year(year)

    # Search Logic
    q = request.GET.get("q", "").strip()

    teachers = (
        Teacher.objects.filter(
            school_memberships__school=active_school,
            school_memberships__is_active=True,
        )
        .distinct()
        .only("id", "name", "phone", "national_id")
        .order_by("name")
    )
    
    if q:
        from django.db.models import Q
        from django.db.models import Prefetch
        teachers = teachers.filter(
            Q(name__icontains=q)
            | Q(phone__icontains=q)
            | Q(national_id__icontains=q)
        )

    files_by_teacher_id = {}
    if year:
        files = (
            TeacherAchievementFile.objects.filter(school=active_school, academic_year=year)
            .select_related("teacher")
            .only("id", "teacher_id", "status", "academic_year")
        )
        if q:
            # تصفية الملفات أيضاً لتحسين الأداء
            files = files.filter(teacher__in=teachers)

        files_by_teacher_id = {f.teacher_id: f for f in files}

    rows = [{"teacher": t, "file": files_by_teacher_id.get(t.id)} for t in teachers]

    return render(
        request,
        "reports/achievement_school_files.html",
        {
            "rows": rows,
            "year": year,
            "year_choices": year_choices,
            "current_school": active_school,
            "is_manager": _can_manage_achievement(request.user, active_school),
            "q": q,
        },
    )


@login_required(login_url="reports:login")
@require_http_methods(["GET", "POST"])
def achievement_school_teachers(request: HttpRequest) -> HttpResponse:
    """Alias قديم: توجيه إلى صفحة المدرسة الموحدة."""
    params = {}
    year = (request.GET.get("year") or request.POST.get("year") or "").strip()
    if year:
        params["year"] = year
    url = reverse("reports:achievement_school_files")
    if params:
        return redirect(f"{url}?{urlencode(params)}")
    return redirect("reports:achievement_school_files")


@login_required(login_url="reports:login")
@require_http_methods(["GET", "POST"])
def achievement_file_detail(request: HttpRequest, pk: int) -> HttpResponse:
    active_school = _get_active_school(request)
    ach_file = get_object_or_404(TeacherAchievementFile, pk=pk)
    user = request.user

    if not getattr(user, "is_superuser", False):
        if active_school is None or ach_file.school_id != getattr(active_school, "id", None):
            raise Http404

    is_manager = _can_manage_achievement(user, active_school)
    is_viewer = _is_report_viewer(user, active_school)
    is_owner = (ach_file.teacher_id == getattr(user, "id", None))
    is_platform = bool(is_platform_admin(user) and platform_can_access_school(user, active_school))

    if not (getattr(user, "is_superuser", False) or is_manager or is_viewer or is_owner or is_platform):
        return HttpResponse(status=403)

    _ensure_achievement_sections(ach_file)
    try:
        from django.db.models import Prefetch

        ev_reports_qs = AchievementEvidenceReport.objects.select_related(
            "report",
            "report__category",
        ).order_by("id")
        sections = (
            AchievementSection.objects.filter(file=ach_file)
            .prefetch_related("evidence_images", Prefetch("evidence_reports", queryset=ev_reports_qs))
            .order_by("code", "id")
        )
    except Exception:
        sections = (
            AchievementSection.objects.filter(file=ach_file)
            .prefetch_related("evidence_images", "evidence_reports")
            .order_by("code", "id")
        )

    can_edit_teacher = bool(is_owner and ach_file.status in {TeacherAchievementFile.Status.DRAFT, TeacherAchievementFile.Status.RETURNED})
    can_post = bool((can_edit_teacher or is_manager) and not is_viewer)

    general_form = TeacherAchievementFileForm(request.POST or None, instance=ach_file)
    manager_notes_form = AchievementManagerNotesForm(request.POST or None, instance=ach_file)
    year_form = AchievementCreateYearForm()
    upload_form = AchievementEvidenceUploadForm()

    # تعليقات خاصة (يراها المعلم + أصحاب الصلاحية داخل المدرسة/المنصة)
    is_staff_user = _is_staff(user)
    can_add_private_comment = bool(is_platform or is_manager or is_staff_user or getattr(user, "is_superuser", False))
    show_private_comments = bool(is_owner or can_add_private_comment)
    private_comments = (
        TeacherPrivateComment.objects.select_related("created_by")
        .filter(achievement_file=ach_file, teacher=ach_file.teacher)
        .order_by("-created_at", "-id")
        if show_private_comments
        else TeacherPrivateComment.objects.none()
    )
    private_comment_form = PrivateCommentForm(request.POST or None) if can_add_private_comment else None

    # الرجوع ثابت حسب الدور: المعلّم -> ملفاتي، غير ذلك -> ملفات المدرسة
    if is_owner:
        back_url = reverse("reports:achievement_my_files")
    else:
        url = reverse("reports:achievement_school_files")
        back_url = f"{url}?{urlencode({'year': ach_file.academic_year})}"

    if request.method == "POST":
        action = (request.POST.get("action") or "").strip()
        section_id = request.POST.get("section_id")

        # ===== تعليقات خاصة (لا تظهر في الطباعة أو المشاركة) =====
        # توافق خلفي: platform_comment
        if action in {"platform_comment", "private_comment_create", "private_comment_update", "private_comment_delete"}:
            if not can_add_private_comment:
                return HttpResponse(status=403)

            # create
            if action in {"platform_comment", "private_comment_create"}:
                if private_comment_form is not None and private_comment_form.is_valid():
                    body = private_comment_form.cleaned_data["body"]
                    try:
                        with transaction.atomic():
                            TeacherPrivateComment.objects.create(
                                teacher=ach_file.teacher,
                                created_by=user,
                                school=active_school,
                                achievement_file=ach_file,
                                body=body,
                            )
                            n = Notification.objects.create(
                                title="تعليق خاص على ملف الإنجاز",
                                message=body,
                                is_important=True,
                                school=active_school,
                                created_by=user,
                            )
                            NotificationRecipient.objects.create(notification=n, teacher=ach_file.teacher)
                        messages.success(request, "تم إرسال التعليق الخاص للمعلّم ✅")
                        return redirect("reports:achievement_file_detail", pk=ach_file.pk)
                    except Exception:
                        logger.exception("Failed to create private achievement comment")
                        messages.error(request, "تعذر حفظ التعليق. حاول مرة أخرى.")
                else:
                    messages.error(request, "تحقق من نص التعليق وأعد المحاولة.")
                return redirect("reports:achievement_file_detail", pk=ach_file.pk)

            # update/delete (only comment owner, or superuser)
            comment_id = request.POST.get("comment_id")
            try:
                comment_id_int = int(comment_id) if comment_id else None
            except (TypeError, ValueError):
                comment_id_int = None

            if not comment_id_int:
                messages.error(request, "تعذر تحديد التعليق.")
                return redirect("reports:achievement_file_detail", pk=ach_file.pk)

            comment = TeacherPrivateComment.objects.filter(
                pk=comment_id_int,
                achievement_file=ach_file,
                teacher=ach_file.teacher,
            ).first()
            if comment is None:
                messages.error(request, "التعليق غير موجود.")
                return redirect("reports:achievement_file_detail", pk=ach_file.pk)

            is_owner_of_comment = getattr(comment, "created_by_id", None) == getattr(user, "id", None)

            if action == "private_comment_update":
                # تعديل: لصاحب التعليق فقط
                if not is_owner_of_comment:
                    return HttpResponse(status=403)
                body = (request.POST.get("body") or "").strip()
                if not body:
                    messages.error(request, "نص التعليق مطلوب.")
                    return redirect("reports:achievement_file_detail", pk=ach_file.pk)
                try:
                    TeacherPrivateComment.objects.filter(pk=comment.pk).update(body=body)
                    messages.success(request, "تم تعديل التعليق ✅")
                except Exception:
                    messages.error(request, "تعذر تعديل التعليق.")
                return redirect("reports:achievement_file_detail", pk=ach_file.pk)

            if action == "private_comment_delete":
                # حذف: لصاحب التعليق فقط، والسوبر يمكنه حذف أي تعليق
                if not (is_owner_of_comment or getattr(user, "is_superuser", False)):
                    return HttpResponse(status=403)
                try:
                    comment.delete()
                    messages.success(request, "تم حذف التعليق ✅")
                except Exception:
                    messages.error(request, "تعذر حذف التعليق.")
                return redirect("reports:achievement_file_detail", pk=ach_file.pk)

        if not can_post:
            return HttpResponse(status=403)

        if action == "save_general" and can_edit_teacher:
            if general_form.is_valid():
                general_form.save()
                messages.success(request, "تم حفظ البيانات العامة ✅")
                return redirect("reports:achievement_file_detail", pk=ach_file.pk)
            messages.error(request, "تحقق من الحقول وأعد المحاولة.")

        elif action == "save_section" and can_edit_teacher and section_id:
            sec = get_object_or_404(AchievementSection, pk=int(section_id), file=ach_file)
            sec_form = AchievementSectionNotesForm(request.POST, instance=sec)
            if sec_form.is_valid():
                sec_form.save()
                messages.success(request, "تم حفظ ملاحظات المحور ✅")
                return redirect("reports:achievement_file_detail", pk=ach_file.pk)
            messages.error(request, "تحقق من ملاحظات المحور وأعد المحاولة.")

        elif action == "upload_evidence" and can_edit_teacher and section_id:
            sec = get_object_or_404(AchievementSection, pk=int(section_id), file=ach_file)
            imgs = request.FILES.getlist("images")
            if not imgs:
                messages.error(request, "اختر صورًا للرفع.")
                return redirect("reports:achievement_file_detail", pk=ach_file.pk)
            existing_count = AchievementEvidenceImage.objects.filter(section=sec).count()
            remaining = max(0, 8 - existing_count)
            if remaining <= 0:
                messages.error(request, "لا يمكن إضافة أكثر من 8 صور لهذا المحور.")
                return redirect("reports:achievement_file_detail", pk=ach_file.pk)
            imgs = imgs[:remaining]
            for f in imgs:
                AchievementEvidenceImage.objects.create(section=sec, image=f)
            messages.success(request, "تم رفع الشواهد ✅")
            return redirect("reports:achievement_file_detail", pk=ach_file.pk)

        elif action == "delete_evidence" and can_edit_teacher:
            img_id = request.POST.get("image_id")
            if img_id:
                img = get_object_or_404(AchievementEvidenceImage, pk=int(img_id), section__file=ach_file)
                try:
                    img.delete()
                except Exception:
                    pass
                messages.success(request, "تم حذف الصورة ✅")
            return redirect("reports:achievement_file_detail", pk=ach_file.pk)

        elif action == "add_report_evidence" and can_edit_teacher and section_id:
            sec = get_object_or_404(AchievementSection, pk=int(section_id), file=ach_file)
            report_id = request.POST.get("report_id")
            try:
                report_id_int = int(report_id) if report_id else None
            except (TypeError, ValueError):
                report_id_int = None
            if not report_id_int:
                messages.error(request, "تعذر تحديد التقرير.")
                return redirect("reports:achievement_file_detail", pk=ach_file.pk)

            rep_qs = Report.objects.select_related("category").filter(teacher=request.user)
            try:
                if active_school is not None and _model_has_field(Report, "school"):
                    rep_qs = rep_qs.filter(school=active_school)
            except Exception:
                pass
            r = get_object_or_404(rep_qs, pk=report_id_int)

            try:
                add_report_evidence(section=sec, report=r)
                messages.success(request, "تم إضافة التقرير كشاهِد ✅")
            except Exception:
                messages.error(request, "تعذر إضافة التقرير. ربما تمت إضافته مسبقاً.")
            return redirect("reports:achievement_file_detail", pk=ach_file.pk)

        elif action == "delete_report_evidence" and can_edit_teacher and section_id:
            sec = get_object_or_404(AchievementSection, pk=int(section_id), file=ach_file)
            evidence_id = request.POST.get("evidence_id")
            try:
                evidence_id_int = int(evidence_id) if evidence_id else None
            except (TypeError, ValueError):
                evidence_id_int = None
            if not evidence_id_int:
                messages.error(request, "تعذر تحديد الشاهد.")
                return redirect("reports:achievement_file_detail", pk=ach_file.pk)

            try:
                ok = remove_report_evidence(section=sec, evidence_id=evidence_id_int)
                if ok:
                    messages.success(request, "تم إزالة التقرير من الشواهد ✅")
                else:
                    messages.error(request, "الشاهد غير موجود.")
            except Exception:
                messages.error(request, "تعذر إزالة الشاهد.")
            return redirect("reports:achievement_file_detail", pk=ach_file.pk)

        elif action == "import_prev" and can_edit_teacher:
            prev_year = (request.POST.get("prev_year") or "").strip()
            if prev_year:
                prev = TeacherAchievementFile.objects.filter(
                    teacher=ach_file.teacher,
                    school=ach_file.school,
                    academic_year=prev_year,
                ).first()
            else:
                prev = (
                    TeacherAchievementFile.objects.filter(
                        teacher=ach_file.teacher, school=ach_file.school
                    )
                    .exclude(pk=ach_file.pk)
                    .order_by("-academic_year", "-id")
                    .first()
                )
            if not prev:
                messages.error(request, "لا يوجد ملف سابق للاستيراد.")
                return redirect("reports:achievement_file_detail", pk=ach_file.pk)

            # استيراد الحقول الثابتة فقط
            ach_file.qualifications = prev.qualifications
            ach_file.professional_experience = prev.professional_experience
            ach_file.specialization = prev.specialization
            ach_file.teaching_load = prev.teaching_load
            ach_file.subjects_taught = prev.subjects_taught
            ach_file.contact_info = prev.contact_info
            ach_file.save(update_fields=[
                "qualifications",
                "professional_experience",
                "specialization",
                "teaching_load",
                "subjects_taught",
                "contact_info",
                "updated_at",
            ])
            messages.success(request, "تم استيراد البيانات الثابتة من ملف سابق ✅")
            return redirect("reports:achievement_file_detail", pk=ach_file.pk)

        elif action == "submit" and can_edit_teacher:
            now = timezone.now()
            try:
                with transaction.atomic():
                    ach_file.status = TeacherAchievementFile.Status.SUBMITTED
                    ach_file.submitted_at = now
                    ach_file.save(update_fields=["status", "submitted_at", "updated_at"])

                    frozen = freeze_achievement_report_evidences(ach_file=ach_file)
                if frozen:
                    messages.success(request, f"تم إرسال الملف للاعتماد ✅ (تم تجميد {frozen} تقرير/تقارير كشواهد)")
                else:
                    messages.success(request, "تم إرسال الملف للاعتماد ✅")
            except Exception:
                # حتى لو فشل التجميد لأي سبب، لا نكسر تجربة المستخدم
                messages.success(request, "تم إرسال الملف للاعتماد ✅")
            return redirect("reports:achievement_file_detail", pk=ach_file.pk)

        elif action == "approve" and is_manager:
            ach_file.status = TeacherAchievementFile.Status.APPROVED
            ach_file.decided_at = timezone.now()
            ach_file.decided_by = request.user
            ach_file.save(update_fields=["status", "decided_at", "decided_by", "updated_at"])
            messages.success(request, "تم اعتماد ملف الإنجاز ✅")
            return redirect("reports:achievement_file_detail", pk=ach_file.pk)

        elif action == "return" and is_manager:
            if manager_notes_form.is_valid():
                manager_notes_form.save()
            ach_file.status = TeacherAchievementFile.Status.RETURNED
            ach_file.decided_at = timezone.now()
            ach_file.decided_by = request.user
            ach_file.save(update_fields=["status", "decided_at", "decided_by", "updated_at", "manager_notes"])
            messages.success(request, "تم إرجاع الملف للمعلّم مع الملاحظات ✅")
            return redirect("reports:achievement_file_detail", pk=ach_file.pk)

        messages.error(request, "تعذر تنفيذ العملية.")

    try:
        if show_private_comments and private_comments is not None:
            for c in private_comments:
                try:
                    c.created_by_role_label = _private_comment_role_label(getattr(c, "created_by", None), active_school)
                except Exception:
                    c.created_by_role_label = ""
    except Exception:
        pass

    return render(
        request,
        "reports/achievement_file.html",
        {
            "file": ach_file,
            "sections": sections,
            "general_form": general_form,
            "upload_form": upload_form,
            "manager_notes_form": manager_notes_form,
            "can_edit_teacher": can_edit_teacher,
            "is_manager": is_manager,
            "is_viewer": is_viewer,
            "is_owner": is_owner,
            "show_private_comments": show_private_comments,
            "private_comments": private_comments,
            "private_comment_form": private_comment_form,
            "can_add_private_comment": can_add_private_comment,
            "current_user_id": getattr(user, "id", None),
            "is_superuser": bool(getattr(user, "is_superuser", False)),
            "year_form": year_form,
            "current_school": active_school,
            "back_url": back_url,
        },
    )


@login_required(login_url="reports:login")
@require_http_methods(["GET"])
def achievement_report_picker(request: HttpRequest, pk: int) -> HttpResponse:
    """Return a partial HTML list to pick teacher reports as evidence for a section."""

    active_school = _get_active_school(request)
    ach_file = get_object_or_404(TeacherAchievementFile, pk=pk)
    user = request.user

    is_owner = (ach_file.teacher_id == getattr(user, "id", None))
    if not is_owner:
        return HttpResponse(status=403)
    if not getattr(user, "is_superuser", False):
        if active_school is None or ach_file.school_id != getattr(active_school, "id", None):
            raise Http404

    if ach_file.status not in {
        TeacherAchievementFile.Status.DRAFT,
        TeacherAchievementFile.Status.RETURNED,
    }:
        return HttpResponse(status=403)

    section_id = request.GET.get("section_id")
    try:
        section_id_int = int(section_id) if section_id else None
    except (TypeError, ValueError):
        section_id_int = None
    if not section_id_int:
        return HttpResponse(status=400)

    section = get_object_or_404(AchievementSection, pk=section_id_int, file=ach_file)
    q = (request.GET.get("q") or "").strip()

    qs = achievement_picker_reports_qs(teacher=user, active_school=active_school, q=q).select_related("category")
    reports = list(qs[:50])
    already_ids = set(
        AchievementEvidenceReport.objects.filter(section=section, report__isnull=False).values_list(
            "report_id", flat=True
        )
    )

    return render(
        request,
        "reports/partials/achievement_report_picker_list.html",
        {
            "file": ach_file,
            "section": section,
            "reports": reports,
            "q": q,
            "already_ids": already_ids,
        },
    )


@login_required(login_url="reports:login")
@require_http_methods(["GET"])
def achievement_file_pdf(request: HttpRequest, pk: int) -> HttpResponse:
    active_school = _get_active_school(request)
    ach_file = get_object_or_404(TeacherAchievementFile, pk=pk)
    if not getattr(request.user, "is_superuser", False):
        if active_school is None or ach_file.school_id != getattr(active_school, "id", None):
            raise Http404
    if not (_can_view_achievement(request.user, active_school) or ach_file.teacher_id == getattr(request.user, "id", None)):
        return HttpResponse(status=403)

    # توليد PDF عند الطلب
    try:
        from .pdf_achievement import generate_achievement_pdf

        pdf_bytes, filename = generate_achievement_pdf(request=request, ach_file=ach_file)
    except OSError as ex:
        # WeasyPrint on Windows يحتاج مكتبات نظام (GTK/Pango/Cairo) مثل libgobject.
        msg = str(ex) or ""
        if "libgobject" in msg or "gobject-2.0" in msg:
            # أفضل UX: لا نعرض صفحة خطأ/نص؛ نرجع لنفس صفحة الملف برسالة واضحة.
            messages.error(
                request,
                "تعذر توليد PDF محليًا لأن مكتبات الطباعة غير مثبتة على هذا الجهاز. "
                "أفضل حل: شغّل المشروع على Render/Docker/WSL (Linux) أو ثبّت GTK runtime على Windows.",
            )
            logger.warning("WeasyPrint native deps missing: %s", msg)
            return redirect("reports:achievement_file_detail", pk=ach_file.pk)

        if settings.DEBUG:
            raise
        messages.error(request, "تعذر توليد ملف PDF حاليًا.")
        return redirect("reports:achievement_file_detail", pk=ach_file.pk)
    except Exception:
        if settings.DEBUG:
            raise
        messages.error(request, "تعذر توليد ملف PDF حاليًا.")
        return redirect("reports:achievement_file_detail", pk=ach_file.pk)

    resp = HttpResponse(pdf_bytes, content_type="application/pdf")
    resp["Content-Disposition"] = f'attachment; filename="{filename}"'
    return resp


@login_required(login_url="reports:login")
@require_http_methods(["GET"])
def achievement_file_print(request: HttpRequest, pk: int) -> HttpResponse:
    """صفحة طباعة ملف الإنجاز (مثل طباعة التقارير).

    تعتمد على الطباعة من المتصفح (Save as PDF) لتجنّب مشاكل WeasyPrint على Windows.
    """

    active_school = _get_active_school(request)
    ach_file = get_object_or_404(TeacherAchievementFile, pk=pk)

    if not getattr(request.user, "is_superuser", False):
        if active_school is None or ach_file.school_id != getattr(active_school, "id", None):
            raise Http404

    if not (_can_view_achievement(request.user, active_school) or ach_file.teacher_id == getattr(request.user, "id", None)):
        return HttpResponse(status=403)

    _ensure_achievement_sections(ach_file)
    try:
        from django.db.models import Prefetch

        ev_reports_qs = AchievementEvidenceReport.objects.select_related(
            "report",
            "report__category",
        ).order_by("id")
        sections = (
            AchievementSection.objects.filter(file=ach_file)
            .prefetch_related("evidence_images", Prefetch("evidence_reports", queryset=ev_reports_qs))
            .order_by("code", "id")
        )
        has_evidence_reports = AchievementEvidenceReport.objects.filter(section__file=ach_file).exists()
    except Exception:
        sections = (
            AchievementSection.objects.filter(file=ach_file)
            .prefetch_related("evidence_images", "evidence_reports")
            .order_by("code", "id")
        )
        has_evidence_reports = False

    school = ach_file.school
    primary = (getattr(school, "print_primary_color", None) or "").strip() or "#2563eb"

    # تم حذف شعارات المدارس (logo_file/logo_url) نهائيًا من النظام
    school_logo_url = ""

    try:
        from .pdf_achievement import _static_png_as_data_uri

        ministry_logo_src = _static_png_as_data_uri("img/UntiTtled-1.png")
    except Exception:
        ministry_logo_src = None

    # تحديد URL الرجوع الذكي حسب دور المستخدم
    back_url = "reports:achievement_my_files"  # الافتراضي للمعلم
    is_manager = _is_manager_in_school(request.user, active_school)
    is_staff_user = _is_staff(request.user)
    is_superuser_val = bool(getattr(request.user, "is_superuser", False))
    
    if is_superuser_val or is_manager or is_staff_user:
        back_url = "reports:achievement_school_files"
    
    return render(
        request,
        "reports/pdf/achievement_file.html",
        {
            "file": ach_file,
            "school": school,
            "sections": sections,
            "has_evidence_reports": has_evidence_reports,
            "theme": {"brand": primary},
            "now": timezone.localtime(timezone.now()),
            "school_logo_url": school_logo_url,
            "ministry_logo_src": ministry_logo_src,
            "back_url": back_url,
        },
    )


@login_required(login_url="reports:login")
@role_required({"manager"})
@require_http_methods(["GET", "POST"])
def report_viewer_create(request: HttpRequest) -> HttpResponse:
    """مدير المدرسة ينشئ حساب مشرف تقارير (عرض فقط) داخل المدرسة النشطة."""
    active_school = _get_active_school(request)
    if active_school is None:
        messages.error(request, "فضلاً اختر مدرسة أولاً.")
        return redirect("reports:select_school")

    if (not request.user.is_superuser) and active_school not in _user_manager_schools(request.user):
        messages.error(request, "ليست لديك صلاحية على هذه المدرسة.")
        return redirect("reports:select_school")

    form = ManagerCreateForm(request.POST or None)
    if request.method == "POST":
        if form.is_valid():
            try:
                with transaction.atomic():
                    # ✅ حد أقصى: 2 مشرفي تقارير نشطين لكل مدرسة
                    active_viewers = SchoolMembership.objects.filter(
                        school=active_school,
                        role_type=SchoolMembership.RoleType.REPORT_VIEWER,
                        is_active=True,
                    ).count()
                    if active_viewers >= 2:
                        messages.error(request, "لا يمكن إضافة أكثر من 2 مشرف تقارير (عرض فقط) لهذه المدرسة.")
                        raise ValidationError("viewer_limit")

                    viewer = form.save(commit=True)

                    # تأكيد: لا نعطي صلاحيات موظف لوحة ولا دور manager
                    try:
                        viewer_role = Role.objects.filter(slug="teacher").first()
                        viewer.role = viewer_role
                        viewer.is_staff = False
                        viewer.save(update_fields=["role", "is_staff"])
                    except Exception:
                        try:
                            viewer.is_staff = False
                            viewer.save(update_fields=["is_staff"])
                        except Exception:
                            viewer.save()

                    SchoolMembership.objects.update_or_create(
                        school=active_school,
                        teacher=viewer,
                        role_type=SchoolMembership.RoleType.REPORT_VIEWER,
                        defaults={"is_active": True},
                    )

                messages.success(request, "تم إنشاء حساب مشرف التقارير وربطه بالمدرسة بنجاح.")
                return redirect("reports:manage_teachers")
            except ValidationError as e:
                # رسائل الحد/التحقق
                if "viewer_limit" not in " ".join(getattr(e, "messages", []) or [str(e)]):
                    messages.error(request, " ".join(getattr(e, "messages", []) or [str(e)]))
            except Exception:
                logger.exception("report_viewer_create failed")
                messages.error(request, "تعذّر إنشاء مشرف التقارير. تحقّق من البيانات وحاول مرة أخرى.")
        else:
            messages.error(request, "فضلاً تحقق من الحقول وأعد المحاولة.")

    return render(
        request,
        "reports/add_teacher.html",
        {
            "form": form,
            "page_title": "إضافة مشرف  (عرض فقط)",
            "page_subtitle": "هذا الحساب يستطيع الاطلاع على تقارير المدرسة و ملفات الإنجاز فقط",
            "save_label": "حفظ المشرف",
            "back_url": "reports:manage_teachers",
            "back_label": "رجوع لإدارة المعلمين",
        },
    )


@login_required(login_url="reports:login")
@role_required({"manager"})
@require_http_methods(["GET", "POST"])
def report_viewer_update(request: HttpRequest, pk: int) -> HttpResponse:
    """تعديل بيانات مشرف التقارير (عرض فقط) داخل المدرسة النشطة."""
    active_school = _get_active_school(request)
    if active_school is None:
        messages.error(request, "فضلاً اختر مدرسة أولاً.")
        return redirect("reports:select_school")

    if (not request.user.is_superuser) and active_school not in _user_manager_schools(request.user):
        messages.error(request, "ليست لديك صلاحية على هذه المدرسة.")
        return redirect("reports:select_school")

    viewer = get_object_or_404(Teacher, pk=pk)
    has_membership = SchoolMembership.objects.filter(
        school=active_school,
        teacher=viewer,
        role_type=SchoolMembership.RoleType.REPORT_VIEWER,
    ).exists()
    if not has_membership:
        messages.error(request, "هذا المستخدم ليس مشرف تقارير في المدرسة الحالية.")
        return redirect("reports:manage_teachers")

    form = ManagerCreateForm(request.POST or None, instance=viewer)
    if request.method == "POST":
        if form.is_valid():
            try:
                with transaction.atomic():
                    updated = form.save(commit=True)
                    # ضمان عدم منحه صلاحيات موظف لوحة
                    try:
                        updated.is_staff = False
                        if getattr(getattr(updated, "role", None), "slug", None) == MANAGER_SLUG:
                            updated.role = Role.objects.filter(slug="teacher").first()
                        updated.save(update_fields=["is_staff", "role"])
                    except Exception:
                        pass
                messages.success(request, "✏️ تم تحديث بيانات مشرف التقارير.")
                return redirect("reports:manage_teachers")
            except Exception:
                logger.exception("report_viewer_update failed")
                messages.error(request, "تعذّر تحديث البيانات. حاول لاحقًا.")
        else:
            messages.error(request, "الرجاء تصحيح الأخطاء الظاهرة.")

    return render(
        request,
        "reports/add_teacher.html",
        {
            "form": form,
            "page_title": "تعديل مشرف تقارير (عرض فقط)",
            "page_subtitle": "تعديل بيانات الحساب دون تغيير صلاحياته",
            "save_label": "حفظ التعديلات",
            "back_url": "reports:manage_teachers",
            "back_label": "رجوع لإدارة المعلمين",
        },
    )


@login_required(login_url="reports:login")
@role_required({"manager"})
@require_http_methods(["POST"])
def report_viewer_toggle(request: HttpRequest, pk: int) -> HttpResponse:
    """تفعيل/إيقاف مشرف التقارير داخل المدرسة النشطة."""
    active_school = _get_active_school(request)
    if active_school is None:
        messages.error(request, "فضلاً اختر مدرسة أولاً.")
        return redirect("reports:select_school")

    if (not request.user.is_superuser) and active_school not in _user_manager_schools(request.user):
        messages.error(request, "ليست لديك صلاحية على هذه المدرسة.")
        return redirect("reports:select_school")

    viewer = get_object_or_404(Teacher, pk=pk)
    membership = SchoolMembership.objects.filter(
        school=active_school,
        teacher=viewer,
        role_type=SchoolMembership.RoleType.REPORT_VIEWER,
    ).first()
    if membership is None:
        messages.error(request, "هذا المستخدم ليس مشرف تقارير في المدرسة الحالية.")
        return redirect("reports:manage_teachers")

    try:
        with transaction.atomic():
            target_active = not bool(membership.is_active)
            if target_active:
                # حد 2 مشرفين نشطين
                active_viewers = SchoolMembership.objects.filter(
                    school=active_school,
                    role_type=SchoolMembership.RoleType.REPORT_VIEWER,
                    is_active=True,
                ).exclude(pk=membership.pk).count()
                if active_viewers >= 2:
                    raise ValidationError("لا يمكن تفعيل أكثر من 2 مشرف تقارير (عرض فقط) لهذه المدرسة.")

            membership.is_active = target_active
            membership.save(update_fields=["is_active"])

            viewer.is_active = target_active
            viewer.save(update_fields=["is_active"])

        messages.success(request, "✅ تم تفعيل الحساب." if target_active else "⛔ تم إيقاف الحساب.")
    except ValidationError as e:
        messages.error(request, " ".join(getattr(e, "messages", []) or [str(e)]))
    except Exception:
        logger.exception("report_viewer_toggle failed")
        messages.error(request, "تعذّر تغيير حالة الحساب. حاول لاحقًا.")

    return redirect("reports:manage_teachers")


@login_required(login_url="reports:login")
@role_required({"manager"})
@require_http_methods(["POST"])
def report_viewer_delete(request: HttpRequest, pk: int) -> HttpResponse:
    """حذف (آمن) لمشرف التقارير من المدرسة: تعطيل الحساب وإزالة العضوية من المدرسة."""
    active_school = _get_active_school(request)
    if active_school is None:
        messages.error(request, "فضلاً اختر مدرسة أولاً.")
        return redirect("reports:select_school")

    if (not request.user.is_superuser) and active_school not in _user_manager_schools(request.user):
        messages.error(request, "ليست لديك صلاحية على هذه المدرسة.")
        return redirect("reports:select_school")

    viewer = get_object_or_404(Teacher, pk=pk)
    membership_qs = SchoolMembership.objects.filter(
        school=active_school,
        teacher=viewer,
        role_type=SchoolMembership.RoleType.REPORT_VIEWER,
    )
    if not membership_qs.exists():
        messages.error(request, "هذا المستخدم ليس مشرف تقارير في المدرسة الحالية.")
        return redirect("reports:manage_teachers")

    try:
        with transaction.atomic():
            viewer.is_active = False
            viewer.save(update_fields=["is_active"])
            # إزالة الربط حتى يختفي من القائمة
            membership_qs.delete()
        messages.success(request, "🗑️ تم حذف مشرف التقارير من المدرسة.")
    except Exception:
        logger.exception("report_viewer_delete failed")
        messages.error(request, "تعذّر حذف المستخدم. حاول لاحقًا.")
    return redirect("reports:manage_teachers")

# =========================
# لوحة تقارير المسؤول (Officer)
# =========================
@login_required(login_url="reports:login")
@require_http_methods(["GET"])
def officer_reports(request: HttpRequest) -> HttpResponse:
    active_school = _get_active_school(request)
    user = request.user
    if user.is_superuser:
        return redirect("reports:admin_reports")

    if not (Department is not None and DepartmentMembership is not None):
        messages.error(request, "صلاحيات المسؤول تتطلب تفعيل الأقسام وعضوياتها.")
        return redirect("reports:home")

    if active_school is None:
        messages.error(request, "فضلاً اختر مدرسة أولاً.")
        return redirect("reports:select_school")

    officer_memberships_qs = DepartmentMembership.objects.select_related("department").filter(
        teacher=user,
        role_type=DM_OFFICER,
        department__is_active=True,
        department__school=active_school,
    )
    membership = officer_memberships_qs.first()

    # ✅ يلزم أن تكون مسؤولاً داخل المدرسة النشطة نفسها (بدون fallback عبر مدرسة أخرى)
    if membership is None:
        messages.error(request, "لا تملك صلاحية مسؤول قسم.")
        return redirect("reports:home")

    dept = membership.department if membership else None

    # ✅ الأنواع المسموحة لمسؤول القسم = اتحاد reporttypes لأقسامه داخل المدرسة النشطة
    allowed_cats_qs = None
    if HAS_RTYPE and ReportType is not None:
        allowed_cats_qs = (
            ReportType.objects.filter(
                is_active=True,
                departments__memberships__teacher=user,
                departments__memberships__role_type=DM_OFFICER,
                departments__school=active_school,
            )
            .distinct()
            .order_by("order", "name")
        )

    if allowed_cats_qs is None or not allowed_cats_qs.exists():
        messages.info(request, "لم يتم ربط قسمك بأي أنواع تقارير بعد.")
        empty_page = Paginator(Report.objects.none(), 25).get_page(1)
        return render(
            request,
            "reports/officer_reports.html",
            {
                "reports": empty_page,
                "categories": [],
                "category": "",
                "teacher_name": "",
                "start_date": "",
                "end_date": "",
                "department": dept,
            },
        )

    start_date = request.GET.get("start_date") or ""
    end_date = request.GET.get("end_date") or ""
    teacher_name = request.GET.get("teacher_name", "").strip()
    category = request.GET.get("category") or ""

    qs = Report.objects.select_related("teacher", "category", "school").filter(category__in=allowed_cats_qs)
    qs = _filter_by_school(qs, active_school)

    if start_date:
        qs = qs.filter(report_date__gte=start_date)
    if end_date:
        qs = qs.filter(report_date__lte=end_date)
    if teacher_name:
        qs = qs.filter(Q(teacher__name__icontains=teacher_name) | Q(teacher_name__icontains=teacher_name))
    if category:
        qs = qs.filter(category_id=category)

    qs = qs.order_by("-report_date", "-created_at")

    paginator = Paginator(qs, 25)
    page_number = request.GET.get("page")
    page_obj = paginator.get_page(page_number)
    
    # ✅ إضافة صلاحيات الحذف والمشاركة لكل تقرير (بدون N+1 على قاعدة البيانات)
    is_superuser = bool(getattr(user, "is_superuser", False))
    is_platform = bool(is_platform_admin(user))

    allowed_category_ids = set()
    try:
        allowed_category_ids = set(allowed_cats_qs.values_list("id", flat=True))
    except Exception:
        allowed_category_ids = set()

    manager_school_ids = set()
    if (not is_superuser) and (not is_platform):
        try:
            manager_school_ids = set(
                SchoolMembership.objects.filter(
                    teacher=user,
                    role_type=SchoolMembership.RoleType.MANAGER,
                    is_active=True,
                ).values_list("school_id", flat=True)
            )
        except Exception:
            manager_school_ids = set()

    for report in page_obj:
        if is_superuser:
            allowed = True
        elif is_platform:
            allowed = False
        else:
            allowed = bool(
                getattr(report, "teacher_id", None) == getattr(user, "id", None)
                or (getattr(report, "school_id", None) in manager_school_ids)
                or (getattr(report, "category_id", None) in allowed_category_ids)
            )
        report.user_can_delete = allowed
        report.user_can_share = allowed

    categories_choices = [(str(c.pk), c.name) for c in allowed_cats_qs.order_by("order", "name")]

    return render(
        request,
        "reports/officer_reports.html",
        {
            "reports": page_obj,
            "categories": categories_choices,
            "category": category,
            "teacher_name": teacher_name,
            "start_date": start_date,
            "end_date": end_date,
            "department": dept,
        },
    )


# =========================
# تقارير القسم للأعضاء (عرض + طباعة فقط)
# =========================
@login_required(login_url="reports:login")
@require_http_methods(["GET"])
def department_reports(request: HttpRequest) -> HttpResponse:
    """تقارير القسم لأعضاء القسم (TEACHER) - بدون حذف/مشاركة."""
    active_school = _get_active_school(request)
    user = request.user

    if getattr(user, "is_superuser", False):
        return redirect("reports:admin_reports")

    if not (Department is not None and DepartmentMembership is not None):
        messages.error(request, "عرض تقارير الأقسام يتطلب تفعيل الأقسام وعضوياتها.")
        return redirect("reports:home")

    if active_school is None:
        messages.error(request, "فضلاً اختر مدرسة أولاً.")
        return redirect("reports:select_school")

    # لو كان مسؤول قسم، نوجهه للوحة المسؤول الحالية (تدعم الحذف/المشاركة حسب الصلاحية)
    officer_memberships_qs = DepartmentMembership.objects.select_related("department").filter(
        teacher=user,
        role_type=DM_OFFICER,
        department__is_active=True,
        department__school=active_school,
    )
    if officer_memberships_qs.exists():
        return redirect("reports:officer_reports")

    member_memberships_qs = DepartmentMembership.objects.select_related("department").filter(
        teacher=user,
        role_type=DM_TEACHER,
        department__is_active=True,
        department__school=active_school,
    )
    membership = member_memberships_qs.first()

    if membership is None:
        messages.error(request, "لا تملك صلاحية عضو قسم ضمن المدرسة الحالية.")
        return redirect("reports:home")

    dept = membership.department

    allowed_cats_qs = None
    if HAS_RTYPE and ReportType is not None:
        allowed_cats_qs = (
            ReportType.objects.filter(
                is_active=True,
                departments__memberships__teacher=user,
                departments__memberships__role_type=DM_TEACHER,
                departments__school=active_school,
            )
            .distinct()
            .order_by("order", "name")
        )

    if allowed_cats_qs is None or not allowed_cats_qs.exists():
        messages.info(request, "لم يتم ربط قسمك بأي أنواع تقارير بعد.")
        empty_page = Paginator(Report.objects.none(), 25).get_page(1)
        return render(
            request,
            "reports/officer_reports.html",
            {
                "page_title": "📄 تقارير قسمي (عرض فقط)",
                "reports": empty_page,
                "categories": [],
                "category": "",
                "teacher_name": "",
                "start_date": "",
                "end_date": "",
                "department": dept,
                "can_delete": False,
            },
        )

    start_date = request.GET.get("start_date") or ""
    end_date = request.GET.get("end_date") or ""
    teacher_name = request.GET.get("teacher_name", "").strip()
    category = request.GET.get("category") or ""

    qs = Report.objects.select_related("teacher", "category", "school").filter(category__in=allowed_cats_qs)
    qs = _filter_by_school(qs, active_school)

    if start_date:
        qs = qs.filter(report_date__gte=start_date)
    if end_date:
        qs = qs.filter(report_date__lte=end_date)
    if teacher_name:
        qs = qs.filter(Q(teacher__name__icontains=teacher_name) | Q(teacher_name__icontains=teacher_name))
    if category:
        qs = qs.filter(category_id=category)

    qs = qs.order_by("-report_date", "-created_at")

    paginator = Paginator(qs, 25)
    page_number = request.GET.get("page")
    page_obj = paginator.get_page(page_number)

    categories_choices = [(str(c.pk), c.name) for c in allowed_cats_qs.order_by("order", "name")]

    return render(
        request,
        "reports/officer_reports.html",
        {
            "page_title": "📄 تقارير قسمي (عرض فقط)",
            "reports": page_obj,
            "categories": categories_choices,
            "category": category,
            "teacher_name": teacher_name,
            "start_date": start_date,
            "end_date": end_date,
            "department": dept,
            "can_delete": False,
        },
    )

# =========================
# حذف تقرير (لوحة المدير)
# =========================
@login_required(login_url="reports:login")
@require_http_methods(["POST"])
def admin_delete_report(request: HttpRequest, pk: int) -> HttpResponse:
    """
    حذف تقرير مع التحقق من الصلاحيات.
    يسمح للأشخاص التالية بالحذف:
    - السوبر
    - مدير المدرسة
    - رئيس القسم (OFFICER) للتقارير المرتبطة بقسمه
    - صاحب التقرير نفسه
    
    ✅ عضو القسم (TEACHER) لا يستطيع الحذف (عرض فقط)
    ✅ مشرف المنصة لا يستطيع الحذف (عرض فقط)
    """
    active_school = _get_active_school(request)
    user = request.user
    
    try:
        # جلب التقرير مع احترام المدرسة النشطة
        qs = Report.objects.all()
        qs = _filter_by_school(qs, active_school)
        report = get_object_or_404(qs, pk=pk)
        
        # التحقق من صلاحية الحذف
        if not can_delete_report(user, report, active_school=active_school):
            messages.error(request, "لا تملك صلاحية حذف هذا التقرير.")
            return _safe_redirect(request, "reports:admin_reports")
        
        report.delete()
        messages.success(request, "🗑️ تم حذف التقرير بنجاح.")
    except Exception:
        messages.error(request, "تعذّر حذف التقرير.")
    
    return _safe_redirect(request, "reports:admin_reports")

# =========================
# حذف تقرير (لوحة المسؤول Officer)
# =========================
@login_required(login_url="reports:login")
@user_passes_test(_is_staff_or_officer, login_url="reports:login")
@require_http_methods(["POST"])
def officer_delete_report(request: HttpRequest, pk: int) -> HttpResponse:
    """
    حذف تقرير من قبل:
    - رئيس القسم (OFFICER) للتقارير المرتبطة بقسمه
    - مدير المدرسة
    - السوبر
    
    ✅ عضو القسم (TEACHER) لا يستطيع الحذف (عرض فقط)
    """
    active_school = _get_active_school(request)
    user = request.user
    
    try:
        r = _get_report_for_user_or_404(request, pk)
        
        # التحقق من الصلاحية
        if not can_delete_report(user, r, active_school=active_school):
            messages.error(request, "لا تملك صلاحية حذف هذا التقرير.")
            return _safe_redirect(request, "reports:admin_reports")
        
        r.delete()
        messages.success(request, "🗑️ تم حذف التقرير بنجاح.")
    except Exception:
        messages.error(request, "تعذّر حذف التقرير أو لا تملك صلاحية لذلك.")
    
    return _safe_redirect(request, "reports:admin_reports")

# =========================
# الوصول إلى تقرير معيّن (مع احترام المدرسة النشطة)
# =========================
def _get_report_for_user_or_404(request: HttpRequest, pk: int):
    active_school = _get_active_school(request)
    return svc_get_report_for_user_or_404(user=request.user, pk=pk, active_school=active_school)

from .utils import _resolve_department_for_category, _build_head_decision

# =========================
# طباعة التقرير (نسخة مُحسّنة)
# =========================
@login_required(login_url="reports:login")
@require_http_methods(["GET", "POST"])
def report_print(request: HttpRequest, pk: int) -> HttpResponse:
    try:
        active_school = _get_active_school(request)
        user = request.user

        # ✅ المدير/الموظف/السوبر يجب أن يستطيع طباعة أي تقرير ضمن نطاق المدرسة النشطة
        if getattr(user, "is_superuser", False) or _is_staff(user):
            qs = Report.objects.select_related("teacher", "category")
            if (not getattr(user, "is_superuser", False)) and active_school is None:
                messages.error(request, "فضلاً اختر مدرسة أولاً.")
                return redirect("reports:select_school")
            if active_school is not None and _model_has_field(Report, "school"):
                qs = qs.filter(school=active_school)
            r = get_object_or_404(qs, pk=pk)
        elif is_platform_admin(user):
            qs = Report.objects.select_related("teacher", "category")
            if _model_has_field(Report, "school"):
                allowed_ids = list(platform_allowed_schools_qs(user).values_list("id", flat=True))
                qs = qs.filter(school_id__in=allowed_ids)
                if active_school is not None:
                    qs = qs.filter(school=active_school)
            r = get_object_or_404(qs, pk=pk)
        else:
            r = _get_report_for_user_or_404(request, pk)

        school_scope = getattr(r, "school", None) or active_school

        # ===== تعليقات خاصة (تظهر للمعلم فقط) =====
        show_comments = False
        is_report_owner = False
        can_add_private_comment = False
        private_comments = TeacherPrivateComment.objects.none()
        comment_form = None
        try:
            is_report_owner = getattr(r, "teacher_id", None) == getattr(user, "id", None)
            is_allowed_platform = bool(is_platform_admin(user) and platform_can_access_school(user, school_scope))
            is_manager = _is_manager_in_school(user, school_scope)
            is_staff_user = _is_staff(user)
            can_add_private_comment = bool(is_allowed_platform or is_manager or is_staff_user)
            show_comments = bool(is_report_owner or can_add_private_comment)

            # عرض سجل التعليقات للمعلم + أصحاب الصلاحية (ولا تظهر في الطباعة/المشاركة)
            if is_report_owner or can_add_private_comment:
                private_comments = (
                    TeacherPrivateComment.objects.select_related("created_by")
                    .filter(report=r, teacher=getattr(r, "teacher", None))
                    .order_by("-created_at", "-id")
                )

            try:
                if private_comments is not None:
                    for c in private_comments:
                        try:
                            c.created_by_role_label = _private_comment_role_label(getattr(c, "created_by", None), school_scope)
                        except Exception:
                            c.created_by_role_label = ""
            except Exception:
                pass

            # السماح بإضافة تعليق (يصل للمعلم فقط)
            if can_add_private_comment:
                if request.method == "POST":
                    action = (request.POST.get("action") or "").strip() or "private_comment_create"

                    # create (default)
                    if action == "private_comment_create":
                        comment_form = PrivateCommentForm(request.POST)
                        if comment_form.is_valid():
                            body = comment_form.cleaned_data["body"]
                            with transaction.atomic():
                                TeacherPrivateComment.objects.create(
                                    teacher=r.teacher,
                                    created_by=user,
                                    school=school_scope,
                                    report=r,
                                    body=body,
                                )
                                n = Notification.objects.create(
                                    title="تعليق خاص على تقرير",
                                    message=body,
                                    is_important=True,
                                    school=school_scope,
                                    created_by=user,
                                )
                                NotificationRecipient.objects.create(notification=n, teacher=r.teacher)
                            return redirect(request.get_full_path())

                    # update/delete (only comment author, or superuser)
                    if action in {"private_comment_update", "private_comment_delete"}:
                        comment_id = request.POST.get("comment_id")
                        try:
                            comment_id_int = int(comment_id) if comment_id else None
                        except (TypeError, ValueError):
                            comment_id_int = None

                        if not comment_id_int:
                            return redirect(request.get_full_path())

                        comment = TeacherPrivateComment.objects.filter(
                            pk=comment_id_int,
                            report=r,
                            teacher=getattr(r, "teacher", None),
                        ).first()
                        if comment is None:
                            return redirect(request.get_full_path())

                        is_owner_of_comment = getattr(comment, "created_by_id", None) == getattr(user, "id", None)

                        if action == "private_comment_update":
                            # تعديل: لصاحب التعليق فقط
                            if not is_owner_of_comment:
                                return HttpResponse(status=403)

                        if action == "private_comment_delete":
                            # حذف: لصاحب التعليق فقط، والسوبر يمكنه حذف أي تعليق
                            if not (is_owner_of_comment or getattr(user, "is_superuser", False)):
                                return HttpResponse(status=403)

                        if action == "private_comment_delete":
                            try:
                                comment.delete()
                            except Exception:
                                pass
                            return redirect("reports:report_print", pk=r.pk)

                        body = (request.POST.get("body") or "").strip()
                        if body:
                            try:
                                TeacherPrivateComment.objects.filter(pk=comment.pk).update(body=body)
                            except Exception:
                                pass
                        return redirect(request.get_full_path())
                else:
                    comment_form = PrivateCommentForm()
        except Exception:
            show_comments = False
            is_report_owner = False
            can_add_private_comment = False
            private_comments = TeacherPrivateComment.objects.none()
            comment_form = None

        # اختيار القسم يدويًا عبر ?dept=slug-or-id (اختياري)
        dept = None
        if Department is not None:
            pref = request.GET.get("dept")
            if pref:
                dept_qs = Department.objects.all()
                try:
                    if school_scope is not None and "school" in [f.name for f in Department._meta.get_fields()]:
                        dept_qs = dept_qs.filter(school=school_scope)
                except Exception:
                    pass

                dept = dept_qs.filter(Q(slug=pref) | Q(id=pref)).first() or dept

                # لا نسمح باختيار قسم لا يرتبط بتصنيف التقرير
                cat = getattr(r, "category", None)
                if dept is not None and cat is not None:
                    try:
                        if hasattr(dept, "reporttypes") and getattr(cat, "pk", None) is not None:
                            if not dept.reporttypes.filter(pk=cat.pk).exists():
                                dept = None
                    except Exception:
                        dept = None

        if dept is None:
            cat = getattr(r, "category", None)
            dept = _resolve_department_for_category(cat, school_scope)
            # حماية إضافية: تأكد أن القسم من نفس مدرسة التقرير/المدرسة النشطة
            if dept is not None and school_scope is not None:
                try:
                    dept_school = getattr(dept, "school", None)
                    if dept_school is not None and dept_school != school_scope:
                        dept = None
                except Exception:
                    dept = None

        head_decision = _build_head_decision(dept)

        # اسم مدير المدرسة
        school_principal = ""
        try:
            school_for_principal = getattr(r, "school", None) or _get_active_school(request)
            if school_for_principal is not None:
                principal_membership = (
                    SchoolMembership.objects.select_related("teacher")
                    .filter(
                        school=school_for_principal,
                        role_type=SchoolMembership.RoleType.MANAGER,
                        is_active=True,
                    )
                    .order_by("-id")
                    .first()
                )
                if principal_membership and principal_membership.teacher:
                    school_principal = getattr(principal_membership.teacher, "name", "") or ""
        except Exception:
            school_principal = ""

        if not school_principal:
            school_principal = getattr(settings, "SCHOOL_PRINCIPAL", "")

        # إعدادات المدرسة (الاسم + المرحلة + الشعار)
        school_name = getattr(school_scope, "name", "") if school_scope else getattr(settings, "SCHOOL_NAME", "منصة التقارير المدرسية")
        school_stage = ""
        school_logo_url = ""
        if school_scope:
            try:
                school_stage = getattr(school_scope, "get_stage_display", lambda: "")() or ""
            except Exception:
                school_stage = getattr(school_scope, "stage", "") or ""
            # تم حذف شعارات المدارس (logo_file/logo_url) نهائيًا من النظام
            school_logo_url = ""

        moe_logo_url = (getattr(settings, "MOE_LOGO_URL", "") or "").strip()
        # Optional fallback: allow providing a static path via env/settings
        if not moe_logo_url:
            try:
                moe_logo_static_path = (getattr(settings, "MOE_LOGO_STATIC", "") or "").strip()
                if moe_logo_static_path:
                    moe_logo_url = static(moe_logo_static_path)
            except Exception:
                moe_logo_url = ""

        # Final fallback: always use the bundled ministry logo for printing
        if not moe_logo_url:
            moe_logo_url = static("img/UntiTtled-1.png")

        # تحديد URL الرجوع الذكي حسب دور المستخدم
        back_url = "reports:my_reports"  # الافتراضي للمعلم
        is_manager = _is_manager_in_school(user, school_scope)
        is_staff_user = _is_staff(user)
        is_superuser_val = bool(getattr(user, "is_superuser", False))
        
        if is_superuser_val or is_manager or is_staff_user:
            back_url = "reports:admin_reports"
        
        return render(
            request,
            "reports/report_print.html",
            {
                "r": r,
                "head_decision": head_decision,
                "SCHOOL_PRINCIPAL": school_principal,
                "SCHOOL_NAME": school_name,
                "SCHOOL_STAGE": school_stage,
                "SCHOOL_LOGO_URL": school_logo_url,
                "MOE_LOGO_URL": moe_logo_url,
                "show_comments": show_comments,
                "is_report_owner": is_report_owner,
                "can_add_private_comment": can_add_private_comment,
                "current_user_id": getattr(user, "id", None),
                "is_superuser": is_superuser_val,
                "private_comments": private_comments,
                "comment_form": comment_form,
                "back_url": back_url,
            },
        )
    except Http404:
        raise
    except Exception as e:
        logger.exception(f"Error in report_print view for report {pk}: {e}")
        return render(request, "500.html", {"error": str(e)}, status=500)


def _valid_sharelink_or_404(token: str, *, kind: str) -> ShareLink:
    link = (
        ShareLink.objects.select_related("report", "achievement_file", "school")
        .filter(token=token, kind=kind)
        .first()
    )
    if not link or (not link.is_active) or link.is_expired:
        raise Http404
    return link


@login_required(login_url="reports:login")
@require_http_methods(["GET", "POST"])
def report_share_manage(request: HttpRequest, pk: int) -> HttpResponse:
    """
    تفعيل/إلغاء مشاركة تقرير عبر رابط عام صالح لمدة محددة.
    
    الصلاحيات:
    - صاحب التقرير
    - مدير المدرسة
    - رئيس القسم (OFFICER) للتقارير المرتبطة بقسمه
    - السوبر
    
    ✅ عضو القسم (TEACHER) لا يستطيع المشاركة (عرض فقط)
    """
    active_school = _get_active_school(request)
    user = request.user
    
    report = get_object_or_404(Report.objects.select_related("school"), pk=pk)
    
    # التحقق من الصلاحية
    if not can_share_report(user, report, active_school=active_school):
        messages.error(request, "لا تملك صلاحية مشاركة هذا التقرير.")
        return redirect("reports:admin_reports" if _is_staff(user) else "reports:my_reports")

    expiry_days = get_share_link_default_days(school=report.school)

    now = timezone.now()
    active_link = (
        ShareLink.objects.filter(
            kind=ShareLink.Kind.REPORT,
            report=report,
            is_active=True,
            expires_at__gt=now,
        )
        .order_by("-id")
        .first()
    )

    if request.method == "POST":
        action = (request.POST.get("action") or "").strip().lower()
        if action == "enable":
            with transaction.atomic():
                ShareLink.objects.filter(kind=ShareLink.Kind.REPORT, report=report, is_active=True).update(is_active=False)

                created = None
                for _ in range(6):
                    token = ShareLink.generate_token()
                    try:
                        created = ShareLink.objects.create(
                            token=token,
                            kind=ShareLink.Kind.REPORT,
                            created_by=request.user,
                            school=getattr(report, "school", None),
                            report=report,
                            expires_at=ShareLink.default_expires_at(),
                            is_active=True,
                        )
                        break
                    except IntegrityError:
                        created = None
                        continue

                if created is None:
                    messages.error(request, "تعذر إنشاء رابط مشاركة الآن. حاول مرة أخرى.")
                    return redirect("reports:report_share_manage", pk=report.pk)

            public_url = request.build_absolute_uri(reverse("reports:share_public", args=[created.token]))
            messages.success(
                request,
                f"تم تفعيل مشاركة التقرير ✅ (الرابط صالح لمدة {expiry_days} أيام حتى {timezone.localtime(created.expires_at).strftime('%Y-%m-%d %H:%M')})",
            )
            messages.info(request, f"رابط المشاركة: {public_url}")
            return redirect("reports:report_share_manage", pk=report.pk)

        if action == "disable" and active_link is not None:
            ShareLink.objects.filter(pk=active_link.pk).update(is_active=False)
            messages.success(request, "تم إلغاء رابط المشاركة ✅")
            return redirect("reports:report_share_manage", pk=report.pk)

        messages.error(request, "طلب غير صالح.")
        return redirect("reports:report_share_manage", pk=report.pk)

    public_url = ""
    expires_at_str = ""
    if active_link is not None:
        public_url = request.build_absolute_uri(reverse("reports:share_public", args=[active_link.token]))
        expires_at_str = timezone.localtime(active_link.expires_at).strftime("%Y-%m-%d %H:%M")

    return render(
        request,
        "reports/report_share_manage.html",
        {
            "report": report,
            "active_link": active_link,
            "public_url": public_url,
            "expires_at_str": expires_at_str,
            "expiry_days": expiry_days,
        },
    )


@login_required(login_url="reports:login")
@require_http_methods(["GET", "POST"])
def achievement_share_manage(request: HttpRequest, pk: int) -> HttpResponse:
    """تفعيل/إلغاء مشاركة ملف الإنجاز (PDF) عبر رابط عام صالح لمدة محددة (اختياري للمعلم)."""
    ach_file = get_object_or_404(TeacherAchievementFile.objects.select_related("school"), pk=pk, teacher=request.user)

    expiry_days = get_share_link_default_days(school=ach_file.school)

    now = timezone.now()
    active_link = (
        ShareLink.objects.filter(
            kind=ShareLink.Kind.ACHIEVEMENT,
            achievement_file=ach_file,
            is_active=True,
            expires_at__gt=now,
        )
        .order_by("-id")
        .first()
    )

    if request.method == "POST":
        action = (request.POST.get("action") or "").strip().lower()
        if action == "enable":
            with transaction.atomic():
                ShareLink.objects.filter(kind=ShareLink.Kind.ACHIEVEMENT, achievement_file=ach_file, is_active=True).update(is_active=False)

                created = None
                for _ in range(6):
                    token = ShareLink.generate_token()
                    try:
                        created = ShareLink.objects.create(
                            token=token,
                            kind=ShareLink.Kind.ACHIEVEMENT,
                            created_by=request.user,
                            school=getattr(ach_file, "school", None),
                            achievement_file=ach_file,
                            expires_at=ShareLink.default_expires_at(),
                            is_active=True,
                        )
                        break
                    except IntegrityError:
                        created = None
                        continue

                if created is None:
                    messages.error(request, "تعذر إنشاء رابط مشاركة الآن. حاول مرة أخرى.")
                    return redirect("reports:achievement_share_manage", pk=ach_file.pk)

            public_url = request.build_absolute_uri(reverse("reports:share_public", args=[created.token]))
            messages.success(
                request,
                f"تم تفعيل مشاركة ملف الإنجاز ✅ (الرابط صالح لمدة {expiry_days} أيام حتى {timezone.localtime(created.expires_at).strftime('%Y-%m-%d %H:%M')})",
            )
            messages.info(request, f"رابط المشاركة: {public_url}")
            return redirect("reports:achievement_share_manage", pk=ach_file.pk)

        if action == "disable" and active_link is not None:
            ShareLink.objects.filter(pk=active_link.pk).update(is_active=False)
            messages.success(request, "تم إلغاء رابط المشاركة ✅")
            return redirect("reports:achievement_share_manage", pk=ach_file.pk)

        messages.error(request, "طلب غير صالح.")
        return redirect("reports:achievement_share_manage", pk=ach_file.pk)

    public_url = ""
    expires_at_str = ""
    if active_link is not None:
        public_url = request.build_absolute_uri(reverse("reports:share_public", args=[active_link.token]))
        expires_at_str = timezone.localtime(active_link.expires_at).strftime("%Y-%m-%d %H:%M")

    return render(
        request,
        "reports/achievement_share_manage.html",
        {
            "file": ach_file,
            "active_link": active_link,
            "public_url": public_url,
            "expires_at_str": expires_at_str,
            "expiry_days": expiry_days,
        },
    )


@require_http_methods(["GET"])
def share_public(request: HttpRequest, token: str) -> HttpResponse:
    """عرض عام حسب توكن: تقرير كامل + الصور، أو صفحة تحميل PDF لملف الإنجاز."""
    link = ShareLink.objects.select_related("report", "achievement_file", "school").filter(token=token).first()
    if not link or (not link.is_active) or link.is_expired:
        return render(request, "reports/share_invalid.html", status=404)

    ShareLink.objects.filter(pk=link.pk).update(last_accessed_at=timezone.now())

    if link.kind == ShareLink.Kind.REPORT:
        r = link.report
        if r is None:
            return render(request, "reports/share_invalid.html", status=404)

        school_scope = getattr(r, "school", None) or getattr(link, "school", None)
        cat = getattr(r, "category", None)
        dept = _resolve_department_for_category(cat, school_scope)
        head_decision = _build_head_decision(dept)

        # اسم مدير المدرسة
        school_principal = ""
        try:
            if school_scope is not None:
                principal_membership = (
                    SchoolMembership.objects.select_related("teacher")
                    .filter(
                        school=school_scope,
                        role_type=SchoolMembership.RoleType.MANAGER,
                        is_active=True,
                    )
                    .order_by("-id")
                    .first()
                )
                if principal_membership and principal_membership.teacher:
                    school_principal = getattr(principal_membership.teacher, "name", "") or ""
        except Exception:
            school_principal = ""
        if not school_principal:
            school_principal = getattr(settings, "SCHOOL_PRINCIPAL", "")

        # إعدادات المدرسة
        school_name = getattr(school_scope, "name", "") if school_scope else getattr(settings, "SCHOOL_NAME", "منصة التقارير المدرسية")
        school_stage = ""
        school_logo_url = ""
        if school_scope:
            try:
                school_stage = getattr(school_scope, "get_stage_display", lambda: "")() or ""
            except Exception:
                school_stage = getattr(school_scope, "stage", "") or ""
            # تم حذف شعارات المدارس (logo_file/logo_url) نهائيًا من النظام
            school_logo_url = ""

        moe_logo_url = (getattr(settings, "MOE_LOGO_URL", "") or "").strip()
        if not moe_logo_url:
            try:
                moe_logo_static_path = (getattr(settings, "MOE_LOGO_STATIC", "") or "").strip()
                if moe_logo_static_path:
                    moe_logo_url = static(moe_logo_static_path)
            except Exception:
                moe_logo_url = ""

        # Final fallback: always use the bundled ministry logo for printing
        if not moe_logo_url:
            moe_logo_url = static("img/UntiTtled-1.png")

        return render(
            request,
            "reports/report_print.html",
            {
                "r": r,
                "head_decision": head_decision,
                "SCHOOL_PRINCIPAL": school_principal,
                "SCHOOL_NAME": school_name,
                "SCHOOL_STAGE": school_stage,
                "SCHOOL_LOGO_URL": school_logo_url,
                "MOE_LOGO_URL": moe_logo_url,
                "show_comments": False,
                "private_comments": [],
                "comment_form": None,
                "image1_url": reverse("reports:share_report_image", args=[token, 1]),
                "image2_url": reverse("reports:share_report_image", args=[token, 2]),
                "image3_url": reverse("reports:share_report_image", args=[token, 3]),
                "image4_url": reverse("reports:share_report_image", args=[token, 4]),
            },
        )

    if link.kind == ShareLink.Kind.ACHIEVEMENT:
        ach_file = link.achievement_file
        if ach_file is None:
            return render(request, "reports/share_invalid.html", status=404)

        # نفس تجربة مشاركة التقارير: فتح الرابط يعرض "الملف" مباشرة (صفحة طباعة/معاينة)، مع خيار تنزيل PDF.
        _ensure_achievement_sections(ach_file)
        try:
            from django.db.models import Prefetch

            ev_reports_qs = AchievementEvidenceReport.objects.select_related(
                "report",
                "report__category",
            ).order_by("id")
            sections = (
                AchievementSection.objects.filter(file=ach_file)
                .prefetch_related("evidence_images", Prefetch("evidence_reports", queryset=ev_reports_qs))
                .order_by("code", "id")
            )
            has_evidence_reports = AchievementEvidenceReport.objects.filter(section__file=ach_file).exists()
        except Exception:
            sections = (
                AchievementSection.objects.filter(file=ach_file)
                .prefetch_related("evidence_images", "evidence_reports")
                .order_by("code", "id")
            )
            has_evidence_reports = False

        school = ach_file.school
        primary = (getattr(school, "print_primary_color", None) or "").strip() or "#2563eb"

        # تم حذف شعارات المدارس (logo_file/logo_url) نهائيًا من النظام
        school_logo_url = ""

        try:
            from .pdf_achievement import _static_png_as_data_uri

            ministry_logo_src = _static_png_as_data_uri("img/UntiTtled-1.png")
        except Exception:
            ministry_logo_src = None

        download_url = request.build_absolute_uri(reverse("reports:share_achievement_pdf", args=[token]))
        return render(
            request,
            "reports/pdf/achievement_file.html",
            {
                "file": ach_file,
                "school": school,
                "sections": sections,
                "has_evidence_reports": has_evidence_reports,
                "theme": {"brand": primary},
                "now": timezone.localtime(timezone.now()),
                "public_mode": True,
                "public_download_url": download_url,
                "school_logo_url": school_logo_url,
                "ministry_logo_src": ministry_logo_src,
            },
        )

    return render(request, "reports/share_invalid.html", status=404)


@require_http_methods(["GET"])
def share_report_image(request: HttpRequest, token: str, slot: int) -> HttpResponse:
    link = _valid_sharelink_or_404(token, kind=ShareLink.Kind.REPORT)
    r = link.report
    if r is None:
        raise Http404

    if slot not in (1, 2, 3, 4):
        raise Http404
    field = getattr(r, f"image{slot}", None)
    if not field:
        raise Http404

    try:
        f = field.open("rb")
        resp = FileResponse(f)
        try:
            filename = os.path.basename(getattr(field, "name", "") or "") or f"image{slot}"
            resp["Content-Disposition"] = f'inline; filename="{filename}"'
        except Exception:
            pass
        return resp
    except Exception:
        url = getattr(field, "url", None)
        if url:
            return redirect(url)
        raise


@require_http_methods(["GET"])
def share_achievement_pdf(request: HttpRequest, token: str) -> HttpResponse:
    link = _valid_sharelink_or_404(token, kind=ShareLink.Kind.ACHIEVEMENT)
    ach_file = link.achievement_file
    if ach_file is None:
        raise Http404

    # إذا لم يكن الـ PDF مخزنًا بعد، ولّدْه عند الطلب واحتفظ به لتعمل المشاركة دائمًا.
    if not getattr(ach_file, "pdf_file", None):
        try:
            from django.core.files.base import ContentFile
            from .pdf_achievement import generate_achievement_pdf

            pdf_bytes, filename = generate_achievement_pdf(request=request, ach_file=ach_file)

            try:
                ach_file.pdf_file.save(filename, ContentFile(pdf_bytes), save=False)
                ach_file.pdf_generated_at = timezone.now()
                ach_file.save(update_fields=["pdf_file", "pdf_generated_at"])
            except Exception:
                # حتى لو فشل التخزين (S3/permissions..)، نُرجع الملف للمستخدم.
                pass

            resp = HttpResponse(pdf_bytes, content_type="application/pdf")
            resp["Content-Disposition"] = f'inline; filename="{filename}"'
            return resp
        except OSError as ex:
            # WeasyPrint قد يفشل بسبب مكتبات النظام (خصوصًا على Windows).
            msg = str(ex) or ""
            if "libgobject" in msg or "gobject-2.0" in msg:
                return HttpResponse(
                    "تعذر توليد ملف PDF حاليًا بسبب نقص مكتبات الطباعة على الخادم.",
                    status=503,
                    content_type="text/plain; charset=utf-8",
                )
            if settings.DEBUG:
                raise
            return HttpResponse(
                "تعذر توليد ملف PDF حاليًا.",
                status=503,
                content_type="text/plain; charset=utf-8",
            )
        except Exception:
            if settings.DEBUG:
                raise
            return HttpResponse(
                "تعذر توليد ملف PDF حاليًا.",
                status=503,
                content_type="text/plain; charset=utf-8",
            )

    try:
        f = ach_file.pdf_file.open("rb")
        resp = FileResponse(f, content_type="application/pdf")
        try:
            filename = os.path.basename(getattr(ach_file.pdf_file, "name", "") or "") or "achievement.pdf"
            resp["Content-Disposition"] = f'inline; filename="{filename}"'
        except Exception:
            pass
        return resp
    except Exception:
        url = getattr(ach_file.pdf_file, "url", None)
        if url:
            return redirect(url)
        raise

@login_required(login_url="reports:login")
@require_http_methods(["GET"])
# =========================
# إدارة المعلّمين (مدير فقط)
# =========================
@login_required(login_url="reports:login")
@role_required({"manager"})  # إن كنت تبغى السماح للسوبر دائمًا، خلي role_required يتجاوز للسوبر أو أضف دور admin
@require_http_methods(["GET"])
def manage_teachers(request: HttpRequest) -> HttpResponse:
    active_school = _get_active_school(request)

    # ✅ اجبار اختيار مدرسة لغير السوبر (أوضح وأأمن)
    if not request.user.is_superuser:
        if active_school is None:
            messages.error(request, "فضلاً اختر مدرسة أولاً.")
            return redirect("reports:select_school")

        if active_school not in _user_manager_schools(request.user):
            messages.error(request, "ليست لديك صلاحية على هذه المدرسة.")
            return redirect("reports:select_school")

    term = (request.GET.get("q") or "").strip()

    qs = Teacher.objects.select_related("role").order_by("-id")

    # ✅ عزل حسب المدرسة (نُظهر المعلمين + مشرفي التقارير المرتبطين بالمدرسة)
    if active_school is not None:
        qs = qs.filter(
            school_memberships__school=active_school,
            school_memberships__role_type__in=[
                SchoolMembership.RoleType.TEACHER,
                SchoolMembership.RoleType.REPORT_VIEWER,
            ],
        ).distinct()

    # ✅ بحث
    if term:
        qs = qs.filter(
            Q(name__icontains=term) |
            Q(phone__icontains=term) |
            Q(national_id__icontains=term)
        )

    # ✅ annotate: role_slug/label
    qs = qs.annotate(
        role_slug=F("role__slug"),
        role_label=F("role__name"),
    )

    # ✅ تمييز مشرف التقارير داخل المدرسة النشطة
    if active_school is not None:
        try:
            title_sq = (
                SchoolMembership.objects.filter(
                    school=active_school,
                    teacher=OuterRef("pk"),
                    role_type=SchoolMembership.RoleType.TEACHER,
                )
                .values("job_title")[:1]
            )
            viewer_m = SchoolMembership.objects.filter(
                school=active_school,
                teacher=OuterRef("pk"),
                role_type=SchoolMembership.RoleType.REPORT_VIEWER,
            )
            qs = qs.annotate(
                is_report_viewer=Exists(viewer_m),
                school_job_title=Subquery(title_sq),
            )
        except Exception:
            pass

    # ✅ اسم القسم من Department حسب slug مع تقييد المدرسة (إن كان Department فيه FK school)
    if Department is not None:
        dept_qs = Department.objects.filter(slug=OuterRef("role__slug"))
        if active_school is not None and _model_has_field(Department, "school"):
            dept_qs = dept_qs.filter(Q(school=active_school) | Q(school__isnull=True))
        dept_name_sq = dept_qs.values("name")[:1]
        qs = qs.annotate(role_dept_name=Subquery(dept_name_sq))

    # ✅ منع N+1: Prefetch عضويات الأقسام مرة واحدة وبحقول أقل
    if DepartmentMembership is not None:
        dm_qs = (
            DepartmentMembership.objects
            .select_related("department")
            .only("id", "teacher_id", "role_type", "department__id", "department__name", "department__slug")
            .order_by("department__name")
        )
        if active_school is not None and _model_has_field(Department, "school"):
            dm_qs = dm_qs.filter(Q(department__school=active_school) | Q(department__school__isnull=True))

        qs = qs.prefetch_related(Prefetch("dept_memberships", queryset=dm_qs))

    page = Paginator(qs, 20).get_page(request.GET.get("page"))
    return render(request, "reports/manage_teachers.html", {"teachers_page": page, "term": term})

@login_required(login_url="reports:login")
@role_required({"manager"})
@require_http_methods(["GET", "POST"])
def bulk_import_teachers(request: HttpRequest) -> HttpResponse:
    active_school = _get_active_school(request)
    if active_school is None:
        messages.error(request, "فضلاً اختر مدرسة أولاً.")
        return redirect("reports:select_school")

    # Defense-in-depth: تأكد أن المدير يملك صلاحية على المدرسة النشطة
    try:
        if (not request.user.is_superuser) and active_school not in _user_manager_schools(request.user):
            messages.error(request, "ليست لديك صلاحية على هذه المدرسة.")
            return redirect("reports:select_school")
    except Exception:
        pass

    # الاستيراد يُنشئ عضويات TEACHER؛ نتحقق من وجود اشتراك فعّال لتجنب ValidationError العام
    sub = getattr(active_school, "subscription", None)
    try:
        if sub is None or bool(getattr(sub, "is_expired", True)):
            messages.error(request, "لا يوجد اشتراك فعّال لهذه المدرسة.")
            return redirect("reports:my_subscription")
    except Exception:
        messages.error(request, "لا يوجد اشتراك فعّال لهذه المدرسة.")
        return redirect("reports:my_subscription")

    if request.method == "POST":
        excel_file = request.FILES.get("excel_file")
        if not excel_file:
            messages.error(request, "الرجاء اختيار ملف Excel.")
            return render(request, "reports/bulk_import_teachers.html")

        # تحقق بسيط من الامتداد لتقليل أخطاء المستخدم
        try:
            fname = (getattr(excel_file, "name", "") or "").lower()
            if not fname.endswith(".xlsx"):
                messages.error(request, "الملف غير صالح. الرجاء اختيار ملف بصيغة .xlsx")
                return render(request, "reports/bulk_import_teachers.html")
        except Exception:
            pass

        try:
            import re
            from django.core.exceptions import ValidationError

            def _norm_str(v) -> str:
                return (str(v).strip() if v is not None else "").strip()

            def _normalize_phone(v) -> str:
                if v is None:
                    return ""
                # openpyxl يعيد int/float للأرقام
                try:
                    if isinstance(v, bool):
                        return ""
                    if isinstance(v, int):
                        s = str(v)
                    elif isinstance(v, float):
                        s = str(int(v)) if float(v).is_integer() else str(v)
                    else:
                        s = str(v)
                except Exception:
                    s = str(v)
                s = s.strip()
                # إزالة المسافات والرموز الشائعة (نحتفظ بالأرقام فقط)
                digits = re.sub(r"\D+", "", s)
                if not digits:
                    return s

                # تطبيع أرقام الجوال الشائعة (السعودية)
                # - 9665XXXXXXXX  -> 05XXXXXXXX
                # - 5XXXXXXXX     -> 05XXXXXXXX
                try:
                    if digits.startswith("966"):
                        digits = digits[3:]
                    if digits.startswith("5") and len(digits) == 9:
                        digits = "0" + digits
                    if digits.startswith("5") and len(digits) == 10:
                        digits = "0" + digits[-9:]
                except Exception:
                    pass
                return digits

            def _normalize_national_id(v) -> str:
                s = _norm_str(v)
                if not s:
                    return ""
                digits = re.sub(r"\D+", "", s)
                return digits or s

            wb = openpyxl.load_workbook(excel_file, read_only=True, data_only=True)
            sheet = wb.active

            def _norm_header(v) -> str:
                s = _norm_str(v).lower()
                s = re.sub(r"\s+", "", s)
                s = re.sub(r"[\-_/\\]+", "", s)
                return s

            header_row = next(sheet.iter_rows(min_row=1, max_row=1, values_only=True), None) or ()
            header_norm = [_norm_header(h) for h in (header_row or ())]

            def _find_col_idx(candidates: tuple[str, ...]) -> int | None:
                for i, h in enumerate(header_norm):
                    if not h:
                        continue
                    for c in candidates:
                        if c and c in h:
                            return i
                return None

            # نحدد الأعمدة حسب العناوين لتفادي ملفات فيها أعمدة فارغة/غير متجاورة
            name_idx = _find_col_idx(("الاسمالكامل", "اسم", "الاسم"))
            phone_idx = _find_col_idx(("رقمالجوال", "الجوال", "رقمالهاتف", "الهاتف"))
            nat_idx = _find_col_idx(("رقمالهوية", "الهوية", "السجلالمدني", "رقمالسجل"))

            # توقع الأعمدة: الاسم، رقم الجوال، رقم الهوية (اختياري)
            # الصف الأول عناوين
            parsed_rows: list[tuple[int, str, str, str | None]] = []
            phones_in_file: set[str] = set()
            nat_ids_in_file: set[str] = set()

            max_rows_guard = 2000
            for row_idx, row in enumerate(sheet.iter_rows(min_row=2, values_only=True), start=2):
                if len(parsed_rows) >= max_rows_guard:
                    messages.error(request, f"الملف يحتوي على عدد كبير من الصفوف (>{max_rows_guard}). الرجاء تقسيم الملف.")
                    return render(request, "reports/bulk_import_teachers.html")

                row = row or ()
                # إن استطعنا تحديد الأعمدة من العناوين: نقرأ حسب الفهارس
                if name_idx is not None or phone_idx is not None or nat_idx is not None:
                    name = row[name_idx] if name_idx is not None and name_idx < len(row) else None
                    phone = row[phone_idx] if phone_idx is not None and phone_idx < len(row) else None
                    national_id = row[nat_idx] if nat_idx is not None and nat_idx < len(row) else None
                else:
                    # fallback للملفات التي بلا عناوين واضحة
                    name, phone, national_id = (row + (None, None, None))[:3]

                name_s = _norm_str(name)
                phone_s = _normalize_phone(phone)
                nat_s = _normalize_national_id(national_id) or None

                if nat_s:
                    nat_ids_in_file.add(nat_s)

                if not name_s or not phone_s:
                    # نؤجل الأخطاء إلى مرحلة الرسائل حتى لا نقطع المعالجة مبكرًا
                    parsed_rows.append((row_idx, name_s, phone_s, nat_s))
                    continue

                if phone_s in phones_in_file:
                    parsed_rows.append((row_idx, name_s, phone_s, nat_s))
                    continue
                phones_in_file.add(phone_s)
                parsed_rows.append((row_idx, name_s, phone_s, nat_s))

            if not parsed_rows:
                messages.error(request, "الملف فارغ أو لا يحتوي على بيانات.")
                return render(request, "reports/bulk_import_teachers.html")

            # التحقق من حد الباقة (نحسب فقط العضويات الجديدة الفعلية)
            max_teachers = int(getattr(getattr(sub, "plan", None), "max_teachers", 0) or 0)
            current_count = SchoolMembership.objects.filter(
                school=active_school,
                role_type=SchoolMembership.RoleType.TEACHER,
            ).count()

            phones_unique = {p for p in phones_in_file if p}
            existing_phones_in_school: set[str] = set()
            if phones_unique:
                existing_phones_in_school = set(
                    SchoolMembership.objects.filter(
                        school=active_school,
                        role_type=SchoolMembership.RoleType.TEACHER,
                        teacher__phone__in=phones_unique,
                    ).values_list("teacher__phone", flat=True)
                )

            expected_new = len([p for p in phones_unique if p not in existing_phones_in_school])
            if max_teachers > 0 and (current_count + expected_new) > max_teachers:
                remaining = max_teachers - current_count
                messages.error(request, f"لا يمكن استيراد {expected_new} معلّم جديد. الحد المتبقي في باقتك هو {remaining}.")
                return render(request, "reports/bulk_import_teachers.html")

            created_count = 0
            updated_count = 0
            reactivated_count = 0
            errors: list[str] = []
            seen_phone_rows: set[str] = set()

            # ملاحظة مهمة: أي IntegrityError داخل atomic قد يكسر المعاملة في Postgres.
            # لذلك نستخدم savepoint لكل صف حتى نستمر في الصفوف التالية.
            for row_idx, name_s, phone_s, nat_s in parsed_rows:
                with transaction.atomic():
                    if not name_s or not phone_s:
                        errors.append(f"الصف {row_idx}: الاسم ورقم الجوال مطلوبان.")
                        continue

                    if phone_s in seen_phone_rows:
                        # نعتبره تكرار داخل الملف ونتجاهله بدون تحذير (سلوك متوقع حسب الطلب)
                        continue
                    seen_phone_rows.add(phone_s)

                    # Upsert: تحديد المعلم الموجود ثم تحديث بياناته عند اللزوم
                    teacher = Teacher.objects.filter(phone=phone_s).first()
                    if teacher is None and nat_s:
                        teacher = Teacher.objects.filter(national_id=nat_s).first()

                    if teacher is None:
                        try:
                            teacher = Teacher.objects.create(
                                name=name_s,
                                phone=phone_s,
                                national_id=nat_s,
                                password=make_password(phone_s),  # كلمة المرور الافتراضية هي رقم الجوال
                            )
                            # ضبط الدور الافتراضي للتوافق
                            try:
                                teacher.role = Role.objects.filter(slug="teacher").first()
                                teacher.save(update_fields=["role"])
                            except Exception:
                                pass
                            created_count += 1
                        except (IntegrityError, ValidationError):
                            errors.append(f"الصف {row_idx}: تعذّر إنشاء المستخدم بسبب تعارض في رقم الجوال/الهوية.")
                            continue
                    else:
                        changed_fields: list[str] = []

                        # ✅ تحديث الاسم إذا اختلف
                        try:
                            if name_s and (getattr(teacher, "name", "") or "").strip() != name_s:
                                teacher.name = name_s
                                changed_fields.append("name")
                        except Exception:
                            pass

                        # ✅ تحديث رقم الهوية (إن وُجد)
                        if nat_s:
                            try:
                                current_nat = (getattr(teacher, "national_id", None) or "").strip() or None
                                if current_nat != nat_s:
                                    # تأكد أن الهوية ليست مرتبطة بمستخدم آخر
                                    nat_owner = Teacher.objects.filter(national_id=nat_s).exclude(pk=teacher.pk).first()
                                    if nat_owner is None:
                                        teacher.national_id = nat_s
                                        changed_fields.append("national_id")
                                    else:
                                        errors.append(
                                            f"الصف {row_idx}: رقم الهوية مرتبط بمستخدم آخر، لا يمكن تحديثه تلقائياً."
                                        )
                                        continue
                            except Exception:
                                pass

                        # ✅ تحديث رقم الجوال إذا تم العثور على المعلم عبر الهوية (أو اختلاف الجوال)
                        try:
                            current_phone = (getattr(teacher, "phone", "") or "").strip()
                            if phone_s and current_phone != phone_s:
                                phone_owner = Teacher.objects.filter(phone=phone_s).exclude(pk=teacher.pk).first()
                                if phone_owner is None:
                                    teacher.phone = phone_s
                                    changed_fields.append("phone")
                                    # لو تغير الجوال نحدّث كلمة المرور الافتراضية لتبقى متوافقة (اختياري)
                                    try:
                                        teacher.password = make_password(phone_s)
                                        changed_fields.append("password")
                                    except Exception:
                                        pass
                                else:
                                    errors.append(
                                        f"الصف {row_idx}: رقم الجوال مرتبط بمستخدم آخر، لا يمكن تحديثه تلقائياً."
                                    )
                                    continue
                        except Exception:
                            pass

                        if changed_fields:
                            try:
                                # حفظ الحقول التي تغيرت فقط
                                teacher.save(update_fields=list(dict.fromkeys(changed_fields)))
                                updated_count += 1
                            except Exception:
                                # إن فشل التحديث لسبب غير متوقع، نعتبره خطأ صف ونكمل
                                errors.append(f"الصف {row_idx}: تعذّر تحديث بيانات المستخدم.")
                                continue

                    # ربط المعلم بالمدرسة
                    try:
                        membership, created = SchoolMembership.objects.get_or_create(
                            school=active_school,
                            teacher=teacher,
                            role_type=SchoolMembership.RoleType.TEACHER,
                            defaults={
                                "is_active": True,
                            },
                        )
                    except ValidationError as ve:
                        msg = " ".join(getattr(ve, "messages", []) or []) or str(ve)
                        errors.append(f"الصف {row_idx}: {msg}")
                        continue

                    if not created:
                        # إن كانت العضوية موجودة لكنها غير نشطة، فعّلها
                        try:
                            if hasattr(membership, "is_active") and not bool(getattr(membership, "is_active", True)):
                                membership.is_active = True
                                membership.save(update_fields=["is_active"])
                                reactivated_count += 1
                        except Exception:
                            pass

            if created_count > 0:
                messages.success(request, f"✅ تم إنشاء {created_count} معلّم جديد.")
            if updated_count > 0:
                messages.info(request, f"تم تحديث بيانات {updated_count} معلّم موجود.")
            if reactivated_count > 0:
                messages.info(request, f"تم تفعيل {reactivated_count} عضوية موجودة سابقاً.")
            if errors:
                for err in errors[:10]:
                    messages.warning(request, err)
                if len(errors) > 10:
                    messages.warning(request, f"... وهناك {len(errors)-10} أخطاء أخرى.")

            return redirect("reports:manage_teachers")

        except Exception:
            logger.exception("Bulk import failed")
            messages.error(request, "تعذّر معالجة الملف. تأكد أنه ملف .xlsx صحيح ومطابق للتعليمات.")

    return render(request, "reports/bulk_import_teachers.html")

@login_required(login_url="reports:login")
@role_required({"manager"})
@require_http_methods(["GET", "POST"])
def add_teacher(request: HttpRequest) -> HttpResponse:
    # كل معلم جديد يُربط تلقائياً بالمدرسة النشطة لهذا المدير
    active_school = _get_active_school(request)
    if School.objects.filter(is_active=True).exists():
        if active_school is None:
            messages.error(request, "فضلاً اختر مدرسة أولاً.")
            return redirect("reports:select_school")
        if (not request.user.is_superuser) and active_school not in _user_manager_schools(request.user):
            messages.error(request, "ليست لديك صلاحية على هذه المدرسة.")
            return redirect("reports:select_school")

    if request.method == "POST":
        # إنشاء معلّم فقط: بدون قسم/بدون دور داخل قسم. التكاليف تتم من صفحة أعضاء القسم.
        form = TeacherCreateForm(request.POST)
        job_title = None
        try:
            # يحدد المسمى الوظيفي داخل المدرسة (بنفس الصلاحيات)
            job_title = (request.POST.get("job_title") or "").strip() or None
        except Exception:
            job_title = None

        # ✅ إذا كان رقم الجوال موجودًا مسبقًا: لا ننشئ مستخدمًا جديدًا، بل نربطه بهذه المدرسة
        try:
            phone_raw = (request.POST.get("phone") or "").strip()
            existing_teacher = None
            if phone_raw:
                existing_teacher = Teacher.objects.filter(phone=phone_raw).first()
            if existing_teacher is not None and active_school is not None:
                # تأكيد: صفحة "إضافة معلم" يجب أن تجعل الدور Teacher
                # (لتوافق عرض "القسم/الدور" في شاشة إدارة المعلمين)
                try:
                    if getattr(existing_teacher, "role_id", None) is None:
                        role_obj, _ = Role.objects.get_or_create(
                            slug="teacher",
                            defaults={
                                "name": "المعلم",
                                "is_staff_by_default": False,
                                "can_view_all_reports": False,
                                "is_active": True,
                            },
                        )
                        existing_teacher.role = role_obj
                        existing_teacher.save(update_fields=["role"])
                except Exception:
                    pass

                # هل هو مرتبط فعلاً بهذه المدرسة كـ TEACHER؟
                already = SchoolMembership.objects.filter(
                    school=active_school,
                    teacher=existing_teacher,
                    role_type=SchoolMembership.RoleType.TEACHER,
                    is_active=True,
                ).exists()
                if already:
                    messages.info(request, "المستخدم مرتبط بالفعل بهذه المدرسة.")
                    next_url = _safe_next_url(request.POST.get("next") or request.GET.get("next"))
                    return redirect(next_url or "reports:manage_teachers")

                # نفس منطق حد الباقة الحالي (مع ترك الضمان النهائي للموديل)
                try:
                    sub = getattr(active_school, "subscription", None)
                    if sub is None or bool(getattr(sub, "is_expired", True)):
                        messages.error(request, "لا يوجد اشتراك فعّال لهذه المدرسة.")
                        return render(request, "reports/add_teacher.html", {"form": form, "title": "إضافة مستخدم"})

                    max_teachers = int(getattr(getattr(sub, "plan", None), "max_teachers", 0) or 0)
                    if max_teachers > 0:
                        current_count = SchoolMembership.objects.filter(
                            school=active_school,
                            role_type=SchoolMembership.RoleType.TEACHER,
                        ).count()
                        if current_count >= max_teachers:
                            messages.error(request, f"لا يمكن إضافة أكثر من {max_teachers} معلّم لهذه المدرسة حسب الباقة.")
                            return render(request, "reports/add_teacher.html", {"form": form, "title": "إضافة مستخدم"})
                except Exception:
                    pass

                try:
                    with transaction.atomic():
                        SchoolMembership.objects.update_or_create(
                            school=active_school,
                            teacher=existing_teacher,
                            role_type=SchoolMembership.RoleType.TEACHER,
                            defaults={
                                "is_active": True,
                                **({"job_title": job_title} if job_title else {}),
                            },
                        )
                    messages.success(request, "✅ تم ربط المستخدم الموجود بهذه المدرسة بنجاح (بدون إنشاء حساب جديد).")
                    next_url = _safe_next_url(request.POST.get("next") or request.GET.get("next"))
                    return redirect(next_url or "reports:manage_teachers")
                except ValidationError as e:
                    messages.error(request, " ".join(getattr(e, "messages", []) or [str(e)]))
                except Exception:
                    logger.exception("add_teacher link existing failed")
                    messages.error(request, "حدث خطأ غير متوقع أثناء الربط. جرّب لاحقًا.")
        except Exception:
            # لو فشل هذا المسار لأي سبب نكمل التدفق الطبيعي (وقد يظهر خطأ unique من الفورم)
            pass

        # ✅ منع إضافة معلّم إذا تجاوزت المدرسة حد الباقة (يشمل غير النشط)
        try:
            if active_school is not None:
                sub = getattr(active_school, "subscription", None)
                if sub is None or bool(getattr(sub, "is_expired", True)):
                    messages.error(request, "لا يوجد اشتراك فعّال لهذه المدرسة.")
                    return render(request, "reports/add_teacher.html", {"form": form, "title": "إضافة مستخدم"})

                max_teachers = int(getattr(getattr(sub, "plan", None), "max_teachers", 0) or 0)
                if max_teachers > 0:
                    current_count = SchoolMembership.objects.filter(
                        school=active_school,
                        role_type=SchoolMembership.RoleType.TEACHER,
                    ).count()
                    if current_count >= max_teachers:
                        messages.error(request, f"لا يمكن إضافة أكثر من {max_teachers} معلّم لهذه المدرسة حسب الباقة.")
                        return render(request, "reports/add_teacher.html", {"form": form, "title": "إضافة مستخدم"})
        except Exception:
            # في حال خطأ غير متوقع، نكمل المسار الطبيعي (وسيمنعنا model validation عند الحفظ)
            pass

        if form.is_valid():
            try:
                with transaction.atomic():
                    teacher = form.save(commit=True)
                    # ربط المعلّم بالمدرسة الحالية كـ TEACHER
                    if active_school is not None:
                        SchoolMembership.objects.update_or_create(
                            school=active_school,
                            teacher=teacher,
                            role_type=SchoolMembership.RoleType.TEACHER,
                            defaults={
                                "is_active": True,
                                **({"job_title": job_title} if job_title else {}),
                            },
                        )
                messages.success(request, "✅ تم إضافة المستخدم بنجاح.")
                next_url = _safe_next_url(request.POST.get("next") or request.GET.get("next"))
                return redirect(next_url or "reports:manage_teachers")
            except IntegrityError:
                messages.error(request, "تعذّر الحفظ: قد يكون رقم الجوال أو الهوية مستخدمًا مسبقًا.")
            except ValidationError as e:
                # مثال: تجاوز حد المعلمين حسب الباقة أو عدم وجود اشتراك فعّال
                messages.error(request, " ".join(getattr(e, "messages", []) or [str(e)]))
            except Exception:
                logger.exception("add_teacher failed")
                messages.error(request, "حدث خطأ غير متوقع أثناء الحفظ. جرّب لاحقًا.")
        else:
            messages.error(request, "الرجاء تصحيح الأخطاء الظاهرة.")
    else:
        form = TeacherCreateForm()
    return render(request, "reports/add_teacher.html", {"form": form, "title": "إضافة مستخدم"})

@login_required(login_url="reports:login")
@role_required({"manager"})
@require_http_methods(["GET", "POST"])
def edit_teacher(request: HttpRequest, pk: int) -> HttpResponse:
    active_school = _get_active_school(request)
    teacher = get_object_or_404(Teacher, pk=pk)

    # لا يُسمح للمدير بتعديل معلّم غير مرتبط بمدرسته
    if not getattr(request.user, "is_superuser", False) and active_school is not None:
        has_membership = SchoolMembership.objects.filter(
            school=active_school,
            teacher=teacher,
            role_type=SchoolMembership.RoleType.TEACHER,
            is_active=True,
        ).exists()
        if not has_membership:
            messages.error(request, "لا يمكنك تعديل هذا المعلّم لأنه غير مرتبط بمدرستك.")
            return redirect("reports:manage_teachers")
    if request.method == "POST":
        # تعديل بيانات المعلّم فقط — التكاليف تتم من صفحة أعضاء القسم
        form = TeacherEditForm(request.POST, instance=teacher, active_school=active_school)
        if form.is_valid():
            try:
                with transaction.atomic():
                    form.save(commit=True)
                messages.success(request, "✏️ تم تحديث بيانات المستخدم بنجاح.")
                next_url = _safe_next_url(request.POST.get("next") or request.GET.get("next"))
                return redirect(next_url or "reports:manage_teachers")
            except Exception:
                logger.exception("edit_teacher failed")
                messages.error(request, "حدث خطأ غير متوقع أثناء التحديث.")
        else:
            messages.error(request, "الرجاء تصحيح الأخطاء الظاهرة.")
    else:
        form = TeacherEditForm(instance=teacher, active_school=active_school)

    return render(request, "reports/edit_teacher.html", {"form": form, "teacher": teacher, "title": "تعديل مستخدم"})

@login_required(login_url="reports:login")
@role_required({"manager"})
@require_http_methods(["POST"])
def delete_teacher(request: HttpRequest, pk: int) -> HttpResponse:
    active_school = _get_active_school(request)
    teacher = get_object_or_404(Teacher, pk=pk)

    # لا يُسمح للمدير بحذف معلّم غير مرتبط بمدرسته
    if not getattr(request.user, "is_superuser", False) and active_school is not None:
        has_membership = SchoolMembership.objects.filter(
            school=active_school,
            teacher=teacher,
            role_type=SchoolMembership.RoleType.TEACHER,
            is_active=True,
        ).exists()
        if not has_membership:
            messages.error(request, "لا يمكنك حذف هذا المعلّم لأنه غير مرتبط بمدرستك.")
            return redirect("reports:manage_teachers")
    try:
        with transaction.atomic():
            if active_school is not None and not getattr(request.user, "is_superuser", False):
                # ✅ في وضع تعدد المدارس: لا نحذف الحساب عالميًا، بل نفصل عضويته عن هذه المدرسة فقط
                SchoolMembership.objects.filter(
                    school=active_school,
                    teacher=teacher,
                    role_type__in=[
                        SchoolMembership.RoleType.TEACHER,
                        SchoolMembership.RoleType.REPORT_VIEWER,
                    ],
                ).delete()
                messages.success(request, "🗑️ تم إزالة المستخدم من المدرسة الحالية.")
            else:
                teacher.delete()
                messages.success(request, "🗑️ تم حذف المستخدم.")
    except Exception:
        logger.exception("delete_teacher failed")
        messages.error(request, "تعذّر حذف المستخدم. حاول لاحقًا.")
    next_url = _safe_next_url(request.POST.get("next") or request.GET.get("next"))
    return redirect(next_url or "reports:manage_teachers")

# =========================
# التذاكر (Tickets)
# =========================
def _can_act(user, ticket: Ticket) -> bool:
    if not getattr(user, "is_authenticated", False):
        return False

    # 1. المشرف العام (تذاكر المنصة)
    if ticket.is_platform and getattr(user, "is_superuser", False):
        return True

    # 2. المستلم المباشر (Assignee)
    if ticket.assignee_id == user.id:
        return True

    # 2.1 المستلمون (Recipients)
    try:
        rel = getattr(ticket, "recipients", None)
        if rel is not None and rel.filter(id=user.id).exists():
            return True
    except Exception:
        pass

    # 3. مدير المدرسة (لتذاكر المدرسة)
    # يحق للمدير التحكم في أي تذكرة تابعة لمدرسته
    if not ticket.is_platform and ticket.school_id:
        if SchoolMembership.objects.filter(
            school_id=ticket.school_id,
            teacher=user,
            role_type=SchoolMembership.RoleType.MANAGER,
            is_active=True
        ).exists():
            return True

    # 3.1 المشرف العام (ضمن نطاقه) - لتذاكر المدرسة فقط
    if not ticket.is_platform and ticket.school_id:
        try:
            if is_platform_admin(user) and platform_allowed_schools_qs(user).filter(id=ticket.school_id).exists():
                return True
        except Exception:
            pass

    # 4. مسؤول القسم (Officer)
    # إذا كانت التذكرة تابعة لقسم، فمسؤول القسم يملك صلاحية عليها
    if ticket.department_id and DepartmentMembership is not None:
        if DepartmentMembership.objects.filter(
            department_id=ticket.department_id,
            teacher=user,
            role_type=DepartmentMembership.OFFICER
        ).exists():
            # عزل المدرسة: لمسؤول القسم، لا نسمح بالتعامل مع تذكرة مدرسة أخرى
            try:
                if (not ticket.is_platform) and ticket.school_id:
                    if not SchoolMembership.objects.filter(
                        teacher=user,
                        school_id=ticket.school_id,
                        is_active=True,
                    ).exists():
                        return False
            except Exception:
                pass
            return True

    return False

@login_required(login_url="reports:login")
@require_http_methods(["GET", "POST"])
def request_create(request: HttpRequest) -> HttpResponse:
    active_school = _get_active_school(request)

    # إذا كانت هناك مدارس مفعّلة، نلزم اختيار مدرسة لإنشاء تذكرة مدرسة
    if School.objects.filter(is_active=True).exists() and active_school is None:
        messages.error(request, "فضلاً اختر مدرسة أولاً.")
        return redirect("reports:select_school")

    if request.method == "POST":
        form = TicketCreateForm(request.POST, request.FILES, user=request.user, active_school=active_school)
        if form.is_valid():
            ticket: Ticket = form.save(commit=True, user=request.user)  # يحفظ التذكرة والصور
            if hasattr(ticket, "school") and active_school is not None:
                ticket.school = active_school
                ticket.save(update_fields=["school"])
            messages.success(request, "✅ تم إرسال الطلب بنجاح.")
            return redirect("reports:my_requests")
        messages.error(request, "فضلاً تحقّق من الحقول.")
    else:
        form = TicketCreateForm(user=request.user, active_school=active_school)
    return render(request, "reports/request_create.html", {"form": form})

@login_required(login_url="reports:login")
@role_required({"manager"})
@require_http_methods(["GET", "POST"])
def support_ticket_create(request: HttpRequest) -> HttpResponse:
    """إنشاء تذكرة دعم فني للمنصة (للمدراء فقط)"""
    from .forms import SupportTicketForm
    
    active_school = _get_active_school(request)

    if School.objects.filter(is_active=True).exists():
        if active_school is None:
            messages.error(request, "فضلاً اختر مدرسة أولاً.")
            return redirect("reports:select_school")
        if (not request.user.is_superuser) and active_school not in _user_manager_schools(request.user):
            messages.error(request, "ليست لديك صلاحية على هذه المدرسة.")
            return redirect("reports:select_school")
    
    if request.method == "POST":
        form = SupportTicketForm(request.POST, request.FILES)
        if form.is_valid():
            ticket = form.save(commit=False, user=request.user)
            if active_school:
                ticket.school = active_school
            ticket.save()
            messages.success(request, "✅ تم إرسال طلب الدعم الفني بنجاح.")
            return redirect("reports:my_support_tickets")
        messages.error(request, "فضلاً تحقّق من الحقول.")
    else:
        form = SupportTicketForm()
        
    return render(request, "reports/support_ticket_create.html", {"form": form})


@login_required(login_url="reports:login")
@role_required({"manager"})
def my_support_tickets(request: HttpRequest) -> HttpResponse:
    """عرض تذاكر الدعم الفني الخاصة بالمدير"""
    active_school = _get_active_school(request)

    if School.objects.filter(is_active=True).exists():
        if active_school is None:
            messages.error(request, "فضلاً اختر مدرسة أولاً.")
            return redirect("reports:select_school")
        if (not request.user.is_superuser) and active_school not in _user_manager_schools(request.user):
            messages.error(request, "ليست لديك صلاحية على هذه المدرسة.")
            return redirect("reports:select_school")

    tickets = Ticket.objects.filter(
        creator=request.user, 
        is_platform=True,
        school=active_school,
    ).order_by("-created_at")
    
    return render(request, "reports/my_support_tickets.html", {"tickets": tickets})


@login_required(login_url="reports:login")
def my_requests(request: HttpRequest) -> HttpResponse:
    user = request.user
    active_school = _get_active_school(request)

    if School.objects.filter(is_active=True).exists() and active_school is None:
        messages.error(request, "فضلاً اختر مدرسة أولاً.")
        return redirect("reports:select_school")

    notes_qs = (
        TicketNote.objects.filter(is_public=True)
        .select_related("author")
        .only("id", "body", "created_at", "author__name")
        .order_by("-created_at", "-id")
    )
    base_qs = _filter_by_school(
        Ticket.objects.select_related("assignee", "department")
        .prefetch_related("recipients")
        .prefetch_related(Prefetch("notes", queryset=notes_qs, to_attr="pub_notes"))
        .only("id", "title", "status", "department", "created_at", "assignee__name")
        .filter(creator=user, is_platform=False),
        active_school,
    )

    q = (request.GET.get("q") or "").strip()
    if q:
        base_qs = base_qs.filter(
            Q(title__icontains=q)
            | Q(id__icontains=q)
            | Q(assignee__name__icontains=q)
            | Q(recipients__name__icontains=q)
        ).distinct()

    counts = dict(base_qs.values("status").annotate(c=Count("id")).values_list("status", "c"))
    stats = {
        "open": counts.get("open", 0),
        "in_progress": counts.get("in_progress", 0),
        "done": counts.get("done", 0),
        "rejected": counts.get("rejected", 0),
    }

    status = request.GET.get("status")
    qs = base_qs
    if status in {"open", "in_progress", "done", "rejected"}:
        qs = qs.filter(status=status)

    order = request.GET.get("order") or "-created_at"
    allowed_order = {"-created_at", "created_at", "-id", "id"}
    if order not in allowed_order:
        order = "-created_at"
    if order in {"created_at", "-created_at"}:
        qs = qs.order_by(order, "-id")
    else:
        qs = qs.order_by(order)

    page = Paginator(qs, 12).get_page(request.GET.get("page") or 1)
    view_mode = request.GET.get("view", "list")

    return render(
        request,
        "reports/my_requests.html",
        {"tickets": page, "page_obj": page, "stats": stats, "view_mode": view_mode},
    )

@login_required(login_url="reports:login")
@require_http_methods(["GET", "POST"])
def ticket_detail(request: HttpRequest, pk: int) -> HttpResponse:
    active_school = _get_active_school(request)
    user = request.user

    # احضر التذكرة مع الحقول المطلوبة مع احترام المدرسة النشطة
    base_qs = Ticket.objects.select_related("creator", "assignee", "department").prefetch_related("recipients").only(
        "id", "title", "body", "status", "department", "created_at",
        "creator__name", "assignee__name", "assignee_id", "creator_id", "is_platform", "school_id"
    )
    
    # إذا كانت التذكرة للمنصة، لا نفلتر بالمدرسة (لأنها قد لا تكون مرتبطة بمدرسة أو نريد السماح للمدير برؤيتها)
    # لكن يجب التأكد أن المستخدم هو المنشئ أو مشرف نظام
    # سنحاول جلب التذكرة أولاً بدون فلتر المدرسة إذا كانت is_platform=True
    
    # الحل الأبسط: نعدل _filter_by_school ليتجاهل الفلتر إذا كانت التذكرة is_platform=True
    # لكن _filter_by_school تعمل على QuerySet.
    
    # لذا سنقوم بالتالي:
    # 1. نحاول جلب التذكرة بـ PK فقط
    # 2. نتحقق من الصلاحية يدوياً
    
    t = get_object_or_404(base_qs, pk=pk)
    
    # التحقق من الوصول
    if t.is_platform:
        # تذاكر المنصة: مسموحة للمنشئ (المدير) أو المشرف العام
        if not (user.is_superuser or t.creator_id == user.id):
             raise Http404("ليس لديك صلاحية لعرض هذه التذكرة.")
    else:
        # تذاكر المدرسة: نلزم عضوية المستخدم في مدرسة التذكرة
        if not user.is_superuser:
            if not t.school_id:
                raise Http404("هذه التذكرة غير مرتبطة بمدرسة.")
            if is_platform_admin(user):
                if not platform_allowed_schools_qs(user).filter(id=t.school_id).exists():
                    raise Http404("ليس لديك صلاحية لعرض هذه التذكرة.")
            else:
                if not SchoolMembership.objects.filter(
                    teacher=user,
                    school_id=t.school_id,
                    is_active=True,
                ).exists():
                    raise Http404("ليس لديك صلاحية لعرض هذه التذكرة.")

            # عند تعدد المدارس: نلزم توافق المدرسة النشطة مع مدرسة التذكرة
            if active_school is not None and t.school_id != active_school.id:
                raise Http404("هذه التذكرة تابعة لمدرسة أخرى.")

    is_owner = (t.creator_id == user.id)
    can_act = _can_act(user, t)

    if request.method == "POST":
        status_val = (request.POST.get("status") or "").strip()
        note_txt   = (request.POST.get("note") or "").strip()
        changed = False

        status_label = dict(getattr(Ticket.Status, "choices", [])).get

        locked_statuses = {Ticket.Status.DONE, Ticket.Status.REJECTED}
        is_locked_now_or_will_be = (t.status in locked_statuses) or (status_val in locked_statuses)

        # إضافة ملاحظة (المرسل أو من يملك الصلاحية)
        # يسمح للمرسل بإضافة ملاحظات (للتواصل) ولكن لا يملك صلاحية تغيير الحالة إلا إذا كان من ضمن المستلمين/الإدارة
        can_comment = False
        if is_owner or can_act:
            can_comment = True

        if note_txt and can_comment and is_locked_now_or_will_be:
            messages.warning(request, "لا يمكن إضافة ملاحظة عندما تكون حالة الطلب مكتمل أو مرفوض.")

        if note_txt and can_comment and (not is_locked_now_or_will_be):
            try:
                with transaction.atomic():
                    TicketNote.objects.create(
                        ticket=t, author=request.user, body=note_txt, is_public=True
                    )

                    # خيار: إعادة الفتح تلقائيًا عند ملاحظة المرسل (إن كانت مفعّلة)
                    if AUTO_REOPEN_ON_SENDER_NOTE and is_owner and t.status in {
                        Ticket.Status.DONE, Ticket.Status.REJECTED, Ticket.Status.IN_PROGRESS
                    }:
                        old_status = t.status
                        t.status = Ticket.Status.OPEN
                        try:
                            t.save(update_fields=["status"])
                        except Exception:
                            t.save()
                        TicketNote.objects.create(
                            ticket=t,
                            author=request.user,
                            body=f"تغيير الحالة تلقائيًا بسبب ملاحظة المرسل: {status_label(old_status, old_status)} → {status_label(Ticket.Status.OPEN, Ticket.Status.OPEN)}",
                            is_public=True,
                        )
                changed = True
            except Exception:
                logger.exception("Failed to create note")
                messages.error(request, "تعذّر حفظ الملاحظة.")

        # تغيير الحالة (لمن له صلاحية فقط)
        if status_val:
            if not can_act:
                messages.warning(request, "لا يمكنك تغيير حالة هذا الطلب. يمكنك فقط إضافة ملاحظة.")
            else:
                valid_statuses = {k for k, _ in Ticket.Status.choices}
                if status_val in valid_statuses and status_val != t.status:
                    old = t.status
                    t.status = status_val
                    try:
                        t.save(update_fields=["status"])
                    except Exception:
                        t.save()
                    changed = True
                    try:
                        TicketNote.objects.create(
                            ticket=t,
                            author=request.user,
                            body="تغيير الحالة: {} → {}".format(status_label(old, old), status_label(status_val, status_val)),
                            is_public=True,
                        )
                    except Exception:
                        logger.exception("Failed to create system note")

        if changed:
            messages.success(request, "تم حفظ التغييرات.")
        else:
            messages.info(request, "لا يوجد تغييرات.")
        return redirect("reports:ticket_detail", pk=pk)

    # ===== صور التذكرة (بغض النظر عن related_name) =====
    images_manager = getattr(t, "images", None)  # لو related_name='images'
    if images_manager is None:
        images_manager = getattr(t, "ticketimage_set", None)  # الاسم الافتراضي إن وُجد

    if images_manager is not None and hasattr(images_manager, "all"):
        images = list(images_manager.all().only("id", "image"))
    else:
        # fallback مضمون
        images = list(TicketImage.objects.filter(ticket_id=t.id).only("id", "image"))

    # سجلّ الملاحظات + نموذج الإجراء (إن وُجدت صلاحية)
    notes_qs = (
        t.notes.select_related("author")
        .only("id", "body", "created_at", "author__name")
        .order_by("-created_at", "-id")
    )
    form = TicketActionForm(initial={"status": t.status}) if can_act else None

    ctx = {
        "t": t,
        "images": images,     # ← استخدم هذا في القالب
        "notes": notes_qs,
        "form": form,
        "can_act": can_act,
        "is_owner": is_owner,
    }
    return render(request, "reports/ticket_detail.html", ctx)


@login_required(login_url="reports:login")
@require_http_methods(["GET", "POST"])
def ticket_note_edit(request: HttpRequest, pk: int) -> HttpResponse:
    """تعديل ملاحظة طلب: فقط صاحب الملاحظة."""
    active_school = _get_active_school(request)
    user = request.user

    note = get_object_or_404(
        TicketNote.objects.select_related("ticket", "ticket__school", "author"),
        pk=pk,
    )
    t = note.ticket

    if note.author_id != user.id:
        raise Http404("ليس لديك صلاحية لتعديل هذه الملاحظة.")

    # لا نسمح بالتعديل بعد اكتمال/رفض الطلب
    if getattr(t, "status", None) in {Ticket.Status.DONE, Ticket.Status.REJECTED}:
        messages.warning(request, "لا يمكن تعديل الملاحظات بعد تحويل الطلب إلى مكتمل أو مرفوض.")
        return redirect("reports:ticket_detail", pk=t.id)

    # تحقق الوصول للتذكرة (نفس منطق ticket_detail)
    if getattr(t, "is_platform", False):
        if not (user.is_superuser or t.creator_id == user.id or note.author_id == user.id):
            raise Http404("ليس لديك صلاحية لعرض هذه التذكرة.")
    else:
        if not user.is_superuser:
            if not getattr(t, "school_id", None):
                raise Http404("هذه التذكرة غير مرتبطة بمدرسة.")
            if is_platform_admin(user):
                if not platform_allowed_schools_qs(user).filter(id=t.school_id).exists():
                    raise Http404("ليس لديك صلاحية لعرض هذه التذكرة.")
            else:
                if not SchoolMembership.objects.filter(teacher=user, school_id=t.school_id, is_active=True).exists():
                    raise Http404("ليس لديك صلاحية لعرض هذه التذكرة.")
            if active_school is not None and t.school_id != active_school.id:
                raise Http404("هذه التذكرة تابعة لمدرسة أخرى.")

    next_url = (request.GET.get("next") or "").strip()
    if next_url and not url_has_allowed_host_and_scheme(next_url, allowed_hosts={request.get_host()}):
        next_url = ""

    if request.method == "POST":
        form = TicketNoteEditForm(request.POST, instance=note)
        if form.is_valid():
            form.save()
            messages.success(request, "تم تعديل الملاحظة.")
            return redirect(next_url or "reports:ticket_detail", pk=t.id)
    else:
        form = TicketNoteEditForm(instance=note)

    return render(
        request,
        "reports/ticket_note_edit.html",
        {"t": t, "note": note, "form": form, "next": next_url or ""},
    )


@login_required(login_url="reports:login")
@require_http_methods(["GET"])
def ticket_print(request: HttpRequest, pk: int) -> HttpResponse:
    """طباعة رسمية للطلب (A4) بنفس أسلوب طباعة التقارير."""
    active_school = _get_active_school(request)
    user = request.user

    base_qs = Ticket.objects.select_related("creator", "assignee", "department", "school").prefetch_related("recipients").only(
        "id",
        "title",
        "body",
        "status",
        "department",
        "created_at",
        "creator__name",
        "assignee__name",
        "assignee_id",
        "creator_id",
        "is_platform",
        "school_id",
        "attachment",
        "school__name",
        "school__stage",
    )

    t = get_object_or_404(base_qs, pk=pk)

    # نفس منطق الصلاحيات في ticket_detail
    if t.is_platform:
        if not (user.is_superuser or t.creator_id == user.id):
            raise Http404("ليس لديك صلاحية لعرض هذه التذكرة.")
    else:
        if not user.is_superuser:
            if not t.school_id:
                raise Http404("هذه التذكرة غير مرتبطة بمدرسة.")
            if is_platform_admin(user):
                if not platform_allowed_schools_qs(user).filter(id=t.school_id).exists():
                    raise Http404("ليس لديك صلاحية لعرض هذه التذكرة.")
            else:
                if not SchoolMembership.objects.filter(
                    teacher=user,
                    school_id=t.school_id,
                    is_active=True,
                ).exists():
                    raise Http404("ليس لديك صلاحية لعرض هذه التذكرة.")

            if active_school is not None and t.school_id != active_school.id:
                raise Http404("هذه التذكرة تابعة لمدرسة أخرى.")

    # المرفقات/الصور
    images_manager = getattr(t, "images", None)
    if images_manager is None:
        images_manager = getattr(t, "ticketimage_set", None)
    if images_manager is not None and hasattr(images_manager, "all"):
        images = list(images_manager.all().only("id", "image"))
    else:
        images = list(TicketImage.objects.filter(ticket_id=t.id).only("id", "image"))

    notes_qs = (
        t.notes.select_related("author")
        .only("id", "body", "created_at", "author__name")
        .order_by("-created_at", "-id")
    )

    # إعدادات المدرسة/الشعارات مثل report_print
    school_scope = getattr(t, "school", None) or active_school
    school_name = getattr(school_scope, "name", "") if school_scope else getattr(settings, "SCHOOL_NAME", "منصة التقارير المدرسية")
    school_stage = ""
    school_logo_url = ""
    if school_scope:
        try:
            school_stage = getattr(school_scope, "get_stage_display", lambda: "")() or ""
        except Exception:
            school_stage = getattr(school_scope, "stage", "") or ""
        # تم حذف شعارات المدارس (logo_file/logo_url) نهائيًا من النظام
        school_logo_url = ""

    moe_logo_url = (getattr(settings, "MOE_LOGO_URL", "") or "").strip()
    if not moe_logo_url:
        try:
            moe_logo_static_path = (getattr(settings, "MOE_LOGO_STATIC", "") or "").strip()
            if moe_logo_static_path:
                moe_logo_url = static(moe_logo_static_path)
        except Exception:
            moe_logo_url = ""
    if not moe_logo_url:
        moe_logo_url = static("img/UntiTtled-1.png")

    # خصائص المرفق الرئيسي
    attachment_name_lower = (getattr(getattr(t, "attachment", None), "name", "") or "").lower()
    attachment_is_image = attachment_name_lower.endswith((".jpg", ".jpeg", ".png", ".webp"))
    attachment_is_pdf = attachment_name_lower.endswith(".pdf")

    now_local = timezone.localtime(timezone.now())

    return render(
        request,
        "reports/ticket_print.html",
        {
            "t": t,
            "notes": notes_qs,
            "images": images,
            "now": now_local,
            "SCHOOL_NAME": school_name,
            "SCHOOL_STAGE": school_stage,
            "SCHOOL_LOGO_URL": school_logo_url,
            "MOE_LOGO_URL": moe_logo_url,
            "attachment_is_image": attachment_is_image,
            "attachment_is_pdf": attachment_is_pdf,
        },
    )

@login_required(login_url="reports:login")
@user_passes_test(_is_staff, login_url="reports:login")
@require_http_methods(["GET", "POST"])
def admin_request_update(request: HttpRequest, pk: int) -> HttpResponse:
    return ticket_detail(request, pk)

# ========= دعم الأقسام =========
def _dept_code_for(dept_obj_or_code) -> str:
    if hasattr(dept_obj_or_code, "slug") and getattr(dept_obj_or_code, "slug"):
        return getattr(dept_obj_or_code, "slug")
    if hasattr(dept_obj_or_code, "code") and getattr(dept_obj_or_code, "code"):
        return getattr(dept_obj_or_code, "code")
    return str(dept_obj_or_code or "").strip()

def _arabic_label_for_in_school(dept_obj_or_code, active_school: Optional[School] = None) -> str:
    """نسخة آمنة من _arabic_label_for تربط التسمية بالمدرسة النشطة لتجنب تداخل slugs بين المدارس."""
    if hasattr(dept_obj_or_code, "name") and getattr(dept_obj_or_code, "name"):
        return dept_obj_or_code.name
    code = (
        getattr(dept_obj_or_code, "slug", None)
        or getattr(dept_obj_or_code, "code", None)
        or (dept_obj_or_code if isinstance(dept_obj_or_code, str) else "")
    )
    return _role_display_map(active_school).get(code, code or "—")

def _resolve_department_by_code_or_pk(code_or_pk: str, school: Optional[School] = None) -> Tuple[Optional[object], str, str]:
    dept_obj = None
    dept_code = (code_or_pk or "").strip()

    if Department is not None:
        try:
            qs = Department.objects.all()
            if school is not None and _model_has_field(Department, "school"):
                qs = qs.filter(school=school)
            dept_obj = qs.filter(slug__iexact=dept_code).first()
            if not dept_obj:
                try:
                    dept_obj = qs.filter(pk=int(dept_code)).first()
                except (ValueError, TypeError):
                    dept_obj = None
        except Exception:
            dept_obj = None

        if dept_obj:
            dept_code = getattr(dept_obj, "slug", dept_code)

    dept_label = _arabic_label_for_in_school(dept_obj or dept_code, school)
    return dept_obj, dept_code, dept_label

def _members_for_department(dept_code: str, school: Optional[School] = None):
    if not dept_code:
        return Teacher.objects.none()
    if DepartmentMembership is None:
        return Teacher.objects.none()

    mem_qs = DepartmentMembership.objects.filter(department__slug__iexact=dept_code)
    if school is not None:
        mem_qs = mem_qs.filter(department__school=school)
    member_ids = mem_qs.values_list("teacher_id", flat=True)

    qs = Teacher.objects.filter(is_active=True, id__in=member_ids).distinct()
    if school is not None:
        qs = qs.filter(
            school_memberships__school=school,
        )
    return qs.order_by("name")

def _user_department_codes(user, active_school: Optional[School] = None) -> list[str]:
    codes = set()

    # في وضع تعدد المدارس، يجب تحديد المدرسة النشطة لتجنب تداخل slugs بين المدارس
    try:
        if active_school is None and School.objects.filter(is_active=True).count() > 1:
            return []
    except Exception:
        # fail-closed إذا تعذر تحديد عدد المدارس
        if active_school is None:
            return []

    if DepartmentMembership is not None:
        try:
            mem_qs = DepartmentMembership.objects.filter(teacher=user)
            if active_school is not None:
                mem_qs = mem_qs.filter(department__school=active_school)
            mem_codes = mem_qs.values_list("department__slug", flat=True)
            for c in mem_codes:
                if c:
                    codes.add(c)
        except Exception:
            logger.exception("Failed to fetch user department codes")

    return list(codes)

def _tickets_stats_for_department(dept_code: str, school: Optional[School] = None) -> dict:
    qs = Ticket.objects.filter(department__slug=dept_code)
    qs = _filter_by_school(qs, school)
    return {
        "open": qs.filter(status="open").count(),
        "in_progress": qs.filter(status="in_progress").count(),
        "done": qs.filter(status="done").count(),
    }

def _all_departments(active_school: Optional[School] = None):
    items = []
    if Department is not None:
        qs = Department.objects.all().order_by("id")
        if active_school is not None and hasattr(Department, "school"):
            qs = qs.filter(school=active_school)
        for d in qs:
            code = _dept_code_for(d)
            stats = _tickets_stats_for_department(code, active_school)
            # ملاحظة للتوسع: الاعتماد على role.slug وحده يخلط المدارس عند تكرار slugs.
            # نُقيّد العد داخل المدرسة النشطة عبر SchoolMembership.
            if active_school is not None:
                role_ids = set(
                    SchoolMembership.objects.filter(
                        school=active_school,
                        is_active=True,
                        teacher__is_active=True,
                        teacher__role__slug=code,
                    ).values_list("teacher_id", flat=True)
                )
            else:
                role_ids = set(Teacher.objects.filter(role__slug=code, is_active=True).values_list("id", flat=True))
            member_ids = set()
            if DepartmentMembership is not None:
                member_ids = set(DepartmentMembership.objects.filter(department=d).values_list("teacher_id", flat=True))
            members_count = len(role_ids | member_ids)
            items.append(
                {
                    "pk": d.pk,
                    "code": code,
                    "name": _arabic_label_for_in_school(d, active_school),
                    "is_active": getattr(d, "is_active", True),
                    "members_count": members_count,
                    "stats": stats,
                }
            )
    else:
        items = []
    return items

class _DepartmentForm(forms.ModelForm):
    class Meta:
        model = Department
        fields: list[str] = []
        if model is not None:
            for fname in ("name", "slug", "role_label", "is_active"):
                if hasattr(model, fname):
                    fields.append(fname)

    def clean(self):
        cleaned = super().clean()
        return cleaned

def get_department_form():
    if Department is not None and 'DepartmentForm' in globals() and (DepartmentForm is not None):
        return DepartmentForm
    if Department is not None:
        return _DepartmentForm
    return None


# ---- إعدادات المدرسة الحالية (للمدير أو المشرف العام) ----
class _SchoolSettingsForm(forms.ModelForm):
    years_text = forms.CharField(
        label="السنوات الدراسية المتاحة (هجري)",
        required=False,
        widget=forms.TextInput(attrs={"class": "input", "placeholder": "1446-1447, 1447-1448 ..."}),
        help_text="أدخل السنوات المسموحة مفصولة بفاصلة. اتركها فارغة لاستخدام الوضع الافتراضي."
    )

    class Meta:
        model = School
        fields = [
            "name",
            "stage",
            "gender",
            "city",
            "phone",
            "share_link_default_days",
        ]

    def __init__(self, *args, **kwargs):
        super().__init__(*args, **kwargs)
        if self.instance.pk and self.instance.allowed_academic_years:
            self.fields["years_text"].initial = ", ".join(self.instance.allowed_academic_years)

    def clean_years_text(self):
        import re
        data = self.cleaned_data.get("years_text", "")
        if not data:
            return []
        years = []
        for part in data.replace("،", ",").split(","):
            p = part.strip()
            if not p:
                continue
            if not re.match(r"^\d{4}-\d{4}$", p):
                 # يمكن تجاهل غير الصالح أو رفع خطأ. سنرفض الخطأ لتنبيه المستخدم.
                pass 
            years.append(p)
        
        # ترتيبها
        years.sort()
        return years

    def save(self, commit=True):
        self.instance.allowed_academic_years = self.cleaned_data["years_text"]
        return super().save(commit=commit)


@login_required(login_url="reports:login")
@require_http_methods(["GET", "POST"])
def school_settings(request: HttpRequest) -> HttpResponse:
    """إعدادات المدرسة الحالية (الاسم، الشعار...).

    - متاحة لمدير المدرسة على مدرسته النشطة فقط.
    - متاحة للمشرف العام على أي مدرسة بعد اختيارها كـ active_school.
    """
    active_school = _get_active_school(request)
    if active_school is None:
        messages.error(request, "فضلاً اختر مدرسة أولاً.")
        return redirect("reports:select_school")

    # تحقق من الصلاحيات
    if not (getattr(request.user, "is_superuser", False) or active_school in _user_manager_schools(request.user)):
        messages.error(request, "لا تملك صلاحية تعديل إعدادات هذه المدرسة.")
        return redirect("reports:admin_dashboard")

    form = _SchoolSettingsForm(request.POST or None, request.FILES or None, instance=active_school)
    if request.method == "POST":
        if form.is_valid():
            form.save()
            messages.success(request, "تم تحديث إعدادات المدرسة بنجاح.")
            return redirect("reports:admin_dashboard")
        # في حال وجود أخطاء نعرضها للمستخدم ليسهل معرفة سبب الفشل
        messages.error(request, "تعذّر الحفظ. تحقّق من الحقول.")
        try:
            for field, errors in form.errors.items():
                label = form.fields.get(field).label if field in form.fields else field
                joined = "; ".join(errors)
                messages.error(request, f"{label}: {joined}")
        except Exception:
            # لا نكسر الصفحة إن حدث خطأ أثناء بناء الرسالة
            pass

    return render(request, "reports/school_settings.html", {"form": form, "school": active_school})


# ---- إدارة المدارس (إنشاء/تعديل/حذف) للمشرف العام ----
class _SchoolAdminForm(forms.ModelForm):
    class Meta:
        model = School
        fields = [
            "name",
            "code",
            "stage",
            "gender",
            "city",
            "phone",
            "is_active",
        ]


@login_required(login_url="reports:login")
@user_passes_test(lambda u: getattr(u, "is_superuser", False), login_url="reports:login")
@require_http_methods(["GET", "POST"])
def school_create(request: HttpRequest) -> HttpResponse:
    form = _SchoolAdminForm(request.POST or None)
    if request.method == "POST":
        if form.is_valid():
            form.save()
            messages.success(request, "تم إنشاء المدرسة بنجاح.")
            return redirect("reports:schools_admin_list")
        messages.error(request, "تعذّر الحفظ. تحقّق من الحقول.")
    return render(request, "reports/school_form.html", {"form": form, "mode": "create"})


@login_required(login_url="reports:login")
@user_passes_test(lambda u: getattr(u, "is_superuser", False), login_url="reports:login")
@require_http_methods(["GET", "POST"])
def school_update(request: HttpRequest, pk: int) -> HttpResponse:
    school = get_object_or_404(School, pk=pk)
    form = _SchoolAdminForm(request.POST or None, instance=school)
    if request.method == "POST":
        if form.is_valid():
            form.save()
            messages.success(request, "تم تحديث بيانات المدرسة.")
            return redirect("reports:schools_admin_list")
        messages.error(request, "تعذّر الحفظ. تحقّق من الحقول.")
    return render(request, "reports/school_form.html", {"form": form, "mode": "edit", "school": school})


@login_required(login_url="reports:login")
@user_passes_test(lambda u: getattr(u, "is_superuser", False), login_url="reports:login")
@require_http_methods(["POST"])
def school_delete(request: HttpRequest, pk: int) -> HttpResponse:
    school = get_object_or_404(School, pk=pk)
    name = school.name
    from .middleware import set_audit_logging_suppressed

    try:
        set_audit_logging_suppressed(True)
        school.delete()
        messages.success(request, f"🗑️ تم حذف المدرسة «{name}» وكل بياناتها المرتبطة.")
    except Exception:
        logger.exception("school_delete failed")
        messages.error(request, "تعذّر حذف المدرسة. ربما توجد قيود على البيانات المرتبطة.")
    finally:
        set_audit_logging_suppressed(False)
    return redirect("reports:schools_admin_list")


# ---- لوحة إدارة المدارس ومدراء المدارس (للسوبر أدمن) ----
@login_required(login_url="reports:login")
@user_passes_test(lambda u: getattr(u, "is_superuser", False), login_url="reports:login")
@require_http_methods(["GET"])
def schools_admin_list(request: HttpRequest) -> HttpResponse:
    schools = (
        School.objects.all()
        .select_related("subscription")
        .order_by("name")
        .prefetch_related(
            Prefetch(
                "memberships",
                queryset=SchoolMembership.objects.select_related("teacher").filter(
                    role_type=SchoolMembership.RoleType.MANAGER,
                    is_active=True,
                ),
                to_attr="manager_memberships",
            )
        )
    )

    items = []
    for s in schools:
        managers = [m.teacher for m in getattr(s, "manager_memberships", []) if m.teacher]
        items.append({"school": s, "managers": managers})

    return render(request, "reports/schools_admin_list.html", {"schools": items})


@login_required(login_url="reports:login")
@require_http_methods(["GET"])
def school_profile(request: HttpRequest, pk: int) -> HttpResponse:
    """بروفايل تفصيلي لمدرسة واحدة.

    - السوبر أدمن يمكنه عرض أي مدرسة.
    - مدير المدرسة يمكنه عرض المدارس التي يديرها فقط.
    """
    school = get_object_or_404(School, pk=pk)

    user = request.user
    allowed = False
    if getattr(user, "is_superuser", False):
        allowed = True
    elif _is_staff(user) and school in _user_manager_schools(user):
        allowed = True

    if not allowed:
        messages.error(request, "لا تملك صلاحية عرض هذه المدرسة.")
        return redirect("reports:admin_dashboard")

    # إحصائيات بسيطة للمدرسة
    reports_count = Report.objects.filter(school=school).count()

    tickets_qs = Ticket.objects.filter(school=school)
    tickets_total = tickets_qs.count()
    tickets_open = tickets_qs.filter(status__in=["open", "in_progress"]).count()
    tickets_done = tickets_qs.filter(status="done").count()
    tickets_rejected = tickets_qs.filter(status="rejected").count()

    teachers_qs = (
        Teacher.objects.filter(
            school_memberships__school=school,
            school_memberships__is_active=True,
        )
        .distinct()
        .order_by("name")
    )
    teachers_count = teachers_qs.count()

    departments_count = 0
    departments = []
    if Department is not None:
        try:
            depts_qs = Department.objects.filter(is_active=True)
            if DepartmentMembership is not None:
                depts_qs = (
                    depts_qs.filter(
                        memberships__teacher__school_memberships__school=school,
                        memberships__teacher__school_memberships__is_active=True,
                    )
                    .distinct()
                    .order_by("name")
                )
            departments_count = depts_qs.count()
            departments = list(depts_qs[:20])  # عرض عينات محدودة في القالب إن لزم
        except Exception:
            logger.exception("school_profile departments stats failed")

    context = {
        "school": school,
        "reports_count": reports_count,
        "tickets_total": tickets_total,
        "tickets_open": tickets_open,
        "tickets_done": tickets_done,
        "tickets_rejected": tickets_rejected,
        "teachers_count": teachers_count,
        "teachers": list(teachers_qs[:20]),  # أقصى 20 للعرض السريع
        "departments_count": departments_count,
        "departments": departments,
    }
    return render(request, "reports/school_profile.html", context)


@login_required(login_url="reports:login")
@user_passes_test(lambda u: getattr(u, "is_superuser", False), login_url="reports:login")
@require_http_methods(["GET", "POST"])
def school_managers_manage(request: HttpRequest, pk: int) -> HttpResponse:
    school = get_object_or_404(School, pk=pk)

    if request.method == "POST":
        action = request.POST.get("action")
        teacher_id = request.POST.get("teacher_id")
        if not teacher_id:
            messages.error(request, "الرجاء اختيار معلّم.")
            return redirect("reports:school_managers_manage", pk=school.pk)
        try:
            teacher = Teacher.objects.get(pk=teacher_id)
        except Teacher.DoesNotExist:
            messages.error(request, "المعلّم غير موجود.")
            return redirect("reports:school_managers_manage", pk=school.pk)

        if action == "add":
            # لا نسمح بأكثر من مدير نشط واحد لكل مدرسة
            other_manager_exists = SchoolMembership.objects.filter(
                school=school,
                role_type=SchoolMembership.RoleType.MANAGER,
                is_active=True,
            ).exclude(teacher=teacher).exists()
            if other_manager_exists:
                messages.error(request, "لا يمكن تعيين أكثر من مدير نشط لنفس المدرسة. قم بإلغاء تعيين المدير الحالي أولاً.")
                return redirect("reports:school_managers_manage", pk=school.pk)

            SchoolMembership.objects.update_or_create(
                school=school,
                teacher=teacher,
                role_type=SchoolMembership.RoleType.MANAGER,
                defaults={"is_active": True},
            )
            messages.success(request, f"تم تعيين {teacher.name} مديراً للمدرسة.")
        elif action == "remove":
            SchoolMembership.objects.filter(
                school=school,
                teacher=teacher,
                role_type=SchoolMembership.RoleType.MANAGER,
            ).update(is_active=False)
            messages.success(request, f"تم إلغاء إدارة {teacher.name} لهذه المدرسة.")
        else:
            messages.error(request, "إجراء غير معروف.")

        return redirect("reports:school_managers_manage", pk=school.pk)

    managers = (
        Teacher.objects.filter(
            school_memberships__school=school,
            school_memberships__role_type=SchoolMembership.RoleType.MANAGER,
            school_memberships__is_active=True,
        )
        .distinct()
        .order_by("name")
    )

    # في قائمة الإضافة نظهر فقط المستخدمين الذين هم "مديرون" على مستوى النظام
    # (إما أن يكون لهم الدور manager أو لديهم أي عضوية SchoolMembership كمدير).
    teachers = (
        Teacher.objects.filter(is_active=True)
        .filter(
            Q(role__slug__iexact=MANAGER_SLUG)
            | Q(
                school_memberships__role_type=SchoolMembership.RoleType.MANAGER,
                school_memberships__is_active=True,
            )
        )
        .distinct()
        .order_by("name")
    )

    context = {
        "school": school,
        "managers": managers,
        "teachers": teachers,
    }
    return render(request, "reports/school_managers_manage.html", context)

# ---- لوحة المدير المجمعة ----
@login_required(login_url="reports:login")
@user_passes_test(_is_staff, login_url="reports:login")
@role_required({"manager"})
@require_http_methods(["GET"])

def admin_dashboard(request: HttpRequest) -> HttpResponse:
    """لوحة تحكم مدير المدرسة - تحديث Premium 2026"""
    from django.core.cache import cache
    from django.db.models.functions import TruncWeek, TruncMonth
    import json
    
    # إذا لم يكن هناك مدرسة مختارة نوجّه لاختيار مدرسة أولاً
    active_school = _get_active_school(request)
    # السوبر يوزر يمكنه رؤية أي مدرسة، المدير مقيد بمدارسه فقط
    if School.objects.filter(is_active=True).exists():
        if active_school is None:
            return redirect("reports:select_school")
        if (not request.user.is_superuser) and active_school not in _user_manager_schools(request.user):
            messages.error(request, "ليست لديك صلاحية كمدير على هذه المدرسة.")
            return redirect("reports:select_school")

    # محاولة جلب البيانات من الكاش
    cache_key = f"admin_stats_v2_{active_school.id if active_school else 'global'}"
    try:
        stats = cache.get(cache_key)
    except Exception:
        stats = None

    if not stats:
        # عدد المعلّمين داخل المدرسة النشطة فقط (عزل حسب المدرسة)
        teachers_qs = Teacher.objects.all()
        if active_school is not None:
            teachers_qs = teachers_qs.filter(
                school_memberships__school=active_school,
                school_memberships__is_active=True,
            ).distinct()

        stats = {
            "reports_count": _filter_by_school(Report.objects.all(), active_school).count(),
            "teachers_count": teachers_qs.count(),
            "tickets_total": _filter_by_school(Ticket.objects.filter(is_platform=False), active_school).count(),
            "tickets_open": _filter_by_school(Ticket.objects.filter(status__in=["open", "in_progress"], is_platform=False), active_school).count(),
            "tickets_done": _filter_by_school(Ticket.objects.filter(status="done", is_platform=False), active_school).count(),
            "tickets_rejected": _filter_by_school(Ticket.objects.filter(status="rejected", is_platform=False), active_school).count(),
        }
        # تخزين في الكاش لمدة 5 دقائق
        try:
            cache.set(cache_key, stats, 300)
        except Exception:
            pass

    ctx = {
        **stats,
        "has_dept_model": Department is not None,
        "active_school": active_school,
    }

    has_reporttype = False
    reporttypes_count = 0
    try:
        from .models import ReportType  # type: ignore
        has_reporttype = True

        # في لوحة المدير نعرض عدد أنواع التقارير المستخدمة داخل المدرسة الحالية فقط
        if active_school is not None:
            reporttypes_count = (
                Report.objects.filter(school=active_school, category__isnull=False)
                .values("category_id")
                .distinct()
                .count()
            )
        else:
            # في حال عدم وجود مدرسة نشطة نرجع للعدّاد الكلي
            if hasattr(ReportType, "is_active"):
                reporttypes_count = ReportType.objects.filter(is_active=True).count()
            else:
                reporttypes_count = ReportType.objects.count()
    except Exception:
        pass

    ctx.update({
        "has_reporttype": has_reporttype,
        "reporttypes_count": reporttypes_count,
    })
    
    # إضافة بيانات الرسوم البيانية والتحليلات المتقدمة
    if active_school:
        charts_cache_key = f"admin_charts_v2_{active_school.id}"
        charts = cache.get(charts_cache_key)
        
        if not charts:
            now = timezone.now()
            
            # تقارير آخر 8 أسابيع
            eight_weeks_ago = now - timedelta(weeks=8)
            reports_by_week = _filter_by_school(
                Report.objects.filter(created_at__gte=eight_weeks_ago), 
                active_school
            ).annotate(
                week=TruncWeek('created_at')
            ).values('week').annotate(
                count=Count('id')
            ).order_by('week')
            
            reports_labels = []
            reports_data = []
            for item in reports_by_week:
                if item['week']:
                    # عرض تاريخ بداية الأسبوع (أوضح من التاريخ الكامل)
                    week_label = item['week'].strftime('%d/%m')
                    reports_labels.append(week_label)
                    reports_data.append(item['count'])
            
            # تقارير حسب التصنيف/النوع
            reports_by_category = _filter_by_school(
                Report.objects.all(), 
                active_school
            ).values('category__name').annotate(
                count=Count('id')
            ).order_by('-count')[:6]
            
            dept_labels = []
            dept_data = []
            for item in reports_by_category:
                category_name = item['category__name'] or 'غير محدد'
                dept_labels.append(category_name)
                dept_data.append(item['count'])
            
            # معلمين حسب القسم
            teachers_by_dept = []
            if Department is not None:
                teachers_by_dept_qs = Department.objects.filter(
                    school=active_school
                ).annotate(
                    teacher_count=Count('memberships__teacher', distinct=True)
                ).order_by('-teacher_count')[:6]
                
                teachers_labels = []
                teachers_data = []
                for dept in teachers_by_dept_qs:
                    teachers_labels.append(dept.name)
                    teachers_data.append(dept.teacher_count)
            else:
                teachers_labels = []
                teachers_data = []
            
            charts = {
                "reports_labels": json.dumps(reports_labels),
                "reports_data": json.dumps(reports_data),
                "dept_labels": json.dumps(dept_labels),
                "dept_data": json.dumps(dept_data),
                "teachers_labels": json.dumps(teachers_labels),
                "teachers_data": json.dumps(teachers_data),
            }
            
            try:
                cache.set(charts_cache_key, charts, 600)  # 10 دقائق
            except Exception:
                pass
        
        ctx.update(charts)
        
        # بيانات الاشتراك والتنبيهات
        subscription_warning = None
        try:
            from .models import SchoolSubscription
            active_subscription = SchoolSubscription.objects.filter(
                school=active_school,
                is_active=True
            ).first()
            
            if active_subscription:
                days_remaining = (active_subscription.end_date - now.date()).days
                ctx['subscription'] = active_subscription
                ctx['days_remaining'] = days_remaining
                
                if days_remaining <= 7:
                    subscription_warning = 'critical'
                elif days_remaining <= 30:
                    subscription_warning = 'warning'
            else:
                subscription_warning = 'expired'
        except Exception:
            pass
        
        ctx['subscription_warning'] = subscription_warning
        
        # آخر الأنشطة
        recent_activities = []
        try:
            recent_reports = _filter_by_school(
                Report.objects.all(),
                active_school
            ).select_related('teacher').order_by('-created_at')[:5]
            
            for report in recent_reports:
                recent_activities.append({
                    'type': 'report',
                    'icon': 'fa-file-alt',
                    'color': 'primary',
                    'title': 'تقرير جديد',
                    'description': f"{report.teacher.name if report.teacher else 'معلم'} - {report.department.name if report.department else 'قسم'}",
                    'time': report.created_at,
                })
            
            recent_tickets = _filter_by_school(
                Ticket.objects.filter(is_platform=False),
                active_school
            ).order_by('-created_at')[:3]
            
            for ticket in recent_tickets:
                recent_activities.append({
                    'type': 'ticket',
                    'icon': 'fa-ticket-alt',
                    'color': 'warning',
                    'title': 'طلب جديد',
                    'description': ticket.subject[:50],
                    'time': ticket.created_at,
                })
            
            recent_activities.sort(key=lambda x: x['time'], reverse=True)
            recent_activities = recent_activities[:8]
        except Exception:
            pass
        
        ctx['recent_activities'] = recent_activities

    return render(request, "reports/admin_dashboard.html", ctx)

@login_required(login_url="reports:login")
@user_passes_test(_is_staff, login_url="reports:login")
@role_required({"manager"})
def school_audit_logs(request: HttpRequest) -> HttpResponse:
    """عرض سجل العمليات الخاص بالمدرسة للمدير."""
    active_school = _get_active_school(request)
    if active_school is None:
        return redirect("reports:select_school")
    
    if (not request.user.is_superuser) and active_school not in _user_manager_schools(request.user):
        messages.error(request, "ليست لديك صلاحية كمدير على هذه المدرسة.")
        return redirect("reports:select_school")

    # ملاحظة: في بعض بيئات النشر قد لا تكون ترحيلات AuditLog مطبّقة بعد.
    # بدلاً من 500، نظهر الصفحة مع تنبيه واضح.
    try:
        from django.db.utils import OperationalError, ProgrammingError
    except Exception:  # pragma: no cover
        OperationalError = Exception  # type: ignore
        ProgrammingError = Exception  # type: ignore

    logs_qs = None
    try:
        logs_qs = AuditLog.objects.filter(school=active_school).select_related("teacher")
    except (OperationalError, ProgrammingError):
        messages.error(
            request,
            "ميزة سجل العمليات غير مفعّلة حالياً (لم يتم تطبيق الترحيلات بعد). "
            "يرجى تشغيل migrate ثم إعادة المحاولة.",
        )

    # تصفية/عرض السجلات (لو كانت متاحة)
    teacher_id = request.GET.get("teacher")
    action = request.GET.get("action")
    start_date = request.GET.get("start_date")
    end_date = request.GET.get("end_date")

    if logs_qs is not None:
        if teacher_id:
            logs_qs = logs_qs.filter(teacher_id=teacher_id)
        if action:
            logs_qs = logs_qs.filter(action=action)
        if start_date:
            logs_qs = logs_qs.filter(timestamp__date__gte=start_date)
        if end_date:
            logs_qs = logs_qs.filter(timestamp__date__lte=end_date)

        paginator = Paginator(logs_qs, 50)
        page = request.GET.get("page")
        logs = paginator.get_page(page)
    else:
        # لا نستخدم QuerySet هنا حتى لا نلمس قاعدة البيانات.
        logs = Paginator([], 50).get_page(1)

    # قائمة المعلمين في المدرسة للتصفية
    try:
        teachers = Teacher.objects.filter(
            school_memberships__school=active_school,
            school_memberships__is_active=True,
        ).distinct()
    except Exception:
        teachers = Teacher.objects.none()

    ctx = {
        "logs": logs,
        "teachers": teachers,
        "actions": AuditLog.Action.choices,
        "active_school": active_school,
        "q_teacher": teacher_id,
        "q_action": action,
        "q_start": start_date,
        "q_end": end_date,
    }
    return render(request, "reports/audit_logs.html", ctx)


@login_required(login_url="reports:login")
@user_passes_test(lambda u: getattr(u, "is_superuser", False), login_url="reports:login")
def platform_admin_dashboard(request: HttpRequest) -> HttpResponse:
    """لوحة تحكم خاصة بالمشرف العام لإدارة المنصة بالكامل - تحديث 2026."""
    from django.core.cache import cache
    from django.db.models.functions import TruncMonth
    import json
    
    now = timezone.now()
    
    # البيانات الحرجة (بدون كاش أو كاش قصير جداً)
    pending_payments = Payment.objects.filter(status=Payment.Status.PENDING).count()
    tickets_open = Ticket.objects.filter(status__in=["open", "in_progress"], is_platform=True).count()
    
    # البيانات الإحصائية (كاش 5 دقائق)
    stats_cache_key = "platform_stats_v2"
    stats = cache.get(stats_cache_key)
    
    if not stats:
        reports_count = Report.objects.count()
        teachers_count = Teacher.objects.count()
        
        tickets_total = Ticket.objects.filter(is_platform=True).count()
        tickets_done = Ticket.objects.filter(status="done", is_platform=True).count()
        tickets_rejected = Ticket.objects.filter(status="rejected", is_platform=True).count()

        # تحسين الاستعلامات باستخدام aggregate
        school_stats = School.objects.aggregate(
            total=Count('id'),
            active=Count('id', filter=Q(is_active=True))
        )
        
        platform_managers_count = (
            Teacher.objects.filter(
                school_memberships__role_type=SchoolMembership.RoleType.MANAGER,
                school_memberships__is_active=True,
            )
            .distinct()
            .count()
        )

        has_reporttype = False
        reporttypes_count = 0
        try:
            from .models import ReportType  # type: ignore
            has_reporttype = True
            if hasattr(ReportType, "is_active"):
                reporttypes_count = ReportType.objects.filter(is_active=True).count()
            else:
                reporttypes_count = ReportType.objects.count()
        except Exception:
            pass

        stats = {
            "reports_count": reports_count,
            "teachers_count": teachers_count,
            "tickets_total": tickets_total,
            "tickets_done": tickets_done,
            "tickets_rejected": tickets_rejected,
            "platform_schools_total": school_stats['total'],
            "platform_schools_active": school_stats['active'],
            "platform_managers_count": platform_managers_count,
            "has_reporttype": has_reporttype,
            "reporttypes_count": reporttypes_count,
        }
        
        try:
            cache.set(stats_cache_key, stats, 300)  # 5 دقائق
        except Exception:
            pass
    
    # بيانات الاشتراكات والمالية (كاش 3 دقائق)
    financial_cache_key = "platform_financial_v2"
    financial = cache.get(financial_cache_key)
    
    if not financial:
        subscriptions_active = SchoolSubscription.objects.filter(is_active=True, end_date__gte=now.date()).count()
        subscriptions_expired = SchoolSubscription.objects.filter(Q(is_active=False) | Q(end_date__lt=now.date())).count()
        subscriptions_expiring_soon = SchoolSubscription.objects.filter(
            is_active=True,
            end_date__gte=now.date(),
            end_date__lte=now.date() + timedelta(days=30)
        ).count()
        
        total_revenue = Payment.objects.filter(status=Payment.Status.APPROVED).aggregate(total=Sum('amount'))['total'] or 0
        
        # قائمة الاشتراكات المنتهية قريباً (للجدول)
        subscriptions_expiring_list = SchoolSubscription.objects.filter(
            is_active=True,
            end_date__gte=now.date(),
            end_date__lte=now.date() + timedelta(days=30)
        ).select_related('school', 'plan').order_by('end_date')[:10]
        
        financial = {
            "subscriptions_active": subscriptions_active,
            "subscriptions_expired": subscriptions_expired,
            "subscriptions_expiring_soon": subscriptions_expiring_soon,
            "total_revenue": total_revenue,
            "subscriptions_expiring_list": list(subscriptions_expiring_list),
        }
        
        try:
            cache.set(financial_cache_key, financial, 180)  # 3 دقائق
        except Exception:
            pass
    
    # بيانات الرسوم البيانية (كاش 10 دقائق)
    charts_cache_key = "platform_charts_v2"
    charts = cache.get(charts_cache_key)
    
    if not charts:
        # بيانات الإيرادات الشهرية (آخر 6 أشهر)
        six_months_ago = now - timedelta(days=180)
        revenue_by_month = Payment.objects.filter(
            status=Payment.Status.APPROVED,
            created_at__gte=six_months_ago
        ).annotate(
            month=TruncMonth('created_at')
        ).values('month').annotate(
            total=Sum('amount')
        ).order_by('month')
        
        revenue_labels = []
        revenue_data = []
        for item in revenue_by_month:
            month_name = item['month'].strftime('%Y-%m')
            revenue_labels.append(month_name)
            revenue_data.append(float(item['total']))
        
        # بيانات التقارير الأسبوعية (آخر 8 أسابيع)
        eight_weeks_ago = now - timedelta(weeks=8)
        reports_by_week = Report.objects.filter(
            created_at__gte=eight_weeks_ago
        ).annotate(
            week=TruncWeek('created_at')
        ).values('week').annotate(
            count=Count('id')
        ).order_by('week')
        
        # تسمية الأسابيع بالتاريخ بدلاً من رقم الأسبوع (أوضح للمستخدم)
        reports_labels = []
        reports_data = []
        for item in reports_by_week:
            if item['week']:
                # عرض تاريخ بداية الأسبوع (الأحد)
                week_start = item['week'].strftime('%d/%m')
                reports_labels.append(week_start)
                reports_data.append(item['count'])
        
        # توزيع المدارس حسب المرحلة
        schools_by_stage = School.objects.values('stage').annotate(
            count=Count('id')
        ).order_by('stage')
        
        stage_labels = []
        stage_data = []
        stage_colors = []
        color_map = {
            'primary': '#3b82f6',
            'middle': '#10b981', 
            'secondary': '#f59e0b',
            'all': '#8b5cf6'
        }
        
        for item in schools_by_stage:
            stage_name = dict(School.Stage.choices).get(item['stage'], item['stage'])
            stage_labels.append(stage_name)
            stage_data.append(item['count'])
            stage_colors.append(color_map.get(item['stage'], '#6b7280'))
        
        charts = {
            "revenue_labels": json.dumps(revenue_labels),
            "revenue_data": json.dumps(revenue_data),
            "reports_labels": json.dumps(reports_labels),
            "reports_data": json.dumps(reports_data),
            "stage_labels": json.dumps(stage_labels),
            "stage_data": json.dumps(stage_data),
            "stage_colors": json.dumps(stage_colors),
        }
        
        try:
            cache.set(charts_cache_key, charts, 600)  # 10 دقائق
        except Exception:
            pass
    
    # آخر الأنشطة (بدون كاش)
    recent_activities = []
    try:
        recent_payments = Payment.objects.filter(
            status=Payment.Status.APPROVED
        ).select_related('school').order_by('-updated_at')[:5]
        
        for payment in recent_payments:
            recent_activities.append({
                'type': 'payment',
                'icon': 'fa-check-circle',
                'color': 'emerald',
                'title': 'تمت الموافقة على دفعة',
                'description': f"{payment.school.name if payment.school else 'مدرسة'} - {payment.amount} ر.س",
                'time': payment.updated_at,
            })
        
        recent_subscriptions = SchoolSubscription.objects.filter(
            is_active=True
        ).select_related('school', 'plan').order_by('-created_at')[:3]
        
        for sub in recent_subscriptions:
            recent_activities.append({
                'type': 'subscription',
                'icon': 'fa-star',
                'color': 'indigo',
                'title': 'اشتراك جديد',
                'description': f"{sub.school.name} - {sub.plan.name}",
                'time': sub.created_at,
            })
        
        # ترتيب حسب الوقت
        recent_activities.sort(key=lambda x: x['time'], reverse=True)
        recent_activities = recent_activities[:8]
    except Exception:
        pass
    
    # دمج جميع البيانات
    ctx = {
        **stats,
        **financial,
        **charts,
        "pending_payments": pending_payments,
        "tickets_open": tickets_open,
        "recent_activities": recent_activities,
    }

    return render(request, "reports/platform_admin_dashboard.html", ctx)


# =========================
# صفحات إدارة المنصة المخصصة (بديلة للآدمن)
# =========================

@login_required(login_url="reports:login")
@user_passes_test(lambda u: getattr(u, "is_superuser", False), login_url="reports:login")
def platform_audit_logs(request: HttpRequest) -> HttpResponse:
    """عرض سجل العمليات للنظام بالكامل (للمشرف العام)."""
    
    logs_qs = AuditLog.objects.all().select_related("teacher", "school").order_by("-timestamp")

    teacher_id = request.GET.get("teacher")
    action = request.GET.get("action")
    start_date = request.GET.get("start_date")
    end_date = request.GET.get("end_date")

    if teacher_id:
        logs_qs = logs_qs.filter(teacher_id=teacher_id)
    if action:
        logs_qs = logs_qs.filter(action=action)
    if start_date:
        logs_qs = logs_qs.filter(timestamp__date__gte=start_date)
    if end_date:
        logs_qs = logs_qs.filter(timestamp__date__lte=end_date)

    paginator = Paginator(logs_qs, 50)
    page = request.GET.get("page")
    logs = paginator.get_page(page)

    ctx = {
        "logs": logs,
        "actions": AuditLog.Action.choices,
        "is_platform": True,
        "q_teacher": teacher_id,
        "q_action": action,
        "q_start": start_date,
        "q_end": end_date,
    }
    return render(request, "reports/audit_logs.html", ctx)


@login_required(login_url="reports:login")
@user_passes_test(lambda u: getattr(u, "is_superuser", False), login_url="reports:login")
def platform_subscriptions_list(request: HttpRequest) -> HttpResponse:
    today = timezone.localdate()
    status = (request.GET.get("status") or "all").strip().lower()
    plan_id = (request.GET.get("plan") or "").strip()
    q = (request.GET.get("q") or "").strip()

    base_qs = SchoolSubscription.objects.select_related("school", "plan")

    stats = base_qs.aggregate(
        total=Count("id"),
        active=Count("id", filter=Q(is_active=True, end_date__gte=today)),
        cancelled=Count("id", filter=Q(is_active=False, canceled_at__isnull=False)),
        expired=Count(
            "id",
            filter=Q(
                Q(end_date__lt=today, canceled_at__isnull=True)
                | Q(is_active=False, canceled_at__isnull=True)
            ),
        ),
    )

    subscriptions = base_qs
    if status == "active":
        subscriptions = subscriptions.filter(is_active=True, end_date__gte=today)
    elif status == "cancelled":
        subscriptions = subscriptions.filter(is_active=False, canceled_at__isnull=False)
    elif status == "expired":
        subscriptions = subscriptions.filter(
            Q(end_date__lt=today, canceled_at__isnull=True)
            | Q(is_active=False, canceled_at__isnull=True)
        )

    if plan_id:
        subscriptions = subscriptions.filter(plan_id=plan_id)

    if q:
        subscriptions = subscriptions.filter(school__name__icontains=q)

    subscriptions = subscriptions.order_by("-start_date")

    # ✅ لتفادي N+1: نجلب المدفوعات المرتبطة بكل اشتراك
    subscriptions = subscriptions.prefetch_related(
        Prefetch(
            "payments",
            queryset=Payment.objects.filter(
                status__in=[Payment.Status.PENDING, Payment.Status.APPROVED]
            ).only("id", "subscription_id", "payment_date"),
            to_attr="_prefetched_active_payments",
        )
    )

    # ✅ استرجاعات (refunds): مدفوعات approved بمبالغ سالبة
    subscriptions = subscriptions.prefetch_related(
        Prefetch(
            "payments",
            queryset=Payment.objects.filter(
                status=Payment.Status.APPROVED,
                amount__lt=0,
            ).only("id", "subscription_id", "payment_date", "amount"),
            to_attr="_prefetched_refunds",
        )
    )

    # ✅ حساب بسيط: هل يوجد دفع ضمن فترة الاشتراك الحالية؟
    # نستخدم payment_date >= start_date لتحديد أنه يخص نفس الفترة.
    from decimal import Decimal

    for sub in subscriptions:
        try:
            pref = getattr(sub, "_prefetched_active_payments", []) or []
            sub.has_payment_for_period = any(
                (getattr(p, "payment_date", None) is not None and p.payment_date >= sub.start_date)
                for p in pref
            )
        except Exception:
            sub.has_payment_for_period = False

        # مبلغ الاسترجاع لهذه الفترة (مجموع القيم السالبة كقيمة موجبة)
        try:
            refunds = getattr(sub, "_prefetched_refunds", []) or []
            total = Decimal("0")
            for p in refunds:
                if getattr(p, "payment_date", None) is not None and p.payment_date >= sub.start_date:
                    amt = getattr(p, "amount", None)
                    if amt is not None:
                        total += (-amt)
            sub.refund_amount_for_period = total
        except Exception:
            sub.refund_amount_for_period = Decimal("0")

    plans = SubscriptionPlan.objects.all().order_by("price", "name")

    ctx = {
        "subscriptions": subscriptions,
        "status": status,
        "plans": plans,
        "plan_id": plan_id,
        "q": q,
        "stats_total": stats.get("total") or 0,
        "stats_active": stats.get("active") or 0,
        "stats_cancelled": stats.get("cancelled") or 0,
        "stats_expired": stats.get("expired") or 0,
        "results_count": subscriptions.count(),
    }

    return render(request, "reports/platform_subscriptions.html", ctx)


def _record_subscription_payment_if_missing(
    *,
    subscription: SchoolSubscription,
    actor,
    note: str,
    force: bool = False,
) -> bool:
    """تسجيل عملية دفع (approved) لاشتراك مدرسة في حال عدم وجود دفعة للفترة الحالية.

    نستخدم ذلك لحالات "الاشتراك أُضيف/فُعِّل يدويًا" حتى يظهر في صفحة المالية.
    """
    try:
        if not bool(getattr(subscription, "is_active", False)):
            return False

        plan = getattr(subscription, "plan", None)
        price = getattr(plan, "price", None)
        if price is None:
            return False
        try:
            if float(price) <= 0:
                return False
        except Exception:
            pass

        today = timezone.localdate()
        period_start = getattr(subscription, "start_date", None) or today

        # ✅ تحصين: عند التفعيل/التجديد اليدوي (force=True) لا نريد منع التسجيل
        # بسبب وجود دفعات قديمة، لكن نمنع تكرار نفس العملية في نفس اليوم.
        if force:
            dup_qs = Payment.objects.filter(
                subscription=subscription,
                status__in=[Payment.Status.PENDING, Payment.Status.APPROVED],
                created_at__date=today,
                requested_plan=subscription.plan,
                amount=subscription.plan.price,
            )
            if dup_qs.exists():
                return False
        else:
            existing_qs = Payment.objects.filter(
                subscription=subscription,
                status__in=[Payment.Status.PENDING, Payment.Status.APPROVED],
            )
            # نعتمد created_at بدلاً من payment_date لأن payment_date قد تكون "اليوم" دائماً
            # في التسجيلات اليدوية، مما يمنع تسجيل دفعة جديدة عند التجديد في نفس اليوم.
            existing_qs = existing_qs.filter(created_at__date__gte=period_start)
            if existing_qs.exists():
                return False

        Payment.objects.create(
            school=subscription.school,
            subscription=subscription,
            requested_plan=subscription.plan,
            amount=subscription.plan.price,
            receipt_image=None,
            payment_date=today,
            status=Payment.Status.APPROVED,
            notes=(note or "").strip(),
            created_by=actor,
        )
        return True
    except Exception:
        logger.exception("Failed to record manual payment for subscription")
        return False


@login_required(login_url="reports:login")
@user_passes_test(lambda u: getattr(u, "is_superuser", False), login_url="reports:login")
@require_http_methods(["POST"])
def platform_subscription_delete(request: HttpRequest, pk: int) -> HttpResponse:
    subscription = get_object_or_404(SchoolSubscription.objects.select_related("school", "plan"), pk=pk)

    reason = (request.POST.get("reason") or "").strip()
    refund_raw = (request.POST.get("refund_amount") or "").strip()
    if not reason:
        messages.error(request, "سبب الإلغاء مطلوب لإلغاء الاشتراك.")
        next_url = _safe_next_url(request.POST.get("next") or request.GET.get("next"))
        return redirect(next_url or "reports:platform_subscriptions_list")

    try:
        today = timezone.localdate()
        school_name = subscription.school.name

        with transaction.atomic():
            subscription.is_active = False
            subscription.end_date = today
            if getattr(subscription, "canceled_at", None) is None:
                subscription.canceled_at = timezone.now()
            subscription.cancel_reason = reason

            subscription.save(update_fields=["is_active", "end_date", "canceled_at", "cancel_reason", "updated_at"])

            # ✅ سجل مالي/سجل عمليات المدرسة:
            # نُسجل حدث الإلغاء نفسه كعملية (cancelled) حتى يظهر في:
            # - صفحة المالية (ضمن تبويب cancelled)
            # - صفحة "سجل العمليات السابقة" للمدرسة
            # ولا يؤثر على إجمالي الإيرادات (لأنه مبلغ 0 وبحالة cancelled).
            try:
                exists_cancel_event = Payment.objects.filter(
                    subscription=subscription,
                    status=Payment.Status.CANCELLED,
                    payment_date=today,
                    amount=0,
                ).exists()
                if not exists_cancel_event:
                    Payment.objects.create(
                        school=subscription.school,
                        subscription=subscription,
                        requested_plan=subscription.plan,
                        amount=0,
                        receipt_image=None,
                        payment_date=today,
                        status=Payment.Status.CANCELLED,
                        notes=(
                            "تم إلغاء الاشتراك بواسطة إدارة المنصة.\n"
                            f"سبب الإلغاء: {reason}"
                        ),
                        created_by=request.user,
                    )
            except Exception:
                logger.exception("Failed to record subscription cancellation event")

            # ✅ المالية:
            # - عند الإلغاء: نُلغي فقط المدفوعات المعلّقة لهذه الفترة حتى لا يتم اعتمادها لاحقاً بالخطأ.
            # - خيار إضافي: "استرجاع مبلغ" (كامل/جزئي) عبر تسجيل عملية مالية سالبة (approved)
            #   بحيث يظهر الاسترجاع ويخصم من إجمالي المالية.
            try:
                period_start = getattr(subscription, "start_date", None)

                # 1) إلغاء المعلّق فقط
                pending_qs = Payment.objects.filter(
                    subscription=subscription,
                    status=Payment.Status.PENDING,
                )
                if period_start:
                    pending_qs = pending_qs.filter(payment_date__gte=period_start)

                cancel_note = f"تم إلغاء الاشتراك: {reason}"
                for p in pending_qs.only("id", "status", "notes"):
                    p.status = Payment.Status.CANCELLED
                    p.notes = (f"{p.notes}\n" if (p.notes or "").strip() else "") + cancel_note
                    p.save(update_fields=["status", "notes", "updated_at"])

                # 2) استرجاع مبلغ (اختياري)
                if refund_raw:
                    from decimal import Decimal, InvalidOperation
                    from django.db.models import Sum
                    from django.db.models.functions import Coalesce

                    raw = refund_raw.strip().lower()

                    approved_qs = Payment.objects.filter(
                        subscription=subscription,
                        status=Payment.Status.APPROVED,
                    )
                    if period_start:
                        approved_qs = approved_qs.filter(payment_date__gte=period_start)

                    net_paid = approved_qs.aggregate(total=Coalesce(Sum("amount"), Decimal("0"))).get("total")
                    try:
                        net_paid = Decimal(str(net_paid or "0"))
                    except Exception:
                        net_paid = Decimal("0")

                    max_refund = net_paid if net_paid > 0 else Decimal("0")
                    refund_amount = Decimal("0")

                    if raw in {"full", "كامل", "كاملًا", "كاملا", "استرجاع كامل", "استرجاع كاملًا"}:
                        refund_amount = max_refund
                    else:
                        # السماح بأرقام مثل 100 أو 100.50 أو 100,50
                        try:
                            normalized = raw.replace(",", ".")
                            refund_amount = Decimal(normalized)
                        except (InvalidOperation, ValueError):
                            refund_amount = Decimal("0")

                    if refund_amount < 0:
                        refund_amount = Decimal("0")
                    if refund_amount > max_refund:
                        refund_amount = max_refund

                    # منع الاسترجاع المكرر لنفس اليوم/المبلغ (تحصين بسيط)
                    if refund_amount > 0:
                        exists_refund = Payment.objects.filter(
                            subscription=subscription,
                            status=Payment.Status.APPROVED,
                            amount=-refund_amount,
                            payment_date=today,
                        ).exists()

                        if not exists_refund:
                            Payment.objects.create(
                                school=subscription.school,
                                subscription=subscription,
                                requested_plan=subscription.plan,
                                amount=-refund_amount,
                                receipt_image=None,
                                payment_date=today,
                                status=Payment.Status.APPROVED,
                                notes=(
                                    f"استرجاع مبلغ: {refund_amount} ريال.\n"
                                    f"سبب الإلغاء: {reason}"
                                ),
                                created_by=request.user,
                            )
            except Exception:
                logger.exception("Failed to cancel payments for cancelled subscription")

        messages.success(request, f"تم إلغاء اشتراك مدرسة {school_name}.")
    except Exception:
        logger.exception("platform_subscription_delete failed")
        messages.error(request, "حدث خطأ غير متوقع أثناء إلغاء الاشتراك.")

    next_url = _safe_next_url(request.POST.get("next") or request.GET.get("next"))
    return redirect(next_url or "reports:platform_subscriptions_list")

@login_required(login_url="reports:login")
@user_passes_test(lambda u: getattr(u, "is_superuser", False), login_url="reports:login")
def platform_plans_list(request: HttpRequest) -> HttpResponse:
    plans = SubscriptionPlan.objects.all().order_by('price')
    return render(request, "reports/platform_plans.html", {"plans": plans})

@login_required(login_url="reports:login")
@user_passes_test(lambda u: getattr(u, "is_superuser", False), login_url="reports:login")
def platform_payments_list(request: HttpRequest) -> HttpResponse:
    status = (request.GET.get("status") or "active").strip().lower()

    base_qs = Payment.objects.select_related('school').order_by('-created_at')

    # ✅ افتراضيًا: لا نعرض (cancelled) ضمن المالية.
    # ملاحظة: الاسترجاعات = عمليات مقبولة بمبلغ سالب.
    if status == "refunds":
        payments = base_qs.filter(status=Payment.Status.APPROVED, amount__lt=0)
    elif status == "cancelled":
        payments = base_qs.filter(status=Payment.Status.CANCELLED)
    elif status == "all":
        payments = base_qs
    else:
        status = "active"
        payments = base_qs.exclude(status=Payment.Status.CANCELLED)
    
    # حساب الإحصائيات لعرضها في الكروت العلوية
    stats = payments.aggregate(
        total=Count('id'),
        pending=Count('id', filter=Q(status=Payment.Status.PENDING)),
        approved=Count('id', filter=Q(status=Payment.Status.APPROVED)),
        rejected=Count('id', filter=Q(status=Payment.Status.REJECTED)),
        cancelled=Count('id', filter=Q(status=Payment.Status.CANCELLED)),
        refunds=Count('id', filter=Q(status=Payment.Status.APPROVED, amount__lt=0)),
    )

    ctx = {
        "payments": payments,
        "status": status,
        "payments_total": stats['total'] or 0,
        "payments_pending": stats['pending'] or 0,
        "payments_approved": stats['approved'] or 0,
        "payments_rejected": stats['rejected'] or 0,
        "payments_cancelled": stats['cancelled'] or 0,
        "payments_refunds": stats['refunds'] or 0,
    }
    return render(request, "reports/platform_payments.html", ctx)

@login_required(login_url="reports:login")
@user_passes_test(lambda u: getattr(u, "is_superuser", False), login_url="reports:login")
def platform_payment_detail(request: HttpRequest, pk: int) -> HttpResponse:
    payment = get_object_or_404(
        Payment.objects.select_related("school", "subscription", "requested_plan"),
        pk=pk,
    )
    
    if request.method == "POST":
        prev_status = payment.status
        new_status = request.POST.get("status")
        notes = request.POST.get("notes")
        
        if new_status in Payment.Status.values:
            payment.status = new_status
        
        if notes is not None:
            payment.notes = notes

        with transaction.atomic():
            payment.save()

            # ✅ عند اعتماد الدفع لأول مرة: حدّث/جدّد اشتراك المدرسة.
            # ملاحظة: تم إلغاء تغيير الباقة من النظام؛ عند وجود اشتراك قائم نقوم بالتجديد فقط.
            if prev_status != Payment.Status.APPROVED and payment.status == Payment.Status.APPROVED:
                plan_to_apply = payment.requested_plan
                subscription = getattr(payment.school, "subscription", None)

                today = timezone.localdate()

                if subscription is None:
                    if plan_to_apply is not None:
                        subscription = SchoolSubscription(
                            school=payment.school,
                            plan=plan_to_apply,
                            start_date=today,
                            end_date=today,
                            is_active=True,
                        )
                        subscription.save()
                    else:
                        messages.warning(
                            request,
                            "تم اعتماد الدفع، لكن لا توجد باقة مطلوبة لتفعيل الاشتراك تلقائياً.",
                        )
                else:
                    subscription.is_active = True

                    # ✅ تجديد بنفس الباقة فقط (بدون تغيير الباقة)
                    days = int(getattr(subscription.plan, "days_duration", 0) or 0)
                    subscription.start_date = today
                    subscription.end_date = today if days <= 0 else today + timedelta(days=days - 1)
                    subscription.save(update_fields=["start_date", "end_date", "is_active", "updated_at"])

                if subscription is not None and payment.subscription_id != subscription.id:
                    payment.subscription = subscription
                    payment.save(update_fields=["subscription"])

        messages.success(request, "تم تحديث حالة الدفع بنجاح.")
        return redirect("reports:platform_payment_detail", pk=pk)

    return render(request, "reports/platform_payment_detail.html", {"payment": payment})

@login_required(login_url="reports:login")
@user_passes_test(lambda u: getattr(u, "is_superuser", False), login_url="reports:login")
def platform_tickets_list(request: HttpRequest) -> HttpResponse:
    query = request.GET.get("q", "").strip()
    status_filter = request.GET.get("status", "").strip()

    # تذاكر الدعم الفني فقط (platform tickets)
    tickets = (
        Ticket.objects.filter(is_platform=True)
        .select_related("creator", "school")
        .order_by("-created_at")
    )

    if status_filter and status_filter != "all":
        tickets = tickets.filter(status=status_filter)

    if query:
        tickets = tickets.filter(
            Q(school__name__icontains=query) |
            Q(school__code__icontains=query) |
            Q(title__icontains=query) |
            Q(id__icontains=query)
        )

    return render(request, "reports/platform_tickets.html", {
        "tickets": tickets,
        "search_query": query,
        "current_status": status_filter,
    })


# ---- الأقسام: عرض/إنشاء/تعديل/حذف ----
@login_required(login_url="reports:login")
@role_required({"manager"})
@require_http_methods(["GET"])
def departments_list(request: HttpRequest) -> HttpResponse:
    active_school = _get_active_school(request)
    if School.objects.filter(is_active=True).exists():
        if active_school is None:
            messages.error(request, "فضلاً اختر مدرسة أولاً.")
            return redirect("reports:select_school")
        if (not request.user.is_superuser) and active_school not in _user_manager_schools(request.user):
            messages.error(request, "ليست لديك صلاحية على هذه المدرسة.")
            return redirect("reports:select_school")

    depts = _all_departments(active_school)
    return render(
        request,
        "reports/departments_list.html",
        {"departments": depts, "has_dept_model": Department is not None},
    )

@login_required(login_url="reports:login")
@role_required({"manager"})
@require_http_methods(["GET", "POST"])
def department_create(request: HttpRequest) -> HttpResponse:
    active_school = _get_active_school(request)
    if School.objects.filter(is_active=True).exists():
        if active_school is None:
            messages.error(request, "فضلاً اختر مدرسة أولاً.")
            return redirect("reports:select_school")
        if (not request.user.is_superuser) and active_school not in _user_manager_schools(request.user):
            messages.error(request, "ليست لديك صلاحية على هذه المدرسة.")
            return redirect("reports:select_school")

    FormCls = get_department_form()
    if not (Department is not None and FormCls is not None):
        messages.error(request, "إنشاء الأقسام يتطلب تفعيل موديل Department.")
        return redirect("reports:departments_list")

    form = FormCls(request.POST or None, active_school=active_school)
    if request.method == "POST":
        if form.is_valid():
            dep = form.save(commit=False)
            if hasattr(dep, "school") and active_school is not None:
                dep.school = active_school
            dep.save()
            # حفظ علاقات M2M بعد الحفظ الأولي
            if hasattr(form, "save_m2m"):
                form.save_m2m()
            messages.success(request, "✅ تم إنشاء القسم.")
            return redirect("reports:departments_list")
        messages.error(request, "تعذّر الحفظ. تحقّق من الحقول.")
    return render(request, "reports/department_form.html", {"form": form, "mode": "create"})

@login_required(login_url="reports:login")
@role_required({"manager"})
@require_http_methods(["GET", "POST"])
def department_update(request: HttpRequest, pk: int) -> HttpResponse:
    active_school = _get_active_school(request)
    if School.objects.filter(is_active=True).exists():
        if active_school is None:
            messages.error(request, "فضلاً اختر مدرسة أولاً.")
            return redirect("reports:select_school")
        if (not request.user.is_superuser) and active_school not in _user_manager_schools(request.user):
            messages.error(request, "ليست لديك صلاحية على هذه المدرسة.")
            return redirect("reports:select_school")
    FormCls = get_department_form()
    if not (Department is not None and FormCls is not None):
        messages.error(request, "نموذج الأقسام غير مُعد بعد.")
        return redirect("reports:departments_list")
    dep = get_object_or_404(Department, pk=pk, school=active_school)  # type: ignore[arg-type]
    form = FormCls(request.POST or None, instance=dep, active_school=active_school)
    if request.method == "POST" and form.is_valid():
        form.save()
        messages.success(request, "✏️ تم تحديث بيانات القسم.")
        return redirect("reports:departments_list")
    return render(request, "reports/department_form.html", {"form": form, "title": "تعديل قسم", "dep": dep})

@login_required(login_url="reports:login")
@role_required({"manager"})
@require_http_methods(["GET", "POST"])
def department_edit(request: HttpRequest, code: str) -> HttpResponse:
    active_school = _get_active_school(request)
    if School.objects.filter(is_active=True).exists():
        if active_school is None:
            messages.error(request, "فضلاً اختر مدرسة أولاً.")
            return redirect("reports:select_school")
        if (not request.user.is_superuser) and active_school not in _user_manager_schools(request.user):
            messages.error(request, "ليست لديك صلاحية على هذه المدرسة.")
            return redirect("reports:select_school")
    if Department is None:
        messages.error(request, "تعديل الأقسام غير متاح بدون موديل Department.")
        return redirect("reports:departments_list")

    obj, _, label = _resolve_department_by_code_or_pk(str(code), active_school)
    if not obj:
        messages.error(request, "القسم غير موجود.")
        return redirect("reports:departments_list")

    # عزل المدرسة: منع تعديل قسم يخص مدرسة أخرى.
    # الأقسام العامة (school is NULL) يسمح بها للسوبر فقط.
    try:
        if not getattr(request.user, "is_superuser", False) and hasattr(obj, "school_id"):
            if getattr(obj, "school_id", None) is None:
                messages.error(request, "لا يمكنك تعديل قسم عام.")
                return redirect("reports:departments_list")
            if active_school is None or getattr(obj, "school_id", None) != getattr(active_school, "id", None):
                messages.error(request, "لا يمكنك تعديل قسم من مدرسة أخرى.")
                return redirect("reports:departments_list")
    except Exception:
        pass

    FormCls = get_department_form()
    if not FormCls:
        messages.error(request, "DepartmentForm غير متاح.")
        return redirect("reports:departments_list")

    form = FormCls(request.POST or None, instance=obj, active_school=active_school)
    if request.method == "POST":
        if form.is_valid():
            form.save()
            messages.success(request, f"✏️ تم تحديث قسم «{label}».")
            return redirect("reports:departments_list")
        messages.error(request, "تعذّر الحفظ. تحقّق من الحقول.")
    return render(request, "reports/department_form.html", {"form": form, "mode": "edit", "department": obj})

def _assign_role_by_slug(teacher: Teacher, slug: str) -> bool:
    role_obj = Role.objects.filter(slug=slug).first()
    if not role_obj:
        return False
    teacher.role = role_obj
    try:
        teacher.save(update_fields=["role"])
    except Exception:
        teacher.save()
    return True


@login_required(login_url="reports:login")
@user_passes_test(lambda u: getattr(u, "is_superuser", False), login_url="reports:login")
@require_http_methods(["GET", "POST"])
def school_manager_create(request: HttpRequest) -> HttpResponse:
    """إنشاء حساب مدير مدرسة وربطه بمدرسة واحدة على الأقل (ويمكن بأكثر من مدرسة).

    - يستخدم نموذج مبسّط (ManagerCreateForm) لإنشاء مستخدم مدير.
    - بعد الإنشاء يتم إسناد الدور "manager" وضبط عضويات SchoolMembership كمدير.
    """
    # مدارس متاحة للاختيار
    schools = School.objects.filter(is_active=True).order_by("name")
    initial_school_id = request.GET.get("school_id")

    form = ManagerCreateForm(request.POST or None)
    selected_ids = request.POST.getlist("schools") if request.method == "POST" else ([] if not initial_school_id else [initial_school_id])

    if request.method == "POST":
        if not selected_ids:
            messages.error(request, "يجب ربط المدير بمدرسة واحدة على الأقل.")
        if form.is_valid() and selected_ids:
            try:
                with transaction.atomic():
                    teacher = form.save(commit=True)
                    # ضمان أن يكون الدور "manager" إن وُجد
                    _assign_role_by_slug(teacher, MANAGER_SLUG)

                    valid_schools = School.objects.filter(id__in=selected_ids, is_active=True)
                    if not valid_schools:
                        raise ValidationError("لا توجد مدارس صالحة للربط.")

                    # منع أكثر من مدير نشط واحد لكل مدرسة
                    conflict_exists = SchoolMembership.objects.filter(
                        school__in=valid_schools,
                        role_type=SchoolMembership.RoleType.MANAGER,
                        is_active=True,
                    ).exists()
                    if conflict_exists:
                        raise ValidationError("إحدى المدارس المختارة لديها مدير نشط بالفعل. لا يمكن تعيين أكثر من مدير واحد للمدرسة.")

                    for s in valid_schools:
                        SchoolMembership.objects.update_or_create(
                            school=s,
                            teacher=teacher,
                            role_type=SchoolMembership.RoleType.MANAGER,
                            defaults={"is_active": True},
                        )
                messages.success(request, "تم إنشاء مدير المدرسة وربطه بالمدارس المحددة.")
                return redirect("reports:schools_admin_list")
            except ValidationError as e:
                messages.error(request, " ".join(e.messages))
            except Exception:
                logger.exception("school_manager_create failed")
                messages.error(request, "تعذّر إنشاء مدير المدرسة. تحقّق من البيانات وحاول مرة أخرى.")

    context = {
        "form": form,
        "schools": schools,
        "selected_ids": [str(i) for i in selected_ids],
        "mode": "create",
        "manager": None,
    }
    return render(request, "reports/school_manager_create.html", context)


@login_required(login_url="reports:login")
@user_passes_test(lambda u: getattr(u, "is_superuser", False), login_url="reports:login")
@require_http_methods(["GET"])
def school_managers_list(request: HttpRequest) -> HttpResponse:
    """قائمة مدراء المدارس على مستوى المنصة."""
    # نعتبر أي مستخدم مدير منصة إذا:
    # - كان دوره role.slug يطابق MANAGER_SLUG
    #   أو
    # - لديه عضوية SchoolMembership كمدير في أي مدرسة.
    managers_qs = (
        Teacher.objects.filter(
            Q(role__slug__iexact=MANAGER_SLUG)
            | Q(
                school_memberships__role_type=SchoolMembership.RoleType.MANAGER
            )
        )
        .distinct()
        .order_by("name")
        .prefetch_related("school_memberships__school")
    )

    items: list[dict] = []
    for t in managers_qs:
        # نعرض المدارس التي ارتبط بها كمدير (سواء كانت العضوية نشطة أم لا في هذا السياق، 
        # لكننا نفضل عرض المدارس التي كان مديراً لها)
        schools = [
            m.school
            for m in t.school_memberships.all()
            if m.school and m.role_type == SchoolMembership.RoleType.MANAGER
        ]
        items.append({"manager": t, "schools": schools})

    return render(request, "reports/school_managers_list.html", {"managers": items})


@login_required(login_url="reports:login")
@user_passes_test(lambda u: getattr(u, "is_superuser", False), login_url="reports:login")
@require_http_methods(["POST"])
def school_manager_delete(request: HttpRequest, pk: int) -> HttpResponse:
    """تبديل حالة مدير مدرسة (تفعيل/تعطيل).

    لا نحذف السجل نهائيًا للحفاظ على السجلات المرتبطة.
    """

    manager = get_object_or_404(Teacher, pk=pk)

    try:
        with transaction.atomic():
            if manager.is_active:
                manager.is_active = False
                msg = "🗑️ تم إيقاف حساب المدير وإلغاء صلاحياته في المدارس."
                # عند التعطيل، نعطّل العضويات أيضاً
                SchoolMembership.objects.filter(
                    teacher=manager,
                    role_type=SchoolMembership.RoleType.MANAGER,
                ).update(is_active=False)
            else:
                manager.is_active = True
                msg = "✅ تم إعادة تفعيل حساب المدير بنجاح."
                # ملاحظة: لا نفعّل العضويات تلقائياً هنا لأننا لا نعرف أي مدرسة يجب تفعيلها 
                # يفضل أن يقوم المدير بتعديل المدارس من صفحة التعديل.

            manager.save(update_fields=["is_active"])

        messages.success(request, msg)
    except Exception:
        logger.exception("school_manager_toggle failed")
        messages.error(request, "تعذّر تغيير حالة المدير. حاول لاحقًا.")

    return redirect("reports:school_managers_list")


@login_required(login_url="reports:login")
@user_passes_test(lambda u: getattr(u, "is_superuser", False), login_url="reports:login")
@require_http_methods(["GET", "POST"])
def school_manager_update(request: HttpRequest, pk: int) -> HttpResponse:
    """تعديل بيانات مدير مدرسة موجود باستخدام نفس نموذج الإنشاء.

    - يمكن ترك كلمة المرور فارغة للإبقاء على الحالية.
    - يمكن تغيير المدارس المرتبطة بالمدير.
    """

    manager = get_object_or_404(
        Teacher.objects.prefetch_related("school_memberships__school"),
        pk=pk,
    )

    schools = School.objects.filter(is_active=True).order_by("name")

    if request.method == "POST":
        form = ManagerCreateForm(request.POST or None, instance=manager)
        selected_ids = request.POST.getlist("schools")
        if not selected_ids:
            messages.error(request, "يجب ربط المدير بمدرسة واحدة على الأقل.")
        if form.is_valid() and selected_ids:
            try:
                with transaction.atomic():
                    teacher = form.save(commit=True)
                    _assign_role_by_slug(teacher, MANAGER_SLUG)

                    valid_schools = School.objects.filter(id__in=selected_ids, is_active=True)

                    # منع أكثر من مدير نشط واحد لكل مدرسة: نسمح فقط إن كانت المدرسة
                    # بدون مدير أو أن المدير الحالي هو نفس المستخدم الجاري تعديله.
                    conflict_exists = SchoolMembership.objects.filter(
                        school__in=valid_schools,
                        role_type=SchoolMembership.RoleType.MANAGER,
                        is_active=True,
                    ).exclude(teacher=teacher).exists()
                    if conflict_exists:
                        raise ValidationError("إحدى المدارس المختارة لديها مدير آخر نشط بالفعل. لا يمكن تعيين أكثر من مدير واحد للمدرسة.")

                    # تعطيل أي عضويات إدارة مدارس لم تعد مختارة
                    SchoolMembership.objects.filter(
                        teacher=teacher,
                        role_type=SchoolMembership.RoleType.MANAGER,
                    ).exclude(school__in=valid_schools).update(is_active=False)

                    # تفعيل/إنشاء العضويات المختارة
                    for s in valid_schools:
                        SchoolMembership.objects.update_or_create(
                            school=s,
                            teacher=teacher,
                            role_type=SchoolMembership.RoleType.MANAGER,
                            defaults={"is_active": True},
                        )
                messages.success(request, "تم تحديث بيانات مدير المدرسة بنجاح.")
                return redirect("reports:school_managers_list")
            except ValidationError as e:
                messages.error(request, " ".join(e.messages))
            except Exception:
                logger.exception("school_manager_update failed")
                messages.error(request, "تعذّر تحديث بيانات مدير المدرسة. تحقّق من البيانات وحاول مرة أخرى.")
        # في حال وجود أخطاء نمرّر selected_ids كما هي
    else:
        existing_ids = SchoolMembership.objects.filter(
            teacher=manager,
            role_type=SchoolMembership.RoleType.MANAGER,
            is_active=True,
        ).values_list("school_id", flat=True)
        selected_ids = [str(i) for i in existing_ids]
        form = ManagerCreateForm(instance=manager)

    context = {
        "form": form,
        "schools": schools,
        "selected_ids": [str(i) for i in selected_ids],
        "mode": "edit",
        "manager": manager,
    }
    return render(request, "reports/school_manager_create.html", context)

@login_required(login_url="reports:login")
@role_required({"manager"})
@require_http_methods(["POST"])
def department_delete(request: HttpRequest, code: str) -> HttpResponse:
    active_school = _get_active_school(request)
    if School.objects.filter(is_active=True).exists():
        if active_school is None:
            messages.error(request, "فضلاً اختر مدرسة أولاً.")
            return redirect("reports:select_school")
        if (not request.user.is_superuser) and active_school not in _user_manager_schools(request.user):
            messages.error(request, "ليست لديك صلاحية على هذه المدرسة.")
            return redirect("reports:select_school")

    if Department is None:
        messages.error(request, "حذف الأقسام غير متاح بدون موديل Department.")
        return redirect("reports:departments_list")

    obj, _, label = _resolve_department_by_code_or_pk(str(code), active_school)
    if not obj:
        messages.error(request, "القسم غير موجود.")
        return redirect("reports:departments_list")

    # عزل المدرسة: لا يُسمح بحذف قسم يخص مدرسة أخرى.
    # الأقسام العامة (school is NULL) يُسمح بها للسوبر فقط (باستثناء قسم المدير الدائم الذي يمنع حذفه أصلاً).
    try:
        dep_school_id = getattr(obj, "school_id", None)
        if dep_school_id is None:
            if not getattr(request.user, "is_superuser", False):
                messages.error(request, "لا يمكنك حذف قسم عام على مستوى المنصة.")
                return redirect("reports:departments_list")
        elif active_school is not None and dep_school_id != active_school.id:
            messages.error(request, "لا يمكنك حذف قسم يخص مدرسة أخرى.")
            return redirect("reports:departments_list")
    except Exception:
        pass

    try:
        obj.delete()
        messages.success(request, f"🗑️ تم حذف قسم «{label}».")
    except ProtectedError:
        messages.error(request, f"لا يمكن حذف «{label}» لوجود سجلات مرتبطة به. عطّل القسم أو احذف السجلات المرتبطة أولاً.")
    except Exception:
        logger.exception("department_delete failed")
        messages.error(request, "تعذّر حذف القسم.")
    return redirect("reports:departments_list")

def _dept_m2m_field_name_to_teacher(dep_obj) -> str | None:
    try:
        if dep_obj is None:
            return None
        for f in dep_obj._meta.get_fields():
            if isinstance(f, ManyToManyField) and getattr(f.remote_field, "model", None) is Teacher:
                return f.name
    except Exception:
        logger.exception("Failed to detect forward M2M Department→Teacher")
    return None

def _deptmember_field_names() -> tuple[str | None, str | None]:
    dep_field = tea_field = None
    try:
        if DepartmentMembership is None:
            return (None, None)

        for f in DepartmentMembership._meta.get_fields():
            if isinstance(f, ForeignKey):
                if getattr(f.remote_field, "model", None) is Department and dep_field is None:
                    dep_field = f.name
                elif getattr(f.remote_field, "model", None) is Teacher and tea_field is None:
                    tea_field = f.name
            if dep_field and tea_field:
                break

        if dep_field is None:
            for n in ("department", "dept", "dept_fk"):
                if hasattr(DepartmentMembership, n):
                    dep_field = n
                    break
        if tea_field is None:
            for n in ("teacher", "member", "user", "teacher_fk"):
                if hasattr(DepartmentMembership, n):
                    tea_field = n
                    break
    except Exception:
        logger.exception("Failed to detect DepartmentMembership FKs")

    return (dep_field, tea_field)

def _dept_add_member(dep, teacher: Teacher) -> bool:
    try:
        m2m_name = _dept_m2m_field_name_to_teacher(dep)
        if m2m_name:
            getattr(dep, m2m_name).add(teacher)
            return True
    except Exception:
        logger.exception("Add via Department M2M failed")

    try:
        if DepartmentMembership is not None and Department is not None:
            dep_field, tea_field = _deptmember_field_names()
            if dep_field and tea_field:
                kwargs = {dep_field: dep, tea_field: teacher}
                DepartmentMembership.objects.get_or_create(**kwargs)
                return True
    except Exception:
        logger.exception("Add via DepartmentMembership failed")

    return False

def _dept_remove_member(dep, teacher: Teacher) -> bool:
    try:
        m2m_name = _dept_m2m_field_name_to_teacher(dep)
        if m2m_name:
            getattr(dep, m2m_name).remove(teacher)
            return True
    except Exception:
        logger.exception("Remove via Department M2M failed")

    try:
        if DepartmentMembership is not None and Department is not None:
            dep_field, tea_field = _deptmember_field_names()
            if dep_field and tea_field:
                kwargs = {dep_field: dep, tea_field: teacher}
                deleted, _ = DepartmentMembership.objects.filter(**kwargs).delete()
                return deleted > 0
    except Exception:
        logger.exception("Remove via DepartmentMembership failed")

    return False

def _dept_set_member_role(dep, teacher: Teacher, role_type: str) -> bool:
    try:
        if DepartmentMembership is None or Department is None:
            return False

        dep_field, tea_field = _deptmember_field_names()
        if not dep_field or not tea_field:
            return False

        if not hasattr(DepartmentMembership, "role_type"):
            return False

        kwargs = {dep_field: dep, tea_field: teacher}
        obj, created = DepartmentMembership.objects.get_or_create(
            defaults={"role_type": role_type},
            **kwargs,
        )
        if (not created) and getattr(obj, "role_type", None) != role_type:
            obj.role_type = role_type
            obj.save(update_fields=["role_type"])
        return True
    except Exception:
        logger.exception("Failed to set DepartmentMembership role_type")
        return False

def _dept_set_officer(dep, teacher: Teacher) -> bool:
    try:
        if DepartmentMembership is None or Department is None:
            return False

        dep_field, tea_field = _deptmember_field_names()
        if not dep_field or not tea_field:
            return False

        if not hasattr(DepartmentMembership, "role_type"):
            return False

        qs = DepartmentMembership.objects.filter(**{dep_field: dep})
        qs.filter(role_type=DM_OFFICER).exclude(**{tea_field: teacher}).update(role_type=DM_TEACHER)
        return _dept_set_member_role(dep, teacher, DM_OFFICER)
    except Exception:
        logger.exception("Failed to set department officer")
        return False

@login_required(login_url="reports:login")
@role_required({"manager"})
@require_http_methods(["GET", "POST"])
def department_members(request: HttpRequest, code: str | int) -> HttpResponse:
    active_school = _get_active_school(request)
    if School.objects.filter(is_active=True).exists():
        if active_school is None:
            messages.error(request, "فضلاً اختر مدرسة أولاً.")
            return redirect("reports:select_school")
        if (not request.user.is_superuser) and active_school not in _user_manager_schools(request.user):
            messages.error(request, "ليست لديك صلاحية على هذه المدرسة.")
            return redirect("reports:select_school")

    obj, dept_code, dept_label = _resolve_department_by_code_or_pk(str(code), active_school)
    if not dept_code:
        messages.error(request, "القسم غير موجود.")
        return redirect("reports:departments_list")

    # عزل المدرسة: لا نسمح بإدارة أعضاء قسم تابع لمدرسة أخرى.
    try:
        if obj is not None:
            dep_school_id = getattr(obj, "school_id", None)
            if dep_school_id is None:
                if not getattr(request.user, "is_superuser", False):
                    messages.error(request, "لا يمكنك إدارة قسم عام على مستوى المنصة.")
                    return redirect("reports:departments_list")
            elif active_school is not None and dep_school_id != active_school.id:
                messages.error(request, "لا يمكنك إدارة قسم يخص مدرسة أخرى.")
                return redirect("reports:departments_list")
    except Exception:
        pass

    if request.method == "POST":
        teacher_id = request.POST.get("teacher_id")
        action = (request.POST.get("action") or "").strip()

        allowed_teachers = Teacher.objects.filter(is_active=True)
        if active_school is not None:
            allowed_teachers = allowed_teachers.filter(
                school_memberships__school=active_school,
                school_memberships__is_active=True,
            )
        teacher = allowed_teachers.filter(pk=teacher_id).first()
        if not teacher:
            messages.error(request, "المعلّم غير موجود.")
            return redirect("reports:department_members", code=dept_code)

        if Department is not None and obj:
            try:
                with transaction.atomic():
                    ok = False
                    if action == "add":
                        ok = _dept_set_member_role(obj, teacher, DM_TEACHER) or _dept_add_member(obj, teacher)
                        if ok:
                            messages.success(request, f"تم تكليف {teacher.name} في قسم «{dept_label}».")
                        else:
                            messages.error(request, "تعذّر إسناد المعلّم — تحقّق من بنية DepartmentMembership.")
                    elif action == "set_officer":
                        ok = _dept_set_officer(obj, teacher)
                        if ok:
                            messages.success(request, f"تم تعيين {teacher.name} مسؤولاً لقسم «{dept_label}». ")
                        else:
                            messages.error(request, "تعذّر تعيين مسؤول القسم — تحقّق من دعم role_type.")
                    elif action == "unset_officer":
                        ok = _dept_remove_member(obj, teacher)
                        if ok:
                            messages.success(request, f"تم إلغاء تكليف {teacher.name} من القسم.")
                        else:
                            messages.error(request, "تعذّر إلغاء التكليف.")
                    elif action == "remove":
                        ok = _dept_remove_member(obj, teacher)
                        if ok:
                            messages.success(request, f"تم إلغاء تكليف {teacher.name}.")
                        else:
                            messages.error(request, "تعذّر إلغاء التكليف — تحقق من بنية العلاقات.")
                    else:
                        messages.error(request, "إجراء غير معروف.")
            except Exception:
                logger.exception("department_members mutation failed")
                messages.error(request, "حدث خطأ أثناء حفظ التغييرات.")
        else:
            messages.error(request, "إدارة الأعضاء تتطلب تفعيل موديل Department.")
            return redirect("reports:departments_list")

        return redirect("reports:department_members", code=dept_code)

    members_qs = _members_for_department(dept_code, active_school)

    officers_qs = Teacher.objects.none()
    teachers_qs = Teacher.objects.none()
    assigned_ids_qs = Teacher.objects.none()
    try:
        if DepartmentMembership is not None:
            mem_qs = DepartmentMembership.objects.filter(department__slug__iexact=dept_code)
            if active_school is not None:
                mem_qs = mem_qs.filter(department__school=active_school)
            officer_ids = mem_qs.filter(role_type=DM_OFFICER).values_list("teacher_id", flat=True)
            teacher_ids = mem_qs.filter(role_type=DM_TEACHER).values_list("teacher_id", flat=True)
            assigned_ids = mem_qs.values_list("teacher_id", flat=True)

            officers_qs = Teacher.objects.filter(is_active=True, id__in=officer_ids).distinct().order_by("name")
            teachers_qs = Teacher.objects.filter(is_active=True, id__in=teacher_ids).distinct().order_by("name")
            assigned_ids_qs = Teacher.objects.filter(id__in=assigned_ids)

            if active_school is not None:
                officers_qs = officers_qs.filter(
                    school_memberships__school=active_school,
                    school_memberships__is_active=True,
                )
                teachers_qs = teachers_qs.filter(
                    school_memberships__school=active_school,
                    school_memberships__is_active=True,
                )
                assigned_ids_qs = assigned_ids_qs.filter(
                    school_memberships__school=active_school,
                    school_memberships__is_active=True,
                )
    except Exception:
        logger.exception("Failed to compute officers/teachers memberships")

    all_teachers = Teacher.objects.filter(is_active=True)
    if active_school is not None:
        all_teachers = all_teachers.filter(
            school_memberships__school=active_school,
            school_memberships__is_active=True,
        )
    all_teachers = all_teachers.order_by("name")

    try:
        if hasattr(assigned_ids_qs, "values_list"):
            assigned_ids_list = assigned_ids_qs.values_list("id", flat=True)
            available = all_teachers.exclude(id__in=assigned_ids_list)
        else:
            available = all_teachers
    except Exception:
        available = all_teachers

    return render(
        request,
        "reports/department_members.html",
        {
            "department": obj if obj else {"code": dept_code, "name": dept_label},
            "dept_code": dept_code,
            "dept_label": dept_label,
            "members": members_qs,
            "officers": officers_qs,
            "teachers": teachers_qs,
            "all_teachers": all_teachers,
            "available_teachers": available,
            "has_dept_model": Department is not None,
        },
    )

# ===== ReportType CRUD =====
@login_required(login_url="reports:login")
@role_required({"manager"})
@require_http_methods(["GET"])
def reporttypes_list(request: HttpRequest) -> HttpResponse:
    if not (ReportType is not None):
        messages.error(request, "إدارة الأنواع تتطلب تفعيل موديل ReportType وتشغيل الهجرات.")
        return render(request, "reports/reporttypes_list.html", {"items": [], "db_backed": False})

    active_school = _get_active_school(request)
    if School.objects.filter(is_active=True).exists():
        if active_school is None:
            messages.error(request, "فضلاً اختر مدرسة أولاً.")
            return redirect("reports:select_school")
        if (not request.user.is_superuser) and active_school not in _user_manager_schools(request.user):
            messages.error(request, "ليست لديك صلاحية على هذه المدرسة.")
            return redirect("reports:select_school")

    qs = ReportType.objects.all().order_by("order", "name")
    if active_school is not None and hasattr(ReportType, "school"):
        qs = qs.filter(school=active_school)
    items = []
    for rt in qs:
        cnt_qs = Report.objects.filter(category__code=rt.code)
        try:
            if hasattr(Report, "school"):
                rt_school_id = getattr(rt, "school_id", None)
                if rt_school_id is not None:
                    cnt_qs = cnt_qs.filter(school_id=rt_school_id)
                elif active_school is not None:
                    cnt_qs = cnt_qs.filter(school=active_school)
        except Exception:
            pass
        cnt = cnt_qs.count()
        items.append({"obj": rt, "code": rt.code, "name": rt.name, "is_active": rt.is_active, "order": rt.order, "count": cnt})
    return render(request, "reports/reporttypes_list.html", {"items": items, "db_backed": True})

@login_required(login_url="reports:login")
@role_required({"manager"})
@require_http_methods(["GET", "POST"])
def reporttype_create(request: HttpRequest) -> HttpResponse:
    if ReportType is None:
        messages.error(request, "إنشاء الأنواع يتطلب تفعيل موديل ReportType.")
        return redirect("reports:reporttypes_list")

    active_school = _get_active_school(request)
    if School.objects.filter(is_active=True).exists():
        if active_school is None:
            messages.error(request, "فضلاً اختر مدرسة أولاً.")
            return redirect("reports:select_school")
        if (not request.user.is_superuser) and active_school not in _user_manager_schools(request.user):
            messages.error(request, "ليست لديك صلاحية على هذه المدرسة.")
            return redirect("reports:select_school")

    try:
        from .forms import ReportTypeForm  # type: ignore
        FormCls = ReportTypeForm
    except Exception:
        class _RTForm(forms.ModelForm):
            class Meta:
                model = ReportType
                fields = ("name", "code", "description", "order", "is_active")
        FormCls = _RTForm

    form = FormCls(request.POST or None)
    if request.method == "POST":
        if form.is_valid():
            obj = form.save(commit=False)
            if hasattr(obj, "school") and active_school is not None:
                obj.school = active_school
            obj.save()
            if hasattr(form, "save_m2m"):
                form.save_m2m()
            messages.success(request, "✅ تم إضافة نوع التقرير.")
            return redirect("reports:reporttypes_list")
        messages.error(request, "تعذّر الحفظ. تحقّق من الحقول.")
    return render(request, "reports/reporttype_form.html", {"form": form, "mode": "create"})

@login_required(login_url="reports:login")
@role_required({"manager"})
@require_http_methods(["GET", "POST"])
def reporttype_update(request: HttpRequest, pk: int) -> HttpResponse:
    if ReportType is None:
        messages.error(request, "تعديل الأنواع يتطلب تفعيل موديل ReportType.")
        return redirect("reports:reporttypes_list")

    active_school = _get_active_school(request)
    if School.objects.filter(is_active=True).exists():
        if active_school is None:
            messages.error(request, "فضلاً اختر مدرسة أولاً.")
            return redirect("reports:select_school")
        if (not request.user.is_superuser) and active_school not in _user_manager_schools(request.user):
            messages.error(request, "ليست لديك صلاحية على هذه المدرسة.")
            return redirect("reports:select_school")

    obj = get_object_or_404(ReportType, pk=pk, school=active_school)

    try:
        from .forms import ReportTypeForm  # type: ignore
        FormCls = ReportTypeForm
    except Exception:
        class _RTForm(forms.ModelForm):
            class Meta:
                model = ReportType
                fields = ("name", "code", "description", "order", "is_active")
        FormCls = _RTForm

    form = FormCls(request.POST or None, instance=obj)
    if request.method == "POST":
        if form.is_valid():
            form.save()
            messages.success(request, "✏️ تم تعديل نوع التقرير.")
            return redirect("reports:reporttypes_list")
        messages.error(request, "تعذّر الحفظ. تحقّق من الحقول.")
    return render(request, "reports/reporttype_form.html", {"form": form, "mode": "edit", "obj": obj})

@login_required(login_url="reports:login")
@role_required({"manager"})
@require_http_methods(["POST"])
def reporttype_delete(request: HttpRequest, pk: int) -> HttpResponse:
    if ReportType is None:
        messages.error(request, "حذف الأنواع يتطلب تفعيل موديل ReportType.")
        return redirect("reports:reporttypes_list")

    active_school = _get_active_school(request)
    if School.objects.filter(is_active=True).exists():
        if active_school is None:
            messages.error(request, "فضلاً اختر مدرسة أولاً.")
            return redirect("reports:select_school")
        if (not request.user.is_superuser) and active_school not in _user_manager_schools(request.user):
            messages.error(request, "ليست لديك صلاحية على هذه المدرسة.")
            return redirect("reports:select_school")

    # عزل المدرسة: لا نسمح بحذف نوع تقرير يخص مدرسة أخرى.
    # الأنواع العامة (school is NULL) يسمح بها للسوبر فقط.
    if getattr(request.user, "is_superuser", False):
        obj = get_object_or_404(ReportType, pk=pk)
    else:
        obj = get_object_or_404(ReportType, pk=pk, school=active_school)

    # حساب الاستخدام وفق نطاق المدرسة.
    try:
        if getattr(obj, "school_id", None) is not None:
            used = Report.objects.filter(category__code=obj.code, school_id=obj.school_id).count()
        else:
            used = Report.objects.filter(category__code=obj.code).count()
    except Exception:
        used = Report.objects.filter(category__code=obj.code).count()
    if used > 0:
        messages.error(request, f"لا يمكن حذف «{obj.name}» لوجود {used} تقرير مرتبط. يمكنك تعطيله بدلًا من الحذف.")
        return redirect("reports:reporttypes_list")

    try:
        obj.delete()
        messages.success(request, f"🗑️ تم حذف «{obj.name}».")
    except Exception:
        logger.exception("reporttype_delete failed")
        messages.error(request, "تعذّر حذف نوع التقرير.")

    return redirect("reports:reporttypes_list")

# =========================
# واجهة برمجية مساعدة
# =========================
@login_required(login_url="reports:login")
@require_http_methods(["GET"])
def api_department_members(request: HttpRequest) -> HttpResponse:
    active_school = _get_active_school(request)
    dept = (request.GET.get("department") or "").strip()
    if not dept:
        return JsonResponse({"results": []})

    is_superuser = bool(getattr(request.user, "is_superuser", False))
    is_platform = bool(is_platform_admin(request.user)) and not is_superuser

    # Allow platform admin (and superuser) to specify a school explicitly when needed.
    requested_school_id = (request.GET.get("school") or request.GET.get("target_school") or "").strip()
    selected_school = None
    if is_superuser:
        selected_school = active_school
        if requested_school_id:
            try:
                selected_school = School.objects.filter(pk=int(requested_school_id), is_active=True).first()
            except (TypeError, ValueError):
                selected_school = None
    elif is_platform:
        if requested_school_id:
            try:
                candidate = School.objects.filter(pk=int(requested_school_id), is_active=True).first()
            except (TypeError, ValueError):
                candidate = None
            if candidate is None or not platform_can_access_school(request.user, candidate):
                return JsonResponse({"detail": "forbidden", "results": []}, status=403)
            selected_school = candidate
        else:
            selected_school = active_school
            if selected_school is None:
                return JsonResponse({"detail": "target_school_required", "results": []}, status=403)
            if not platform_can_access_school(request.user, selected_school):
                return JsonResponse({"detail": "forbidden", "results": []}, status=403)
    else:
        selected_school = active_school

    # عزل صارم: في وضع تعدد المدارس يجب أن تكون هناك مدرسة نشطة لغير السوبر.
    try:
        has_active_schools = School.objects.filter(is_active=True).exists()
    except Exception:
        has_active_schools = False

    if has_active_schools and selected_school is None and not is_superuser:
        return JsonResponse({"detail": "active_school_required", "results": []}, status=403)

    # تحقق عضوية المستخدم في المدرسة النشطة (حتى لا تُحقن session لمدرسة لا ينتمي لها المستخدم)
    if selected_school is not None and (not is_superuser) and (not is_platform):
        try:
            if not SchoolMembership.objects.filter(
                teacher=request.user,
                school=selected_school,
                is_active=True,
            ).exists():
                return JsonResponse({"detail": "forbidden", "results": []}, status=403)
        except Exception:
            return JsonResponse({"detail": "forbidden", "results": []}, status=403)

    users = _members_for_department(dept, selected_school).values("id", "name")
    return JsonResponse({"results": list(users)})


@login_required(login_url="reports:login")
@user_passes_test(_is_staff_or_officer, login_url="reports:login")
@require_http_methods(["GET"])
def api_school_departments(request: HttpRequest) -> HttpResponse:
    """Return active departments for selected school.

    - Superuser: may request any school via ?school=<id>
    - Non-superuser: school is forced to the active school (session isolation)
    - Always includes global departments (school IS NULL)
    """
    if Department is None:
        return JsonResponse({"results": []})

    active_school = _get_active_school(request)

    is_superuser = bool(getattr(request.user, "is_superuser", False))
    is_platform = bool(is_platform_admin(request.user)) and not is_superuser

    # عزل صارم: في وضع تعدد المدارس يجب أن تكون هناك مدرسة نشطة لغير السوبر.
    try:
        has_active_schools = School.objects.filter(is_active=True).exists()
    except Exception:
        has_active_schools = False

    if has_active_schools and active_school is None and (not is_superuser) and (not is_platform):
        return JsonResponse({"detail": "active_school_required", "results": []}, status=403)

    requested_school_id = (request.GET.get("school") or request.GET.get("target_school") or "").strip()
    selected_school = None

    if is_superuser:
        if requested_school_id:
            try:
                selected_school = School.objects.filter(pk=int(requested_school_id), is_active=True).first()
            except (TypeError, ValueError):
                selected_school = None
    elif is_platform:
        if requested_school_id:
            try:
                candidate = School.objects.filter(pk=int(requested_school_id), is_active=True).first()
            except (TypeError, ValueError):
                candidate = None
            if candidate is None or not platform_can_access_school(request.user, candidate):
                return JsonResponse({"detail": "forbidden", "results": []}, status=403)
            selected_school = candidate
        else:
            # If the platform admin already entered a school, reuse it.
            selected_school = active_school
            if selected_school is None:
                return JsonResponse({"detail": "target_school_required", "results": []}, status=403)
            if not platform_can_access_school(request.user, selected_school):
                return JsonResponse({"detail": "forbidden", "results": []}, status=403)
    else:
        selected_school = active_school
        # تحقق عضوية المستخدم في المدرسة النشطة
        if selected_school is not None:
            try:
                if not SchoolMembership.objects.filter(
                    teacher=request.user,
                    school=selected_school,
                    is_active=True,
                ).exists():
                    return JsonResponse({"detail": "forbidden", "results": []}, status=403)
            except Exception:
                return JsonResponse({"detail": "forbidden", "results": []}, status=403)

    qs = Department.objects.filter(is_active=True)

    # If no school selected (e.g. superuser scope=all), return all active.
    if selected_school is not None and _model_has_field(Department, "school"):
        qs = qs.filter(Q(school=selected_school) | Q(school__isnull=True))

    qs = qs.order_by("name")
    return JsonResponse({"results": list(qs.values("id", "name"))})


@login_required(login_url="reports:login")
@user_passes_test(_is_staff_or_officer, login_url="reports:login")
@require_http_methods(["GET"])
def api_notification_teachers(request: HttpRequest) -> HttpResponse:
    """Return teachers list for notification create form, filtered by selected school/department.

    This powers the dynamic recipients UI in the notification create template.
    """
    if NotificationCreateForm is None:
        return JsonResponse({"results": []})

    active_school = _get_active_school(request)

    is_superuser = bool(getattr(request.user, "is_superuser", False))
    is_platform = bool(is_platform_admin(request.user)) and not is_superuser

    # عزل صارم: في وضع تعدد المدارس يجب أن تكون هناك مدرسة نشطة لغير السوبر
    # (باستثناء مشرف المنصة لأنه يختار المدرسة من النموذج).
    try:
        has_active_schools = School.objects.filter(is_active=True).exists()
    except Exception:
        has_active_schools = False

    if has_active_schools and active_school is None and (not is_superuser) and (not is_platform):
        return JsonResponse({"detail": "active_school_required", "results": []}, status=403)

    # تحقق عضوية المستخدم في المدرسة النشطة (حتى لا تُحقن session لمدرسة لا ينتمي لها المستخدم)
    # مشرف المنصة لا يملك SchoolMembership، لذا نستخدم تحقق النطاق بدلًا من ذلك.
    if active_school is not None and (not is_superuser) and (not is_platform):
        try:
            if not SchoolMembership.objects.filter(
                teacher=request.user,
                school=active_school,
                is_active=True,
            ).exists():
                return JsonResponse({"detail": "forbidden", "results": []}, status=403)
        except Exception:
            return JsonResponse({"detail": "forbidden", "results": []}, status=403)

    if active_school is not None and is_platform:
        try:
            if not platform_can_access_school(request.user, active_school):
                return JsonResponse({"detail": "forbidden", "results": []}, status=403)
        except Exception:
            return JsonResponse({"detail": "forbidden", "results": []}, status=403)

    data = request.GET.copy()
    mode = (data.get("mode") or "notification").strip().lower()
    if mode not in {"notification", "circular"}:
        mode = "notification"
    is_circular = mode == "circular"

    if is_circular:
        if not is_superuser and not is_platform:
            if active_school is None or not _is_manager_in_school(request.user, active_school):
                return JsonResponse({"detail": "forbidden", "results": []}, status=403)
    # allow alternate query param names
    if "target_school" not in data and data.get("school"):
        data["target_school"] = data.get("school")
    if "target_department" not in data and data.get("department"):
        data["target_department"] = data.get("department")
    if "audience_scope" not in data and data.get("scope"):
        data["audience_scope"] = data.get("scope")

    # Build base queryset using the same constraints as the form.
    form = NotificationCreateForm(data=data, user=request.user, active_school=active_school, mode=mode)
    teachers_qs = form.fields["teachers"].queryset

    dept_val = (data.get("target_department") or "").strip()
    if dept_val:
        selected_school = active_school
        if getattr(request.user, "is_superuser", False):
            selected_school = None
            school_id = (data.get("target_school") or "").strip()
            if school_id:
                try:
                    selected_school = School.objects.filter(pk=int(school_id)).first()
                except (TypeError, ValueError):
                    selected_school = None

        dept_obj, dept_code, _dept_label = _resolve_department_by_code_or_pk(dept_val, selected_school)
        if dept_obj is None and selected_school is None:
            dept_obj, dept_code, _dept_label = _resolve_department_by_code_or_pk(dept_val, None)

        dept_school = selected_school
        try:
            if dept_obj is not None and hasattr(dept_obj, "school"):
                dept_school = getattr(dept_obj, "school", None)
        except Exception:
            pass

        dept_members_qs = _members_for_department(dept_code, dept_school)
        teachers_qs = teachers_qs.filter(pk__in=dept_members_qs.values("pk"))

    return JsonResponse({"results": list(teachers_qs.values("id", "name"))})

# =========================
# صناديق التذاكر بحسب القسم/المُعيّن
# =========================
@login_required(login_url="reports:login")
@user_passes_test(_is_staff, login_url="reports:login")
@require_http_methods(["GET"])
def tickets_inbox(request: HttpRequest) -> HttpResponse:
    active_school = _get_active_school(request)
    if School.objects.filter(is_active=True).exists() and active_school is None:
        messages.error(request, "فضلاً اختر مدرسة أولاً.")
        return redirect("reports:select_school")
    qs = Ticket.objects.select_related("creator", "assignee", "department").prefetch_related("recipients").order_by("-created_at")
    qs = _filter_by_school(qs, active_school)
    
    # استبعاد تذاكر الدعم الفني للمنصة (لأنها خاصة بالإدارة العليا)
    qs = qs.filter(is_platform=False)

    is_manager = _is_manager_in_school(request.user, active_school)
    if not is_manager:
        user_codes = _user_department_codes(request.user, active_school)
        qs = qs.filter(Q(assignee=request.user) | Q(recipients=request.user) | Q(department__slug__in=user_codes)).distinct()

    status = (request.GET.get("status") or "").strip()
    q = (request.GET.get("q") or "").strip()
    mine = request.GET.get("mine") == "1"

    if status:
        qs = qs.filter(status=status)
    if mine:
        qs = qs.filter(Q(assignee=request.user) | Q(recipients=request.user)).distinct()
    if q:
        for kw in q.split():
            qs = qs.filter(Q(title__icontains=kw) | Q(body__icontains=kw))

    ctx = {
        "tickets": qs[:200],
        "status": status,
        "q": q,
        "mine": mine,
        "status_choices": Ticket.Status.choices,
    }
    return render(request, "reports/tickets_inbox.html", ctx)

@login_required(login_url="reports:login")
@require_http_methods(["GET"])
def assigned_to_me(request: HttpRequest) -> HttpResponse:
    user = request.user
    active_school = _get_active_school(request)

    if School.objects.filter(is_active=True).exists() and active_school is None:
        messages.error(request, "فضلاً اختر مدرسة أولاً.")
        return redirect("reports:select_school")

    user_codes = _user_department_codes(user, active_school)

    qs = Ticket.objects.select_related("creator", "assignee", "department").prefetch_related("recipients").filter(
        Q(assignee=user)
        | Q(recipients=user)
        | Q(assignee__isnull=True, department__slug__in=user_codes)
    ).distinct()
    qs = _filter_by_school(qs, active_school)
    
    # استبعاد تذاكر الدعم الفني للمنصة
    qs = qs.filter(is_platform=False)

    q = (request.GET.get("q") or "").strip()
    if q:
        qs = qs.filter(Q(title__icontains=q) | Q(creator__name__icontains=q) | Q(id__icontains=q))

    status = request.GET.get("status")
    if status in {"open", "in_progress", "done", "rejected"}:
        qs = qs.filter(status=status)

    order = request.GET.get("order") or "-created_at"
    allowed_order = {"-created_at", "created_at", "-id", "id"}
    if order not in allowed_order:
        order = "-created_at"
    if order in {"created_at", "-created_at"}:
        qs = qs.order_by(order, "-id")
    else:
        qs = qs.order_by(order)

    raw_counts = dict(qs.values("status").annotate(c=Count("id")).values_list("status", "c"))
    stats = {
        "open": raw_counts.get("open", 0),
        "in_progress": raw_counts.get("in_progress", 0),
        "done": raw_counts.get("done", 0),
        "rejected": raw_counts.get("rejected", 0),
    }

    page_obj = Paginator(qs, 12).get_page(request.GET.get("page") or 1)
    view_mode = request.GET.get("view", "list")

    return render(request, "reports/assigned_to_me.html", {"page_obj": page_obj, "stats": stats, "view_mode": view_mode})

# =========================
# تقارير: تعديل/حذف للمستخدم الحالي
# =========================
@login_required(login_url="reports:login")
@require_http_methods(["GET", "POST"])
def edit_my_report(request: HttpRequest, pk: int) -> HttpResponse:
    """
    تعديل تقرير مع التحقق من الصلاحيات.
    يسمح للأشخاص التالية بالتعديل:
    - السوبر
    - مدير المدرسة
    - رئيس القسم (OFFICER) للتقارير المرتبطة بقسمه
    - صاحب التقرير نفسه
    
    ✅ عضو القسم (TEACHER) لا يستطيع التعديل (عرض فقط)
    ✅ مشرف المنصة لا يستطيع التعديل (عرض فقط)
    """
    user = request.user
    active_school = _get_active_school(request)

    # جلب التقرير باستخدام restrict_queryset (للتأكد من أن المستخدم يستطيع رؤيته)
    qs = restrict_queryset_for_user(Report.objects.all(), user, active_school)
    qs = _filter_by_school(qs, active_school)
    r = get_object_or_404(qs, pk=pk)
    
    # التحقق من صلاحية التعديل
    if not can_edit_report(user, r, active_school=active_school):
        messages.error(request, "لا تملك صلاحية تعديل هذا التقرير.")
        return redirect("reports:admin_reports")

    # لا نجبر تغيير المدرسة النشطة بالجَلسة، لكن نستخدم مدرسة التقرير لتصفية الأنواع عند الحاجة.
    form_school = active_school
    if form_school is None and _model_has_field(Report, "school"):
        try:
            form_school = getattr(r, "school", None)
        except Exception:
            form_school = active_school

    if request.method == "POST":
        form = ReportForm(request.POST, request.FILES, instance=r, active_school=form_school)
        if form.is_valid():
            form.save()
            messages.success(request, "✏️ تم تحديث التقرير بنجاح.")
            nxt = request.POST.get("next") or request.GET.get("next")
            if nxt:
                return redirect(nxt)
            # إذا كان المستخدم ليس صاحب التقرير، يعود لـ admin_reports
            if getattr(r, "teacher_id", None) != getattr(user, "id", None):
                return redirect("reports:admin_reports")
            return redirect("reports:my_reports")
        messages.error(request, "تحقّق من الحقول.")
    else:
        form = ReportForm(instance=r, active_school=form_school)

    return render(request, "reports/edit_report.html", {"form": form, "report": r})

@login_required(login_url="reports:login")
@require_http_methods(["POST"])
def delete_my_report(request: HttpRequest, pk: int) -> HttpResponse:
    active_school = _get_active_school(request)
    qs = Report.objects.filter(teacher=request.user)
    qs = _filter_by_school(qs, active_school)
    r = get_object_or_404(qs, pk=pk)
    r.delete()
    messages.success(request, "🗑️ تم حذف التقرير.")
    nxt = request.POST.get("next") or request.GET.get("next")
    return redirect(nxt or "reports:my_reports")

# =========================
# الإشعارات (إرسال/استقبال)
# =========================
@login_required(login_url="reports:login")
@user_passes_test(_is_staff_or_officer, login_url="reports:login")
@require_http_methods(["GET", "POST"])
def notifications_create(request: HttpRequest, mode: str = "notification") -> HttpResponse:
    if NotificationCreateForm is None:
        messages.error(request, "نموذج إنشاء الإشعار غير متوفر.")
        return redirect("reports:home")

    mode = (mode or "notification").strip().lower()
    if mode not in {"notification", "circular"}:
        mode = "notification"
    is_circular = mode == "circular"

    # نربط الإشعارات بمدرسة معيّنة للمدير/الضابط عبر المدرسة النشطة
    active_school = None
    try:
        active_school = _get_active_school(request)
    except Exception:
        active_school = None

    is_superuser = bool(getattr(request.user, "is_superuser", False))
    is_platform = bool(is_platform_admin(request.user)) and not is_superuser

    # حماية: مدير المدرسة/الضابط يحتاج مدرسة نشطة. المشرف العام يختار المدرسة من النموذج.
    if (not is_superuser) and (not is_platform) and active_school is None:
        messages.error(request, "يرجى اختيار المدرسة أولاً قبل إرسال الإشعارات.")
        return redirect("reports:home")

    # التعميمات: مدير المدرسة، مدير النظام، والمشرف العام (ضمن نطاقه فقط).
    if is_circular:
        if not is_superuser and not is_platform:
            if active_school is None or not _is_manager_in_school(request.user, active_school):
                messages.error(request, "التعاميم متاحة لمدير المدرسة فقط.")
                return redirect("reports:home")

    initial = {}
    if request.method == "GET" and is_circular:
        initial["requires_signature"] = True

    form = NotificationCreateForm(
        request.POST or None,
        request.FILES or None,
        user=request.user,
        active_school=active_school,
        initial=initial,
        mode=mode,
    )
    if request.method == "POST":
        if form.is_valid():
            try:
                with transaction.atomic():
                    form.save(
                        creator=request.user,
                        default_school=active_school,
                        force_requires_signature=True if is_circular else False,
                    )
                messages.success(request, "✅ تم إرسال التعميم." if is_circular else "✅ تم إرسال الإشعار.")
                return redirect("reports:circulars_sent" if is_circular else "reports:notifications_sent")
            except Exception:
                logger.exception("notifications_create failed")
                messages.error(request, "تعذّر الإرسال. جرّب لاحقًا.")
        else:
            messages.error(request, "الرجاء تصحيح الأخطاء.")

    return render(
        request,
        "reports/circulars_create.html" if is_circular else "reports/notifications_create.html",
        {
            "form": form,
            "mode": mode,
            "title": "إنشاء تعميم" if is_circular else "إنشاء إشعار",
        },
    )

@login_required(login_url="reports:login")
@user_passes_test(_is_staff_or_officer, login_url="reports:login")
@require_http_methods(["POST"])
def notification_delete(request: HttpRequest, pk: int) -> HttpResponse:
    if Notification is None:
        messages.error(request, "نموذج الإشعار غير متاح.")
        return redirect("reports:notifications_sent")

    active_school = _get_active_school(request)
    is_superuser = bool(getattr(request.user, "is_superuser", False))
    is_platform = bool(is_platform_admin(request.user)) and not is_superuser
    if (not is_superuser) and (not is_platform) and active_school is None:
        messages.error(request, "يرجى اختيار المدرسة أولاً.")
        return redirect("reports:home")

    n = get_object_or_404(Notification, pk=pk)
    sent_list_url = "reports:circulars_sent" if bool(getattr(n, "requires_signature", False)) else "reports:notifications_sent"

    # التعميمات: سماح لمدير المدرسة/مدير النظام/المشرف العام (ضمن نطاقه)
    if bool(getattr(n, "requires_signature", False)):
        if is_platform:
            if getattr(n, "created_by_id", None) != request.user.id:
                messages.error(request, "لا تملك صلاحية التعامل مع هذا التعميم.")
                return redirect(sent_list_url)
            try:
                if getattr(n, "school", None) is not None and not platform_can_access_school(request.user, getattr(n, "school", None)):
                    messages.error(request, "لا تملك صلاحية التعامل مع تعميم خارج نطاقك.")
                    return redirect(sent_list_url)
            except Exception:
                pass
        elif not is_superuser and not _is_manager_in_school(request.user, active_school):
            messages.error(request, "لا تملك صلاحية التعامل مع التعاميم.")
            return redirect(sent_list_url)
    is_owner = getattr(n, "created_by_id", None) == request.user.id
    is_manager = _is_manager_in_school(request.user, active_school)
    if is_platform:
        if not is_owner:
            messages.error(request, "لا تملك صلاحية حذف هذا الإشعار.")
            return redirect(sent_list_url)
    elif not (is_manager or is_owner):
        messages.error(request, "لا تملك صلاحية حذف هذا الإشعار.")
        return redirect(sent_list_url)

    # عزل حسب المدرسة النشطة (غير السوبر)
    try:
        if (not is_superuser) and (not is_platform) and hasattr(n, "school_id"):
            if getattr(n, "school_id", None) is None:
                messages.error(request, "لا تملك صلاحية حذف إشعار عام.")
                return redirect(sent_list_url)
            if getattr(n, "school_id", None) != getattr(active_school, "id", None):
                messages.error(request, "لا تملك صلاحية حذف إشعار من مدرسة أخرى.")
                return redirect(sent_list_url)
    except Exception:
        pass
    try:
        n.delete()
        messages.success(request, "🗑️ تم حذف الإشعار.")
    except Exception:
        logger.exception("notification_delete failed")
        messages.error(request, "تعذّر حذف الإشعار.")
    return redirect(sent_list_url)

def _recipient_is_read(rec) -> tuple[bool, str | None]:
    for flag in ("is_read", "read", "seen", "opened"):
        if hasattr(rec, flag):
            try:
                return (bool(getattr(rec, flag)), None)
            except Exception:
                pass
    for dt in ("read_at", "seen_at", "opened_at"):
        if hasattr(rec, dt):
            try:
                val = getattr(rec, dt)
                return (bool(val), getattr(val, "strftime", lambda fmt: None)("%Y-%m-%d %H:%M") if val else None)
            except Exception:
                pass
    if hasattr(rec, "status"):
        try:
            st = str(getattr(rec, "status") or "").lower()
            if st in {"read", "seen", "opened", "done"}:
                return (True, None)
        except Exception:
            pass
    return (False, None)

def _arabic_role_label(role_slug: str, active_school: Optional[School] = None) -> str:
    return _role_display_map(active_school).get((role_slug or "").lower(), role_slug or "")


def _digits_only(val: str) -> str:
    return "".join(ch for ch in str(val or "") if ch.isdigit())


def _phone_key(val: str) -> str:
    """Normalize phone for comparison.

    We compare by the last 9 digits to support common Saudi formats:
    - 05xxxxxxxx
    - 5xxxxxxxx
    - 9665xxxxxxxx
    """
    d = _digits_only(val)
    if len(d) >= 9:
        return d[-9:]
    return d


def _mask_phone(val: str) -> str:
    d = _digits_only(val)
    if not d:
        return ""
    if len(d) <= 4:
        return "*" * len(d)
    return ("*" * (len(d) - 4)) + d[-4:]

@login_required(login_url="reports:login")
@user_passes_test(_is_staff_or_officer, login_url="reports:login")
@require_http_methods(["GET"])
def notification_detail(request: HttpRequest, pk: int) -> HttpResponse:
    if Notification is None:
        messages.error(request, "نموذج الإشعار غير متاح.")
        return redirect("reports:notifications_sent")

    active_school = _get_active_school(request)
    is_superuser = bool(getattr(request.user, "is_superuser", False))
    is_platform = bool(is_platform_admin(request.user)) and not is_superuser
    if (not is_superuser) and (not is_platform) and active_school is None:
        messages.error(request, "يرجى اختيار المدرسة أولاً.")
        return redirect("reports:home")

    n = get_object_or_404(Notification, pk=pk)
    sent_list_url = "reports:circulars_sent" if bool(getattr(n, "requires_signature", False)) else "reports:notifications_sent"

    # التعميمات: سماح لمدير المدرسة/مدير النظام/المشرف العام (ضمن نطاقه)
    if bool(getattr(n, "requires_signature", False)):
        if is_platform:
            if getattr(n, "created_by_id", None) != request.user.id:
                messages.error(request, "لا تملك صلاحية عرض هذا التعميم.")
                return redirect(sent_list_url)
            try:
                if getattr(n, "school", None) is not None and not platform_can_access_school(request.user, getattr(n, "school", None)):
                    messages.error(request, "لا تملك صلاحية عرض تعميم خارج نطاقك.")
                    return redirect(sent_list_url)
            except Exception:
                pass
        elif (not is_superuser) and (not _is_manager_in_school(request.user, active_school)):
            messages.error(request, "لا تملك صلاحية عرض التعاميم.")
            return redirect(sent_list_url)

    # عزل حسب المدرسة النشطة (غير السوبر)
    try:
        if (not is_superuser) and (not is_platform) and hasattr(n, "school_id"):
            if getattr(n, "school_id", None) is None:
                messages.error(request, "لا تملك صلاحية عرض إشعار عام.")
                return redirect(sent_list_url)
            if getattr(n, "school_id", None) != getattr(active_school, "id", None):
                messages.error(request, "لا تملك صلاحية عرض إشعار من مدرسة أخرى.")
                return redirect(sent_list_url)
    except Exception:
        pass

    if not _is_manager_in_school(request.user, active_school):
        if getattr(n, "created_by_id", None) != request.user.id:
            messages.error(request, "لا تملك صلاحية عرض هذا الإشعار.")
            return redirect(sent_list_url)

    body = (
        getattr(n, "message", None) or getattr(n, "body", None) or
        getattr(n, "content", None) or getattr(n, "text", None) or
        getattr(n, "details", None) or ""
    )

    recipients = []
    sig_total = 0
    sig_signed = 0
    if NotificationRecipient is not None:
        # اكتشف اسم FK للإشعار
        notif_fk = None
        for f in NotificationRecipient._meta.get_fields():
            if getattr(getattr(f, "remote_field", None), "model", None) is Notification:
                notif_fk = f.name
                break

        # اسم حقل الشخص
        user_fk = None
        for cand in ("teacher", "user", "recipient"):
            if hasattr(NotificationRecipient, cand):
                user_fk = cand
                break

        if notif_fk:
            qs = NotificationRecipient.objects.filter(**{f"{notif_fk}": n})
            if user_fk:
                qs = qs.select_related(f"{user_fk}", f"{user_fk}__role")
            qs = qs.order_by("id")

            for r in qs:
                t = getattr(r, user_fk) if user_fk else None
                if not t:
                    continue
                name = getattr(t, "name", None) or getattr(t, "phone", None) or getattr(t, "username", None) or f"مستخدم #{getattr(t, 'pk', '')}"
                rslug = getattr(getattr(t, "role", None), "slug", "") or ""
                role_label = _arabic_role_label(rslug, active_school)
                is_read, read_at_str = _recipient_is_read(r)

                signed = bool(getattr(r, "is_signed", False))
                signed_at_str = None
                try:
                    v = getattr(r, "signed_at", None)
                    signed_at_str = v.strftime("%Y-%m-%d %H:%M") if v else None
                except Exception:
                    signed_at_str = None

                if bool(getattr(n, "requires_signature", False)):
                    sig_total += 1
                    if signed:
                        sig_signed += 1

                recipients.append({
                    "name": str(name),
                    "role": role_label,
                    "read": bool(is_read),
                    "read_at": read_at_str,
                    "signed": signed,
                    "signed_at": signed_at_str,
                })

    ctx = {
        "n": n,
        "body": body,
        "recipients": recipients,
        "signature_stats": {
            "total": int(sig_total),
            "signed": int(sig_signed),
            "unsigned": int(max(sig_total - sig_signed, 0)),
        },
    }
    template_name = "reports/circular_detail.html" if bool(getattr(n, "requires_signature", False)) else "reports/notification_detail.html"
    return render(request, template_name, ctx)


@login_required(login_url="reports:login")
@require_http_methods(["POST"])
def notification_sign(request: HttpRequest, pk: int) -> HttpResponse:
    """Teacher signs a circular (NotificationRecipient.pk) using phone re-entry + acknowledgement."""
    if NotificationRecipient is None:
        messages.error(request, "نظام الإشعارات غير متاح حالياً.")
        return redirect(request.POST.get("next") or "reports:my_notifications")

    rec = get_object_or_404(
        NotificationRecipient.objects.select_related(
            "notification",
            "notification__created_by",
        ),
        pk=pk,
        teacher=request.user,
    )

    n = getattr(rec, "notification", None)
    if n is None:
        messages.error(request, "تعذّر العثور على التعميم.")
        return redirect("reports:my_circulars")

    if not bool(getattr(n, "requires_signature", False)):
        messages.error(request, "هذا الإشعار لا يتطلب توقيعاً.")
        return redirect("reports:my_notification_detail", pk=rec.pk)

    if bool(getattr(rec, "is_signed", False)):
        messages.info(request, "تم تسجيل توقيعك مسبقاً على هذا التعميم.")
        return redirect("reports:my_circular_detail", pk=rec.pk)

    now = timezone.now()
    max_attempts = 5
    window = timedelta(minutes=15)

    try:
        attempts = int(getattr(rec, "signature_attempt_count", 0) or 0)
    except Exception:
        attempts = 0
    last_attempt = getattr(rec, "signature_last_attempt_at", None)

    # Reset attempts after window
    if last_attempt and (now - last_attempt) > window:
        attempts = 0

    if last_attempt and (now - last_attempt) <= window and attempts >= max_attempts:
        minutes_left = int(max(1, (window - (now - last_attempt)).total_seconds() // 60))
        messages.error(request, f"تم تجاوز عدد المحاولات. حاول مرة أخرى بعد {minutes_left} دقيقة.")
        return redirect("reports:my_circular_detail", pk=rec.pk)

    entered_phone = (request.POST.get("phone") or "").strip()
    ack = request.POST.get("ack") in {"1", "on", "true", "yes"}

    # Register an attempt (best-effort)
    try:
        rec.signature_attempt_count = attempts + 1
        rec.signature_last_attempt_at = now
        rec.save(update_fields=["signature_attempt_count", "signature_last_attempt_at"])
    except Exception:
        pass

    if not ack:
        messages.error(request, "يلزم الموافقة على الإقرار قبل اعتماد التوقيع.")
        return redirect("reports:my_circular_detail", pk=rec.pk)

    if not entered_phone:
        messages.error(request, "يرجى إدخال رقم الجوال المسجل للتوقيع.")
        return redirect("reports:my_circular_detail", pk=rec.pk)

    if _phone_key(entered_phone) != _phone_key(getattr(request.user, "phone", "")):
        messages.error(request, "رقم الجوال غير مطابق للرقم المسجل. تأكد وحاول مرة أخرى.")
        return redirect("reports:my_circular_detail", pk=rec.pk)

    # Sign + mark read
    try:
        update_fields: list[str] = []
        if hasattr(rec, "is_signed"):
            rec.is_signed = True
            update_fields.append("is_signed")
        if hasattr(rec, "signed_at"):
            rec.signed_at = now
            update_fields.append("signed_at")
        if hasattr(rec, "is_read") and not bool(getattr(rec, "is_read", False)):
            rec.is_read = True
            update_fields.append("is_read")
        if hasattr(rec, "read_at") and getattr(rec, "read_at", None) is None:
            rec.read_at = now
            update_fields.append("read_at")
        if update_fields:
            try:
                rec.save(update_fields=update_fields)
            except Exception:
                rec.save()
    except Exception:
        logger.exception("notification_sign failed")
        messages.error(request, "تعذّر تسجيل التوقيع. جرّب لاحقًا.")
        return redirect("reports:my_circular_detail", pk=rec.pk)

    messages.success(request, "✅ تم تسجيل توقيعك على التعميم بنجاح.")
    return redirect("reports:my_circular_detail", pk=rec.pk)


@login_required(login_url="reports:login")
@user_passes_test(_is_staff_or_officer, login_url="reports:login")
@require_http_methods(["GET"])
def notification_signatures_print(request: HttpRequest, pk: int) -> HttpResponse:
    if Notification is None or NotificationRecipient is None:
        messages.error(request, "نظام الإشعارات غير متاح.")
        return redirect("reports:notifications_sent")

    active_school = _get_active_school(request)
    is_superuser = bool(getattr(request.user, "is_superuser", False))
    is_platform = bool(is_platform_admin(request.user)) and not is_superuser
    if (not is_superuser) and (not is_platform) and active_school is None:
        messages.error(request, "يرجى اختيار المدرسة أولاً.")
        return redirect("reports:home")

    n = get_object_or_404(Notification, pk=pk)
    sent_list_url = "reports:circulars_sent" if bool(getattr(n, "requires_signature", False)) else "reports:notifications_sent"

    # هذا التقرير خاص بالتعاميم فقط
    if not bool(getattr(n, "requires_signature", False)):
        messages.error(request, "هذا التقرير متاح للتعاميم فقط.")
        return redirect(sent_list_url)

    # سماح للمشرف العام بتقارير التعاميم التي أنشأها ضمن نطاقه
    if is_platform:
        if getattr(n, "created_by_id", None) != request.user.id:
            messages.error(request, "لا تملك صلاحية عرض تقرير هذا التعميم.")
            return redirect(sent_list_url)
        try:
            if getattr(n, "school", None) is not None and not platform_can_access_school(request.user, getattr(n, "school", None)):
                messages.error(request, "لا تملك صلاحية عرض تعميم خارج نطاقك.")
                return redirect(sent_list_url)
        except Exception:
            pass

    # Permission: manager in school or creator
    if (not is_platform) and (not _is_manager_in_school(request.user, active_school)):
        if getattr(n, "created_by_id", None) != request.user.id:
            messages.error(request, "لا تملك صلاحية عرض تقرير هذا التعميم.")
            return redirect(sent_list_url)

    # School isolation
    try:
        if (not is_superuser) and (not is_platform) and hasattr(n, "school_id"):
            if getattr(n, "school_id", None) != getattr(active_school, "id", None):
                messages.error(request, "لا تملك صلاحية عرض تعميم من مدرسة أخرى.")
                return redirect(sent_list_url)
    except Exception:
        pass

    qs = (
        NotificationRecipient.objects
        .filter(notification=n)
        .select_related("teacher", "teacher__role")
        .order_by("teacher__name", "id")
    )

    rows = []
    signed = 0
    total = 0
    for r in qs:
        t = getattr(r, "teacher", None)
        if not t:
            continue
        total += 1
        is_signed = bool(getattr(r, "is_signed", False))
        if is_signed:
            signed += 1
        rows.append({
            "name": getattr(t, "name", "") or str(t),
            "role": _arabic_role_label(getattr(getattr(t, "role", None), "slug", "") or "", active_school),
            "phone": _mask_phone(getattr(t, "phone", "")),
            "read": bool(getattr(r, "is_read", False)),
            "read_at": getattr(r, "read_at", None),
            "signed": is_signed,
            "signed_at": getattr(r, "signed_at", None),
        })

    ctx = {
        "n": n,
        "rows": rows,
        "stats": {"total": total, "signed": signed, "unsigned": max(total - signed, 0)},
    }
    return render(request, "reports/notification_signatures_print.html", ctx)


@login_required(login_url="reports:login")
@user_passes_test(_is_staff_or_officer, login_url="reports:login")
@require_http_methods(["GET"])
def notification_signatures_csv(request: HttpRequest, pk: int) -> HttpResponse:
    if Notification is None or NotificationRecipient is None:
        return HttpResponse("unavailable", status=400)

    active_school = _get_active_school(request)
    is_superuser = bool(getattr(request.user, "is_superuser", False))
    is_platform = bool(is_platform_admin(request.user)) and not is_superuser
    if (not is_superuser) and (not is_platform) and active_school is None:
        return HttpResponse("active_school_required", status=403)

    n = get_object_or_404(Notification, pk=pk)

    if not bool(getattr(n, "requires_signature", False)):
        return HttpResponse("forbidden", status=403)

    if is_platform:
        if getattr(n, "created_by_id", None) != request.user.id:
            return HttpResponse("forbidden", status=403)
        try:
            if getattr(n, "school", None) is not None and not platform_can_access_school(request.user, getattr(n, "school", None)):
                return HttpResponse("forbidden", status=403)
        except Exception:
            pass

    if (not is_platform) and (not _is_manager_in_school(request.user, active_school)):
        if getattr(n, "created_by_id", None) != request.user.id:
            return HttpResponse("forbidden", status=403)

    try:
        if (not is_superuser) and (not is_platform) and hasattr(n, "school_id"):
            if getattr(n, "school_id", None) != getattr(active_school, "id", None):
                return HttpResponse("forbidden", status=403)
    except Exception:
        pass

    import csv
    from io import StringIO

    out = StringIO()
    writer = csv.writer(out)
    writer.writerow([
        "الاسم",
        "الدور",
        "الجوال (مخفي)",
        "الحالة (مقروء)",
        "وقت القراءة",
        "الحالة (موقّع)",
        "وقت التوقيع",
    ])

    qs = (
        NotificationRecipient.objects
        .filter(notification=n)
        .select_related("teacher", "teacher__role")
        .order_by("teacher__name", "id")
    )

    for r in qs:
        t = getattr(r, "teacher", None)
        if not t:
            continue
        role_label = _arabic_role_label(getattr(getattr(t, "role", None), "slug", "") or "", active_school)
        writer.writerow([
            getattr(t, "name", "") or str(t),
            role_label,
            _mask_phone(getattr(t, "phone", "")),
            "نعم" if bool(getattr(r, "is_read", False)) else "لا",
            getattr(getattr(r, "read_at", None), "strftime", lambda fmt: "")("%Y-%m-%d %H:%M") if getattr(r, "read_at", None) else "",
            "نعم" if bool(getattr(r, "is_signed", False)) else "لا",
            getattr(getattr(r, "signed_at", None), "strftime", lambda fmt: "")("%Y-%m-%d %H:%M") if getattr(r, "signed_at", None) else "",
        ])

    resp = HttpResponse(out.getvalue(), content_type="text/csv; charset=utf-8")
    safe_title = (getattr(n, "title", "") or "notification").strip().replace("\n", " ").replace("\r", " ")
    resp["Content-Disposition"] = f'attachment; filename="signatures_{pk}_{safe_title[:40]}.csv"'
    return resp

@require_http_methods(["GET"])
def unread_notifications_count(request: HttpRequest) -> HttpResponse:
    """إرجاع عدد الإشعارات غير المقروءة بتنسيق JSON لاستخدامه في الـ Polling.

    ملاحظة: لا نُعيد توجيه المستخدمين غير المسجلين لصفحة الدخول لأن هذا المسار يُستدعى بشكل دوري
    من الواجهة (Polling)، وإعادة التوجيه قد تسبب ضغطاً وتداخل مع RateLimit.
    """
    if not getattr(request.user, "is_authenticated", False):
        return JsonResponse({"count": 0, "authenticated": False})

    if NotificationRecipient is None:
        return JsonResponse({"count": 0, "unread": 0, "signatures_pending": 0, "authenticated": True})

    # Short-TTL cache per user + school to cut repeated aggregate queries.
    try:
        ttl = int(getattr(settings, "UNREAD_COUNT_CACHE_TTL_SECONDS", 15) or 0)
    except Exception:
        ttl = 15

    cache_key = None
    if ttl > 0:
        try:
            sid_raw = request.session.get("active_school_id")
            sid_for_key = str(int(sid_raw)) if sid_raw else "none"
        except Exception:
            sid_for_key = "none"
        try:
            uid = int(getattr(request.user, "id", 0) or 0)
            cache_key = f"unreadcnt:v1:u{uid}:s{sid_for_key}"
            cached = cache.get(cache_key)
            if isinstance(cached, dict):
                return JsonResponse(cached)
        except Exception:
            cache_key = None

    active_school = _get_active_school(request)
    now = timezone.now()

    qs = NotificationRecipient.objects.filter(teacher=request.user)

    # عزل حسب المدرسة النشطة (مع السماح بإشعارات عامة school=NULL)
    try:
        if active_school is not None and Notification is not None and hasattr(Notification, "school"):
            qs = qs.filter(Q(notification__school=active_school) | Q(notification__school__isnull=True))
    except Exception:
        pass

    # استبعاد المنتهي
    try:
        if Notification is not None and hasattr(Notification, "expires_at"):
            qs = qs.filter(Q(notification__expires_at__gt=now) | Q(notification__expires_at__isnull=True))
    except Exception:
        pass

    # unread = unread notifications only (exclude circulars)
    unread_q = Q(is_read=False)
    try:
        if Notification is not None and hasattr(Notification, "requires_signature"):
            unread_q &= Q(notification__requires_signature=False)
    except Exception:
        pass

    # signatures_pending = unsigned circulars
    pending_sig_q = Q(pk__in=[])
    try:
        if Notification is not None and hasattr(Notification, "requires_signature") and hasattr(NotificationRecipient, "is_signed"):
            pending_sig_q = Q(notification__requires_signature=True, is_signed=False)
    except Exception:
        pending_sig_q = Q(pk__in=[])

    # count = items needing attention (backward compatible): unread notifications OR pending circular signatures
    attention_q = unread_q | pending_sig_q

    agg = qs.aggregate(
        count=Count("id", filter=attention_q),
        unread=Count("id", filter=unread_q),
        signatures_pending=Count("id", filter=pending_sig_q),
    )

    payload = {
        "count": int(agg.get("count") or 0),
        "unread": int(agg.get("unread") or 0),
        "signatures_pending": int(agg.get("signatures_pending") or 0),
        "authenticated": True,
    }

    if cache_key and ttl > 0:
        try:
            cache.set(cache_key, payload, ttl)
        except Exception:
            pass

    return JsonResponse(payload)

@login_required(login_url="reports:login")
@require_http_methods(["GET"])
def my_notifications(request: HttpRequest) -> HttpResponse:
    if NotificationRecipient is None:
        return render(request, "reports/my_notifications.html", {"page_obj": Paginator([], 12).get_page(1)})

    active_school = _get_active_school(request)

    qs = (
        NotificationRecipient.objects
        .select_related("notification", "notification__created_by", "notification__created_by__role")
        .filter(teacher=request.user)
        .order_by("-created_at", "-id")
    )

    # فصل: هذه الصفحة للإشعارات فقط (بدون التعاميم)
    try:
        if Notification is not None and hasattr(Notification, "requires_signature"):
            qs = qs.filter(notification__requires_signature=False)
    except Exception:
        pass

    # عزل حسب المدرسة النشطة (مع السماح بإشعارات عامة school=NULL)
    try:
        if active_school is not None and Notification is not None and hasattr(Notification, "school"):
            qs = qs.filter(Q(notification__school=active_school) | Q(notification__school__isnull=True))
    except Exception:
        pass

    # إخفاء المنتهية بحسب الحقول المتاحة
    now = timezone.now()
    try:
        if Notification is not None and hasattr(Notification, "expires_at"):
            qs = qs.exclude(notification__expires_at__lt=now)
        elif Notification is not None and hasattr(Notification, "ends_at"):
            qs = qs.exclude(notification__ends_at__lt=now)
    except Exception:
        pass

    page = Paginator(qs, 12).get_page(request.GET.get("page") or 1)

    # عند فتح تبويب "إشعاراتي" غالباً يتوقع المستخدم أن تصبح الإشعارات المعروضة كمقروءة.
    # لا يمكن الاعتماد على "إغلاق التبويب" كإشارة مؤكدة من المتصفح، لذا نُحدّثها هنا.
    try:
        items = list(page.object_list)
        unread_ids = [x.pk for x in items if hasattr(x, "is_read") and not bool(getattr(x, "is_read", False))]
        if unread_ids:
            now = timezone.now()
            fields = {f.name for f in NotificationRecipient._meta.get_fields()}
            upd: dict = {}
            if "is_read" in fields:
                upd["is_read"] = True
            if "read_at" in fields:
                upd["read_at"] = now
            if upd:
                NotificationRecipient.objects.filter(pk__in=unread_ids, teacher=request.user).update(**upd)

                # Bulk update won't trigger post_save; request a one-off WS resync.
                try:
                    from .realtime_notifications import push_force_resync

                    push_force_resync(teacher_id=int(getattr(request.user, "id", 0) or 0))
                except Exception:
                    pass

                for x in items:
                    if x.pk in unread_ids:
                        if "is_read" in upd:
                            setattr(x, "is_read", True)
                        if "read_at" in upd:
                            setattr(x, "read_at", now)
            page.object_list = items
    except Exception:
        pass

    # اسم المرسل + الدور الصحيح (مُوحّد)
    try:
        items = list(page.object_list)
        for rr in items:
            n = getattr(rr, "notification", None)
            sender = getattr(n, "created_by", None) if n is not None else None
            school_scope = (getattr(n, "school", None) if n is not None else None) or active_school
            rr.sender_name = _canonical_sender_name(sender)
            rr.sender_role_label = _canonical_role_label(sender, school_scope)
        page.object_list = items
    except Exception:
        pass
    return render(request, "reports/my_notifications.html", {"page_obj": page})


@login_required(login_url="reports:login")
@require_http_methods(["GET"])
def my_circulars(request: HttpRequest) -> HttpResponse:
    """قائمة التعاميم للمستخدم (التي تتطلب توقيعاً)."""
    if NotificationRecipient is None:
        return render(request, "reports/my_circulars.html", {"page_obj": Paginator([], 12).get_page(1)})

    active_school = _get_active_school(request)

    try:
        qs = (
            NotificationRecipient.objects
            .select_related("notification")
            .filter(teacher=request.user)
            .order_by("-created_at", "-id")
        )
    except Exception:
        logger.exception("my_circulars: failed to build base queryset")
        messages.error(request, "تعذر تحميل التعاميم حالياً. سيتم تسجيل المشكلة تلقائياً.")
        return render(request, "reports/my_circulars.html", {"page_obj": Paginator([], 12).get_page(1)})

    # فصل: هذه الصفحة للتعاميم فقط
    try:
        if Notification is not None and hasattr(Notification, "requires_signature"):
            qs = qs.filter(notification__requires_signature=True)
    except Exception:
        pass

    # عزل حسب المدرسة النشطة (مع السماح بإشعارات عامة school=NULL)
    try:
        if active_school is not None and Notification is not None and hasattr(Notification, "school"):
            qs = qs.filter(Q(notification__school=active_school) | Q(notification__school__isnull=True))
    except Exception:
        pass

    # إخفاء المنتهية بحسب الحقول المتاحة
    now = timezone.now()
    try:
        if Notification is not None and hasattr(Notification, "expires_at"):
            qs = qs.exclude(notification__expires_at__lt=now)
        elif Notification is not None and hasattr(Notification, "ends_at"):
            qs = qs.exclude(notification__ends_at__lt=now)
    except Exception:
        pass

    try:
        page = Paginator(qs, 12).get_page(request.GET.get("page") or 1)
    except Exception:
        logger.exception("my_circulars: failed to paginate")
        messages.error(request, "تعذر تحميل التعاميم حالياً. سيتم تسجيل المشكلة تلقائياً.")
        return render(request, "reports/my_circulars.html", {"page_obj": Paginator([], 12).get_page(1)})

    # مهم: QuerySet داخل Page قد يبقى كسولاً، وقد يحدث الخطأ أثناء عرض القالب.
    # هنا نجبر التقييم داخل الـ view حتى نلتقط أخطاء قاعدة البيانات (مثل نقص migrations) ونمنع 500.
    try:
        page.object_list = list(page.object_list)
    except Exception:
        logger.exception("my_circulars: failed to evaluate page object_list")
        messages.error(request, "تعذر تحميل التعاميم حالياً. سيتم تسجيل المشكلة تلقائياً.")
        return render(request, "reports/my_circulars.html", {"page_obj": Paginator([], 12).get_page(1)})

    # عند فتح تبويب "تعاميمي" غالباً يتوقع المستخدم أن تصبح العناصر المعروضة كمقروءة.
    try:
        items = list(page.object_list)
        unread_ids = [x.pk for x in items if hasattr(x, "is_read") and not bool(getattr(x, "is_read", False))]
        if unread_ids:
            now = timezone.now()
            fields = {f.name for f in NotificationRecipient._meta.get_fields()}
            upd: dict = {}
            if "is_read" in fields:
                upd["is_read"] = True
            if "read_at" in fields:
                upd["read_at"] = now
            if upd:
                NotificationRecipient.objects.filter(pk__in=unread_ids, teacher=request.user).update(**upd)
                for x in items:
                    if x.pk in unread_ids:
                        if "is_read" in upd:
                            setattr(x, "is_read", True)
                        if "read_at" in upd:
                            setattr(x, "read_at", now)
            page.object_list = items
    except Exception:
        pass

    return render(request, "reports/my_circulars.html", {"page_obj": page})


@login_required(login_url="reports:login")
@require_http_methods(["GET"])
def my_notification_detail(request: HttpRequest, pk: int) -> HttpResponse:
    """Show a single notification (for the current user) in a dedicated page.

    pk here refers to NotificationRecipient.pk.
    """
    if NotificationRecipient is None:
        messages.error(request, "نموذج الإشعار غير متاح.")
        return redirect("reports:my_notifications")

    try:
        r = get_object_or_404(
            NotificationRecipient.objects.select_related(
                "notification",
                "notification__created_by",
                "notification__created_by__role",
            ),
            pk=pk,
            teacher=request.user,
        )
    except Exception:
        logger.exception("my_notification_detail: failed to load recipient row", extra={"pk": pk})
        messages.error(request, "تعذر فتح التعميم/الإشعار حالياً. سيتم تسجيل المشكلة تلقائياً.")
        return redirect("reports:my_circulars")

    n = getattr(r, "notification", None)
    if n is None:
        messages.error(request, "تعذّر العثور على الإشعار.")
        return redirect("reports:my_notifications")

    is_circular = bool(getattr(n, "requires_signature", False))

    # منع الخلط 100%: إذا كان الرابط من تبويب خاطئ نعيد توجيهه للرابط الصحيح
    try:
        url_name = getattr(getattr(request, "resolver_match", None), "url_name", "") or ""
        if is_circular and url_name == "my_notification_detail":
            return redirect("reports:my_circular_detail", pk=r.pk)
        if (not is_circular) and url_name == "my_circular_detail":
            return redirect("reports:my_notification_detail", pk=r.pk)
    except Exception:
        pass

    body = (
        getattr(n, "message", None)
        or getattr(n, "body", None)
        or getattr(n, "content", None)
        or getattr(n, "text", None)
        or getattr(n, "details", None)
        or ""
    )

    # اسم/دور المرسل (موحّد)
    try:
        sender = getattr(n, "created_by", None)
        school_scope = getattr(n, "school", None) or _get_active_school(request)
        sender_name = _canonical_sender_name(sender)
        sender_role_label = _canonical_role_label(sender, school_scope)
    except Exception:
        sender_name = "الإدارة"
        sender_role_label = ""

    # Mark as read on open (best-effort, supports different schemas)
    try:
        updated_fields: list[str] = []
        if hasattr(r, "is_read") and not bool(getattr(r, "is_read", False)):
            setattr(r, "is_read", True)
            updated_fields.append("is_read")
        if hasattr(r, "read_at") and getattr(r, "read_at", None) is None:
            setattr(r, "read_at", timezone.now())
            updated_fields.append("read_at")
        if updated_fields:
            try:
                r.save(update_fields=updated_fields)
            except Exception:
                r.save()
    except Exception:
        pass

    return render(
        request,
        "reports/my_circular_detail.html" if is_circular else "reports/my_notification_detail.html",
        {
            "r": r,
            "n": n,
            "body": body,
            "sender_name": sender_name,
            "sender_role_label": sender_role_label,
        },
    )

@login_required(login_url="reports:login")
@user_passes_test(_is_staff_or_officer, login_url="reports:login")
@require_http_methods(["GET"])
def notifications_sent(request: HttpRequest, mode: str = "notification") -> HttpResponse:
    mode = (mode or "notification").strip().lower()
    if mode not in {"notification", "circular"}:
        mode = "notification"
    is_circular = mode == "circular"

    is_platform = bool(is_platform_admin(request.user)) and not bool(getattr(request.user, "is_superuser", False))

    if is_circular:
        if not request.user.is_superuser and not is_platform and not _is_manager_in_school(request.user, _get_active_school(request)):
            messages.error(request, "التعاميم متاحة لمدير المدرسة فقط.")
            return redirect("reports:home")

    if Notification is None:
        return render(
            request,
            "reports/circulars_sent.html" if is_circular else "reports/notifications_sent.html",
            {
                "page_obj": Paginator([], 20).get_page(1),
                "stats": {},
                "mode": mode,
                "title": "التعاميم المرسلة" if is_circular else "الإشعارات المرسلة",
            },
        )

    active_school = _get_active_school(request)
    if not request.user.is_superuser and (not is_platform) and active_school is None:
        messages.error(request, "يرجى اختيار المدرسة أولاً.")
        return redirect("reports:home")

    qs = Notification.objects.all().order_by("-created_at", "-id")

    # صفحة "المرسلة" تعرض فقط الإشعارات التي أرسلها مستخدم فعلياً.
    # إشعارات النظام (created_by=NULL) مثل التعليقات الخاصة والتنبيهات الآلية لا تظهر هنا.
    try:
        if hasattr(Notification, "created_by"):
            qs = qs.filter(created_by__isnull=False)
    except Exception:
        pass

    # فصل التعاميم عن الإشعارات
    try:
        if hasattr(Notification, "requires_signature"):
            qs = qs.filter(requires_signature=True) if is_circular else qs.filter(requires_signature=False)
    except Exception:
        pass

    # غير السوبر: لا يرى إلا إشعارات المدرسة النشطة (لا إشعارات عامة)
    try:
        if (not request.user.is_superuser) and (not is_platform) and hasattr(Notification, "school"):
            qs = qs.filter(school=active_school)
    except Exception:
        pass

    # المشرف العام: يرى فقط ما قام بإرساله، وبحد نطاقه إن كانت المدرسة محددة
    if is_platform:
        qs = qs.filter(created_by=request.user)
        try:
            if hasattr(Notification, "school"):
                qs = qs.filter(Q(school__isnull=True) | Q(school__in=platform_allowed_schools_qs(request.user)))
        except Exception:
            pass

    # ✅ صفحة "المرسلة" يجب أن تُظهر ما أرسله المستخدم الحالي فقط
    # (مدير المدرسة كان يرى سابقًا جميع إشعارات المدرسة بما فيها إشعارات المشرفين)
    if not request.user.is_superuser:
        qs = qs.filter(created_by=request.user)

    qs = qs.select_related("created_by")
    page = Paginator(qs, 20).get_page(request.GET.get("page") or 1)

    notif_ids = [n.id for n in page.object_list]
    stats: dict[int, dict] = {}

    # حساب read/total بمرونة على NotificationRecipient
    if NotificationRecipient is not None and notif_ids:
        notif_fk_name = None
        try:
            for f in NotificationRecipient._meta.get_fields():
                if getattr(getattr(f, "remote_field", None), "model", None) is Notification:
                    notif_fk_name = f.name
                    break
        except Exception:
            notif_fk_name = None

        if notif_fk_name:
            fields = {f.name for f in NotificationRecipient._meta.get_fields()}
            if "is_read" in fields:
                read_filter = Q(is_read=True)
            elif "read_at" in fields:
                read_filter = Q(read_at__isnull=False)
            elif "seen_at" in fields:
                read_filter = Q(seen_at__isnull=False)
            elif "status" in fields:
                read_filter = Q(status__in=["read", "seen", "opened", "done"])
            else:
                read_filter = Q(pk__in=[])

            fields = {f.name for f in NotificationRecipient._meta.get_fields()}
            signed_filter = None
            if "is_signed" in fields:
                signed_filter = Q(is_signed=True)
            elif "signed_at" in fields:
                signed_filter = Q(signed_at__isnull=False)

            ann = {
                "total": Count("id"),
                "read": Count("id", filter=read_filter),
            }
            if signed_filter is not None:
                ann["signed"] = Count("id", filter=signed_filter)

            rc = (
                NotificationRecipient.objects
                .filter(**{f"{notif_fk_name}_id__in": notif_ids})
                .values(f"{notif_fk_name}_id")
                .annotate(**ann)
            )
            for row in rc:
                stats[row[f"{notif_fk_name}_id"]] = {
                    "total": row.get("total", 0),
                    "read": row.get("read", 0),
                    "signed": row.get("signed", 0),
                }

    # أسماء مستلمين مختصرة
    rec_names_map: dict[int, list[str]] = {i: [] for i in notif_ids}

    def _name_of(person) -> str:
        return (getattr(person, "name", None) or
                getattr(person, "phone", None) or
                getattr(person, "username", None) or
                getattr(person, "national_id", None) or
                str(person))

    for n in page.object_list:
        names_set = set()
        try:
            rel = getattr(n, "recipients", None)
            if rel is not None:
                for t in rel.all()[:12]:
                    if t:
                        nm = _name_of(t)
                        if nm not in names_set:
                            names_set.add(nm)
        except Exception:
            pass
        rec_names_map[n.id] = list(names_set)

    remaining_ids = [nid for nid, arr in rec_names_map.items() if len(arr) < 5]
    if remaining_ids and NotificationRecipient is not None:
        notif_fk_name = None
        try:
            for f in NotificationRecipient._meta.get_fields():
                if getattr(getattr(f, "remote_field", None), "model", None) is Notification:
                    notif_fk_name = f.name
                    break
        except Exception:
            pass

        if notif_fk_name:
            thr_qs = NotificationRecipient.objects.filter(**{f"{notif_fk_name}_id__in": remaining_ids})
            for r in thr_qs:
                nid = getattr(r, f"{notif_fk_name}_id", None)
                if not nid:
                    continue
                person = (getattr(r, "teacher", None) or
                          getattr(r, "user", None) or
                          getattr(r, "recipient", None))
                if person:
                    nm = _name_of(person)
                    arr = rec_names_map.get(nid, [])
                    if nm and nm not in arr and len(arr) < 12:
                        arr.append(nm)
                        rec_names_map[nid] = arr

    for n in page.object_list:
        n.rec_names = rec_names_map.get(n.id, [])

    return render(
        request,
        "reports/circulars_sent.html" if is_circular else "reports/notifications_sent.html",
        {
            "page_obj": page,
            "stats": stats,
            "mode": mode,
            "title": "التعاميم المرسلة" if is_circular else "الإشعارات المرسلة",
        },
    )

# تعليم الإشعار كمقروء (حسب Recipient pk)
@login_required(login_url="reports:login")
@require_http_methods(["POST"])
def notification_mark_read(request: HttpRequest, pk: int) -> HttpResponse:
    if NotificationRecipient is None:
        return redirect(request.POST.get("next") or "reports:my_notifications")
    item = get_object_or_404(NotificationRecipient, pk=pk, teacher=request.user)
    if not getattr(item, "is_read", False):
        if hasattr(item, "is_read"):
            item.is_read = True
        if hasattr(item, "read_at"):
            item.read_at = timezone.now()
        try:
            if hasattr(item, "is_read") and hasattr(item, "read_at"):
                item.save(update_fields=["is_read", "read_at"])
            else:
                item.save()
        except Exception:
            item.save()
    return redirect(request.POST.get("next") or "reports:my_notifications")

# تحديد الكل كمقروء
@login_required(login_url="reports:login")
@require_http_methods(["POST"])
def notifications_mark_all_read(request: HttpRequest) -> HttpResponse:
    if NotificationRecipient is None:
        return redirect(request.POST.get("next") or "reports:my_notifications")
    qs = NotificationRecipient.objects.filter(teacher=request.user)

    # فصل: هذا الإجراء خاص بالإشعارات فقط (يستبعد التعاميم)
    try:
        if Notification is not None and hasattr(Notification, "requires_signature"):
            qs = qs.filter(notification__requires_signature=False)
    except Exception:
        pass
    try:
        if "is_read" in {f.name for f in NotificationRecipient._meta.get_fields()}:
            qs = qs.filter(is_read=False)
            qs.update(is_read=True, read_at=timezone.now() if hasattr(NotificationRecipient, "read_at") else None)
        elif "read_at" in {f.name for f in NotificationRecipient._meta.get_fields()}:
            qs = qs.filter(read_at__isnull=True)
            qs.update(read_at=timezone.now())
        else:
            pass
    except Exception:
        for x in qs:
            try:
                if hasattr(x, "is_read"):
                    x.is_read = True
                if hasattr(x, "read_at"):
                    x.read_at = timezone.now()
                x.save()
            except Exception:
                continue
    messages.success(request, "تم تحديد جميع الإشعارات كمقروءة.")

    # Bulk update won't trigger signals; ask clients to resync once.
    try:
        from .realtime_notifications import push_force_resync

        push_force_resync(teacher_id=int(getattr(request.user, "id", 0) or 0))
    except Exception:
        pass

    return redirect(request.POST.get("next") or "reports:my_notifications")


@login_required(login_url="reports:login")
@require_http_methods(["POST"])
def circulars_mark_all_read(request: HttpRequest) -> HttpResponse:
    if NotificationRecipient is None:
        return redirect(request.POST.get("next") or "reports:my_circulars")

    qs = NotificationRecipient.objects.filter(teacher=request.user)
    try:
        if Notification is not None and hasattr(Notification, "requires_signature"):
            qs = qs.filter(notification__requires_signature=True)
    except Exception:
        pass

    try:
        if "is_read" in {f.name for f in NotificationRecipient._meta.get_fields()}:
            qs = qs.filter(is_read=False)
            qs.update(is_read=True, read_at=timezone.now() if hasattr(NotificationRecipient, "read_at") else None)
        elif "read_at" in {f.name for f in NotificationRecipient._meta.get_fields()}:
            qs = qs.filter(read_at__isnull=True)
            qs.update(read_at=timezone.now())
    except Exception:
        for x in qs:
            try:
                if hasattr(x, "is_read"):
                    x.is_read = True
                if hasattr(x, "read_at"):
                    x.read_at = timezone.now()
                x.save()
            except Exception:
                continue

    messages.success(request, "تم تحديد جميع التعاميم كمقروءة.")
    return redirect(request.POST.get("next") or "reports:my_circulars")

# تعليم الإشعار كمقروء (حسب رقم الإشعار نفسه لا الـRecipient)
@login_required(login_url="reports:login")
@require_http_methods(["POST"])
def notification_mark_read_by_notification(request: HttpRequest, pk: int) -> HttpResponse:
    if NotificationRecipient is None:
        return JsonResponse({"ok": False}, status=400)
    try:
        item = NotificationRecipient.objects.filter(
            notification_id=pk, teacher=request.user
        ).first()
        if item:
            if hasattr(item, "is_read") and not item.is_read:
                item.is_read = True
            if hasattr(item, "read_at") and getattr(item, "read_at", None) is None:
                item.read_at = timezone.now()
            try:
                if hasattr(item, "is_read") and hasattr(item, "read_at"):
                    item.save(update_fields=["is_read", "read_at"])
                else:
                    item.save()
            except Exception:
                item.save()
        return JsonResponse({"ok": True})
    except Exception:
        return JsonResponse({"ok": False}, status=400)

# إبقاء المسار القديم للتوافق الخلفي: تحويل إلى صفحة الإنشاء
@login_required(login_url="reports:login")
@user_passes_test(_is_staff, login_url="reports:login")
def send_notification(request: HttpRequest) -> HttpResponse:
    return redirect("reports:notifications_create")


# =========================
# إدارة الاشتراكات والمالية
# =========================

def subscription_expired(request):
    """صفحة تظهر عند انتهاء الاشتراك.

    نُمرّر معلومات المدرسة + تاريخ انتهاء الاشتراك إن توفّرت لعرضها في الرسالة.
    """
    school = None
    subscription = None
    is_manager = False

    try:
        user = getattr(request, "user", None)
        if user is not None and getattr(user, "is_authenticated", False):
            active_school = _get_active_school(request)
            memberships = (
                SchoolMembership.objects.filter(teacher=user, is_active=True)
                .select_related("school__subscription__plan", "school")
            )
            membership = None
            if active_school:
                membership = memberships.filter(school=active_school).first()
            if membership is None:
                membership = memberships.first()

            if membership is not None:
                school = membership.school
                is_manager = membership.role_type == SchoolMembership.RoleType.MANAGER
                subscription = getattr(school, "subscription", None)

                # لو أصبحت المدرسة النشطة اشتراكها ساري (بعد التبديل مثلاً)،
                # لا معنى لإظهار صفحة الانتهاء.
                try:
                    if subscription is not None and not bool(subscription.is_expired):
                        if getattr(request.user, "is_superuser", False):
                            return redirect("reports:platform_admin_dashboard")
                        if _is_staff(request.user):
                            return redirect("reports:admin_dashboard")
                        return redirect("reports:home")
                except Exception:
                    pass
    except Exception:
        # لا نكسر الصفحة لو كانت هناك مشكلة في العضويات
        school = None
        subscription = None
        is_manager = False

    return render(
        request,
        "reports/subscription_expired.html",
        {"school": school, "subscription": subscription, "is_manager": is_manager},
    )

@login_required(login_url="reports:login")
def my_subscription(request):
    """صفحة عرض تفاصيل الاشتراك لمدير المدرسة"""
    active_school = _get_active_school(request)
    
    # جلب جميع عضويات الإدارة للمستخدم
    memberships = SchoolMembership.objects.filter(
        teacher=request.user, 
        role_type=SchoolMembership.RoleType.MANAGER,
        is_active=True
    ).select_related('school__subscription__plan')
    
    membership = None
    # محاولة استخدام المدرسة النشطة إذا كان المستخدم مديراً فيها
    if active_school:
        membership = memberships.filter(school=active_school).first()
    
    # إذا لم توجد مدرسة نشطة أو المستخدم ليس مديراً فيها، نأخذ أول مدرسة يديرها
    if not membership:
        membership = memberships.first()
    
    if not membership:
        messages.error(request, "عفواً، هذه الصفحة مخصصة لمدير المدرسة فقط.")
        return redirect('reports:home')

    # ملاحظة: reverse OneToOne (school.subscription) يرفع DoesNotExist إن لم يوجد سجل
    subscription = (
        SchoolSubscription.objects.filter(school=membership.school)
        .select_related("plan")
        .first()
    )
    
    # تظهر آخر 4 عمليات فقط
    payments = Payment.objects.filter(school=membership.school).order_by('-created_at')[:4]
    
    context = {
        "subscription": subscription,
        "school": membership.school,
        # ✅ أظهر كل الخطط (حتى لو غير نشطة) حتى لا تبدو "مفقودة".
        # سيتم تعطيل غير النشطة في القالب.
        "plans": SubscriptionPlan.objects.all().order_by("days_duration", "price"),
        "payments": payments,
    }
    return render(request, 'reports/my_subscription.html', context)

def subscription_history(request):
    """عرض سجل العمليات الكامل للاشتراكات"""
    active_school = _get_active_school(request)
    
    # جلب جميع عضويات الإدارة للمستخدم
    memberships = SchoolMembership.objects.filter(
        teacher=request.user, 
        role_type=SchoolMembership.RoleType.MANAGER,
        is_active=True
    ).select_related('school')
    
    membership = None
    # محاولة استخدام المدرسة النشطة إذا كان المستخدم مديراً فيها
    if active_school:
        membership = memberships.filter(school=active_school).first()
    
    # إذا لم توجد مدرسة نشطة أو المستخدم ليس مديراً فيها، نأخذ أول مدرسة يديرها
    if not membership:
        membership = memberships.first()
    
    if not membership:
        messages.error(request, "عفواً، هذه الصفحة مخصصة لمدير المدرسة فقط.")
        return redirect('reports:home')

    # جلب كامل العمليات
    payments = Payment.objects.filter(school=membership.school).order_by('-created_at')
    
    context = {
        "school": membership.school,
        "payments": payments,
    }
    return render(request, 'reports/subscription_history.html', context)

@login_required(login_url="reports:login")
def payment_create(request):
    """صفحة رفع إيصال الدفع"""
    active_school = _get_active_school(request)
    
    memberships = SchoolMembership.objects.filter(
        teacher=request.user, 
        role_type=SchoolMembership.RoleType.MANAGER,
        is_active=True
    )
    
    membership = None
    if active_school:
        membership = memberships.filter(school=active_school).first()
        
    if not membership:
        membership = memberships.first()
    
    if not membership:
        messages.error(request, "عفواً، هذه الصفحة مخصصة لمدير المدرسة فقط.")
        return redirect('reports:home')

    subscription = (
        SchoolSubscription.objects.filter(school=membership.school)
        .select_related("plan")
        .first()
    )

    if request.method == 'POST':
        receipt = request.FILES.get('receipt_image')
        notes = request.POST.get('notes')
        plan_id = request.POST.get('plan_id')

        requested_plan = None
        
        # 1. محاولة أخذ الباقة من اختيار المستخدم
        if plan_id:
            try:
                requested_plan = SubscriptionPlan.objects.get(pk=plan_id)
            except SubscriptionPlan.DoesNotExist:
                pass
        
        # 2. إذا لم يختر، نأخذ الباقة الحالية
        if not requested_plan and subscription:
            requested_plan = subscription.plan

        # التحقق النهائي
        if not requested_plan:
            messages.error(request, "يرجى اختيار باقة للاشتراك/التجديد.")
            return redirect('reports:my_subscription')

        amount = getattr(requested_plan, "price", None)
        try:
            if amount is None or float(amount) <= 0:
                messages.error(request, "لا يمكن إنشاء طلب دفع لأن الباقة المختارة مجانية/غير صالحة.")
                return redirect('reports:my_subscription')
        except Exception:
            pass

        if not receipt:
            messages.error(request, "يرجى إرفاق صورة الإيصال.")
            return redirect('reports:my_subscription')

        Payment.objects.create(
            school=membership.school,
            subscription=subscription,
            requested_plan=requested_plan,
            amount=amount,
            receipt_image=receipt,
            notes=notes,
            created_by=request.user
        )
        
        msg = f"""
        <div style="text-align: center; line-height: 1.6;">
            <p style="margin-bottom: 0.5rem; font-weight: 700; font-size: 1.1rem;">تم استلام طلبك بنجاح ✅</p>
            <p style="margin-bottom: 0.5rem;">جاري مراجعة الإيصال والتحقق منه، وسيتم تفعيل الباقة التالية فور الاعتماد:</p>
            <div style="background: rgba(255,255,255,0.2); border: 1px solid rgba(255,255,255,0.3); padding: 0.75rem 1rem; border-radius: 12px; display: inline-block; margin-top: 0.5rem; color: #fff;">
                <div style="font-weight: 800; font-size: 1.1rem; margin-bottom: 0.25rem;">{requested_plan.name}</div>
                <div style="font-size: 0.9rem;">
                    السعر: {requested_plan.price} ريال &bull; المدة: {requested_plan.days_duration} يوم
                </div>
            </div>
        </div>
        """
        messages.success(request, mark_safe(msg))
        return redirect('reports:my_subscription')
            
    return redirect('reports:my_subscription')


@login_required(login_url="reports:login")
@user_passes_test(lambda u: getattr(u, "is_superuser", False), login_url="reports:login")
def platform_plan_form(request: HttpRequest, pk: Optional[int] = None) -> HttpResponse:
    """إضافة أو تعديل خطة اشتراك"""
    plan = None
    if pk:
        plan = get_object_or_404(SubscriptionPlan, pk=pk)
    
    if request.method == "POST":
        form = SubscriptionPlanForm(request.POST, instance=plan)
        if form.is_valid():
            form.save()
            messages.success(request, "تم حفظ الخطة بنجاح.")
            return redirect("reports:platform_plans_list")
        else:
            messages.error(request, "يرجى تصحيح الأخطاء أدناه.")
    else:
        form = SubscriptionPlanForm(instance=plan)
    
    return render(request, "reports/platform_plan_form.html", {"form": form, "plan": plan})


@login_required(login_url="reports:login")
@user_passes_test(lambda u: getattr(u, "is_superuser", False), login_url="reports:login")
@require_http_methods(["POST"])
def platform_plan_delete(request: HttpRequest, pk: int) -> HttpResponse:
    plan = get_object_or_404(SubscriptionPlan, pk=pk)

    try:
        plan_name = plan.name
        plan.delete()
        messages.success(request, f"تم حذف الخطة: {plan_name}.")
    except ProtectedError:
        messages.error(request, "لا يمكن حذف هذه الخطة لأنها مرتبطة باشتراكات مدارس حالياً.")
    except Exception:
        logger.exception("platform_plan_delete failed")
        messages.error(request, "حدث خطأ غير متوقع أثناء حذف الخطة.")

    next_url = _safe_next_url(request.POST.get("next") or request.GET.get("next"))
    return redirect(next_url or "reports:platform_plans_list")


@login_required(login_url="reports:login")
@user_passes_test(lambda u: getattr(u, "is_superuser", False), login_url="reports:login")
def platform_subscription_form(request: HttpRequest, pk: Optional[int] = None) -> HttpResponse:
    """إضافة اشتراك مدرسة (تم إلغاء تعديل الباقة/الاشتراك نهائياً)."""
    subscription = None
    # ✅ تم إلغاء التعديل نهائياً: أي محاولة لفتح رابط قديم للتعديل تُرفض.
    if pk is not None:
        raise Http404
    
    if request.method == "POST":
        # ✅ إذا كانت المدرسة لديها اشتراك سابق (ملغي/منتهي) فلا ننشئ سجل جديد (OneToOne)
        # بل نجدد/نفعّل الاشتراك الموجود لتفادي خطأ "المدرسة موجودة مسبقاً".
        school_id_raw = (request.POST.get("school") or "").strip()
        try:
            school_id = int(school_id_raw)
        except Exception:
            school_id = None

        if school_id is not None:
            existing = (
                SchoolSubscription.objects.filter(school_id=school_id)
                .select_related("school", "plan")
                .first()
            )
            if existing is not None:
                # إن كان الاشتراك ملغي/منتهي: نجدد/نفعّل نفس السجل (OneToOne)
                # لكن نسمح بتغيير الباقة حسب اختيار الإدارة (إن لزم).
                if bool(getattr(existing, "is_cancelled", False)) or bool(getattr(existing, "is_expired", False)):
                    from datetime import timedelta

                    today = timezone.localdate()
                    prev_plan_id = getattr(existing, "plan_id", None)
                    form = SchoolSubscriptionForm(request.POST, instance=existing, allow_plan_change=True)
                    if form.is_valid():
                        subscription_obj = form.save(commit=False)

                        # عند التجديد: فعّل وامسح بيانات الإلغاء
                        subscription_obj.is_active = True
                        if getattr(subscription_obj, "canceled_at", None) is not None:
                            subscription_obj.canceled_at = None
                        if (getattr(subscription_obj, "cancel_reason", "") or "").strip():
                            subscription_obj.cancel_reason = ""

                        # إذا لم تتغير الباقة، فاعتبرها تجديداً أيضاً واضبط التواريخ لليوم
                        # (لأن منطق model.save يعيد الحساب فقط عند تغيير plan).
                        if getattr(subscription_obj, "plan_id", None) == prev_plan_id:
                            days = int(getattr(getattr(subscription_obj, "plan", None), "days_duration", 0) or 0)
                            subscription_obj.start_date = today
                            subscription_obj.end_date = today if days <= 0 else today + timedelta(days=days - 1)

                        subscription_obj.save()

                        # تحصين مالي: أي دفعات pending قديمة لا يجب أن تبقى عالقة بعد التجديد.
                        try:
                            Payment.objects.filter(
                                subscription=subscription_obj,
                                status=Payment.Status.PENDING,
                                created_at__date__lt=subscription_obj.start_date,
                            ).update(
                                status=Payment.Status.CANCELLED,
                                notes="تم إلغاء هذه العملية تلقائياً بسبب تجديد/تغيير الاشتراك.",
                            )
                        except Exception:
                            pass

                        _record_subscription_payment_if_missing(
                            subscription=subscription_obj,
                            actor=request.user,
                            note="تم تجديد الاشتراك (مع تحديث الباقة عند الحاجة) وتسجيل الدفعة بواسطة إدارة المنصة.",
                            force=True,
                        )

                        messages.success(
                            request,
                            f"تم تفعيل/تجديد اشتراك مدرسة {subscription_obj.school.name} حتى {subscription_obj.end_date:%Y-%m-%d}.",
                        )
                        return redirect("reports:platform_subscriptions_list")
                    else:
                        messages.error(request, "يرجى تصحيح الأخطاء أدناه.")
                        return render(request, "reports/platform_subscription_add.html", {"form": form})

                messages.info(
                    request,
                    "هذه المدرسة لديها اشتراك قائم بالفعل. استخدم زر (تجديد) من قائمة الاشتراكات.",
                )
                return redirect("reports:platform_subscriptions_list")

        was_existing = bool(subscription and getattr(subscription, "pk", None))
        prev_is_active = bool(getattr(subscription, "is_active", False)) if subscription else False
        form = SchoolSubscriptionForm(request.POST, instance=subscription)
        if form.is_valid():
            subscription_obj = form.save()

            # ✅ المالية:
            # - عند إنشاء اشتراك جديد من لوحة المنصة: نسجّل دفعة (approved) لتظهر في المالية.
            # - عند تعديل اشتراك موجود: لا نسجّل دفعة إلا إذا كان غير نشط ثم تم تفعيله.
            created_payment = False
            try:
                became_active = (not prev_is_active) and bool(getattr(subscription_obj, "is_active", False))
                if (not was_existing) or became_active:
                    created_payment = _record_subscription_payment_if_missing(
                        subscription=subscription_obj,
                        actor=request.user,
                        note="تم تسجيل الدفعة يدويًا بواسطة إدارة المنصة.",
                        force=False,
                    )
            except Exception:
                created_payment = False

            if created_payment:
                messages.success(request, "تم حفظ الاشتراك وتسجيل عملية الدفع بنجاح.")
            else:
                messages.success(request, "تم حفظ الاشتراك بنجاح.")
            return redirect("reports:platform_subscriptions_list")
        else:
            messages.error(request, "يرجى تصحيح الأخطاء أدناه.")
    else:
        form = SchoolSubscriptionForm(instance=subscription)

    return render(request, "reports/platform_subscription_add.html", {"form": form})


@login_required(login_url="reports:login")
@user_passes_test(lambda u: getattr(u, "is_superuser", False), login_url="reports:login")
@require_http_methods(["POST"])
def platform_subscription_renew(request: HttpRequest, pk: int) -> HttpResponse:
    """تجديد اشتراك مدرسة مباشرةً من اليوم (ميلادي).

    - يضبط start_date = اليوم
    - يضبط end_date = اليوم + (plan.days_duration - 1)
    - يفعّل is_active=True

    هذا المسار مخصص للمشرف العام فقط لتسهيل التجديد من صفحة الاشتراكات.
    """
    subscription = get_object_or_404(SchoolSubscription.objects.select_related("plan", "school"), pk=pk)

    from datetime import timedelta

    today = timezone.localdate()
    subscription.start_date = today
    days = int(getattr(subscription.plan, "days_duration", 0) or 0)
    if days <= 0:
        subscription.end_date = today
    else:
        subscription.end_date = today + timedelta(days=days - 1)

    subscription.is_active = True
    # عند التجديد: امسح بيانات الإلغاء
    if getattr(subscription, "canceled_at", None) is not None:
        subscription.canceled_at = None
    if getattr(subscription, "cancel_reason", ""):
        subscription.cancel_reason = ""
    subscription.save()

    created_payment = _record_subscription_payment_if_missing(
        subscription=subscription,
        actor=request.user,
        note="تم تجديد الاشتراك وتسجيل الدفعة يدويًا بواسطة إدارة المنصة.",
        force=True,
    )
    if created_payment:
        messages.success(
            request,
            f"تم تجديد اشتراك مدرسة {subscription.school.name} حتى {subscription.end_date:%Y-%m-%d}، وتم تسجيل عملية الدفع.",
        )
    else:
        messages.success(request, f"تم تجديد اشتراك مدرسة {subscription.school.name} حتى {subscription.end_date:%Y-%m-%d}.")

    next_url = _safe_next_url(request.POST.get("next") or request.GET.get("next"))
    return redirect(next_url or "reports:platform_subscriptions_list")


@login_required(login_url="reports:login")
@user_passes_test(lambda u: getattr(u, "is_superuser", False), login_url="reports:login")
@require_http_methods(["POST"])
def platform_subscription_record_payment(request: HttpRequest, pk: int) -> HttpResponse:
    """تسجيل دفعة يدوية لاشتراك موجود بدون تغيير تواريخه."""
    subscription = get_object_or_404(SchoolSubscription.objects.select_related("plan", "school"), pk=pk)

    ok = _record_subscription_payment_if_missing(
        subscription=subscription,
        actor=request.user,
        note="تم تسجيل الدفعة يدويًا بواسطة إدارة المنصة.",
    )
    if ok:
        messages.success(request, "تم تسجيل عملية الدفع بنجاح.")
    else:
        messages.info(request, "لا يمكن تسجيل دفعة جديدة (يوجد دفع بالفعل أو الاشتراك غير نشط/مجاني).")

    next_url = _safe_next_url(request.POST.get("next") or request.GET.get("next"))
    return redirect(next_url or "reports:platform_subscriptions_list")


# ===== صفحات المحتوى (Footer Links) =====

def user_guide(request: HttpRequest) -> HttpResponse:
    """صفحة دليل الاستخدام"""
    return render(request, "reports/user_guide.html")


def faq(request: HttpRequest) -> HttpResponse:
    """صفحة الأسئلة الشائعة"""
    return render(request, "reports/faq.html")


def privacy_policy(request: HttpRequest) -> HttpResponse:
    """صفحة سياسة الخصوصية"""
    return render(request, "reports/privacy_policy.html")
